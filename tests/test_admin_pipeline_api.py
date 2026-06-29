from __future__ import annotations

import base64
import json
import tempfile
import threading
import time
import unittest
import urllib.request
from http.server import ThreadingHTTPServer
from pathlib import Path
from urllib.parse import quote

import data_admin_server

try:
    from tests.test_business_class_import import BUSINESS_HEADERS, assignment, base_payload, business_row, scheduled_lesson
    from tests.test_pipeline import ORIGINAL_DATA_DIR, write_csv, write_minimal_csv_source
except ModuleNotFoundError:
    from test_business_class_import import BUSINESS_HEADERS, assignment, base_payload, business_row, scheduled_lesson
    from test_pipeline import ORIGINAL_DATA_DIR, write_csv, write_minimal_csv_source


ORIGINAL_OUTPUT_DIR = data_admin_server.OUTPUT_DIR


def request_json(url: str, payload: dict | None = None) -> dict:
    data = None
    headers = {}
    if payload is not None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        headers["Content-Type"] = "application/json"
    request = urllib.request.Request(url, data=data, headers=headers)
    with urllib.request.urlopen(request, timeout=20) as response:
        return json.loads(response.read().decode("utf-8"))


def file_payload(path: Path) -> dict:
    return {
        "name": path.name,
        "content_base64": base64.b64encode(path.read_bytes()).decode("ascii"),
    }


class AdminPipelineApiTest(unittest.TestCase):
    def tearDown(self) -> None:
        data_admin_server.DATA_DIR = ORIGINAL_DATA_DIR
        data_admin_server.OUTPUT_DIR = ORIGINAL_OUTPUT_DIR
        data_admin_server.PIPELINE_JOBS.clear()

    def test_markdown_preview_renders_readable_safe_html(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "report.md"
            path.write_text(
                "\n".join(
                    [
                        "# 排课报告",
                        "",
                        "## 概览",
                        "- 覆盖缺口：0",
                        "- <script>alert(1)</script>",
                        "",
                        "| 班级 | 状态 |",
                        "| --- | --- |",
                        "| C1 | 通过 |",
                        "",
                        "```",
                        "<danger>",
                        "```",
                    ]
                ),
                encoding="utf-8",
            )

            body = data_admin_server.markdown_preview_html(path, '/outputs/report.md?next=<script>"')

        self.assertIn("<title>排课报告</title>", body)
        self.assertIn('<article class="markdown-body">', body)
        self.assertIn("<h2>概览</h2>", body)
        self.assertIn("<li>覆盖缺口：0</li>", body)
        self.assertIn("<table>", body)
        self.assertIn("<th>班级</th>", body)
        self.assertIn("<td>C1</td>", body)
        self.assertIn("&lt;danger&gt;", body)
        self.assertIn("&lt;script&gt;alert(1)&lt;/script&gt;", body)
        self.assertNotIn("<script>alert(1)</script>", body)
        self.assertIn('/outputs/report.md?next=&lt;script&gt;&quot;', body)

    def test_upload_run_job_and_output_file_access(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "incoming"
            data_admin_server.DATA_DIR = root / "data"
            data_admin_server.OUTPUT_DIR = root / "outputs"
            write_minimal_csv_source(source)

            server = ThreadingHTTPServer(("127.0.0.1", 0), data_admin_server.AdminHandler)
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            base_url = f"http://127.0.0.1:{server.server_port}"

            try:
                response = request_json(
                    f"{base_url}/api/pipeline/upload-run",
                    {"files": [file_payload(path) for path in sorted(source.glob("*.csv"))]},
                )
                job_id = response["job_id"]
                self.assertIn(response["status"], {"queued", "running"})

                job = {}
                for _ in range(80):
                    job = request_json(f"{base_url}/api/pipeline/jobs/{job_id}")
                    if job["status"] in {"succeeded", "failed"}:
                        break
                    time.sleep(0.1)

                self.assertEqual(job["status"], "succeeded", job.get("error", ""))
                self.assertTrue(job["report_url"])
                self.assertTrue(job["schedule_csv_url"])
                self.assertTrue(job["schedule_html_url"])

                with urllib.request.urlopen(f"{base_url}{job['schedule_html_url']}", timeout=20) as response:
                    body = response.read().decode("utf-8")
                self.assertIn("班级课表甘特图", body)
            finally:
                server.shutdown()
                thread.join(timeout=5)
                server.server_close()

    def test_markdown_outputs_have_utf8_content_type_and_preview(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_admin_server.DATA_DIR = root / "data"
            data_admin_server.OUTPUT_DIR = root / "outputs"
            data_admin_server.OUTPUT_DIR.mkdir(parents=True)
            report_path = data_admin_server.OUTPUT_DIR / "batch_schedule_maintenance_report.md"
            report_path.write_text("# 排课报告\n\n- 覆盖缺口：0\n", encoding="utf-8")

            server = ThreadingHTTPServer(("127.0.0.1", 0), data_admin_server.AdminHandler)
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            base_url = f"http://127.0.0.1:{server.server_port}"

            try:
                with urllib.request.urlopen(f"{base_url}/outputs/batch_schedule_maintenance_report.md", timeout=20) as response:
                    self.assertIn("text/markdown", response.headers["Content-Type"])
                    self.assertIn("charset=utf-8", response.headers["Content-Type"])
                    self.assertIn("排课报告", response.read().decode("utf-8"))

                with urllib.request.urlopen(f"{base_url}/preview/outputs/batch_schedule_maintenance_report.md", timeout=20) as response:
                    body = response.read().decode("utf-8")
                self.assertIn("text/html", response.headers["Content-Type"])
                self.assertIn("已按 UTF-8 读取原始 Markdown", body)
                self.assertIn("排课报告", body)
                self.assertIn("/outputs/batch_schedule_maintenance_report.md", body)

                head_request = urllib.request.Request(
                    f"{base_url}/preview/outputs/batch_schedule_maintenance_report.md",
                    method="HEAD",
                )
                with urllib.request.urlopen(head_request, timeout=20) as response:
                    self.assertIn("text/html", response.headers["Content-Type"])
                    self.assertEqual(b"", response.read())
            finally:
                server.shutdown()
                thread.join(timeout=5)
                server.server_close()

    def test_results_status_api_reports_downloadable_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_admin_server.DATA_DIR = root / "data"
            data_admin_server.OUTPUT_DIR = root / "outputs"
            data_admin_server.OUTPUT_DIR.mkdir(parents=True)
            template_path = data_admin_server.OUTPUT_DIR / "ai_scheduling_sop_20260625" / "AI排课基础数据模板.xlsx"
            template_path.parent.mkdir(parents=True)
            template_path.write_bytes(b"xlsx")
            (data_admin_server.OUTPUT_DIR / "batch_schedule_maintenance.html").write_text("<h1>课表</h1>", encoding="utf-8")
            (data_admin_server.OUTPUT_DIR / "batch_schedule_maintenance_report.md").write_text("# 排课报告", encoding="utf-8")

            server = ThreadingHTTPServer(("127.0.0.1", 0), data_admin_server.AdminHandler)
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            base_url = f"http://127.0.0.1:{server.server_port}"

            try:
                status = request_json(f"{base_url}/api/results/status")
                self.assertTrue(status["template"]["exists"])
                self.assertEqual(status["template"]["url"], "/outputs/ai_scheduling_sop_20260625/AI排课基础数据模板.xlsx")
                result_by_key = {item["key"]: item for item in status["results"]}
                self.assertTrue(result_by_key["schedule_html"]["exists"])
                self.assertTrue(result_by_key["schedule_report"]["exists"])
                self.assertEqual(result_by_key["schedule_report"]["preview_url"], "/preview/outputs/batch_schedule_maintenance_report.md")
                self.assertFalse(result_by_key["schedule_csv"]["exists"])
                self.assertEqual(result_by_key["schedule_csv"]["url"], "/outputs/batch_schedule_maintenance.csv")
            finally:
                server.shutdown()
                thread.join(timeout=5)
                server.server_close()

    def test_output_file_headers_support_unicode_filenames(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_admin_server.DATA_DIR = root / "data"
            data_admin_server.OUTPUT_DIR = root / "outputs"
            file_path = data_admin_server.OUTPUT_DIR / "ai_scheduling_sop_20260625" / "AI排课基础数据模板.xlsx"
            file_path.parent.mkdir(parents=True)
            file_path.write_bytes(b"xlsx")

            server = ThreadingHTTPServer(("127.0.0.1", 0), data_admin_server.AdminHandler)
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            base_url = f"http://127.0.0.1:{server.server_port}"
            encoded_path = "/outputs/ai_scheduling_sop_20260625/" + quote("AI排课基础数据模板.xlsx")

            try:
                head_request = urllib.request.Request(f"{base_url}{encoded_path}", method="HEAD")
                with urllib.request.urlopen(head_request, timeout=20) as response:
                    self.assertEqual(response.status, 200)
                    self.assertIn("filename*=UTF-8''AI", response.headers["Content-Disposition"])
                    self.assertEqual(b"", response.read())

                with urllib.request.urlopen(f"{base_url}{encoded_path}", timeout=20) as response:
                    self.assertEqual(response.status, 200)
                    self.assertEqual(response.read(), b"xlsx")
            finally:
                server.shutdown()
                thread.join(timeout=5)
                server.server_close()

    def test_product_and_class_import_export_api_round_trip(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_admin_server.DATA_DIR = root / "data"
            data_admin_server.OUTPUT_DIR = root / "outputs"

            server = ThreadingHTTPServer(("127.0.0.1", 0), data_admin_server.AdminHandler)
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            base_url = f"http://127.0.0.1:{server.server_port}"

            try:
                product_csv = (
                    "id,name,project,product_line,sub_product,product_system,season_window_ids,"
                    "applicable_stages,standard_capacity,capacity_type,subject_category,subject,course_nature,notes\n"
                    "P_IMPORT,测试产品,考研,考研无忧,无忧暑,常规体系,WINDOW_SUMMER,基础,30,班课,公共课,英语,正课,测试导入\n"
                )
                product_response = request_json(f"{base_url}/api/products/import", {"csv": product_csv})
                self.assertEqual(product_response["imported"], 1)
                self.assertEqual(product_response["total_products"], 1)

                class_csv = (
                    "id,name,product_id,selected_stages,project,product_line,sub_product,product_system,course_nature,"
                    "subject_category,subject,exam_season,exam_month,suite_code,standard_capacity,capacity_type,size,"
                    "start_date,start_period,first_lesson_date,first_lesson_period,end_date,end_period,"
                    "preferred_teaching_area_ids,preferred_room_ids,preferred_room_is_required,is_manual_schedule_locked,notes\n"
                    "C_IMPORT,测试班级,P_IMPORT,基础,考研,考研无忧,无忧暑,常规体系,正课,"
                    "公共课,英语,27考研,2026-12,9999,30,班课,20,"
                    "2026-07-01,AM,,,2026-08-31,PM,,,否,否,测试导入\n"
                )
                class_response = request_json(f"{base_url}/api/classes/import", {"csv": class_csv})
                self.assertEqual(class_response["imported"], 1)
                self.assertEqual(class_response["total_classes"], 1)

                with urllib.request.urlopen(f"{base_url}/api/products/download", timeout=20) as response:
                    product_download = response.read().decode("utf-8-sig")
                    self.assertIn("P_IMPORT", product_download)
                    self.assertIn("测试产品", product_download)
                with urllib.request.urlopen(f"{base_url}/api/classes/download", timeout=20) as response:
                    class_download = response.read().decode("utf-8-sig")
                    self.assertIn("C_IMPORT", class_download)
                    self.assertIn("测试班级", class_download)
            finally:
                server.shutdown()
                thread.join(timeout=5)
                server.server_close()

    def test_template_generation_and_preflight_api(self) -> None:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl is not installed")

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "incoming"
            data_admin_server.DATA_DIR = root / "data"
            data_admin_server.OUTPUT_DIR = root / "outputs"
            payload = base_payload()
            write_csv(source / "products.csv", ["id", "name", "project", "product_system", "subject", "subject_category", "standard_capacity"], payload["products"])
            write_csv(source / "product_courses.csv", ["product_id", "product_name", "subject_category", "subject", "stage", "course_module", "course_group", "total_hours", "block_hours", "teaching_area_ids"], payload["product_courses"])
            write_csv(source / "20260429班级查询导出.csv", BUSINESS_HEADERS, [business_row("C_REG")])
            write_csv(
                source / "已排课明细.csv",
                ["class_id", "class_name", "date", "start_time", "end_time", "duration_hours", "teacher_id", "teacher_name", "room_id", "business_product_id", "business_product_name", "subject", "stage", "course_module", "course_group"],
                [scheduled_lesson("C_REG", duration="2")],
            )

            server = ThreadingHTTPServer(("127.0.0.1", 0), data_admin_server.AdminHandler)
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            base_url = f"http://127.0.0.1:{server.server_port}"

            try:
                files = [file_payload(path) for path in sorted(source.glob("*.csv"))]
                template_response = request_json(f"{base_url}/api/templates/generate", {"files": files})
                self.assertTrue(template_response["xlsx_url"])
                self.assertTrue(template_response["zip_url"])
                workbook = load_workbook(data_admin_server.OUTPUT_DIR / template_response["xlsx_url"].removeprefix("/outputs/"), read_only=True)
                self.assertIn("business_product_mappings", workbook.sheetnames)
                self.assertIn("class_teacher_assignments", workbook.sheetnames)
                self.assertIn("缺口清单", workbook.sheetnames)
                self.assertNotIn("scope_exceptions", workbook.sheetnames)
                self.assertNotIn("business_product_map", workbook.sheetnames)
                self.assertNotIn("scheduling_scope_overrides", workbook.sheetnames)
                self.assertNotIn("merge_course_details", workbook.sheetnames)
                mapping_header = [
                    cell.value
                    for cell in next(workbook["business_product_mappings"].iter_rows(max_row=1))
                ]
                self.assertIn("local_product_id", mapping_header)
                self.assertNotIn("canonical_product_id", mapping_header)
                product_reference_header = [
                    cell.value
                    for cell in next(workbook["产品目录参考"].iter_rows(max_row=1))
                ]
                self.assertIn("local_product_id", product_reference_header)
                self.assertNotIn("canonical_product_id", product_reference_header)
                teacher_header = [
                    cell.value
                    for cell in next(workbook["class_teacher_assignments"].iter_rows(max_row=1))
                ]
                self.assertIn("class_schedule_mode", teacher_header)
                self.assertIn("actual_scheduled_class_id", teacher_header)
                self.assertNotIn("schedule_mode", teacher_header)
                self.assertNotIn("inherit_from_class_id", teacher_header)

                preflight_response = request_json(f"{base_url}/api/pipeline/preflight", {"files": files})
                self.assertFalse(preflight_response["passed"])
                self.assertIn("未命中产品映射", preflight_response["error"])
                self.assertTrue(preflight_response["report_url"])
                self.assertFalse((data_admin_server.DATA_DIR / "classes.json").exists())

                write_csv(
                    source / "business_product_mappings.csv",
                    ["business_product_id", "business_product_name", "local_product_id"],
                    [{"business_product_id": "100", "business_product_name": "考研英语无忧计划全年班", "local_product_id": "P_REG"}],
                )
                write_csv(source / "class_teacher_assignments.csv", ["class_id", "subject", "stage", "course_module", "course_group", "teacher_id", "teacher_name"], [assignment("C_REG")])
                files = [file_payload(path) for path in sorted(source.glob("*.csv"))]
                preflight_response = request_json(f"{base_url}/api/pipeline/preflight", {"files": files})
                self.assertTrue(preflight_response["passed"], preflight_response.get("error", ""))
            finally:
                server.shutdown()
                thread.join(timeout=5)
                server.server_close()

    def test_preflight_api_returns_missing_teacher_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "incoming"
            data_admin_server.DATA_DIR = root / "data"
            data_admin_server.OUTPUT_DIR = root / "outputs"
            write_minimal_csv_source(source)
            write_csv(
                source / "class_teacher_assignments.csv",
                ["class_id", "subject", "stage", "course_module", "course_group", "teacher_id", "teacher_name"],
                [],
            )

            server = ThreadingHTTPServer(("127.0.0.1", 0), data_admin_server.AdminHandler)
            thread = threading.Thread(target=server.serve_forever, daemon=True)
            thread.start()
            base_url = f"http://127.0.0.1:{server.server_port}"

            try:
                response = request_json(
                    f"{base_url}/api/pipeline/preflight",
                    {"files": [file_payload(path) for path in sorted(source.glob("*.csv"))]},
                )

                self.assertFalse(response["passed"])
                self.assertEqual({item["class_id"] for item in response["missing_teacher_requirements"]}, {"C1", "C2"})
                self.assertEqual({row["class_name"] for row in response["missing_teacher_rows"]}, {"英语1班", "英语2班"})
                self.assertTrue(response["generated_file_urls"])
            finally:
                server.shutdown()
                thread.join(timeout=5)
                server.server_close()


if __name__ == "__main__":
    unittest.main()
