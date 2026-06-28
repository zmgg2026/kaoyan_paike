from __future__ import annotations

import csv
import json
import tempfile
import unittest
from collections import Counter
from datetime import date
from pathlib import Path
from types import SimpleNamespace

import data_admin_server
import scheduler
from generate_time_slots import generate_time_slots, parse_weekdays
from run_scheduling_pipeline import (
    PipelineError,
    SOURCE_TABLES,
    TABLES,
    TABLE_FIELDNAMES,
    build_parser,
    load_source_tables,
    missing_teacher_rows_for_requirements,
    parse_missing_teacher_requirements,
    row_counts_for_tables,
    run_pipeline,
    run_preflight,
    table_name_for,
    write_missing_teacher_rows_template,
    write_missing_teacher_template,
)


ORIGINAL_DATA_DIR = data_admin_server.DATA_DIR


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_minimal_csv_source(source: Path) -> None:
    write_csv(
        source / "teaching_areas.csv",
        ["id", "name"],
        [{"id": "A1", "name": "主校区"}],
    )
    write_csv(
        source / "rooms.csv",
        ["id", "name", "teaching_area_id", "capacity", "is_active"],
        [{"id": "R1", "name": "101", "teaching_area_id": "A1", "capacity": 80, "is_active": "是"}],
    )
    write_csv(
        source / "teachers.csv",
        ["employee_id", "name", "project", "primary_subject", "employment_status"],
        [{"employee_id": "T1", "name": "张老师", "project": "考研", "primary_subject": "英语", "employment_status": "在职"}],
    )
    write_csv(
        source / "products.csv",
        ["id", "name", "subject_category", "subject", "standard_capacity"],
        [{"id": "P1", "name": "考研寒暑营-英语", "subject_category": "公共课", "subject": "英语", "standard_capacity": 50}],
    )
    write_csv(
        source / "product_courses.csv",
        [
            "product_id",
            "product_name",
            "subject_category",
            "subject",
            "stage",
            "course_module",
            "course_group",
            "total_hours",
            "block_hours",
            "teaching_area_ids",
        ],
        [
            {
                "product_id": "P1",
                "product_name": "考研寒暑营-英语",
                "subject_category": "公共课",
                "subject": "英语",
                "stage": "基础",
                "course_module": "词汇",
                "course_group": "阅读类",
                "total_hours": 4,
                "block_hours": 4,
                "teaching_area_ids": "A1",
            }
        ],
    )
    write_csv(
        source / "product_schedule_rules.csv",
        [
            "rule_id",
            "rule_name",
            "scope_type",
            "product_id",
            "subject",
            "start_date",
            "end_date",
            "allowed_periods",
            "allowed_weekdays",
            "block_hours_override",
        ],
        [
            {
                "rule_id": "RULE1",
                "rule_name": "夏季白天",
                "scope_type": "product_ids",
                "product_id": "P1",
                "subject": "英语，政治",
                "start_date": "2026-07-01",
                "end_date": "2026-07-10",
                "allowed_periods": "AM|PM",
                "allowed_weekdays": "周一|周二|周三|周四|周五|周六",
                "block_hours_override": 4,
            },
            {
                "rule_id": "RULE2",
                "rule_name": "周末晚间",
                "scope_type": "product_ids",
                "product_id": "P1",
                "subject": "英语，政治",
                "start_date": "2026-07-11",
                "end_date": "2026-07-12",
                "allowed_periods": "EVENING",
                "allowed_weekdays": "周六|周日",
                "block_hours_override": 2,
            }
        ],
    )
    write_csv(
        source / "classes.csv",
        [
            "id",
            "name",
            "product_id",
            "subject",
            "stages",
            "exam_season",
            "suite_code",
            "size",
            "start_date",
            "start_period",
            "end_date",
            "end_period",
            "preferred_teaching_area_ids",
        ],
        [
            {
                "id": "C1",
                "name": "英语1班",
                "product_id": "P1",
                "subject": "英语",
                "stages": "基础",
                "exam_season": "27考研",
                "suite_code": "S1",
                "size": 30,
                "start_date": "2026-07-01",
                "start_period": "",
                "end_date": "",
                "end_period": "",
                "preferred_teaching_area_ids": "A1",
            },
            {
                "id": "C2",
                "name": "英语2班",
                "product_id": "P1",
                "subject": "英语",
                "stages": "基础",
                "exam_season": "27考研",
                "suite_code": "S1",
                "size": 30,
                "start_date": "2026-07-01",
                "start_period": "",
                "end_date": "",
                "end_period": "",
                "preferred_teaching_area_ids": "A1",
            },
        ],
    )
    write_csv(
        source / "class_teacher_assignments.csv",
        ["class_id", "subject", "stage", "course_module", "course_group", "teacher_id", "teacher_name"],
        [
            {
                "class_id": "C1",
                "subject": "英语",
                "stage": "基础",
                "course_module": "词汇",
                "course_group": "阅读类",
                "teacher_id": "T1",
                "teacher_name": "张老师",
            },
            {
                "class_id": "C2",
                "subject": "英语",
                "stage": "基础",
                "course_module": "词汇",
                "course_group": "阅读类",
                "teacher_id": "T1",
                "teacher_name": "张老师",
            },
        ],
    )


def pipeline_args(source: Path, data_dir: Path, output_dir: Path, timestamp: str = "20260429_120000") -> SimpleNamespace:
    return SimpleNamespace(
        source=str(source),
        data_dir=str(data_dir),
        output_dir=str(output_dir),
        timestamp=timestamp,
        exclude_weekdays="Sun",
        slot_set="all",
    )


class SchedulingPipelineTest(unittest.TestCase):
    def tearDown(self) -> None:
        data_admin_server.DATA_DIR = ORIGINAL_DATA_DIR

    def test_default_stage_order_matches_current_business_stage_sequence(self) -> None:
        expected = ["导学", "基础", "强化", "冲刺", "一轮", "二轮", "三轮", "四轮"]

        self.assertEqual(expected, data_admin_server.DEFAULT_STAGE_ORDER)
        self.assertEqual(expected, data_admin_server.product_stage_order({}))
        self.assertNotIn("专项", data_admin_server.DEFAULT_STAGE_ORDER)
        self.assertNotIn("复试", data_admin_server.DEFAULT_STAGE_ORDER)
        self.assertEqual(
            ["导学1", "导学2", "基础"],
            data_admin_server.sort_stage_values(["基础", "导学2", "导学1"]),
        )

    def test_cli_parser_accepts_preflight_mode(self) -> None:
        args = build_parser().parse_args(["--source", "incoming", "--preflight"])

        self.assertTrue(args.preflight)
        self.assertEqual(args.source, "incoming")

    def test_summer_only_sunday_policy_keeps_fall_sundays(self) -> None:
        slots = generate_time_slots(
            date(2026, 8, 1),
            date(2026, 9, 6),
            parse_weekdays("Sun"),
            "all",
            sunday_policy="summer-only",
        )
        counts_by_date = {
            date_text: sum(1 for slot in slots if slot["date"] == date_text)
            for date_text in {"2026-08-02", "2026-09-06"}
        }
        self.assertEqual(counts_by_date["2026-08-02"], 0)
        self.assertEqual(counts_by_date["2026-09-06"], 5)

    def test_csv_pipeline_generates_schedule_and_dynamic_input(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "incoming"
            data_dir = root / "data"
            output_dir = root / "outputs"
            write_minimal_csv_source(source)

            result = run_pipeline(pipeline_args(source, data_dir, output_dir))

            self.assertTrue(result.schedule_csv_path.exists())
            self.assertTrue(result.schedule_html_path.exists())
            self.assertTrue(result.report_path.exists())
            with result.schedule_csv_path.open(encoding="utf-8") as handle:
                schedule_rows = list(csv.DictReader(handle))
            self.assertEqual({row["room_id"] for row in schedule_rows}, {"R1"})
            self.assertEqual({row["room_name"] for row in schedule_rows}, {"101"})
            self.assertEqual({row["teaching_area_id"] for row in schedule_rows}, {"A1"})

            scheduler_input = json.loads((data_dir / "scheduler_input_draft.json").read_text(encoding="utf-8"))
            self.assertEqual(len(scheduler_input["classes"]), 2)
            self.assertEqual(len(scheduler_input["conflict_groups"]), 1)
            self.assertEqual(set(scheduler_input["conflict_groups"][0]["class_ids"]), {"C1", "C2"})
            self.assertTrue(all(slot["date"] != "2026-07-05" for slot in scheduler_input["time_slots"]))
            self.assertEqual({slot["period"] for slot in scheduler_input["time_slots"]}, {"AM", "PM", "EVENING"})
            self.assertEqual(scheduler_input["classes"][0]["end_date"], "2026-07-12")
            self.assertEqual(scheduler_input["classes"][0]["start_period"], "AM")
            report = result.report_path.read_text(encoding="utf-8")
            self.assertIn("缺少 end_date", report)
            self.assertNotIn("scheduling_scope_overrides", report)
            self.assertNotIn("merge_course_details", report)

    def test_public_csv_minimal_example_runs_preflight_and_pipeline(self) -> None:
        source = Path(__file__).resolve().parents[1] / "examples" / "csv_minimal"
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_dir = root / "data"
            output_dir = root / "outputs"

            preflight = run_preflight(pipeline_args(source, data_dir, output_dir, "csv_minimal_preflight"))
            self.assertTrue(preflight.passed, preflight.error)

            result = run_pipeline(pipeline_args(source, data_dir, output_dir, "csv_minimal_run"))
            with result.schedule_csv_path.open(encoding="utf-8") as handle:
                rows = list(csv.DictReader(handle))

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["class_id"], "C_DEMO_01")
            self.assertEqual(rows[0]["duration_hours"], "4")
            self.assertEqual(rows[0]["period"], "AM")

    def test_preflight_validates_scheduler_input_and_writes_missing_teacher_template(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "incoming"
            data_dir = root / "data"
            output_dir = root / "outputs"
            write_minimal_csv_source(source)
            write_csv(
                source / "class_teacher_assignments.csv",
                ["class_id", "subject", "stage", "course_module", "course_group", "teacher_id", "teacher_name"],
                [],
            )

            result = run_preflight(pipeline_args(source, data_dir, output_dir, timestamp="20260628_020000"))

            self.assertFalse(result.passed)
            self.assertIn("缺少课程老师安排", result.error)
            self.assertEqual({item.class_id for item in result.missing_teacher_requirements}, {"C1", "C2"})
            self.assertFalse((data_dir / "classes.json").exists())
            self.assertEqual(len(result.generated_files), 1)
            with result.generated_files[0].open(encoding="utf-8") as handle:
                rows = list(csv.DictReader(handle))
            self.assertEqual({row["class_id"] for row in rows}, {"C1", "C2"})
            self.assertEqual({row["class_name"] for row in rows}, {"英语1班", "英语2班"})
            self.assertEqual({row["product_name"] for row in rows}, {"考研寒暑营-英语"})
            self.assertEqual({row["class_id"] for row in result.missing_teacher_rows}, {"C1", "C2"})
            self.assertTrue(all(row["class_schedule_mode"] == "本班实际排课" for row in rows))
            report = result.report_path.read_text(encoding="utf-8")
            self.assertIn("## 缺老师补录摘要", report)
            self.assertIn("| 英语1班 | 考研寒暑营-英语 | 英语 | 基础 | 阅读类 |", report)
            self.assertIn("missing_class_teacher_assignments_", report)

    def test_numbered_template_sheets_use_english_header_row(self) -> None:
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl is not installed")

        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "template.xlsx"
            workbook = Workbook()
            window_sheet = workbook.active
            window_sheet.title = "01_年度排课窗口表"
            window_sheet.append(["01 · 年度排课窗口表"])
            window_sheet.append(["用途：定义年度排课窗口"])
            window_sheet.append([])
            window_sheet.append([])
            window_sheet.append(["排课窗口ID", "排课窗口名称", "窗口年份"])
            window_sheet.append(["schedule_window_id", "schedule_window_name", "window_year"])
            window_sheet.append(["2026暑假", "2026暑假", 2026])

            slot_sheet = workbook.create_sheet("02_课节表")
            slot_sheet.append(["02 · 课节表"])
            slot_sheet.append(["用途：定义课节"])
            slot_sheet.append([])
            slot_sheet.append([])
            slot_sheet.append(["唯一ID", "日期", "是否可用"])
            slot_sheet.append(["id", "date", "is_usable"])
            slot_sheet.append(["S1", "2026-07-01", "是"])

            mapping_sheet = workbook.create_sheet("18_ERP产品对应表")
            mapping_sheet.append(["18 · ERP产品对应表"])
            mapping_sheet.append(["用途：关联本地产品和 ERP 标准产品"])
            mapping_sheet.append([])
            mapping_sheet.append([])
            mapping_sheet.append(["本地产品ID", "ERP课程编码"])
            mapping_sheet.append(["local_product_id", "business_product_id"])
            mapping_sheet.append(["P1", "100"])
            workbook.save(workbook_path)

            tables = load_source_tables(workbook_path)

            self.assertIn("schedule_windows", tables)
            self.assertIn("time_slots", tables)
            self.assertIn("business_product_mappings", tables)
            self.assertEqual(tables["schedule_windows"].rows[0]["schedule_window_id"], "2026暑假")
            self.assertNotIn("排课窗口ID", tables["schedule_windows"].rows[0])
            self.assertEqual(tables["time_slots"].rows[0]["is_usable"], "是")
            self.assertEqual(tables["business_product_mappings"].rows[0]["business_product_id"], "100")

    def test_legacy_business_product_map_file_loads_as_current_mapping_table(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp)
            write_csv(
                source / "business_product_map.csv",
                ["business_product_id", "canonical_product_id"],
                [{"business_product_id": "100", "canonical_product_id": "P_REG"}],
            )

            tables = load_source_tables(source)

            self.assertIn("business_product_mappings", tables)
            self.assertNotIn("business_product_map", tables)
            self.assertEqual(tables["business_product_mappings"].rows[0]["canonical_product_id"], "P_REG")

    def test_csv_source_accepts_gb18030_export(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp)
            (source / "products.csv").write_bytes("id,name\nP1,考研产品\n".encode("gb18030"))

            tables = load_source_tables(source)

            self.assertIn("products", tables)
            self.assertEqual(tables["products"].rows[0], {"id": "P1", "name": "考研产品"})

    def test_current_source_table_csv_filenames_are_auto_recognized(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp)
            for table_name in SOURCE_TABLES:
                fieldnames = TABLE_FIELDNAMES.get(table_name, ["id"])
                write_csv(
                    source / f"{table_name}.csv",
                    [fieldnames[0]],
                    [{fieldnames[0]: f"{table_name}_row"}],
                )

            for table_name in SOURCE_TABLES:
                self.assertEqual(table_name_for(f"{table_name}.csv"), table_name)

            tables = load_source_tables(source)

            self.assertEqual(set(SOURCE_TABLES), set(tables))
            self.assertEqual(tables["locked_scheduled_lessons"].rows[0]["id"], "locked_scheduled_lessons_row")
            self.assertEqual(tables["historical_scheduled_lessons"].rows[0]["id"], "historical_scheduled_lessons_row")

    def test_removed_merge_course_details_file_is_not_loaded_as_source_table(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp)
            write_csv(source / "business_classes.csv", ["班级编码"], [{"班级编码": "C1"}])
            write_csv(
                source / "merge_course_details.csv",
                ["source_class_id", "scheduled_class_id", "merge_type"],
                [{"source_class_id": "C1", "scheduled_class_id": "C2", "merge_type": "partial"}],
            )

            tables = load_source_tables(source)

            self.assertIn("business_classes", tables)
            self.assertNotIn("merge_course_details", tables)

    def test_export_scheduler_input_filters_unusable_template_slots(self) -> None:
        state = data_admin_server.normalize_payload(
            {
                "time_slots": [
                    {
                        "id": "S_USABLE",
                        "date": "2026-07-01",
                        "period": "AM",
                        "name": "上午一",
                        "order": 1,
                        "start_time": "08:00",
                        "end_time": "10:00",
                        "duration_hours": 2,
                        "is_usable": "是",
                    },
                    {
                        "id": "S_BLOCKED",
                        "date": "2026-07-02",
                        "period": "AM",
                        "name": "上午一",
                        "order": 1,
                        "start_time": "08:00",
                        "end_time": "10:00",
                        "duration_hours": 2,
                        "is_usable": "否",
                    },
                ],
                "teaching_areas": [],
                "rooms": [],
                "teachers": [],
                "products": [],
                "product_courses": [],
                "product_schedule_rules": [],
                "classes": [],
                "global_blackout_dates": [],
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            export = data_admin_server.export_scheduler_input(state, time_slots=state["time_slots"])
            scheduler_input = json.loads(Path(export["path"]).read_text(encoding="utf-8"))

        self.assertEqual([slot["id"] for slot in scheduler_input["time_slots"]], ["S_USABLE"])

    def test_build_scheduler_input_exports_products_for_schedulable_classes_only(self) -> None:
        state = data_admin_server.normalize_payload(
            {
                "teaching_areas": [],
                "rooms": [],
                "teachers": [],
                "products": [
                    {"id": "P_ACTIVE", "name": "本轮产品", "subject": "英语"},
                    {"id": "P_LOCKED", "name": "锁定产品", "subject": "数学"},
                ],
                "product_courses": [
                    {
                        "product_id": "P_ACTIVE",
                        "product_name": "本轮产品",
                        "subject": "英语",
                        "stage": "基础",
                        "course_module": "词汇",
                        "course_group": "阅读类",
                        "total_hours": 2,
                        "block_hours": 2,
                    },
                    {
                        "product_id": "P_LOCKED",
                        "product_name": "锁定产品",
                        "subject": "数学",
                        "stage": "基础",
                        "course_module": "函数",
                        "course_group": "数学类",
                        "total_hours": 2,
                        "block_hours": 2,
                    },
                ],
                "product_schedule_rules": [
                    {
                        "rule_id": "RULE_ALL",
                        "scope_type": "all",
                        "season_window_id": "WINDOW_SUMMER",
                        "allowed_periods": "AM",
                    }
                ],
                "classes": [
                    {
                        "id": "C_ACTIVE",
                        "name": "本轮班",
                        "product_id": "P_ACTIVE",
                        "subject": "英语",
                        "stages": "基础",
                        "size": 30,
                    },
                    {
                        "id": "C_LOCKED",
                        "name": "锁定班",
                        "product_id": "P_LOCKED",
                        "subject": "数学",
                        "stages": "基础",
                        "size": 30,
                        "is_schedule_locked": "是",
                    },
                ],
                "global_blackout_dates": [],
            }
        )

        scheduler_input = data_admin_server.build_scheduler_input(state, time_slots=[])

        self.assertEqual([item["id"] for item in scheduler_input["products"]], ["P_ACTIVE"])
        self.assertEqual([item["id"] for item in scheduler_input["classes"]], ["C_ACTIVE"])
        self.assertEqual(
            [item["product_id"] for item in scheduler_input["product_schedule_rules"]],
            ["P_ACTIVE"],
        )

    def test_export_scheduler_input_includes_active_teacher_unavailability(self) -> None:
        state = data_admin_server.normalize_payload(
            {
                "time_slots": [
                    {
                        "id": "S1",
                        "date": "2026-07-01",
                        "period": "AM",
                        "name": "上午一",
                        "order": 1,
                        "start_time": "08:00",
                        "end_time": "10:00",
                        "duration_hours": 2,
                        "is_usable": "是",
                    }
                ],
                "teaching_areas": [],
                "rooms": [],
                "teachers": [],
                "teacher_unavailability": [
                    {
                        "unavailable_id": "UNAVAIL_T1_AM",
                        "employee_id": "T1",
                        "teacher_name": "张老师",
                        "unavailable_type": "请假",
                        "start_date": "2026-07-01",
                        "end_date": "2026-07-01",
                        "periods": "AM",
                        "is_active": "是",
                        "reason": "请假",
                    },
                    {
                        "unavailable_id": "UNAVAIL_T2_TBD",
                        "employee_id": "T2",
                        "is_active": "否",
                        "reason": "待确认",
                    },
                ],
                "products": [],
                "product_courses": [],
                "product_schedule_rules": [],
                "classes": [],
                "global_blackout_dates": [],
            }
        )
        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            export = data_admin_server.export_scheduler_input(state, time_slots=state["time_slots"])
            scheduler_input = json.loads(Path(export["path"]).read_text(encoding="utf-8"))

        self.assertEqual(export["counts"]["teacher_unavailability"], 1)
        self.assertEqual(
            scheduler_input["teacher_unavailability"],
            [
                {
                    "unavailable_id": "UNAVAIL_T1_AM",
                    "teacher_id": "T1",
                    "employee_id": "T1",
                    "teacher_name": "张老师",
                    "unavailable_type": "请假",
                    "start_date": "2026-07-01",
                    "end_date": "2026-07-01",
                    "periods": ["AM"],
                    "is_active": True,
                    "reason": "请假",
                }
            ],
        )

    def test_export_scheduler_input_strips_old_teacher_available_slots_from_assignments_and_requirements(self) -> None:
        state = data_admin_server.normalize_payload(
            {
                "time_slots": [
                    {
                        "id": "S1",
                        "date": "2026-07-01",
                        "period": "AM",
                        "name": "上午一",
                        "order": 1,
                        "start_time": "08:00",
                        "end_time": "10:00",
                        "duration_hours": 2,
                        "is_usable": "是",
                    }
                ],
                "teaching_areas": [{"id": "A1", "name": "教学区", "is_active": "是"}],
                "rooms": [{"id": "R1", "name": "101", "teaching_area_id": "A1", "capacity": 80, "is_active": "是"}],
                "teachers": [],
                "products": [{"id": "P1", "name": "测试产品", "subject": "英语"}],
                "product_courses": [
                    {
                        "product_id": "P1",
                        "product_name": "测试产品",
                        "subject": "英语",
                        "stage": "基础",
                        "course_group": "阅读类",
                        "total_hours": 2,
                    }
                ],
                "product_schedule_rules": [
                    {
                        "rule_id": "RULE_P1",
                        "product_id": "P1",
                        "allowed_periods": ["AM"],
                        "allowed_weekdays": ["周三"],
                        "block_hours": 2,
                    }
                ],
                "classes": [
                    {
                        "id": "C1",
                        "name": "英语1班",
                        "product_id": "P1",
                        "subject": "英语",
                        "stages": ["基础"],
                        "size": 30,
                        "preferred_room_ids": ["R1"],
                        "teacher_assignments": [
                            {
                                "subject": "英语",
                                "stage": "基础",
                                "course_group": "阅读类",
                                "class_schedule_mode": "本班实际排课",
                                "actual_scheduled_class_id": "C1",
                                "teacher_id": "T1",
                                "teacher_name": "张老师",
                                "teacher_available_slots": ["OLD_SLOT_ONLY"],
                            }
                        ],
                        "requirements": [
                            {
                                "subject": "英语",
                                "stage": "基础",
                                "course_group": "阅读类",
                                "teacher_id": "T1",
                                "teacher_name": "张老师",
                                "teacher_available_slots": ["OLD_SLOT_ONLY"],
                                "total_hours": 2,
                                "block_hours": 2,
                            }
                        ],
                    }
                ],
                "global_blackout_dates": [],
            }
        )
        normalized_assignment = state["classes"][0]["teacher_assignments"][0]
        normalized_requirement = state["classes"][0]["requirements"][0]
        self.assertNotIn("teacher_available_slots", normalized_assignment)
        self.assertNotIn("teacher_available_slots", normalized_requirement)
        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            export = data_admin_server.export_scheduler_input(state, time_slots=state["time_slots"])
            scheduler_input = json.loads(Path(export["path"]).read_text(encoding="utf-8"))

        assignment = scheduler_input["classes"][0]["teacher_assignments"][0]
        requirement = scheduler_input["classes"][0]["requirements"][0]
        self.assertNotIn("teacher_available_slots", assignment)
        self.assertNotIn("teacher_available_slots", requirement)
        scheduler.schedule(scheduler.load_input_data(scheduler_input))

    def test_export_scheduler_input_strips_legacy_requirement_teaching_area_ids(self) -> None:
        state = data_admin_server.normalize_payload(
            {
                "time_slots": [
                    {
                        "id": "S1",
                        "date": "2026-07-01",
                        "period": "AM",
                        "name": "上午一",
                        "order": 1,
                        "start_time": "08:00",
                        "end_time": "10:00",
                        "duration_hours": 2,
                        "is_usable": "是",
                    }
                ],
                "teaching_areas": [
                    {"id": "A1", "name": "主校区", "is_active": "是"},
                    {"id": "A2", "name": "分校区", "is_active": "是"},
                ],
                "rooms": [
                    {"id": "R1", "name": "101", "teaching_area_id": "A1", "capacity": 80, "is_active": "是"},
                    {"id": "R2", "name": "201", "teaching_area_id": "A2", "capacity": 80, "is_active": "是"},
                ],
                "classes": [
                    {
                        "id": "C1",
                        "name": "直填班",
                        "subject": "数学",
                        "requirements": [
                            {
                                "subject": "数学",
                                "stage": "基础",
                                "course_group": "高数",
                                "teacher_id": "T1",
                                "teacher_name": "张老师",
                                "total_hours": 2,
                                "block_hours": 2,
                                "teaching_area_ids": ["A2"],
                            },
                            {
                                "subject": "数学",
                                "stage": "强化",
                                "course_group": "高数",
                                "teacher_id": "T1",
                                "teacher_name": "张老师",
                                "total_hours": 2,
                                "block_hours": 2,
                                "room_ids": ["R2"],
                                "teaching_area_ids": ["A1"],
                            },
                        ],
                    }
                ],
            }
        )
        normalized_requirements = state["classes"][0]["requirements"]
        self.assertNotIn("teaching_area_ids", normalized_requirements[0])
        self.assertNotIn("teaching_area_ids", normalized_requirements[1])

        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            export = data_admin_server.export_scheduler_input(state, time_slots=state["time_slots"])
            scheduler_input = json.loads(Path(export["path"]).read_text(encoding="utf-8"))

        exported_requirements = {
            item["stage"]: item
            for item in scheduler_input["classes"][0]["requirements"]
        }
        self.assertNotIn("teaching_area_ids", exported_requirements["基础"])
        self.assertNotIn("room_ids", exported_requirements["基础"])
        self.assertEqual(["R2"], exported_requirements["强化"]["room_ids"])

    def test_scheduler_ignores_legacy_available_slot_fields(self) -> None:
        payload = {
            "time_slots": [
                {"id": "2026-07-01-AM-1", "date": "2026-07-01", "period": "AM", "name": "上午一", "order": 1, "start_time": "08:00", "end_time": "10:00", "duration_hours": 2},
                {"id": "2026-07-02-AM-1", "date": "2026-07-02", "period": "AM", "name": "上午一", "order": 1, "start_time": "08:00", "end_time": "10:00", "duration_hours": 2},
            ],
            "rooms": [{"id": "R1", "capacity": 80}],
            "products": [
                {
                    "id": "P1",
                    "name": "测试产品",
                    "requirements": [
                        {
                            "subject_category": "公共课",
                            "subject": "英语",
                            "stage": "基础",
                            "course_group": "阅读类",
                            "total_hours": 2,
                            "block_hours": 2,
                            "room_ids": ["R1"],
                        }
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C1",
                    "name": "英语1班",
                    "product_id": "P1",
                    "subject": "英语",
                    "size": 30,
                    "available_slots": ["OLD_SLOT_ONLY"],
                    "teacher_assignments": [
                        {
                            "subject": "英语",
                            "stage": "基础",
                            "course_group": "阅读类",
                            "teacher_id": "T1",
                            "teacher_name": "张老师",
                            "teacher_available_slots": ["OLD_SLOT_ONLY"],
                        }
                    ],
                }
            ],
        }

        assignments = scheduler.schedule(scheduler.load_input_data(payload))
        generated = [assignment for assignment in assignments if not assignment.task.is_locked]

        self.assertEqual(len(generated), 1)
        self.assertEqual(generated[0].candidate.slots[0].date, "2026-07-01")
        self.assertFalse(hasattr(scheduler.load_input_data(payload).classes["C1"], "available_slots"))

    def test_save_state_writes_current_teacher_assignment_fields(self) -> None:
        payload = {
            "time_slots": [],
            "teaching_areas": [],
            "rooms": [],
            "teachers": [],
            "teacher_unavailability": [],
            "products": [{"id": "P1", "name": "测试产品", "subject": "英语"}],
            "product_courses": [
                {
                    "product_id": "P1",
                    "product_name": "测试产品",
                    "subject": "英语",
                    "stage": "基础",
                    "course_group": "阅读类",
                    "total_hours": 2,
                }
            ],
            "product_schedule_rules": [],
            "classes": [
                {
                    "id": "C1",
                    "name": "英语1班",
                    "product_id": "P1",
                    "subject": "英语",
                    "teacher_assignments": [
                        {
                            "subject": "英语",
                            "stage": "基础",
                            "course_group": "阅读类",
                            "class_schedule_mode": "本班实际排课",
                            "actual_scheduled_class_id": "C1",
                            "schedule_mode": "独立排课",
                            "inherit_from_class_id": "",
                            "teacher_id": "T1",
                            "teacher_name": "张老师",
                            "teacher_available_slots": ["OLD_SLOT_ONLY"],
                            "course_module": "词汇",
                        }
                    ],
                }
            ],
            "class_window_boundaries": [],
            "class_conflict_groups": [],
            "locked_scheduled_lessons": [],
            "teaching_area_links": [],
            "global_blackout_dates": [],
            "historical_scheduled_lessons": [],
            "erp_standard_products": [],
            "business_product_mappings": [],
        }
        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            data_admin_server.save_state(payload)
            classes_doc = json.loads((data_admin_server.DATA_DIR / "classes.json").read_text(encoding="utf-8"))
            assignments_doc = json.loads(
                (data_admin_server.DATA_DIR / "class_teacher_assignments.json").read_text(encoding="utf-8")
            )
            with (data_admin_server.DATA_DIR / "class_teacher_assignments.csv").open(encoding="utf-8") as handle:
                header = next(csv.reader(handle))

        nested_assignment = classes_doc["classes"][0]["teacher_assignments"][0]
        flat_assignment = assignments_doc["class_teacher_assignments"][0]
        for old_field in ("schedule_mode", "inherit_from_class_id", "teacher_available_slots", "course_module"):
            self.assertNotIn(old_field, nested_assignment)
            self.assertNotIn(old_field, flat_assignment)
            self.assertNotIn(old_field, header)
        self.assertIn("class_schedule_mode", header)
        self.assertIn("actual_scheduled_class_id", header)

    def test_teacher_fieldnames_are_shared_by_admin_and_pipeline(self) -> None:
        self.assertEqual(data_admin_server.TEACHER_FIELDNAMES, TABLE_FIELDNAMES["teachers"])
        payload = {
            "teachers": [
                {
                    "employee_id": "100001",
                    "id": "LEGACY_ID",
                    "name": "张老师",
                    "project": "考研",
                    "identity": "教师",
                    "teacher_type": "全职",
                    "primary_subject": "英语",
                    "employment_status": "在职",
                }
            ],
            "products": [],
            "product_courses": [],
            "classes": [],
        }
        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            data_admin_server.save_state(payload)
            teachers_doc = json.loads((data_admin_server.DATA_DIR / "teachers.json").read_text(encoding="utf-8"))
            with (data_admin_server.DATA_DIR / "teachers.csv").open(encoding="utf-8") as handle:
                header = next(csv.reader(handle))

        saved_row = teachers_doc["teachers"][0]
        self.assertEqual(data_admin_server.TEACHER_FIELDNAMES, list(saved_row))
        self.assertEqual(data_admin_server.TEACHER_FIELDNAMES, header)
        self.assertEqual(saved_row["employee_id"], "100001")
        self.assertEqual(saved_row["teacher_role"], "教师")
        self.assertEqual(saved_row["employment_type"], "全职")

    def test_standard_tables_are_shared_by_admin_pipeline_json_and_csv_exports(self) -> None:
        self.assertIs(TABLE_FIELDNAMES, data_admin_server.STANDARD_TABLE_FIELDNAMES)
        self.assertEqual(list(data_admin_server.STANDARD_TABLE_FIELDNAMES), TABLES)
        for table_name, fieldnames in data_admin_server.STANDARD_TABLE_FIELDNAMES.items():
            duplicates = [field for field, count in Counter(fieldnames).items() if count > 1]
            self.assertEqual([], duplicates, table_name)
        payload = {
            "products": [],
            "product_courses": [],
            "classes": [],
        }
        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            data_admin_server.save_state(payload)
            for table_name, fieldnames in data_admin_server.STANDARD_TABLE_FIELDNAMES.items():
                with (data_admin_server.DATA_DIR / f"{table_name}.csv").open(encoding="utf-8") as handle:
                    header = next(csv.reader(handle))
                self.assertEqual(fieldnames, header, table_name)
                document = json.loads((data_admin_server.DATA_DIR / f"{table_name}.json").read_text(encoding="utf-8"))
                self.assertEqual("data_admin_server.py", document["source"], table_name)
                self.assertIn(table_name, document)
                self.assertEqual(len(document[table_name]), document["record_count"], table_name)
                if table_name in data_admin_server.TABLES_WITH_EMPTY_WARNINGS:
                    self.assertEqual([], document["warnings"], table_name)

    def test_saved_json_rows_are_limited_to_current_table_fields(self) -> None:
        payload = {
            "schedule_windows": [{"schedule_window_id": "2026暑假", "season_name": "暑假", "row": 99}],
            "time_slots": [{"id": "S1", "date": "2026-07-01", "row": 99}],
            "teaching_areas": [{"id": "A1", "name": "教学区1", "row": 99}],
            "rooms": [{"id": "R1", "name": "101", "teaching_area_id": "A1", "is_active": "是", "row": 99}],
            "teachers": [{"employee_id": "100001", "name": "张老师", "identity": "教师", "teacher_type": "全职", "row": 99}],
            "teacher_unavailability": [{"unavailable_id": "U1", "employee_id": "100001", "row": 99}],
            "products": [{"id": "P1", "name": "产品1", "row": 99}],
            "product_courses": [{"product_id": "P1", "stage": "基础", "course_module": "词汇", "row": 99}],
            "product_schedule_rules": [
                {
                    "rule_id": "RULE1",
                    "rule_name": "旧规则名",
                    "scope_type": "product_ids",
                    "product_id": "P1",
                    "product_ids": "P1",
                    "product_name_keywords": "产品",
                    "block_hours": 4,
                    "block_hours_override": 4,
                    "row": 99,
                }
            ],
            "classes": [
                {
                    "id": "C1",
                    "name": "班级1",
                    "product_id": "P1",
                    "selected_stages": "基础",
                    "stages": "旧阶段别名",
                    "is_schedule_locked": "是",
                    "is_manual_schedule_locked": "是",
                    "teacher_assignments": [{"subject": "英语", "teacher_id": "100001", "teacher_name": "张老师"}],
                    "requirements": [{"subject": "英语", "total_hours": 2}],
                },
                {"id": "C2", "name": "班级2", "product_id": "P1"},
            ],
            "class_window_boundaries": [{"class_window_id": "CW1", "class_id": "C1", "preferred_room_ids": "R1", "row": 99}],
            "class_conflict_groups": [{"id": "G1", "class_ids": "C1|C2", "is_active": "是", "source": "旧来源", "row": 99}],
            "locked_scheduled_lessons": [{"id": "L1", "class_id": "C1", "room_id": "R1", "row": 99}],
            "teaching_area_links": [{"id": "LINK1", "from_teaching_area_id": "A1", "to_teaching_area_id": "A1", "row": 99}],
            "global_blackout_dates": [{"id": "B1", "name": "停课", "row": 99}],
            "historical_scheduled_lessons": [{"id": "H1", "class_id": "C1", "room_id": "R1", "is_locked": "是", "row": 99}],
            "erp_standard_products": [{"erp_product_key": "ERP1", "row": 99}],
            "business_product_mappings": [{"local_product_id": "P1", "canonical_product_id": "OLD", "row": 99}],
        }
        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            data_admin_server.save_state(payload)
            for table_name, fieldnames in data_admin_server.STANDARD_TABLE_FIELDNAMES.items():
                document = json.loads((data_admin_server.DATA_DIR / f"{table_name}.json").read_text(encoding="utf-8"))
                allowed = set(fieldnames)
                if table_name == "classes":
                    allowed.update(data_admin_server.CLASS_JSON_EXTRA_FIELDNAMES)
                for row in document[table_name]:
                    self.assertLessEqual(set(row), allowed, table_name)

            classes_doc = json.loads((data_admin_server.DATA_DIR / "classes.json").read_text(encoding="utf-8"))
            class_row = next(row for row in classes_doc["classes"] if row["id"] == "C1")
            self.assertIn("teacher_assignments", class_row)
            self.assertIn("requirements", class_row)
            self.assertNotIn("stages", class_row)
            self.assertNotIn("is_schedule_locked", class_row)

            rules_doc = json.loads((data_admin_server.DATA_DIR / "product_schedule_rules.json").read_text(encoding="utf-8"))
            rule_row = rules_doc["product_schedule_rules"][0]
            for old_field in ("rule_name", "scope_type", "product_ids", "product_name_keywords", "block_hours_override", "row"):
                self.assertNotIn(old_field, rule_row)
            self.assertEqual(rule_row["product_id"], "P1")
            self.assertEqual(rule_row["block_hours"], 4)

    def test_global_blackout_fieldnames_are_shared_by_admin_and_pipeline(self) -> None:
        self.assertEqual(data_admin_server.GLOBAL_BLACKOUT_FIELDNAMES, TABLE_FIELDNAMES["global_blackout_dates"])
        payload = {
            "global_blackout_dates": [
                {
                    "id": "BLACKOUT_QM",
                    "name": "清明",
                    "start_date": "2027-04-05",
                    "end_date": "2027-04-05",
                    "is_active": "是",
                    "notes": "法定节假日",
                }
            ],
            "products": [],
            "product_courses": [],
            "classes": [],
        }
        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            data_admin_server.save_state(payload)
            with (data_admin_server.DATA_DIR / "global_blackout_dates.csv").open(encoding="utf-8") as handle:
                header = next(csv.reader(handle))

        self.assertEqual(data_admin_server.GLOBAL_BLACKOUT_FIELDNAMES, header)

    def test_current_conflict_group_active_field_wins_over_stale_legacy_field(self) -> None:
        payload = {
            "teaching_areas": [{"id": "A1", "name": "教学区", "is_active": "是"}],
            "rooms": [{"id": "R1", "name": "101", "teaching_area_id": "A1", "capacity": 80, "is_active": "是"}],
            "classes": [
                {"id": "C1", "name": "一班", "product_id": "P1"},
                {"id": "C2", "name": "二班", "product_id": "P1"},
            ],
            "products": [{"id": "P1", "name": "测试产品"}],
            "class_conflict_groups": [
                {
                    "id": "G_STALE",
                    "name": "旧字段残留互斥组",
                    "class_ids": "C1|C2",
                    "is_conflict_group_active": "否",
                    "is_active": "是",
                    "conflict_source": "当前字段",
                    "source": "旧字段",
                }
            ],
            "product_courses": [],
        }
        state = data_admin_server.normalize_payload(payload)
        group = state["class_conflict_groups"][0]

        self.assertFalse(group["is_conflict_group_active"])
        self.assertEqual("当前字段", group["conflict_source"])
        self.assertNotIn("is_active", group)
        self.assertNotIn("source", group)
        self.assertEqual([], data_admin_server.scheduler_conflict_groups(state, state["classes"], []))

        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            data_admin_server.save_state(payload)
            document = json.loads((data_admin_server.DATA_DIR / "class_conflict_groups.json").read_text(encoding="utf-8"))
            saved_group = document["class_conflict_groups"][0]
            with (data_admin_server.DATA_DIR / "class_conflict_groups.csv").open(encoding="utf-8") as handle:
                header = next(csv.reader(handle))

        self.assertIn("is_conflict_group_active", saved_group)
        self.assertIn("conflict_source", saved_group)
        self.assertNotIn("is_active", saved_group)
        self.assertNotIn("source", saved_group)
        self.assertIn("is_conflict_group_active", header)
        self.assertIn("conflict_source", header)
        self.assertNotIn("is_active", header)
        self.assertNotIn("source", header)

    def test_generated_suite_conflict_groups_use_current_fields_only(self) -> None:
        groups = data_admin_server.build_suite_conflict_groups(
            [
                {"id": "C1", "name": "一班", "suite_code": "2701", "exam_season": "27"},
                {"id": "C2", "name": "二班", "suite_code": "2701", "exam_season": "27"},
            ]
        )

        self.assertEqual(1, len(groups))
        group = groups[0]
        self.assertTrue(group["is_conflict_group_active"])
        self.assertEqual("套班编码", group["conflict_source"])
        self.assertNotIn("is_active", group)
        self.assertNotIn("source", group)

    def test_historical_lesson_fieldnames_are_separate_from_locked_schedule(self) -> None:
        self.assertEqual(data_admin_server.HISTORICAL_SCHEDULED_LESSON_FIELDNAMES, TABLE_FIELDNAMES["historical_scheduled_lessons"])
        self.assertIn("is_locked", data_admin_server.LOCKED_SCHEDULED_LESSON_FIELDNAMES)
        self.assertNotIn("is_locked", data_admin_server.HISTORICAL_SCHEDULED_LESSON_FIELDNAMES)
        payload = {
            "historical_scheduled_lessons": [
                {
                    "id": "HIST_1",
                    "class_id": "C1",
                    "class_name": "英语1班",
                    "date": "2027-07-01",
                    "period": "AM",
                    "duration_hours": "2",
                    "teacher_id": "T1",
                    "teacher_name": "张老师",
                    "room_id": "R1",
                    "subject": "英语",
                    "is_locked": "否",
                }
            ],
            "products": [],
            "product_courses": [],
            "classes": [],
        }
        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            data_admin_server.save_state(payload)
            with (data_admin_server.DATA_DIR / "historical_scheduled_lessons.csv").open(encoding="utf-8") as handle:
                header = next(csv.reader(handle))

        self.assertEqual(data_admin_server.HISTORICAL_SCHEDULED_LESSON_FIELDNAMES, header)

    def test_business_product_mapping_saves_current_local_product_field_only(self) -> None:
        payload = {
            "products": [{"id": "P1", "name": "测试产品", "subject": "英语"}],
            "product_courses": [],
            "business_product_mappings": [
                {
                    "business_product_id": "100",
                    "business_product_name": "ERP产品",
                    "canonical_product_id": "P1",
                    "match_status": "已匹配",
                }
            ],
        }

        normalized = data_admin_server.normalize_payload(payload)
        self.assertEqual(normalized["business_product_mappings"][0]["local_product_id"], "P1")
        self.assertNotIn("canonical_product_id", normalized["business_product_mappings"][0])

        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            data_admin_server.save_state(payload)
            mapping_doc = json.loads(
                (data_admin_server.DATA_DIR / "business_product_mappings.json").read_text(encoding="utf-8")
            )
            with (data_admin_server.DATA_DIR / "business_product_mappings.csv").open(encoding="utf-8") as handle:
                header = next(csv.reader(handle))

        saved_row = mapping_doc["business_product_mappings"][0]
        self.assertEqual(saved_row["local_product_id"], "P1")
        self.assertNotIn("canonical_product_id", saved_row)
        self.assertIn("local_product_id", header)
        self.assertNotIn("canonical_product_id", header)

    def test_class_window_ids_are_not_saved_in_class_base_table(self) -> None:
        payload = {
            "classes": [
                {
                    "id": "C1",
                    "name": "英语1班",
                    "product_id": "P1",
                    "actual_schedule_window_ids": "2026暑假|2026秋季",
                }
            ],
            "products": [{"id": "P1", "name": "测试产品", "subject": "英语"}],
            "product_courses": [],
        }

        normalized = data_admin_server.normalize_payload(payload)
        self.assertNotIn("actual_schedule_window_ids", normalized["classes"][0])

        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            data_admin_server.save_state(payload)
            classes_doc = json.loads((data_admin_server.DATA_DIR / "classes.json").read_text(encoding="utf-8"))
            with (data_admin_server.DATA_DIR / "classes.csv").open(encoding="utf-8") as handle:
                header = next(csv.reader(handle))

        self.assertNotIn("actual_schedule_window_ids", classes_doc["classes"][0])
        self.assertNotIn("actual_schedule_window_ids", header)

    def test_scheduler_rules_export_preserves_season_window(self) -> None:
        rules = [
            data_admin_server.normalize_product_rule(
                {
                    "rule_id": "RULE_P1_SUMMER",
                    "product_id": "P1",
                    "season_window_id": "WINDOW_SUMMER",
                    "window_name": "暑假",
                    "allowed_periods": "AM|PM",
                    "allowed_weekdays": "周一|周二|周三",
                    "block_hours": 4,
                    "max_hours_per_class_per_day": 4,
                    "max_blocks_per_class_per_day": 1,
                }
            )
        ]

        exported = data_admin_server.scheduler_rules(rules, {"P1"}, {"P1": {"name": "测试产品"}})

        self.assertEqual(len(exported), 1)
        self.assertEqual(exported[0]["season_window_id"], "WINDOW_SUMMER")
        self.assertEqual(exported[0]["window_name"], "暑假")
        self.assertEqual(exported[0]["max_hours_per_class_per_day"], 4)
        self.assertEqual(exported[0]["max_blocks_per_class_per_day"], 1)

    def test_start_date_is_window_and_first_lesson_is_optional_anchor(self) -> None:
        base_input = {
            "time_slots": [
                {"id": "2026-07-01-AM-1", "date": "2026-07-01", "period": "AM", "name": "上午一", "order": 1, "start_time": "08:00", "end_time": "10:00", "duration_hours": 2},
                {"id": "2026-07-01-AM-2", "date": "2026-07-01", "period": "AM", "name": "上午二", "order": 2, "start_time": "10:20", "end_time": "12:20", "duration_hours": 2},
                {"id": "2026-07-02-AM-1", "date": "2026-07-02", "period": "AM", "name": "上午一", "order": 1, "start_time": "08:00", "end_time": "10:00", "duration_hours": 2},
                {"id": "2026-07-02-AM-2", "date": "2026-07-02", "period": "AM", "name": "上午二", "order": 2, "start_time": "10:20", "end_time": "12:20", "duration_hours": 2},
            ],
            "rooms": [{"id": "R1", "capacity": 80}],
            "products": [
                {
                    "id": "P1",
                    "name": "英语",
                    "requirements": [
                        {
                            "subject_category": "公共课",
                            "subject": "英语",
                            "stage": "基础",
                            "course_module": "词汇",
                            "course_group": "阅读类",
                            "total_hours": 4,
                            "block_hours": 4,
                            "room_ids": ["R1"],
                        }
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C1",
                    "name": "英语1班",
                    "product_id": "P1",
                    "subject": "英语",
                    "size": 30,
                    "start_date": "2026-07-01",
                    "start_period": "AM",
                    "end_date": "2026-07-02",
                    "end_period": "AM",
                    "teacher_assignments": [
                        {"subject": "英语", "stage": "基础", "course_group": "阅读类", "teacher_id": "T1", "teacher_name": "张老师"}
                    ],
                }
            ],
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "input.json"
            input_without_anchor = json.loads(json.dumps(base_input, ensure_ascii=False))
            input_without_anchor["locked_lessons"] = [
                {
                    "id": "L1",
                    "class_id": "LOCKED",
                    "class_name": "已排课",
                    "date": "2026-07-01",
                    "start_time": "08:00",
                    "end_time": "12:20",
                    "teacher_id": "T1",
                    "teacher_name": "张老师",
                    "room_id": "R1",
                    "subject": "英语",
                }
            ]
            path.write_text(json.dumps(input_without_anchor, ensure_ascii=False), encoding="utf-8")
            assignments = scheduler.schedule(scheduler.load_input(path))
            self.assertEqual(assignments[-1].candidate.slots[0].date, "2026-07-02")

            input_with_anchor = json.loads(json.dumps(base_input, ensure_ascii=False))
            input_with_anchor["classes"][0]["first_lesson_date"] = "2026-07-02"
            input_with_anchor["classes"][0]["first_lesson_period"] = "AM"
            path.write_text(json.dumps(input_with_anchor, ensure_ascii=False), encoding="utf-8")
            assignments = scheduler.schedule(scheduler.load_input(path))
            self.assertEqual(assignments[0].candidate.slots[0].date, "2026-07-02")

    def test_class_scheduling_bounds_normalize_period_values(self) -> None:
        payload = {
            "time_slots": [],
            "rooms": [],
            "classes": [
                {
                    "id": "C1",
                    "name": "英语1班",
                    "start_date": "2026-07-01",
                    "start_period": "pm",
                    "end_date": "2026-07-07",
                    "end_period": " evening ",
                    "first_lesson_date": "2026-07-02",
                    "first_lesson_period": "am",
                    "requirements": [
                        {
                            "subject": "英语",
                            "teacher_id": "T1",
                            "teacher_name": "张老师",
                            "total_hours": 2,
                            "block_hours": 2,
                        }
                    ],
                }
            ],
        }

        parsed_class = scheduler.load_input_data(payload).classes["C1"]

        self.assertEqual(parsed_class.start_period, "PM")
        self.assertEqual(parsed_class.end_period, "EVENING")
        self.assertEqual(parsed_class.first_lesson_period, "AM")

    def test_class_boundary_period_requires_matching_date(self) -> None:
        payload = {
            "time_slots": [],
            "rooms": [],
            "classes": [
                {
                    "id": "C1",
                    "name": "英语1班",
                    "first_lesson_period": "AM",
                    "requirements": [
                        {
                            "subject": "英语",
                            "teacher_id": "T1",
                            "teacher_name": "张老师",
                            "total_hours": 2,
                            "block_hours": 2,
                        }
                    ],
                }
            ],
        }

        with self.assertRaisesRegex(ValueError, "first_lesson_period 时也需要填写 first_lesson_date"):
            scheduler.load_input_data(payload)

    def test_parse_direct_requirements_preserves_filters_and_legacy_group_field(self) -> None:
        requirements = scheduler.parse_direct_requirements(
            "C1",
            [
                {
                    "subject_category": "公共课",
                    "subject": "英语",
                    "quarter": "暑假",
                    "stage": "基础",
                    "course_module": "词汇",
                    "teacher_group": "阅读类",
                    "teacher_id": "T1",
                    "teacher_name": "张老师",
                    "total_hours": 4,
                    "block_hours": 2,
                    "course_code": "ENG001",
                    "course_name": "-",
                    "room_ids": "R1|R2",
                    "start_date": "2026-07-01",
                    "end_date": "2026-07-07",
                    "allowed_periods": "AM|晚上",
                    "allowed_weekdays": "周一|周三",
                    "excluded_weekdays": "周日",
                }
            ],
            {"R2", "R3"},
        )

        requirement = requirements[0]

        self.assertEqual(requirement.course_group, "阅读类")
        self.assertEqual(requirement.room_ids, {"R2"})
        self.assertEqual(requirement.allowed_periods, {"AM", "EVENING"})
        self.assertEqual(requirement.allowed_weekdays, {0, 2})
        self.assertEqual(requirement.excluded_weekdays, {6})
        self.assertEqual(requirement.course_code, "ENG001")
        self.assertEqual(requirement.course_name, "")

    def test_parse_time_slots_sorts_and_preserves_window_aliases(self) -> None:
        slots = scheduler.parse_time_slots(
            [
                {
                    "id": "S2",
                    "date": "2026-07-01",
                    "period": "PM",
                    "name": "下午一",
                    "order": "2",
                    "duration_hours": "2",
                    "window_id": "2026暑假",
                    "window_name": "暑假",
                },
                {
                    "id": "S1",
                    "date": "2026-07-01",
                    "period": "AM",
                    "order": 1,
                    "duration_hours": 4,
                    "schedule_window_id": "2026暑假",
                    "season_window_id": "WINDOW_SUMMER",
                    "season_name": "暑假",
                },
            ]
        )

        self.assertEqual([slot.id for slot in slots], ["S1", "S2"])
        self.assertEqual(slots[0].name, "S1")
        self.assertEqual(slots[0].duration_hours, 4)
        self.assertEqual(slots[0].season_window_id, "WINDOW_SUMMER")
        self.assertEqual(slots[1].schedule_window_id, "2026暑假")
        self.assertEqual(slots[1].season_name, "暑假")

    def test_parse_time_slots_rejects_invalid_rows_and_duration(self) -> None:
        with self.assertRaisesRegex(ValueError, "time_slots 现在需要使用对象格式"):
            scheduler.parse_time_slots(["S1"])

        with self.assertRaisesRegex(ValueError, "重复的课节 id: S1"):
            scheduler.parse_time_slots(
                [
                    {"id": "S1", "date": "2026-07-01", "period": "AM", "order": 1},
                    {"id": "S1", "date": "2026-07-01", "period": "PM", "order": 2},
                ]
            )

        with self.assertRaisesRegex(ValueError, "课节 S1 的 duration_hours 必须大于 0"):
            scheduler.parse_time_slots(
                [{"id": "S1", "date": "2026-07-01", "period": "AM", "order": 1, "duration_hours": 0}]
            )

    def test_parse_teacher_unavailability_accepts_aliases_and_skips_incomplete_rules(self) -> None:
        rules = scheduler.parse_teacher_unavailability(
            [
                "not a row",
                {
                    "unavailable_id": "UNAVAIL_T1_LEAVE",
                    "employee_id": "T1",
                    "start_date": "2026-07-01",
                    "end_date": "2026-07-02",
                    "periods": "AM|晚上",
                    "notes": "请假",
                },
                {
                    "teacher_id": "T1",
                    "weekdays": "周一|周三",
                    "schedule_window_id": "2026暑假",
                    "reason": "兼职限制",
                },
                {
                    "id": "T2",
                    "schedule_window_ids": "2026秋季|WINDOW_AUTUMN",
                    "is_active": True,
                },
                {"teacher_id": "T3", "is_active": "否", "periods": "AM"},
                {"teacher_id": "T4"},
            ]
        )

        self.assertEqual(set(rules), {"T1", "T2"})
        self.assertEqual(len(rules["T1"]), 2)
        self.assertEqual(rules["T1"][0].unavailable_id, "UNAVAIL_T1_LEAVE")
        self.assertEqual(rules["T1"][0].start_date, "2026-07-01")
        self.assertEqual(rules["T1"][0].end_date, "2026-07-02")
        self.assertEqual(rules["T1"][0].periods, {"AM", "EVENING"})
        self.assertEqual(rules["T1"][0].reason, "请假")
        self.assertEqual(rules["T1"][1].weekdays, {0, 2})
        self.assertEqual(rules["T1"][1].schedule_window_ids, {"2026暑假"})
        self.assertEqual(rules["T1"][1].reason, "兼职限制")
        self.assertEqual(rules["T2"][0].schedule_window_ids, {"2026秋季", "WINDOW_AUTUMN"})

    def test_parse_teacher_unavailability_rejects_reversed_date_range(self) -> None:
        with self.assertRaisesRegex(ValueError, "教师不可排规则 UNAVAIL_BAD 的 end_date 不能早于 start_date"):
            scheduler.parse_teacher_unavailability(
                [
                    {
                        "unavailable_id": "UNAVAIL_BAD",
                        "teacher_id": "T1",
                        "start_date": "2026-07-03",
                        "end_date": "2026-07-01",
                    }
                ]
            )

    def test_parse_product_requirements_preserves_shared_filters_and_legacy_group_field(self) -> None:
        products = scheduler.parse_products(
            [
                {
                    "id": "P1",
                    "name": "英语产品",
                    "requirements": [
                        {
                            "subject_category": "公共课",
                            "subject": "英语",
                            "quarter": "暑假",
                            "stage": "基础",
                            "course_module": "词汇",
                            "teacher_group": "阅读类",
                            "total_hours": 4,
                            "block_hours": 2,
                            "course_code": "ENG001",
                            "course_name": "-",
                            "room_ids": "R1|R2",
                            "start_date": "2026-07-01",
                            "end_date": "2026-07-07",
                            "allowed_periods": "AM|晚上",
                            "allowed_weekdays": "周一|周三",
                            "excluded_weekdays": "周日",
                        }
                    ],
                }
            ]
        )

        requirement = products["P1"].requirements[0]

        self.assertEqual(requirement.course_group, "阅读类")
        self.assertEqual(requirement.room_ids, {"R1", "R2"})
        self.assertEqual(requirement.allowed_periods, {"AM", "EVENING"})
        self.assertEqual(requirement.allowed_weekdays, {0, 2})
        self.assertEqual(requirement.excluded_weekdays, {6})
        self.assertEqual(requirement.course_code, "ENG001")
        self.assertEqual(requirement.course_name, "")

    def test_class_requirement_from_product_requirement_preserves_fields_and_rooms(self) -> None:
        rule = scheduler.ScheduleRule(
            subject="英语",
            stage="基础",
            course_module="词汇",
            course_group="阅读类",
            start_date="2026-07-01",
            end_date="2026-07-31",
            allowed_periods={"AM"},
            allowed_weekdays={0, 2},
            excluded_weekdays={6},
            block_hours=4,
            season_window_ids={"WINDOW_SUMMER"},
        )
        product_req = scheduler.ProductRequirement(
            subject_category="公共课",
            subject="英语",
            quarter="暑假",
            stage="基础",
            course_module="词汇",
            course_group="阅读类",
            total_hours=8,
            block_hours=4,
            course_code="ENG001",
            course_name="暑假阅读",
            room_ids={"R1", "R2"},
            start_date="2026-07-01",
            end_date="2026-07-31",
            allowed_periods={"AM", "PM"},
            allowed_weekdays={0, 2},
            excluded_weekdays={6},
            schedule_rules=(rule,),
        )
        teacher = scheduler.TeacherAssignment(
            product_id="P1",
            subject="英语",
            stage="基础",
            course_module="词汇",
            course_group="阅读类",
            teacher_id="T1",
            teacher_name="张老师",
        )

        requirement = scheduler.class_requirement_from_product_requirement(
            "C1",
            product_req,
            teacher,
            {"R2", "R3"},
        )

        self.assertEqual(requirement.subject_category, "公共课")
        self.assertEqual(requirement.subject, "英语")
        self.assertEqual(requirement.quarter, "暑假")
        self.assertEqual(requirement.stage, "基础")
        self.assertEqual(requirement.course_module, "词汇")
        self.assertEqual(requirement.course_group, "阅读类")
        self.assertEqual(requirement.teacher_id, "T1")
        self.assertEqual(requirement.teacher_name, "张老师")
        self.assertEqual(requirement.total_hours, 8)
        self.assertEqual(requirement.block_hours, 4)
        self.assertEqual(requirement.course_code, "ENG001")
        self.assertEqual(requirement.course_name, "暑假阅读")
        self.assertEqual(requirement.room_ids, {"R2"})
        self.assertEqual(requirement.start_date, "2026-07-01")
        self.assertEqual(requirement.end_date, "2026-07-31")
        self.assertEqual(requirement.allowed_periods, {"AM", "PM"})
        self.assertEqual(requirement.allowed_weekdays, {0, 2})
        self.assertEqual(requirement.excluded_weekdays, {6})
        self.assertEqual(requirement.schedule_rules, (rule,))

        with self.assertRaisesRegex(ValueError, "班级 C1/英语/词汇 的班级教室限制与产品/课程教室限制没有交集"):
            scheduler.class_requirement_from_product_requirement("C1", product_req, teacher, {"R9"})

    def test_resolve_product_requirement_teachers_reports_unique_missing_labels(self) -> None:
        requirements = [
            scheduler.ProductRequirement("公共课", "英语", None, "基础", "词汇", "阅读类", 2, 2),
            scheduler.ProductRequirement("公共课", "英语", None, "基础", "语法", "阅读类", 2, 2),
            scheduler.ProductRequirement("公共课", "政治", None, "基础", "毛史", "毛史类", 2, 2),
        ]

        with self.assertRaises(ValueError) as context:
            scheduler.resolve_product_requirement_teachers("C1", "P1", requirements, {})

        message = str(context.exception)

        self.assertEqual(message, "班级 C1 的产品 P1 缺少课程老师安排: 英语/基础/阅读类、政治/基础/毛史类")
        self.assertEqual(message.count("英语/基础/阅读类"), 1)

    def test_parse_class_window_constraints_expands_rooms_sorts_and_validates_periods(self) -> None:
        rooms = {
            "R1": scheduler.Room("R1", teaching_area_id="A1"),
            "R2": scheduler.Room("R2", teaching_area_id="A1"),
            "R3": scheduler.Room("R3", teaching_area_id="A2"),
        }

        by_class = scheduler.parse_class_window_constraints(
            [
                {"class_id": "C1", "class_window_id": "DISABLED", "is_class_window_included": "否"},
                {"class_window_id": "MISSING_CLASS", "earliest_date": "2026-07-01"},
                {
                    "class_id": "C1",
                    "class_window_id": "W2",
                    "earliest_date": "2026-07-10",
                    "latest_date": "2026-07-12",
                    "preferred_teaching_area_ids": "A2",
                    "schedule_window_id": "2026暑假",
                },
                {
                    "class_id": "C1",
                    "class_window_id": "W1",
                    "start_date": "2026-07-01",
                    "start_period": "pm",
                    "end_date": "2026-07-02",
                    "end_period": "evening",
                    "preferred_room_ids": "R1|R2",
                    "schedule_window_name": "暑假",
                },
            ],
            rooms,
        )

        constraints = by_class["C1"]

        self.assertEqual([constraint.class_window_id for constraint in constraints], ["W1", "W2"])
        self.assertEqual(constraints[0].start_period, "PM")
        self.assertEqual(constraints[0].end_period, "EVENING")
        self.assertEqual(constraints[0].season_name, "暑假")
        self.assertEqual(constraints[0].room_ids, {"R1", "R2"})
        self.assertTrue(constraints[0].has_room_constraint)
        self.assertEqual(constraints[1].start_period, "AM")
        self.assertEqual(constraints[1].end_period, "EVENING")
        self.assertEqual(constraints[1].room_ids, {"R3"})
        self.assertTrue(constraints[1].has_room_constraint)

        with self.assertRaisesRegex(ValueError, "earliest_period"):
            scheduler.parse_class_window_constraints(
                [{"class_id": "C1", "earliest_period": "MIDDAY"}],
                rooms,
            )

    def test_build_course_blocks_expands_requirements_with_stable_ids_and_metadata(self) -> None:
        cls = scheduler.SchoolClass(
            id="C1",
            name="英语1班",
            product_id="P1",
            product_name="英语产品",
            size=45,
            room_ids={"R1"},
            start_date="2026-07-01",
            start_period="AM",
            end_date="2026-07-31",
            end_period="PM",
            first_lesson_date=None,
            first_lesson_period=None,
            stage_order={},
            requirements=[
                scheduler.Requirement(
                    subject_category="公共课",
                    subject="英语",
                    quarter="暑假",
                    stage="基础",
                    course_module="词汇",
                    course_group="阅读类",
                    teacher_id="T1",
                    teacher_name="张老师",
                    total_hours=6,
                    block_hours=2,
                    course_code="ENG001",
                    course_name="英语词汇",
                    room_ids={"R1", "R2"},
                    start_date="2026-07-03",
                    end_date="2026-07-20",
                    allowed_periods={"AM"},
                    allowed_weekdays={0, 2},
                    excluded_weekdays={6},
                )
            ],
        )

        tasks = scheduler.build_course_blocks({"C1": cls})

        self.assertEqual([task.task_id for task in tasks], ["C1:英语:1:1", "C1:英语:1:2", "C1:英语:1:3"])
        self.assertEqual({task.block_hours for task in tasks}, {2})
        self.assertEqual({task.course_code for task in tasks}, {"ENG001"})
        self.assertEqual({task.course_name for task in tasks}, {"英语词汇"})
        self.assertEqual(tasks[0].product_id, "P1")
        self.assertEqual(tasks[0].class_size, 45)
        self.assertEqual(tasks[0].room_ids, {"R1", "R2"})
        self.assertEqual(tasks[0].allowed_periods, {"AM"})
        self.assertEqual(tasks[0].allowed_weekdays, {0, 2})
        self.assertEqual(tasks[0].excluded_weekdays, {6})

    def test_teacher_assignment_fallback_requires_same_course_group(self) -> None:
        base_requirement = scheduler.ProductRequirement(
            subject_category="公共课",
            subject="英语",
            quarter=None,
            stage="基础",
            course_module="词汇",
            course_group="阅读类",
            total_hours=2,
            block_hours=2,
        )
        reinforced_requirement = scheduler.ProductRequirement(
            subject_category="公共课",
            subject="英语",
            quarter=None,
            stage="强化",
            course_module="阅读",
            course_group="阅读类",
            total_hours=2,
            block_hours=2,
        )
        writing_requirement = scheduler.ProductRequirement(
            subject_category="公共课",
            subject="英语",
            quarter=None,
            stage="强化",
            course_module="写作",
            course_group="写作类",
            total_hours=2,
            block_hours=2,
        )
        assignments = {
            ("英语", "基础", "", "阅读类"): scheduler.TeacherAssignment(
                product_id="P1",
                subject="英语",
                stage="基础",
                course_module=None,
                course_group="阅读类",
                teacher_id="T_BASE",
                teacher_name="基础老师",
            )
        }
        requirements = [base_requirement, reinforced_requirement, writing_requirement]

        reinforced_assignment = scheduler.resolve_teacher_assignment_for_requirement(
            reinforced_requirement,
            assignments,
            requirements,
        )
        writing_assignment = scheduler.resolve_teacher_assignment_for_requirement(
            writing_requirement,
            assignments,
            requirements,
        )

        self.assertEqual(reinforced_assignment.teacher_id if reinforced_assignment else "", "T_BASE")
        self.assertIsNone(writing_assignment)

    def test_schedule_fallback_backtracking_uses_search_state_assignments(self) -> None:
        payload = {
            "time_slots": [
                {
                    "id": "S1",
                    "date": "2026-07-01",
                    "period": "AM",
                    "name": "上午一",
                    "order": 1,
                    "duration_hours": 2,
                }
            ],
            "rooms": [{"id": "R1", "name": "101", "capacity": 40}],
            "products": [],
            "classes": [
                {
                    "id": "C1",
                    "name": "回溯测试班",
                    "preferred_room_ids": ["R1"],
                    "requirements": [
                        {
                            "subject": "英语",
                            "total_hours": 2,
                            "block_hours": 2,
                            "teacher_id": "T1",
                            "teacher_name": "张老师",
                        }
                    ],
                }
            ],
        }
        schedule_input = scheduler.load_input_data(payload)
        original_greedy_schedule = scheduler.greedy_schedule

        try:
            scheduler.greedy_schedule = lambda *_args, **_kwargs: None
            assignments = scheduler.schedule(schedule_input)
        finally:
            scheduler.greedy_schedule = original_greedy_schedule

        generated = [assignment for assignment in assignments if not assignment.task.is_locked]
        self.assertEqual(len(generated), 1)
        self.assertEqual(generated[0].candidate.slots[0].id, "S1")
        self.assertEqual(generated[0].candidate.room_id, "R1")

    def test_greedy_choice_filters_first_lesson_anchor_options(self) -> None:
        payload = {
            "time_slots": [
                {
                    "id": "S1",
                    "date": "2026-07-01",
                    "period": "AM",
                    "name": "上午一",
                    "order": 1,
                    "duration_hours": 2,
                },
                {
                    "id": "S2",
                    "date": "2026-07-01",
                    "period": "PM",
                    "name": "下午一",
                    "order": 2,
                    "duration_hours": 2,
                },
            ],
            "rooms": [{"id": "R1", "name": "101", "capacity": 40}],
            "products": [],
            "classes": [
                {
                    "id": "C1",
                    "name": "首课测试班",
                    "first_lesson_date": "2026-07-01",
                    "first_lesson_period": "PM",
                    "preferred_room_ids": ["R1"],
                    "requirements": [
                        {
                            "subject": "英语",
                            "total_hours": 2,
                            "block_hours": 2,
                            "teacher_id": "T1",
                            "teacher_name": "张老师",
                        }
                    ],
                }
            ],
        }
        schedule_input = scheduler.load_input_data(payload)
        plan = scheduler.build_schedule_plan(schedule_input)
        search_state = scheduler.ScheduleSearchState(
            schedule_input,
            plan.task_by_id,
            plan.task_ids_by_class,
        )

        choice = scheduler.choose_greedy_task(
            schedule_input,
            plan.task_by_id,
            plan.task_ids_by_class,
            plan.domains,
            search_state,
        )

        self.assertIsNotNone(choice)
        _, options = choice
        self.assertTrue(options)
        self.assertEqual({candidate.slots[0].period for candidate in options}, {"PM"})

    def test_scheduler_respects_teacher_unavailability_date_period(self) -> None:
        payload = {
            "time_slots": [
                {"id": "2026-07-01-AM-1", "date": "2026-07-01", "period": "AM", "name": "上午一", "order": 1, "start_time": "08:00", "end_time": "10:00", "duration_hours": 2},
                {"id": "2026-07-02-AM-1", "date": "2026-07-02", "period": "AM", "name": "上午一", "order": 1, "start_time": "08:00", "end_time": "10:00", "duration_hours": 2},
            ],
            "rooms": [{"id": "R1", "capacity": 80}],
            "products": [
                {
                    "id": "P1",
                    "name": "测试产品",
                    "requirements": [
                        {
                            "subject_category": "公共课",
                            "subject": "英语",
                            "stage": "基础",
                            "course_module": "词汇",
                            "course_group": "阅读类",
                            "total_hours": 2,
                            "block_hours": 2,
                            "room_ids": ["R1"],
                        }
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C1",
                    "name": "英语1班",
                    "product_id": "P1",
                    "subject": "英语",
                    "size": 30,
                    "teacher_assignments": [
                        {"subject": "英语", "stage": "基础", "course_group": "阅读类", "teacher_id": "T1", "teacher_name": "张老师"}
                    ],
                }
            ],
            "teacher_unavailability": [
                {
                    "unavailable_id": "UNAVAIL_T1_AM",
                    "employee_id": "T1",
                    "start_date": "2026-07-01",
                    "end_date": "2026-07-01",
                    "periods": ["AM"],
                    "is_active": True,
                }
            ],
        }

        assignments = scheduler.schedule(scheduler.load_input_data(payload))
        generated = [assignment for assignment in assignments if not assignment.task.is_locked]

        self.assertEqual(len(generated), 1)
        self.assertEqual(generated[0].candidate.slots[0].date, "2026-07-02")

    def test_scheduler_ignores_inactive_teacher_unavailability(self) -> None:
        payload = {
            "time_slots": [
                {"id": "2026-07-01-AM-1", "date": "2026-07-01", "period": "AM", "name": "上午一", "order": 1, "start_time": "08:00", "end_time": "10:00", "duration_hours": 2},
                {"id": "2026-07-02-AM-1", "date": "2026-07-02", "period": "AM", "name": "上午一", "order": 1, "start_time": "08:00", "end_time": "10:00", "duration_hours": 2},
            ],
            "rooms": [{"id": "R1", "capacity": 80}],
            "products": [
                {
                    "id": "P1",
                    "name": "测试产品",
                    "requirements": [
                        {
                            "subject_category": "公共课",
                            "subject": "英语",
                            "stage": "基础",
                            "course_module": "词汇",
                            "course_group": "阅读类",
                            "total_hours": 2,
                            "block_hours": 2,
                            "room_ids": ["R1"],
                        }
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C1",
                    "name": "英语1班",
                    "product_id": "P1",
                    "subject": "英语",
                    "size": 30,
                    "teacher_assignments": [
                        {"subject": "英语", "stage": "基础", "course_group": "阅读类", "teacher_id": "T1", "teacher_name": "张老师"}
                    ],
                }
            ],
            "teacher_unavailability": [
                {
                    "unavailable_id": "UNAVAIL_T1_AM",
                    "employee_id": "T1",
                    "start_date": "2026-07-01",
                    "end_date": "2026-07-01",
                    "periods": ["AM"],
                    "is_active": False,
                }
            ],
        }

        assignments = scheduler.schedule(scheduler.load_input_data(payload))
        generated = [assignment for assignment in assignments if not assignment.task.is_locked]

        self.assertEqual(len(generated), 1)
        self.assertEqual(generated[0].candidate.slots[0].date, "2026-07-01")

    def test_scheduler_respects_class_stage_order(self) -> None:
        payload = {
            "time_slots": [
                {"id": "2026-07-01-AM-1", "date": "2026-07-01", "period": "AM", "name": "上午一", "order": 1, "start_time": "08:00", "end_time": "10:00", "duration_hours": 2},
                {"id": "2026-07-02-AM-1", "date": "2026-07-02", "period": "AM", "name": "上午一", "order": 1, "start_time": "08:00", "end_time": "10:00", "duration_hours": 2},
            ],
            "rooms": [{"id": "R1", "capacity": 80}],
            "products": [
                {
                    "id": "P1",
                    "name": "考研暑假营-英语",
                    "requirements": [
                        {
                            "subject_category": "公共课",
                            "subject": "英语",
                            "stage": "强化",
                            "course_module": "阅读",
                            "course_group": "阅读类",
                            "total_hours": 2,
                            "block_hours": 2,
                            "room_ids": ["R1"],
                        },
                        {
                            "subject_category": "公共课",
                            "subject": "英语",
                            "stage": "基础",
                            "course_module": "词汇",
                            "course_group": "阅读类",
                            "total_hours": 2,
                            "block_hours": 2,
                            "room_ids": ["R1"],
                        },
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C1",
                    "name": "暑假营英语1班",
                    "product_id": "P1",
                    "sub_product": "暑假营",
                    "subject": "英语",
                    "stage_order": ["基础", "强化", "冲刺"],
                    "size": 30,
                    "teacher_assignments": [
                        {"subject": "英语", "stage": "基础", "course_group": "阅读类", "teacher_id": "T1", "teacher_name": "张老师"},
                        {"subject": "英语", "stage": "强化", "course_group": "阅读类", "teacher_id": "T2", "teacher_name": "李老师"},
                    ],
                }
            ],
        }

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "input.json"
            path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
            assignments = scheduler.schedule(scheduler.load_input(path))

        generated = [assignment for assignment in assignments if not assignment.task.is_locked]
        self.assertEqual([assignment.task.stage for assignment in generated], ["基础", "强化"])
        self.assertEqual([assignment.candidate.slots[0].date for assignment in generated], ["2026-07-01", "2026-07-02"])

    def test_teacher_assignments_are_merged_into_classes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "incoming"
            data_dir = root / "data"
            output_dir = root / "outputs"
            write_minimal_csv_source(source)

            run_pipeline(pipeline_args(source, data_dir, output_dir))

            classes = json.loads((data_dir / "classes.json").read_text(encoding="utf-8"))["classes"]
            assignments = {cls["id"]: cls["teacher_assignments"] for cls in classes}
            self.assertEqual(assignments["C1"][0]["teacher_id"], "T1")
            self.assertEqual(assignments["C2"][0]["course_group"], "阅读类")
            self.assertNotIn("course_module", assignments["C2"][0])

    def test_missing_teacher_template_uses_current_assignment_fields(self) -> None:
        error = (
            "班级数据校验失败:\n"
            "班级 C1 的产品 P1 缺少课程老师安排: 英语/基础/阅读类、政治/基础/毛史类\n"
            "班级 C2 缺少课程老师安排: 数学/强化/数学类"
        )
        state = {
            "classes": [
                {"id": "C1", "name": "英语1班", "product_id": "P1"},
                {"id": "C2", "name": "数学2班", "product_id": "P2"},
            ],
            "products": [
                {"id": "P1", "name": "英语产品"},
                {"id": "P2", "name": "数学产品"},
            ],
            "product_courses": [],
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = write_missing_teacher_template(Path(tmp), "20260628_010000", error, state=state)
            self.assertIsNotNone(path)
            with Path(path).open(encoding="utf-8") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[0]["class_id"], "C1")
        self.assertEqual(rows[0]["class_name"], "英语1班")
        self.assertEqual(rows[0]["product_id"], "P1")
        self.assertEqual(rows[0]["product_name"], "英语产品")
        self.assertEqual(rows[0]["class_schedule_mode"], "本班实际排课")
        self.assertEqual(rows[0]["actual_scheduled_class_id"], "C1")
        self.assertIn("assignment_extra_time_requirement", rows[0])
        self.assertNotIn("schedule_mode", rows[0])
        self.assertNotIn("inherit_from_class_id", rows[0])
        self.assertNotIn("teacher_available_slots", rows[0])
        self.assertEqual(rows[1]["subject"], "政治")
        self.assertEqual(rows[2]["class_id"], "C2")
        self.assertEqual(rows[2]["product_name"], "数学产品")

    def test_missing_teacher_rows_template_writes_precomputed_rows(self) -> None:
        rows = [
            {
                "class_id": "C1",
                "class_name": "英语1班",
                "product_id": "P1",
                "product_name": "英语产品",
                "subject": "英语",
                "stage": "基础",
                "course_group": "阅读类",
                "class_schedule_mode": "本班实际排课",
                "actual_scheduled_class_id": "C1",
                "teacher_id": "",
                "teacher_name": "",
                "assignment_extra_time_requirement": "",
                "notes": "上传前校验自动生成，请补齐老师后重新上传",
            }
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = write_missing_teacher_rows_template(Path(tmp), "20260628_020000", rows)
            self.assertIsNotNone(path)
            with Path(path).open(encoding="utf-8") as handle:
                written_rows = list(csv.DictReader(handle))

        self.assertEqual(written_rows, rows)

    def test_missing_teacher_requirements_parse_current_and_legacy_errors_once(self) -> None:
        error = (
            "班级 C1 的产品 P1 缺少课程老师安排: 英语/基础/阅读类、英语/基础/阅读类\n"
            "班级 C1 的产品 P2 缺少课程老师安排: 英语/基础/阅读类\n"
            "班级 C2 缺少 数学/强化/数学类 的老师安排"
        )

        requirements = parse_missing_teacher_requirements(error)

        self.assertEqual(len(requirements), 3)
        self.assertEqual(requirements[0].class_id, "C1")
        self.assertEqual(requirements[0].product_id, "P1")
        self.assertEqual(requirements[0].subject, "英语")
        self.assertEqual(requirements[1].class_id, "C1")
        self.assertEqual(requirements[1].product_id, "P2")
        self.assertEqual(requirements[1].subject, "英语")
        self.assertEqual(requirements[2].class_id, "C2")
        self.assertEqual(requirements[2].product_id, "")

    def test_missing_teacher_rows_are_derived_from_structured_requirements(self) -> None:
        requirements = parse_missing_teacher_requirements(
            "班级 C1 的产品 P1 缺少课程老师安排: 英语/基础/阅读类\n"
            "班级 C1 的产品 P2 缺少课程老师安排: 英语/基础/阅读类"
        )
        state = {
            "classes": [{"id": "C1", "name": "英语1班", "product_id": "P1"}],
            "products": [
                {"id": "P1", "name": "英语产品"},
                {"id": "P2", "name": "英语导学产品"},
            ],
            "product_courses": [],
        }

        rows = missing_teacher_rows_for_requirements(requirements, state)

        self.assertEqual([row["product_id"] for row in rows], ["P1", "P2"])
        self.assertEqual([row["product_name"] for row in rows], ["英语产品", "英语导学产品"])
        self.assertTrue(all(row["class_name"] == "英语1班" for row in rows))

    def test_teaching_area_short_name_is_normalized_and_used_for_rooms(self) -> None:
        state = data_admin_server.normalize_payload(
            {
                "teaching_areas": [{"id": "A1", "name": "蜀山-环球金融校区（安农大）"}],
                "rooms": [{"id": "R1", "name": "101", "teaching_area_id": "A1", "capacity": 80, "is_active": "是"}],
                "teachers": [],
                "products": [],
                "product_courses": [],
                "product_schedule_rules": [],
                "classes": [],
                "teaching_area_links": [],
                "global_blackout_dates": [],
            }
        )

        self.assertEqual(state["teaching_areas"][0]["short_name"], "环球金融")
        self.assertEqual(state["teaching_areas"][0]["region_tag"], "蜀山")
        self.assertEqual(state["rooms"][0]["teaching_area_name"], "环球金融")

    def test_scheduler_room_parser_uses_area_metadata_and_legacy_area_resources(self) -> None:
        rooms, has_explicit_rooms = scheduler.parse_rooms(
            {
                "teaching_areas": [
                    {
                        "id": "A1",
                        "name": "环球金融校区",
                        "short_name": "环球金融",
                        "region_tag": "蜀山",
                    }
                ],
                "rooms": [
                    {
                        "id": "R1",
                        "name": "101",
                        "teaching_area_id": "A1",
                        "capacity": 80,
                        "capacity_unlimited": "否",
                    }
                ],
            }
        )
        legacy_rooms, legacy_has_explicit_rooms = scheduler.parse_rooms(
            {"teaching_areas": [{"id": "A_LEGACY", "name": "旧版资源", "region_tag": "包河"}]}
        )

        self.assertTrue(has_explicit_rooms)
        self.assertEqual(rooms["R1"].teaching_area_id, "A1")
        self.assertEqual(rooms["R1"].teaching_area_name, "环球金融")
        self.assertEqual(rooms["R1"].region_tag, "蜀山")
        self.assertFalse(rooms["R1"].capacity_unlimited)
        self.assertFalse(legacy_has_explicit_rooms)
        self.assertEqual(legacy_rooms["A_LEGACY"].teaching_area_id, "A_LEGACY")
        self.assertEqual(legacy_rooms["A_LEGACY"].teaching_area_name, "旧版资源")

    def test_suite_code_is_inferred_from_class_name_when_empty(self) -> None:
        inferred = data_admin_server.normalize_class({"id": "C1", "name": "考研英语寒暑集训营（27届01班）"})
        preserved = data_admin_server.normalize_class({"id": "C2", "name": "考研英语寒暑集训营（27届01班）", "suite_code": "MANUAL"})

        self.assertEqual(inferred["suite_code"], "2701")
        self.assertEqual(preserved["suite_code"], "MANUAL")

    def test_preferred_room_can_expand_to_same_teaching_area_unless_required(self) -> None:
        state = data_admin_server.normalize_payload(
            {
                "teaching_areas": [{"id": "A1", "name": "主校区"}, {"id": "A2", "name": "分校区"}],
                "rooms": [
                    {"id": "R1", "name": "101", "teaching_area_id": "A1", "capacity": 60, "is_active": "是"},
                    {"id": "R2", "name": "102", "teaching_area_id": "A1", "capacity": 60, "is_active": "是"},
                    {"id": "R3", "name": "201", "teaching_area_id": "A2", "capacity": 60, "is_active": "是"},
                ],
                "teachers": [],
                "products": [{"id": "P1", "name": "考研英语", "subject": "英语"}],
                "product_courses": [
                    {
                        "product_id": "P1",
                        "product_name": "考研英语",
                        "subject": "英语",
                        "stage": "基础",
                        "course_module": "词汇",
                        "course_group": "阅读类",
                        "total_hours": 2,
                        "block_hours": 2,
                        "teaching_area_ids": "A1|A2",
                    }
                ],
                "product_schedule_rules": [],
                "classes": [
                    {
                        "id": "C1",
                        "name": "英语1班",
                        "product_id": "P1",
                        "subject": "英语",
                        "stages": "基础",
                        "size": 30,
                        "preferred_room_ids": "R1",
                    }
                ],
                "teaching_area_links": [],
                "class_conflict_groups": [],
                "global_blackout_dates": [],
            }
        )
        time_slots = [
            {
                "id": "2026-07-01-AM-1",
                "date": "2026-07-01",
                "period": "AM",
                "name": "上午一",
                "order": 1,
                "start_time": "08:00",
                "end_time": "10:00",
                "duration_hours": 2,
            }
        ]

        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            relaxed_export = data_admin_server.export_scheduler_input(state, time_slots=time_slots)
            relaxed_input = json.loads(Path(relaxed_export["path"]).read_text(encoding="utf-8"))
            self.assertEqual(set(relaxed_input["classes"][0]["room_ids"]), {"R1", "R2"})

            state["classes"][0]["preferred_room_is_required"] = True
            fixed_export = data_admin_server.export_scheduler_input(state, time_slots=time_slots)
            fixed_input = json.loads(Path(fixed_export["path"]).read_text(encoding="utf-8"))
            self.assertEqual(fixed_input["classes"][0]["room_ids"], ["R1"])

    def test_export_scheduler_input_uses_class_window_room_when_window_exists(self) -> None:
        state = data_admin_server.normalize_payload(
            {
                "teaching_areas": [{"id": "A1", "name": "默认校区"}, {"id": "A2", "name": "暑假校区"}],
                "rooms": [
                    {"id": "R1", "name": "默认101", "teaching_area_id": "A1", "capacity": 60, "is_active": "是"},
                    {"id": "R2", "name": "暑假201", "teaching_area_id": "A2", "capacity": 60, "is_active": "是"},
                ],
                "teachers": [{"employee_id": "T1", "name": "张老师"}],
                "products": [{"id": "P1", "name": "考研暑假营-英语"}],
                "product_courses": [
                    {
                        "product_id": "P1",
                        "product_name": "考研暑假营-英语",
                        "subject": "英语",
                        "window_name": "暑假",
                        "stage": "基础",
                        "course_module": "词汇",
                        "course_group": "阅读类",
                        "total_hours": 2,
                        "block_hours": 2,
                    }
                ],
                "product_schedule_rules": [],
                "classes": [
                    {
                        "id": "C1",
                        "name": "暑假班",
                        "product_id": "P1",
                        "subject": "英语",
                        "size": 30,
                        "start_date": "2026-07-01",
                        "start_period": "AM",
                        "end_date": "2026-07-10",
                        "end_period": "PM",
                        "preferred_room_ids": "R1",
                        "preferred_room_is_required": "是",
                        "teacher_assignments": [
                            {
                                "subject": "英语",
                                "stage": "基础",
                                "course_group": "阅读类",
                                "teacher_id": "T1",
                                "teacher_name": "张老师",
                            }
                        ],
                    }
                ],
                "class_window_boundaries": [
                    {
                        "class_window_id": "C1_2026暑假",
                        "class_id": "C1",
                        "schedule_window_id": "2026暑假",
                        "season_window_id": "WINDOW_SUMMER",
                        "season_name": "暑假",
                        "earliest_date": "2026-07-05",
                        "earliest_period": "AM",
                        "latest_date": "2026-07-10",
                        "latest_period": "PM",
                        "preferred_teaching_area_ids": "A2",
                        "preferred_room_ids": "",
                        "preferred_room_is_required": "是",
                        "is_class_window_included": "是",
                    }
                ],
                "teaching_area_links": [],
                "class_conflict_groups": [],
                "global_blackout_dates": [],
            }
        )

        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            export = data_admin_server.export_scheduler_input(
                state,
                time_slots=[
                    {
                        "id": "S_SUMMER",
                        "date": "2026-07-06",
                        "period": "AM",
                        "name": "上午",
                        "order": 1,
                        "duration_hours": 2,
                        "schedule_window_id": "2026暑假",
                        "season_window_id": "WINDOW_SUMMER",
                        "season_name": "暑假",
                    }
                ],
            )
            scheduler_input = json.loads(Path(export["path"]).read_text(encoding="utf-8"))

        self.assertNotIn("room_ids", scheduler_input["classes"][0])
        self.assertEqual(scheduler_input["class_window_boundaries"][0]["room_ids"], ["R2"])
        assignments = scheduler.schedule(scheduler.load_input_data(scheduler_input))
        self.assertEqual(assignments[0].candidate.room_id, "R2")

    def test_class_window_area_without_active_rooms_is_validation_error(self) -> None:
        payload = {
            "teaching_areas": [{"id": "A2", "name": "暑假校区"}],
            "rooms": [
                {"id": "R_OFF", "name": "停用教室", "teaching_area_id": "A2", "capacity": 60, "is_active": "否"},
            ],
            "products": [{"id": "P1", "name": "考研暑假营-英语"}],
            "classes": [{"id": "C1", "name": "暑假班", "product_id": "P1", "subject": "英语"}],
            "class_window_boundaries": [
                {
                    "class_window_id": "C1_2026暑假",
                    "class_id": "C1",
                    "schedule_window_id": "2026暑假",
                    "season_name": "暑假",
                    "earliest_date": "2026-07-05",
                    "latest_date": "2026-07-10",
                    "preferred_teaching_area_ids": "A2",
                    "preferred_room_ids": "",
                    "is_class_window_included": "是",
                }
            ],
        }

        with self.assertRaisesRegex(ValueError, "没有启用教室"):
            data_admin_server.normalize_payload(payload)

    def test_class_window_disabled_room_is_validation_error(self) -> None:
        payload = {
            "teaching_areas": [{"id": "A2", "name": "暑假校区"}],
            "rooms": [
                {"id": "R_OFF", "name": "停用教室", "teaching_area_id": "A2", "capacity": 60, "is_active": "否"},
            ],
            "products": [{"id": "P1", "name": "考研暑假营-英语"}],
            "classes": [{"id": "C1", "name": "暑假班", "product_id": "P1", "subject": "英语"}],
            "class_window_boundaries": [
                {
                    "class_window_id": "C1_2026暑假",
                    "class_id": "C1",
                    "schedule_window_id": "2026暑假",
                    "season_name": "暑假",
                    "earliest_date": "2026-07-05",
                    "latest_date": "2026-07-10",
                    "preferred_room_ids": "R_OFF",
                    "is_class_window_included": "是",
                }
            ],
        }

        with self.assertRaisesRegex(ValueError, "未启用"):
            data_admin_server.normalize_payload(payload)

    def test_scheduler_blocks_class_window_area_constraint_with_no_rooms(self) -> None:
        payload = {
            "time_slots": [
                {"id": "S1", "date": "2026-07-06", "period": "AM", "name": "上午", "order": 1, "duration_hours": 2},
            ],
            "rooms": [
                {"id": "R1", "name": "默认101", "teaching_area_id": "A1", "capacity": 60},
            ],
            "products": [
                {
                    "id": "P1",
                    "name": "英语产品",
                    "requirements": [
                        {
                            "subject": "英语",
                            "stage": "基础",
                            "course_group": "阅读类",
                            "total_hours": 2,
                            "block_hours": 2,
                        }
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C1",
                    "name": "暑假班",
                    "product_id": "P1",
                    "subject": "英语",
                    "teacher_assignments": [
                        {"subject": "英语", "stage": "基础", "course_group": "阅读类", "teacher_id": "T1", "teacher_name": "张老师"}
                    ],
                }
            ],
            "class_window_boundaries": [
                {
                    "class_window_id": "C1_2026暑假",
                    "class_id": "C1",
                    "preferred_teaching_area_ids": ["A2"],
                    "is_class_window_included": True,
                }
            ],
        }

        schedule_input = scheduler.load_input_data(payload)
        constraint = schedule_input.class_window_constraints["C1"][0]
        self.assertTrue(constraint.has_room_constraint)
        self.assertIsNone(constraint.room_ids)
        task = scheduler.build_course_blocks(schedule_input.classes)[0]
        self.assertEqual(scheduler.candidate_assignments(task, schedule_input), [])

    def test_candidate_assignments_applies_class_window_rooms_and_capacity_order(self) -> None:
        slot = scheduler.TimeSlot(
            "S1",
            "2026-07-06",
            "AM",
            "上午",
            1,
            duration_hours=2,
            schedule_window_id="2026暑假",
            season_name="暑假",
        )
        cls = scheduler.SchoolClass(
            id="C1",
            name="暑假班",
            product_id="P1",
            product_name="暑假产品",
            size=50,
            room_ids=None,
            start_date=None,
            start_period=None,
            end_date=None,
            end_period=None,
            first_lesson_date=None,
            first_lesson_period=None,
            stage_order={},
            requirements=[],
        )
        task = scheduler.CourseBlock(
            task_id="T1",
            class_id="C1",
            class_name="暑假班",
            product_id="P1",
            product_name="暑假产品",
            class_size=50,
            subject_category="公共课",
            subject="英语",
            quarter=None,
            stage="基础",
            course_module=None,
            course_group=None,
            teacher_id="TEACHER1",
            teacher_name="张老师",
            block_hours=2,
            room_ids={"R1", "R2", "R3"},
            start_date=None,
            end_date=None,
            allowed_periods=None,
            allowed_weekdays=None,
            excluded_weekdays=None,
            schedule_rules=(),
        )
        schedule_input = scheduler.ScheduleInput(
            time_slots=[slot],
            rooms={
                "R1": scheduler.Room("R1", capacity=40),
                "R2": scheduler.Room("R2", capacity=80),
                "R3": scheduler.Room("R3", capacity=120),
            },
            classes={"C1": cls},
            conflict_groups={},
            class_conflict_groups={"C1": set()},
            locked_assignments=[],
            class_window_constraints={
                "C1": [
                    scheduler.ClassWindowConstraint(
                        class_id="C1",
                        start_date=None,
                        start_period=None,
                        end_date=None,
                        end_period=None,
                        schedule_window_id="2026暑假",
                        season_window_id=None,
                        season_name=None,
                        room_ids={"R1", "R2"},
                        has_room_constraint=True,
                    )
                ]
            },
        )

        candidates = scheduler.candidate_assignments(task, schedule_input, [(slot,)])

        self.assertEqual([candidate.room_id for candidate in candidates], ["R2", "R1"])

    def test_candidate_assignments_filters_slot_blocks_and_keeps_room_order(self) -> None:
        too_early_slot = scheduler.TimeSlot("S0", "2026-07-05", "AM", "上午一", 1, duration_hours=2)
        usable_slot = scheduler.TimeSlot("S1", "2026-07-06", "AM", "上午一", 1, duration_hours=2)
        unavailable_slot = scheduler.TimeSlot("S2", "2026-07-07", "AM", "上午一", 1, duration_hours=2)
        cls = scheduler.SchoolClass(
            id="C1",
            name="暑假班",
            product_id="P1",
            product_name="暑假产品",
            size=50,
            room_ids=None,
            start_date="2026-07-06",
            start_period="AM",
            end_date="2026-07-07",
            end_period="AM",
            first_lesson_date=None,
            first_lesson_period=None,
            stage_order={},
            requirements=[],
        )
        task = scheduler.CourseBlock(
            task_id="T1",
            class_id="C1",
            class_name="暑假班",
            product_id="P1",
            product_name="暑假产品",
            class_size=50,
            subject_category="公共课",
            subject="英语",
            quarter=None,
            stage="基础",
            course_module=None,
            course_group=None,
            teacher_id="TEACHER1",
            teacher_name="张老师",
            block_hours=2,
            room_ids={"R1", "R2"},
            start_date=None,
            end_date=None,
            allowed_periods={"AM"},
            allowed_weekdays=None,
            excluded_weekdays=None,
            schedule_rules=(),
        )
        schedule_input = scheduler.ScheduleInput(
            time_slots=[too_early_slot, usable_slot, unavailable_slot],
            rooms={
                "R1": scheduler.Room("R1", capacity=40),
                "R2": scheduler.Room("R2", capacity=80),
            },
            classes={"C1": cls},
            conflict_groups={},
            class_conflict_groups={"C1": set()},
            locked_assignments=[],
            teacher_unavailability={
                "TEACHER1": [
                    scheduler.TeacherUnavailableRule(
                        teacher_id="TEACHER1",
                        start_date="2026-07-07",
                        end_date="2026-07-07",
                        weekdays=None,
                        periods={"AM"},
                        schedule_window_ids=None,
                    )
                ]
            },
        )

        candidates = scheduler.candidate_assignments(
            task,
            schedule_input,
            [(too_early_slot,), (usable_slot,), (unavailable_slot,)],
        )

        self.assertEqual([candidate.slots[0].id for candidate in candidates], ["S1", "S1"])
        self.assertEqual([candidate.room_id for candidate in candidates], ["R2", "R1"])

    def test_hanshuying_export_keeps_class_room_and_ignores_product_course_area(self) -> None:
        state = data_admin_server.normalize_payload(
            {
                "teaching_areas": [{"id": "A1", "name": "寒假校区"}, {"id": "A2", "name": "暑假校区"}],
                "rooms": [
                    {"id": "R1", "name": "寒假101", "teaching_area_id": "A1", "capacity": 60, "is_active": "是"},
                    {"id": "R2", "name": "暑假201", "teaching_area_id": "A2", "capacity": 60, "is_active": "是"},
                ],
                "teachers": [],
                "products": [{"id": "P_HSY", "name": "考研寒暑营-英语", "sub_product": "寒暑营"}],
                "product_courses": [
                    {
                        "product_id": "P_HSY",
                        "product_name": "考研寒暑营-英语",
                        "subject": "英语",
                        "stage": "寒假",
                        "course_module": "词汇",
                        "course_group": "阅读类",
                        "total_hours": 4,
                        "block_hours": 4,
                        "teaching_area_ids": "A1",
                    },
                    {
                        "product_id": "P_HSY",
                        "product_name": "考研寒暑营-英语",
                        "subject": "英语",
                        "stage": "暑假",
                        "course_module": "阅读",
                        "course_group": "阅读类",
                        "total_hours": 4,
                        "block_hours": 4,
                        "teaching_area_ids": "A2",
                    },
                ],
                "product_schedule_rules": [],
                "classes": [
                    {
                        "id": "C_HSY",
                        "name": "考研英语寒暑集训营（27届01班）",
                        "product_id": "P_HSY",
                        "subject": "英语",
                        "size": 30,
                        "preferred_room_ids": "R1",
                    }
                ],
                "teaching_area_links": [],
                "class_conflict_groups": [],
                "global_blackout_dates": [],
            }
        )
        self.assertNotIn("teaching_area_ids", state["product_courses"][0])
        self.assertNotIn("teaching_area_ids", state["product_courses"][1])

        with tempfile.TemporaryDirectory() as tmp:
            data_admin_server.DATA_DIR = Path(tmp) / "data"
            export = data_admin_server.export_scheduler_input(state, time_slots=[])
            scheduler_input = json.loads(Path(export["path"]).read_text(encoding="utf-8"))

        cls = scheduler_input["classes"][0]
        self.assertEqual(cls["room_ids"], ["R1"])
        product_requirements = {
            item["stage"]: item
            for item in scheduler_input["products"][0]["requirements"]
        }
        self.assertNotIn("room_ids", product_requirements["寒假"])
        self.assertNotIn("room_ids", product_requirements["暑假"])

    def test_scheduler_ignores_product_teaching_area_fields(self) -> None:
        payload = {
            "time_slots": [
                {"id": "S1", "date": "2026-07-01", "period": "AM", "name": "上午一", "order": 1, "start_time": "08:00", "end_time": "10:00", "duration_hours": 2},
                {"id": "S2", "date": "2026-07-01", "period": "AM", "name": "上午二", "order": 2, "start_time": "10:20", "end_time": "12:20", "duration_hours": 2},
            ],
            "rooms": [
                {"id": "R1", "capacity": 60},
                {"id": "R2", "capacity": 60},
            ],
            "products": [
                {
                    "id": "P_HSY",
                    "name": "考研寒暑营-英语",
                    "teaching_area_ids": ["R2"],
                    "requirements": [
                        {
                            "subject_category": "公共课",
                            "subject": "英语",
                            "stage": "暑假",
                            "course_module": "阅读",
                            "course_group": "阅读类",
                            "total_hours": 4,
                            "block_hours": 4,
                            "teaching_area_ids": ["R2"],
                        }
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C_HSY",
                    "name": "考研寒暑营英语班",
                    "product_id": "P_HSY",
                    "subject": "英语",
                    "size": 30,
                    "preferred_room_ids": ["R1"],
                    "teacher_assignments": [
                        {"subject": "英语", "stage": "暑假", "course_group": "阅读类", "teacher_id": "T1", "teacher_name": "张老师"}
                    ],
                }
            ],
        }

        schedule_input = scheduler.load_input_data(payload)
        requirement = schedule_input.classes["C_HSY"].requirements[0]
        self.assertEqual(requirement.room_ids, {"R1"})
        assignments = scheduler.schedule(schedule_input)
        self.assertEqual(assignments[0].candidate.room_id, "R1")

    def test_scheduler_does_not_treat_area_fields_as_room_ids_when_rooms_are_explicit(self) -> None:
        payload = {
            "time_slots": [
                {"id": "S1", "date": "2026-07-01", "period": "AM", "name": "上午一", "order": 1, "duration_hours": 2},
            ],
            "rooms": [
                {"id": "R1", "name": "101", "teaching_area_id": "A1", "capacity": 60},
                {"id": "R2", "name": "201", "teaching_area_id": "A2", "capacity": 60},
            ],
            "products": [
                {
                    "id": "P1",
                    "name": "英语产品",
                    "requirements": [
                        {
                            "subject": "英语",
                            "stage": "基础",
                            "course_group": "阅读类",
                            "total_hours": 2,
                            "block_hours": 2,
                            "room_ids": ["R1", "R2"],
                        }
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C_PRODUCT",
                    "name": "产品班",
                    "product_id": "P1",
                    "subject": "英语",
                    "preferred_teaching_area_ids": ["A1"],
                    "teacher_assignments": [
                        {"subject": "英语", "stage": "基础", "course_group": "阅读类", "teacher_id": "T1", "teacher_name": "张老师"}
                    ],
                },
                {
                    "id": "C_DIRECT",
                    "name": "直填班",
                    "subject": "数学",
                    "requirements": [
                        {
                            "subject": "数学",
                            "stage": "基础",
                            "course_group": "高数",
                            "teacher_id": "T2",
                            "teacher_name": "李老师",
                            "total_hours": 2,
                            "block_hours": 2,
                            "teaching_area_ids": ["A1"],
                        }
                    ],
                },
            ],
        }

        schedule_input = scheduler.load_input_data(payload)

        product_requirement = schedule_input.classes["C_PRODUCT"].requirements[0]
        direct_requirement = schedule_input.classes["C_DIRECT"].requirements[0]
        self.assertIsNone(schedule_input.classes["C_PRODUCT"].room_ids)
        self.assertEqual(product_requirement.room_ids, {"R1", "R2"})
        self.assertIsNone(direct_requirement.room_ids)

    def test_sync_class_teacher_assignments_fills_product_courses(self) -> None:
        state = {
            "products": [{"id": "P1", "name": "考研寒暑营-英语"}],
            "product_courses": [
                {"product_id": "P1", "product_name": "考研寒暑营-英语", "subject": "英语", "stage": "基础", "course_module": "词汇", "course_group": "阅读类"},
                {"product_id": "P1", "product_name": "考研寒暑营-英语", "subject": "英语", "stage": "强化", "course_module": "阅读", "course_group": "阅读类"},
                {"product_id": "P1", "product_name": "考研寒暑营-英语", "subject": "英语", "stage": "冲刺", "course_module": "写作", "course_group": "写作类"},
            ],
            "classes": [
                {
                    "id": "C1",
                    "name": "测试班",
                    "product_id": "P1",
                    "subject": "英语",
                    "stages": [],
                    "teacher_assignments": [
                        {"product_id": "P1", "subject": "英语", "stage": "基础", "course_group": "阅读类", "teacher_id": "T1", "teacher_name": "张老师"}
                    ],
                }
            ],
        }

        stats = data_admin_server.sync_class_teacher_assignments(state)

        self.assertEqual(stats, {"classes": 1, "assignments": 3})
        assignments = {
            (item["stage"], item["course_group"]): item
            for item in state["classes"][0]["teacher_assignments"]
        }
        self.assertEqual(assignments[("基础", "阅读类")]["teacher_id"], "T1")
        self.assertEqual(assignments[("强化", "阅读类")]["teacher_id"], "T1")
        self.assertNotIn("teacher_id", assignments[("冲刺", "写作类")])
        self.assertNotIn("course_module", assignments[("基础", "阅读类")])

    def test_sync_class_teacher_assignments_preserves_shared_mode(self) -> None:
        state = {
            "products": [{"id": "P1", "name": "考研寒暑营-英语"}],
            "product_courses": [
                {"product_id": "P1", "product_name": "考研寒暑营-英语", "subject": "英语", "stage": "春季", "course_module": "阅读", "course_group": "阅读类"},
                {"product_id": "P1", "product_name": "考研寒暑营-英语", "subject": "英语", "stage": "暑假", "course_module": "阅读", "course_group": "阅读类"},
            ],
            "classes": [
                {
                    "id": "C1",
                    "name": "共享班",
                    "product_id": "P1",
                    "subject": "英语",
                    "stages": [],
                    "teacher_assignments": [
                        {"product_id": "P1", "subject": "英语", "stage": "春季", "course_group": "阅读类", "schedule_mode": "共享课表", "inherit_from_class_id": "C_MAIN", "teacher_id": "OLD", "teacher_name": "旧老师"},
                        {"product_id": "P1", "subject": "英语", "stage": "暑假", "course_group": "阅读类", "teacher_id": "T1", "teacher_name": "张老师"},
                    ],
                }
            ],
        }

        data_admin_server.sync_class_teacher_assignments(state)

        assignments = {
            item["stage"]: item
            for item in state["classes"][0]["teacher_assignments"]
        }
        self.assertEqual(assignments["春季"]["class_schedule_mode"], "共享实际排课班级")
        self.assertEqual(assignments["春季"]["actual_scheduled_class_id"], "C_MAIN")
        self.assertNotIn("teacher_id", assignments["春季"])
        self.assertEqual(assignments["暑假"]["class_schedule_mode"], "本班实际排课")
        self.assertEqual(assignments["暑假"]["teacher_id"], "T1")

    def test_sync_class_teacher_assignments_preserves_teachers_after_product_change(self) -> None:
        state = {
            "products": [{"id": "P_OLD", "name": "考研无忧春-正课-英语"}, {"id": "P_NEW", "name": "考研无忧秋-正课-英语"}],
            "product_courses": [
                {"product_id": "P_NEW", "product_name": "考研无忧秋-正课-英语", "subject": "英语", "stage": "基础", "course_module": "词汇", "course_group": "阅读类"},
                {"product_id": "P_NEW", "product_name": "考研无忧秋-正课-英语", "subject": "英语", "stage": "强化", "course_module": "阅读", "course_group": "阅读类"},
            ],
            "classes": [
                {
                    "id": "C1",
                    "name": "测试班",
                    "product_id": "P_NEW",
                    "subject": "英语",
                    "stages": [],
                    "teacher_assignments": [
                        {"product_id": "P_OLD", "subject": "英语", "stage": "基础", "course_group": "阅读类", "teacher_id": "T1", "teacher_name": "张老师"}
                    ],
                }
            ],
        }

        data_admin_server.sync_class_teacher_assignments(state)

        assignments = {
            (item["stage"], item["course_group"]): item
            for item in state["classes"][0]["teacher_assignments"]
        }
        self.assertEqual(assignments[("基础", "阅读类")]["product_id"], "P_NEW")
        self.assertEqual(assignments[("基础", "阅读类")]["teacher_id"], "T1")
        self.assertEqual(assignments[("强化", "阅读类")]["teacher_id"], "T1")

    def test_blank_teacher_assignment_placeholder_is_not_schedulable(self) -> None:
        payload = {
            "time_slots": [
                {
                    "id": "2026-07-01-AM-1",
                    "date": "2026-07-01",
                    "period": "AM",
                    "name": "上午一",
                    "order": 1,
                    "start_time": "08:00",
                    "end_time": "12:00",
                    "duration_hours": 4,
                }
            ],
            "rooms": [{"id": "R1", "capacity": 50}],
            "products": [
                {
                    "id": "P1",
                    "name": "英语产品",
                    "requirements": [
                        {
                            "subject": "英语",
                            "stage": "基础",
                            "course_module": "阅读",
                            "course_group": "阅读类",
                            "total_hours": 4,
                            "block_hours": 4,
                            "room_ids": ["R1"],
                        }
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C1",
                    "name": "英语班",
                    "product_id": "P1",
                    "subject": "英语",
                    "size": 20,
                    "teacher_assignments": [
                        {
                            "subject": "英语",
                            "stage": "基础",
                            "course_group": "阅读类",
                            "teacher_id": "",
                            "teacher_name": "",
                        }
                    ],
                },
                {
                    "id": "C2",
                    "name": "英语班2",
                    "product_id": "P1",
                    "subject": "英语",
                    "size": 20,
                    "teacher_assignments": [],
                }
            ],
        }

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "input.json"
            path.write_text(json.dumps(payload), encoding="utf-8")
            with self.assertRaises(ValueError) as context:
                scheduler.load_input(path)
            message = str(context.exception)

        self.assertIn("班级数据校验失败", message)
        self.assertIn("班级 C1 的产品 P1 缺少课程老师安排: 英语/基础/阅读类", message)
        self.assertIn("班级 C2 的产品 P1 缺少课程老师安排: 英语/基础/阅读类", message)

    def test_shared_teacher_assignment_skips_class_course_requirement(self) -> None:
        payload = {
            "time_slots": [
                {
                    "id": "2026-07-01-AM-1",
                    "date": "2026-07-01",
                    "period": "AM",
                    "name": "上午一",
                    "order": 1,
                    "start_time": "08:00",
                    "end_time": "12:00",
                    "duration_hours": 4,
                },
                {
                    "id": "2026-07-02-AM-1",
                    "date": "2026-07-02",
                    "period": "AM",
                    "name": "上午一",
                    "order": 1,
                    "start_time": "08:00",
                    "end_time": "12:00",
                    "duration_hours": 4,
                },
            ],
            "rooms": [{"id": "R1", "capacity": 50}],
            "products": [
                {
                    "id": "P1",
                    "name": "寒暑营英语",
                    "requirements": [
                        {
                            "subject": "英语",
                            "stage": "春季",
                            "course_module": "阅读",
                            "course_group": "阅读类",
                            "total_hours": 4,
                            "block_hours": 4,
                            "room_ids": ["R1"],
                        },
                        {
                            "subject": "英语",
                            "stage": "暑假",
                            "course_module": "阅读",
                            "course_group": "阅读类",
                            "total_hours": 4,
                            "block_hours": 4,
                            "room_ids": ["R1"],
                        },
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C_SHARE",
                    "name": "共享班",
                    "product_id": "P1",
                    "subject": "英语",
                    "stages": ["春季", "暑假"],
                    "size": 20,
                    "teacher_assignments": [
                        {
                            "subject": "英语",
                            "stage": "春季",
                            "course_group": "阅读类",
                            "schedule_mode": "共享课表",
                            "inherit_from_class_id": "C_MAIN",
                        },
                        {
                            "subject": "英语",
                            "stage": "暑假",
                            "course_group": "阅读类",
                            "teacher_id": "T1",
                            "teacher_name": "张老师",
                        },
                    ],
                }
            ],
        }

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "input.json"
            path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
            assignments = scheduler.schedule(scheduler.load_input(path))

        self.assertEqual(len(assignments), 1)
        self.assertEqual(assignments[0].task.stage, "暑假")

    def test_online_room_capacity_can_be_unlimited(self) -> None:
        payload = {
            "time_slots": [
                {
                    "id": "2026-07-01-EVENING-1",
                    "date": "2026-07-01",
                    "period": "EVENING",
                    "name": "晚上",
                    "order": 1,
                    "start_time": "19:00",
                    "end_time": "21:00",
                    "duration_hours": 2,
                }
            ],
            "rooms": [{"id": "R_ONLINE", "capacity": 1, "capacity_unlimited": True}],
            "products": [
                {
                    "id": "P1",
                    "name": "线上产品",
                    "requirements": [
                        {
                            "subject": "英语",
                            "stage": "春季",
                            "course_module": "阅读",
                            "course_group": "阅读类",
                            "total_hours": 2,
                            "block_hours": 2,
                            "room_ids": ["R_ONLINE"],
                        }
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C_ONLINE",
                    "name": "线上班",
                    "product_id": "P1",
                    "subject": "英语",
                    "size": 80,
                    "teacher_assignments": [
                        {
                            "subject": "英语",
                            "stage": "春季",
                            "course_group": "阅读类",
                            "teacher_id": "T1",
                            "teacher_name": "张老师",
                        }
                    ],
                }
            ],
        }

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "input.json"
            path.write_text(json.dumps(payload), encoding="utf-8")
            schedule_input = scheduler.load_input(path)
            assignments = scheduler.schedule(schedule_input)

        self.assertEqual(len(assignments), 1)
        self.assertEqual(assignments[0].candidate.room_id, "R_ONLINE")

    def test_scheduler_uses_course_group_teacher_and_first_stage_fallback(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "incoming"
            data_dir = root / "data"
            output_dir = root / "outputs"
            write_minimal_csv_source(source)
            write_csv(
                source / "product_courses.csv",
                [
                    "product_id",
                    "product_name",
                    "subject_category",
                    "subject",
                    "stage",
                    "course_module",
                    "course_group",
                    "total_hours",
                    "block_hours",
                    "teaching_area_ids",
                ],
                [
                    {
                        "product_id": "P1",
                        "product_name": "考研寒暑营-英语",
                        "subject_category": "公共课",
                        "subject": "英语",
                        "stage": "基础",
                        "course_module": "词汇",
                        "course_group": "阅读类",
                        "total_hours": 4,
                        "block_hours": 4,
                        "teaching_area_ids": "A1",
                    },
                    {
                        "product_id": "P1",
                        "product_name": "考研寒暑营-英语",
                        "subject_category": "公共课",
                        "subject": "英语",
                        "stage": "强化",
                        "course_module": "阅读",
                        "course_group": "阅读类",
                        "total_hours": 4,
                        "block_hours": 4,
                        "teaching_area_ids": "A1",
                    },
                ],
            )
            write_csv(
                source / "classes.csv",
                [
                    "id",
                    "name",
                    "product_id",
                    "subject",
                    "stages",
                    "exam_season",
                    "suite_code",
                    "size",
                    "start_date",
                    "start_period",
                    "end_date",
                    "end_period",
                    "preferred_teaching_area_ids",
                ],
                [
                    {
                        "id": "C1",
                        "name": "英语1班",
                        "product_id": "P1",
                        "subject": "英语",
                        "stages": "基础|强化",
                        "exam_season": "27考研",
                        "suite_code": "",
                        "size": 30,
                        "start_date": "2026-07-01",
                        "start_period": "",
                        "end_date": "",
                        "end_period": "",
                        "preferred_teaching_area_ids": "A1",
                    }
                ],
            )
            write_csv(
                source / "class_teacher_assignments.csv",
                ["class_id", "product_id", "subject", "stage", "course_group", "teacher_id", "teacher_name"],
                [
                    {
                        "class_id": "C1",
                        "product_id": "P1",
                        "subject": "英语",
                        "stage": "基础",
                        "course_group": "阅读类",
                        "teacher_id": "T1",
                        "teacher_name": "张老师",
                    }
                ],
            )

            result = run_pipeline(pipeline_args(source, data_dir, output_dir))
            schedule_csv = result.schedule_csv_path.read_text(encoding="utf-8")

            self.assertIn("基础,词汇,阅读类,T1", schedule_csv)
            self.assertIn("强化,阅读,阅读类,T1", schedule_csv)

    def test_scheduler_groups_short_modules_by_course_group_teacher_and_block_hours(self) -> None:
        payload = {
            "time_slots": [
                {
                    "id": "2026-07-01-AM-1",
                    "date": "2026-07-01",
                    "period": "AM",
                    "name": "上午一",
                    "order": 1,
                    "start_time": "08:00",
                    "end_time": "10:00",
                    "duration_hours": 2,
                },
                {
                    "id": "2026-07-01-AM-2",
                    "date": "2026-07-01",
                    "period": "AM",
                    "name": "上午二",
                    "order": 2,
                    "start_time": "10:00",
                    "end_time": "12:00",
                    "duration_hours": 2,
                },
            ],
            "rooms": [{"id": "R1", "capacity": 50}],
            "products": [
                {
                    "id": "P_POL",
                    "name": "政治暑假营",
                    "requirements": [
                        {
                            "subject": "政治",
                            "stage": "冲刺",
                            "course_module": "马原",
                            "course_group": "马原类",
                            "total_hours": 2,
                            "block_hours": 4,
                            "room_ids": ["R1"],
                        },
                        {
                            "subject": "政治",
                            "stage": "冲刺",
                            "course_module": "思修",
                            "course_group": "马原类",
                            "total_hours": 2,
                            "block_hours": 4,
                            "room_ids": ["R1"],
                        },
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C_POL",
                    "name": "政治班",
                    "product_id": "P_POL",
                    "subject": "政治",
                    "stages": ["冲刺"],
                    "size": 30,
                    "room_ids": ["R1"],
                    "teacher_assignments": [
                        {
                            "subject": "政治",
                            "stage": "冲刺",
                            "course_group": "马原类",
                            "teacher_id": "T1",
                            "teacher_name": "张老师",
                        }
                    ],
                }
            ],
        }

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "input.json"
            path.write_text(json.dumps(payload), encoding="utf-8")
            schedule_input = scheduler.load_input(path)
            assignments = scheduler.schedule(schedule_input)

        self.assertEqual(len(assignments), 1)
        self.assertEqual(assignments[0].task.course_module, "马原+思修")
        self.assertEqual(assignments[0].task.block_hours, 4)
        self.assertEqual(sum(slot.duration_hours for slot in assignments[0].candidate.slots), 4)

    def test_scheduler_uses_product_rule_block_hours_when_course_omits_it(self) -> None:
        payload = {
            "time_slots": [
                {
                    "id": "2026-07-01-AM-1",
                    "date": "2026-07-01",
                    "period": "AM",
                    "name": "上午一",
                    "order": 1,
                    "start_time": "08:00",
                    "end_time": "10:00",
                    "duration_hours": 2,
                },
                {
                    "id": "2026-07-01-AM-2",
                    "date": "2026-07-01",
                    "period": "AM",
                    "name": "上午二",
                    "order": 2,
                    "start_time": "10:00",
                    "end_time": "12:00",
                    "duration_hours": 2,
                },
            ],
            "rooms": [{"id": "R1", "capacity": 50}],
            "product_schedule_rules": [
                {
                    "product_id": "P_POL",
                    "allowed_periods": ["AM"],
                    "allowed_weekdays": ["周三"],
                    "block_hours": 4,
                }
            ],
            "products": [
                {
                    "id": "P_POL",
                    "name": "政治暑假营",
                    "requirements": [
                        {
                            "subject": "政治",
                            "stage": "冲刺",
                            "course_module": "马原",
                            "course_group": "马原类",
                            "total_hours": 2,
                            "room_ids": ["R1"],
                        },
                        {
                            "subject": "政治",
                            "stage": "冲刺",
                            "course_module": "思修",
                            "course_group": "马原类",
                            "total_hours": 2,
                            "room_ids": ["R1"],
                        },
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C_POL",
                    "name": "政治班",
                    "product_id": "P_POL",
                    "subject": "政治",
                    "stages": ["冲刺"],
                    "size": 30,
                    "room_ids": ["R1"],
                    "teacher_assignments": [
                        {
                            "subject": "政治",
                            "stage": "冲刺",
                            "course_group": "马原类",
                            "teacher_id": "T1",
                            "teacher_name": "张老师",
                        }
                    ],
                }
            ],
        }

        schedule_input = scheduler.load_input_data(payload)
        assignments = scheduler.schedule(schedule_input)

        self.assertEqual(len(assignments), 1)
        self.assertEqual(assignments[0].task.course_module, "马原+思修")
        self.assertEqual(assignments[0].task.block_hours, 4)
        self.assertEqual(sum(slot.duration_hours for slot in assignments[0].candidate.slots), 4)

    def test_parse_products_merges_top_level_and_inline_schedule_rules(self) -> None:
        top_level_rules = scheduler.group_schedule_rules_by_product(
            [
                {
                    "product_id": "P1",
                    "season_window_id": "WINDOW_AUTUMN",
                    "window_name": "秋季",
                    "allowed_periods": ["AM"],
                    "block_hours": 4,
                }
            ]
        )
        products = scheduler.parse_products(
            [
                {
                    "id": "P1",
                    "name": "秋季产品",
                    "schedule_rules": [
                        {
                            "season_window_id": "WINDOW_AUTUMN",
                            "window_name": "秋季",
                            "course_group": "写作类",
                            "allowed_periods": ["EVENING"],
                            "block_hours": 2,
                        }
                    ],
                    "requirements": [
                        {
                            "subject": "英语",
                            "quarter": "秋季",
                            "stage": "冲刺",
                            "course_module": "写作",
                            "course_group": "写作类",
                            "total_hours": 4,
                        }
                    ],
                }
            ],
            top_level_rules,
        )

        requirement = products["P1"].requirements[0]

        self.assertEqual(requirement.block_hours, 2)
        self.assertEqual([rule.block_hours for rule in requirement.schedule_rules], [2, 4])
        self.assertEqual(requirement.schedule_rules[0].allowed_periods, {"EVENING"})
        self.assertEqual(requirement.schedule_rules[1].allowed_periods, {"AM"})

    def test_scheduler_matches_product_rules_by_season_window(self) -> None:
        payload = {
            "time_slots": [
                {
                    "id": "2026-07-01-AM-1",
                    "date": "2026-07-01",
                    "period": "AM",
                    "name": "上午一",
                    "order": 1,
                    "start_time": "08:00",
                    "end_time": "10:00",
                    "duration_hours": 2,
                    "season_window_id": "WINDOW_SUMMER",
                    "season_name": "暑假",
                },
                {
                    "id": "2026-07-01-AM-2",
                    "date": "2026-07-01",
                    "period": "AM",
                    "name": "上午二",
                    "order": 2,
                    "start_time": "10:20",
                    "end_time": "12:20",
                    "duration_hours": 2,
                    "season_window_id": "WINDOW_SUMMER",
                    "season_name": "暑假",
                },
                {
                    "id": "2026-09-02-EVENING-1",
                    "date": "2026-09-02",
                    "period": "EVENING",
                    "name": "晚上",
                    "order": 1,
                    "start_time": "19:00",
                    "end_time": "21:00",
                    "duration_hours": 2,
                    "season_window_id": "WINDOW_AUTUMN",
                    "season_name": "秋季",
                },
            ],
            "rooms": [{"id": "R1", "capacity": 50}],
            "product_schedule_rules": [
                {
                    "product_id": "P1",
                    "season_window_id": "WINDOW_SUMMER",
                    "window_name": "暑假",
                    "allowed_periods": ["AM"],
                    "allowed_weekdays": ["周三"],
                    "block_hours": 4,
                },
                {
                    "product_id": "P1",
                    "season_window_id": "WINDOW_AUTUMN",
                    "window_name": "秋季",
                    "allowed_periods": ["EVENING"],
                    "allowed_weekdays": ["周三"],
                    "block_hours": 2,
                },
            ],
            "products": [
                {
                    "id": "P1",
                    "name": "半年营产品",
                    "requirements": [
                        {
                            "subject_category": "公共课",
                            "subject": "英语",
                            "quarter": "秋季",
                            "stage": "冲刺",
                            "course_module": "写作",
                            "course_group": "写作类",
                            "total_hours": 2,
                            "room_ids": ["R1"],
                        }
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C1",
                    "name": "秋季英语班",
                    "product_id": "P1",
                    "subject": "英语",
                    "stages": ["冲刺"],
                    "size": 30,
                    "room_ids": ["R1"],
                    "teacher_assignments": [
                        {
                            "subject": "英语",
                            "stage": "冲刺",
                            "course_group": "写作类",
                            "teacher_id": "T1",
                            "teacher_name": "张老师",
                        }
                    ],
                }
            ],
        }

        schedule_input = scheduler.load_input_data(payload)
        assignments = scheduler.schedule(schedule_input)

        self.assertEqual(len(assignments), 1)
        self.assertEqual(assignments[0].task.block_hours, 2)
        self.assertEqual(assignments[0].candidate.slots[0].date, "2026-09-02")

    def test_scheduler_enforces_product_rule_class_day_limits(self) -> None:
        payload = {
            "time_slots": [
                {
                    "id": "2026-07-01-AM-1",
                    "date": "2026-07-01",
                    "period": "AM",
                    "name": "上午",
                    "order": 1,
                    "start_time": "08:00",
                    "end_time": "10:00",
                    "duration_hours": 2,
                    "season_window_id": "WINDOW_SUMMER",
                    "season_name": "暑假",
                },
                {
                    "id": "2026-07-01-PM-1",
                    "date": "2026-07-01",
                    "period": "PM",
                    "name": "下午",
                    "order": 1,
                    "start_time": "14:00",
                    "end_time": "16:00",
                    "duration_hours": 2,
                    "season_window_id": "WINDOW_SUMMER",
                    "season_name": "暑假",
                },
                {
                    "id": "2026-07-02-AM-1",
                    "date": "2026-07-02",
                    "period": "AM",
                    "name": "上午",
                    "order": 1,
                    "start_time": "08:00",
                    "end_time": "10:00",
                    "duration_hours": 2,
                    "season_window_id": "WINDOW_SUMMER",
                    "season_name": "暑假",
                },
            ],
            "rooms": [{"id": "R1", "capacity": 50}],
            "product_schedule_rules": [
                {
                    "product_id": "P1",
                    "season_window_id": "WINDOW_SUMMER",
                    "window_name": "暑假",
                    "allowed_periods": ["AM", "PM"],
                    "allowed_weekdays": ["周三", "周四"],
                    "block_hours": 2,
                    "max_hours_per_class_per_day": 2,
                    "max_blocks_per_class_per_day": 1,
                }
            ],
            "products": [
                {
                    "id": "P1",
                    "name": "暑假产品",
                    "requirements": [
                        {
                            "subject_category": "公共课",
                            "subject": "英语",
                            "quarter": "暑假",
                            "stage": "基础",
                            "course_module": "词汇",
                            "course_group": "阅读类",
                            "total_hours": 2,
                            "room_ids": ["R1"],
                        },
                        {
                            "subject_category": "公共课",
                            "subject": "政治",
                            "quarter": "暑假",
                            "stage": "基础",
                            "course_module": "马原",
                            "course_group": "马原类",
                            "total_hours": 2,
                            "room_ids": ["R1"],
                        },
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C1",
                    "name": "暑假班",
                    "product_id": "P1",
                    "stages": ["基础"],
                    "size": 30,
                    "room_ids": ["R1"],
                    "teacher_assignments": [
                        {
                            "subject": "英语",
                            "stage": "基础",
                            "course_group": "阅读类",
                            "teacher_id": "T1",
                            "teacher_name": "张老师",
                        },
                        {
                            "subject": "政治",
                            "stage": "基础",
                            "course_group": "马原类",
                            "teacher_id": "T2",
                            "teacher_name": "李老师",
                        },
                    ],
                }
            ],
        }

        schedule_input = scheduler.load_input_data(payload)
        assignments = scheduler.schedule(schedule_input)
        hours_by_date = {}
        blocks_by_date = {}
        for assignment in assignments:
            date_text = assignment.candidate.slots[0].date
            hours_by_date[date_text] = hours_by_date.get(date_text, 0) + sum(
                slot.duration_hours
                for slot in assignment.candidate.slots
            )
            blocks_by_date[date_text] = blocks_by_date.get(date_text, 0) + 1

        self.assertEqual(len(assignments), 2)
        self.assertEqual(set(hours_by_date), {"2026-07-01", "2026-07-02"})
        self.assertTrue(all(hours <= 2 for hours in hours_by_date.values()))
        self.assertTrue(all(blocks <= 1 for blocks in blocks_by_date.values()))

    def test_scheduler_filters_conflict_groups_after_shared_classes_are_skipped(self) -> None:
        payload = {
            "time_slots": [
                {
                    "id": "2026-07-01-AM-1",
                    "date": "2026-07-01",
                    "period": "AM",
                    "name": "上午一",
                    "order": 1,
                    "start_time": "08:00",
                    "end_time": "10:00",
                    "duration_hours": 2,
                }
            ],
            "rooms": [{"id": "R1", "capacity": 50}],
            "products": [
                {
                    "id": "P1",
                    "name": "英语产品",
                    "requirements": [
                        {
                            "subject": "英语",
                            "stage": "基础",
                            "course_module": "词汇",
                            "course_group": "阅读类",
                            "total_hours": 2,
                            "block_hours": 2,
                            "room_ids": ["R1"],
                        }
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C_ACTIVE",
                    "name": "实际排课班",
                    "product_id": "P1",
                    "subject": "英语",
                    "stages": ["基础"],
                    "room_ids": ["R1"],
                    "teacher_assignments": [
                        {
                            "subject": "英语",
                            "stage": "基础",
                            "course_group": "阅读类",
                            "teacher_id": "T1",
                            "teacher_name": "张老师",
                        }
                    ],
                },
                {
                    "id": "C_SHARED",
                    "name": "共享课表班",
                    "product_id": "P1",
                    "subject": "英语",
                    "stages": ["基础"],
                    "room_ids": ["R1"],
                    "teacher_assignments": [
                        {
                            "subject": "英语",
                            "stage": "基础",
                            "course_group": "阅读类",
                            "schedule_mode": "共享课表",
                            "inherit_from_class_id": "C_ACTIVE",
                        }
                    ],
                },
            ],
            "conflict_groups": [
                {
                    "id": "G_HISTORY",
                    "class_ids": ["C_ACTIVE", "C_SHARED", "C_MISSING"],
                }
            ],
        }

        schedule_input = scheduler.load_input_data(payload)

        self.assertEqual(set(schedule_input.classes), {"C_ACTIVE"})
        self.assertEqual(schedule_input.conflict_groups, {})

    def test_write_html_includes_locked_class_not_in_auto_classes(self) -> None:
        slot = scheduler.TimeSlot(
            "2026-07-01-AM-1",
            "2026-07-01",
            "AM",
            "上午一",
            1,
            "08:00",
            "10:00",
            2,
        )
        locked_task = scheduler.CourseBlock(
            task_id="LOCKED:L1",
            class_id="C_LOCKED",
            class_name="锁定班",
            product_id="P_LOCKED",
            product_name="固定课产品",
            class_size=30,
            subject_category="公共课",
            subject="英语",
            quarter="暑假",
            stage="基础",
            course_module="词汇",
            course_group="阅读类",
            teacher_id="T1",
            teacher_name="张老师",
            block_hours=2,
            room_ids={"R1"},
            start_date="2026-07-01",
            end_date="2026-07-01",
            allowed_periods={"AM"},
            allowed_weekdays=None,
            excluded_weekdays=None,
            schedule_rules=(),
            is_locked=True,
        )
        schedule_input = scheduler.ScheduleInput(
            time_slots=[slot],
            rooms={"R1": scheduler.Room("R1", capacity=50)},
            classes={},
            conflict_groups={},
            class_conflict_groups={},
            locked_assignments=[],
        )
        assignment = scheduler.Assignment(locked_task, scheduler.Candidate((slot,), "T1", "张老师", "R1"))
        html_view = scheduler.build_schedule_html_view([assignment], schedule_input)

        self.assertEqual(schedule_input.classes, {})
        self.assertEqual([cls.id for cls in html_view.classes], ["C_LOCKED"])
        self.assertEqual(html_view.slot_index, {"2026-07-01-AM-1": 1})
        self.assertEqual(html_view.subjects, ["英语"])
        self.assertIn("英语", html_view.colors)
        self.assertEqual(html_view.assignments_by_class["C_LOCKED"], [assignment])
        html_document = scheduler.render_schedule_html_document(html_view, [assignment])

        self.assertIn("班级 1 个", html_document)
        self.assertIn("课程块 1 个", html_document)
        self.assertIn("课节 1 个", html_document)
        self.assertIn("min-width: 760px", html_document)
        self.assertEqual(html_document.count('<section class="timeline">'), 1)

        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "schedule.html"
            scheduler.write_html(
                [assignment],
                schedule_input,
                out_path,
            )
            html_text = out_path.read_text(encoding="utf-8")

        self.assertIn("锁定班", html_text)
        self.assertIn("英语", html_text)

    def test_parse_locked_lessons_uses_slot_ids_and_preserves_course_metadata(self) -> None:
        slots = [
            scheduler.TimeSlot("S2", "2026-07-01", "AM", "上午二", 2, "10:20", "12:20", 2),
            scheduler.TimeSlot("S1", "2026-07-01", "AM", "上午一", 1, "08:00", "10:00", 2),
        ]
        rooms = {"R1": scheduler.Room("R1", capacity=50)}
        assignments = scheduler.parse_locked_lessons(
            [
                {
                    "id": "L1",
                    "class_id": "LOCKED_CLASS",
                    "class_name": "锁定班",
                    "date": "2026-07-01",
                    "slot_ids": "S2|S1",
                    "room_id": "R1",
                    "teacher_id": "无",
                    "teacher_name": "暂无",
                    "business_product_id": "ERP_PRODUCT_1",
                    "business_product_name": "ERP 标准产品",
                    "subject_category": "公共课",
                    "subject": "英语",
                    "quarter": "暑假",
                    "stage": "基础",
                    "course_module": "阅读",
                    "course_group": "英语阅读",
                    "course_code": "ENG001",
                    "course_name": "阅读",
                }
            ],
            slots,
            rooms,
        )

        self.assertEqual(len(assignments), 1)
        assignment = assignments[0]
        self.assertEqual([slot.id for slot in assignment.candidate.slots], ["S1", "S2"])
        self.assertEqual(assignment.candidate.teacher_id, "")
        self.assertEqual(assignment.candidate.teacher_name, "")
        self.assertEqual(assignment.task.block_hours, 4)
        self.assertEqual(assignment.task.product_id, "ERP_PRODUCT_1")
        self.assertEqual(assignment.task.product_name, "ERP 标准产品")
        self.assertEqual(assignment.task.subject_category, "公共课")
        self.assertEqual(assignment.task.quarter, "暑假")
        self.assertEqual(assignment.task.stage, "基础")
        self.assertEqual(assignment.task.course_module, "阅读")
        self.assertEqual(assignment.task.course_group, "英语阅读")
        self.assertEqual(assignment.task.course_code, "ENG001")
        self.assertEqual(assignment.task.course_name, "阅读")
        self.assertTrue(assignment.task.is_locked)

    def test_locked_lesson_task_defaults_class_subject_and_time_fields(self) -> None:
        slot = scheduler.TimeSlot("S1", "2026-07-01", "PM", "下午一", 1, "14:00", "16:00", 2)

        task = scheduler.locked_lesson_task(
            {"course_code": "-", "course_name": "暂无"},
            "L2",
            "LOCKED_CLASS",
            "R2",
            (slot,),
        )

        self.assertEqual(task.task_id, "LOCKED:L2")
        self.assertEqual(task.class_name, "LOCKED_CLASS")
        self.assertEqual(task.subject, "已定课程")
        self.assertEqual(task.room_ids, {"R2"})
        self.assertEqual(task.block_hours, 2)
        self.assertEqual(task.start_date, "2026-07-01")
        self.assertEqual(task.end_date, "2026-07-01")
        self.assertEqual(task.allowed_periods, {"PM"})
        self.assertEqual(task.course_code, "")
        self.assertEqual(task.course_name, "")
        self.assertTrue(task.is_locked)

    def test_locked_lesson_slots_match_contiguous_time_range_and_reject_order_gaps(self) -> None:
        slots = [
            scheduler.TimeSlot("S1", "2026-07-01", "AM", "上午一", 1, "08:00", "10:00", 2),
            scheduler.TimeSlot("S2", "2026-07-01", "AM", "上午二", 2, "10:20", "12:20", 2),
            scheduler.TimeSlot("S_GAP", "2026-07-01", "AM", "上午三", 4, "12:40", "14:40", 2),
            scheduler.TimeSlot("S_OTHER", "2026-07-02", "AM", "上午一", 1, "08:00", "10:00", 2),
        ]
        slot_dates, slot_by_id, day_slots_by_date = scheduler.locked_lesson_slot_indexes(slots)

        matched = scheduler.locked_lesson_slots(
            {"id": "LOCK_RANGE", "date": "2026-07-01", "start_time": "8:00", "end_time": "12:20"},
            slots,
            slot_dates,
            slot_by_id,
            day_slots_by_date,
        )

        self.assertEqual([slot.id for slot in matched], ["S1", "S2"])
        self.assertEqual(
            scheduler.locked_lesson_slots(
                {"id": "OUT_OF_RANGE", "date": "2026-07-03"},
                slots,
                slot_dates,
                slot_by_id,
                day_slots_by_date,
            ),
            (),
        )
        with self.assertRaisesRegex(ValueError, "无法匹配当前课节"):
            scheduler.locked_lesson_slots(
                {"id": "LOCK_GAP", "date": "2026-07-01", "start_time": "08:00", "end_time": "14:40"},
                slots,
                slot_dates,
                slot_by_id,
                day_slots_by_date,
            )

    def test_locked_lessons_block_room_slots_but_not_existing_teachers(self) -> None:
        payload = {
            "time_slots": [
                {
                    "id": "2026-07-13-AM-2",
                    "date": "2026-07-13",
                    "period": "AM",
                    "name": "上午二",
                    "order": 2,
                    "start_time": "10:20",
                    "end_time": "12:20",
                    "duration_hours": 2,
                }
            ],
            "rooms": [{"id": "R1", "capacity": 50}, {"id": "R2", "capacity": 50}],
            "products": [
                {
                    "id": "P1",
                    "name": "测试产品",
                    "requirements": [
                        {
                            "subject": "英语",
                            "stage": "基础",
                            "course_module": "词汇",
                            "course_group": "阅读类",
                            "total_hours": 2,
                            "block_hours": 2,
                            "room_ids": ["R1", "R2"],
                        }
                    ],
                }
            ],
            "classes": [
                {
                    "id": "C1",
                    "name": "待排班",
                    "product_id": "P1",
                    "subject": "英语",
                    "stages": ["基础"],
                    "size": 30,
                    "teacher_assignments": [
                        {
                            "subject": "英语",
                            "stage": "基础",
                            "course_group": "阅读类",
                            "teacher_id": "T1",
                            "teacher_name": "张老师",
                        }
                    ],
                }
            ],
            "locked_lessons": [
                {
                    "id": "LOCK1",
                    "class_id": "LOCKED_CLASS",
                    "class_name": "已定班",
                    "date": "2026-07-13",
                    "start_time": "10:20",
                    "end_time": "12:20",
                    "room_id": "R1",
                    "subject": "西医",
                    "stage": "基础",
                    "course_module": "生理学",
                    "course_group": "西医A类",
                }
            ],
        }

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "input.json"
            path.write_text(json.dumps(payload), encoding="utf-8")
            schedule_input = scheduler.load_input(path)
            assignments = scheduler.schedule(schedule_input)

        generated = [assignment for assignment in assignments if not assignment.task.is_locked]
        locked = [assignment for assignment in assignments if assignment.task.is_locked]
        self.assertEqual(len(locked), 1)
        self.assertEqual(len(generated), 1)
        self.assertEqual(locked[0].candidate.room_id, "R1")
        self.assertEqual(generated[0].candidate.room_id, "R2")

    def test_schedule_search_state_tracks_locked_placed_and_unplaced_constraints(self) -> None:
        morning = scheduler.TimeSlot("2026-07-13-AM-1", "2026-07-13", "AM", "上午一", 1)
        afternoon = scheduler.TimeSlot("2026-07-13-PM-1", "2026-07-13", "PM", "下午一", 1)

        def school_class(class_id: str) -> scheduler.SchoolClass:
            return scheduler.SchoolClass(
                id=class_id,
                name=class_id,
                product_id=None,
                product_name=None,
                size=30,
                room_ids={"R1", "R2"},
                start_date=None,
                start_period=None,
                end_date=None,
                end_period=None,
                first_lesson_date=None,
                first_lesson_period=None,
                stage_order={},
                requirements=[],
            )

        def task(task_id: str, class_id: str, teacher_id: str) -> scheduler.CourseBlock:
            return scheduler.CourseBlock(
                task_id=task_id,
                class_id=class_id,
                class_name=class_id,
                product_id=None,
                product_name=None,
                class_size=30,
                subject_category="公共课",
                subject="英语",
                quarter=None,
                stage="基础",
                course_module=None,
                course_group="阅读类",
                teacher_id=teacher_id,
                teacher_name=teacher_id,
                block_hours=2,
                room_ids={"R1", "R2"},
                start_date=None,
                end_date=None,
                allowed_periods=None,
                allowed_weekdays=None,
                excluded_weekdays=None,
                schedule_rules=(),
            )

        locked_task = task("LOCKED", "LOCKED_CLASS", "T_LOCK")
        locked_assignment = scheduler.Assignment(
            locked_task,
            scheduler.Candidate((morning,), "T_LOCK", "锁定老师", "R1"),
        )
        task_a = task("TASK_A", "C1", "T1")
        task_b = task("TASK_B", "C2", "T1")
        schedule_input = scheduler.ScheduleInput(
            time_slots=[morning, afternoon],
            rooms={"R1": scheduler.Room("R1"), "R2": scheduler.Room("R2")},
            classes={"C1": school_class("C1"), "C2": school_class("C2")},
            conflict_groups={"G1": {"LOCKED_CLASS", "C1", "C2"}},
            class_conflict_groups={"LOCKED_CLASS": {"G1"}, "C1": {"G1"}, "C2": {"G1"}},
            locked_assignments=[locked_assignment],
        )
        state = scheduler.ScheduleSearchState(
            schedule_input,
            {"TASK_A": task_a, "TASK_B": task_b},
            {"C1": ["TASK_A"], "C2": ["TASK_B"]},
        )

        self.assertFalse(state.is_valid(task_a, scheduler.Candidate((morning,), "T1", "张老师", "R1")))

        placed = scheduler.Candidate((afternoon,), "T1", "张老师", "R1")
        other_same_teacher = scheduler.Candidate((afternoon,), "T1", "张老师", "R2")
        self.assertTrue(state.is_valid(task_a, placed))
        state.place(task_a, placed)
        self.assertFalse(state.is_valid(task_b, other_same_teacher))
        state.unplace(task_a, placed)
        self.assertTrue(state.is_valid(task_b, other_same_teacher))

    def test_duplicate_table_sources_are_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp)
            write_csv(source / "products.csv", ["id", "name"], [{"id": "P1", "name": "A"}])
            write_csv(source / "product.csv", ["id", "name"], [{"id": "P2", "name": "B"}])

            with self.assertRaises(PipelineError):
                load_source_tables(source)

    def test_source_row_counts_cover_all_standard_tables(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp)
            write_csv(source / "products.csv", ["id", "name"], [{"id": "P1", "name": "A"}])

            row_counts = row_counts_for_tables(load_source_tables(source))

        self.assertEqual(row_counts["products"], 1)
        self.assertEqual(set(SOURCE_TABLES), set(row_counts))
        self.assertTrue(all(row_counts[table] == 0 for table in SOURCE_TABLES if table != "products"))

    def test_excel_source_is_supported(self) -> None:
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl is not installed")

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            csv_source = root / "csv_source"
            xlsx_source = root / "incoming"
            data_dir = root / "data"
            output_dir = root / "outputs"
            write_minimal_csv_source(csv_source)
            xlsx_source.mkdir()
            workbook = Workbook()
            default = workbook.active
            workbook.remove(default)
            for csv_path in sorted(csv_source.glob("*.csv")):
                sheet = workbook.create_sheet(csv_path.stem)
                with csv_path.open(encoding="utf-8") as handle:
                    for row in csv.reader(handle):
                        sheet.append(row)
            workbook.save(xlsx_source / "正式排课数据.xlsx")

            result = run_pipeline(pipeline_args(xlsx_source, data_dir, output_dir))

            self.assertTrue(result.schedule_csv_path.exists())
            self.assertTrue((data_dir / "scheduler_input_draft.json").exists())


if __name__ == "__main__":
    unittest.main()
