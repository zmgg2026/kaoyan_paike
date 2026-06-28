#!/usr/bin/env python3
"""Annotate ERP lesson import failure rows with normalized error types."""

from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scripts.csv_utils import read_csv_rows, write_csv_rows


DEFAULT_OUTPUT_DIR = Path("outputs")
DEFAULT_ROOMS = Path("data/rooms.csv")


def clean(value: object) -> str:
    return str(value or "").strip()


def read_rows(path: Path) -> List[Dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    worksheet.reset_dimensions()
    rows = worksheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    result: List[Dict[str, str]] = []
    for raw in rows:
        if not any(raw):
            continue
        result.append({headers[index]: clean(raw[index]) if index < len(raw) else "" for index in range(len(headers))})
    return result


def room_lookup(path: Path) -> Dict[str, str]:
    if not path.exists():
        return {}
    return {clean(row.get("id")): clean(row.get("name")) for row in read_csv_rows(path) if clean(row.get("id"))}


def add_type(types: List[str], values: List[str], error_type: str, value: str = "") -> None:
    if error_type not in types:
        types.append(error_type)
    if value and value not in values:
        values.append(value)


def classify_error(reason: str) -> Tuple[List[str], List[str], str]:
    types: List[str] = []
    values: List[str] = []

    for match in re.finditer(r"教室'([^']+)'不存在", reason):
        add_type(types, values, "教室不存在", match.group(1))

    for match in re.finditer(r"([^\s:;；,，\[\]]+)课次所在班级无此科目", reason):
        add_type(types, values, "班级科目不匹配", match.group(1))

    for match in re.finditer(r"([A-Za-z0-9_|]+)无此课程", reason):
        add_type(types, values, "课程不存在", match.group(1))

    if "教师课程冲突" in reason:
        add_type(types, values, "老师时间冲突")
    if "教室课程冲突" in reason:
        add_type(types, values, "教室时间冲突")
    if "班级课次时段冲突" in reason:
        add_type(types, values, "班级课次时段冲突")
    if "教室座位数不够" in reason:
        add_type(types, values, "教室容量不足")
    if "课次ID" in reason and ("不存在" in reason or "无效" in reason):
        add_type(types, values, "课次ID异常")

    if not types:
        add_type(types, values, "其他")

    return types, values, "；".join(types)


def action_for(types: Sequence[str], row: Dict[str, str], rooms: Dict[str, str]) -> str:
    suggestions: List[str] = []
    room_id = clean(row.get("教室"))
    if "教室不存在" in types:
        room_name = rooms.get(room_id, "")
        if room_id.startswith("RMONLINE"):
            suggestions.append("ERP维护线上教室编码，或改为ERP已存在的线上教室ID")
        else:
            suggestions.append(f"ERP维护教室ID {room_id}" + (f"（{room_name}）" if room_name else "") + "，或改为ERP已存在教室ID")
    if "班级科目不匹配" in types:
        suggestions.append("核对该课次ID归属班级的ERP科目配置；若课次建错班级/科目，需ERP侧调整课次或重新匹配")
    if "课程不存在" in types:
        suggestions.append("核对课程编码是否在ERP课程目录中启用；必要时维护课程编码或替换为ERP有效课程")
    if "老师时间冲突" in types:
        suggestions.append("需要调课或确认ERP已有课次占用；同一老师同一时段不能重复")
    if "教室时间冲突" in types:
        suggestions.append("需要更换教室或调课；同一教室同一时段已有课")
    if "班级课次时段冲突" in types:
        suggestions.append("该班级同一时段已有课次，需调整课次时间或确认是否重复导入")
    if "教室容量不足" in types:
        suggestions.append("ERP按容量做强校验，需换大教室、调整班级人数/教室容量，或请ERP放宽容量校验")
    if not suggestions:
        suggestions.append("需人工查看ERP原始错误")
    return "；".join(suggestions)


def row_priority(types: Sequence[str]) -> str:
    hard_data = {"教室不存在", "班级科目不匹配", "课程不存在", "课次ID异常"}
    if any(item in hard_data for item in types):
        return "先修基础数据后重导"
    if any(item in {"老师时间冲突", "教室时间冲突", "班级课次时段冲突"} for item in types):
        return "需调课/处理冲突后重导"
    if "教室容量不足" in types:
        return "需处理容量后重导"
    return "需人工确认"


def annotate_rows(rows: Sequence[Dict[str, str]], rooms: Dict[str, str]) -> List[Dict[str, str]]:
    annotated: List[Dict[str, str]] = []
    for row in rows:
        new_row = dict(row)
        types, values, type_text = classify_error(clean(row.get("错误原因")))
        room_id = clean(row.get("教室"))
        new_row["错误类型"] = type_text
        new_row["错误对象"] = "；".join(values)
        new_row["本系统教室名称"] = rooms.get(room_id, "")
        new_row["处理优先级"] = row_priority(types)
        new_row["处理建议"] = action_for(types, row, rooms)
        annotated.append(new_row)
    return annotated


def write_xlsx(path: Path, fieldnames: Sequence[str], rows: Sequence[Dict[str, str]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "未成功明细标注"
    worksheet.append(list(fieldnames))
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        worksheet.append([row.get(field, "") for field in fieldnames])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    width_limits = {
        "错误原因": 90,
        "处理建议": 70,
        "错误类型": 28,
        "错误对象": 35,
    }
    for index, field in enumerate(fieldnames, start=1):
        values = [clean(field)] + [clean(row.get(field)) for row in rows[:300]]
        max_len = max((len(value) for value in values), default=8)
        width = min(width_limits.get(field, 28), max(10, max_len + 2))
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def build_summary(annotated: Sequence[Dict[str, str]]) -> Tuple[List[Dict[str, str]], List[Dict[str, str]], List[Dict[str, str]]]:
    type_counter: Counter[str] = Counter()
    combination_counter: Counter[str] = Counter()
    class_counter: Counter[Tuple[str, str]] = Counter()
    room_counter: Counter[Tuple[str, str, str]] = Counter()

    for row in annotated:
        type_text = clean(row.get("错误类型"))
        combination_counter[type_text] += 1
        for error_type in [item for item in type_text.split("；") if item]:
            type_counter[error_type] += 1
            class_counter[(clean(row.get("班级编码")), error_type)] += 1
        room_id = clean(row.get("教室"))
        if room_id:
            room_counter[(room_id, clean(row.get("本系统教室名称")), type_text)] += 1

    type_rows = [{"错误类型": key, "涉及课次数": value} for key, value in type_counter.most_common()]
    combo_rows = [{"错误类型组合": key, "涉及课次数": value} for key, value in combination_counter.most_common()]
    room_rows = [
        {"教室": room_id, "本系统教室名称": room_name, "错误类型": error_type, "涉及课次数": count}
        for (room_id, room_name, error_type), count in room_counter.most_common()
    ]
    class_rows = [
        {"班级编码": class_id, "错误类型": error_type, "涉及课次数": count}
        for (class_id, error_type), count in class_counter.most_common()
    ]
    return type_rows, combo_rows, room_rows + class_rows


def write_report(path: Path, input_path: Path, annotated_path: Path, type_rows: Sequence[Dict[str, str]], combo_rows: Sequence[Dict[str, str]], rows: Sequence[Dict[str, str]]) -> None:
    missing_room_counter: Counter[Tuple[str, str]] = Counter()
    subject_mismatch_counter: Counter[Tuple[str, str]] = Counter()
    class_conflict_counter: Counter[str] = Counter()
    capacity_room_counter: Counter[Tuple[str, str]] = Counter()
    for row in rows:
        error_type = clean(row.get("错误类型"))
        if "教室不存在" in error_type:
            missing_room_counter[(clean(row.get("教室")), clean(row.get("本系统教室名称")))] += 1
        if "班级科目不匹配" in error_type:
            subject_mismatch_counter[(clean(row.get("班级编码")), clean(row.get("课节科目")))] += 1
        if "班级课次时段冲突" in error_type:
            class_conflict_counter[clean(row.get("班级编码"))] += 1
        if "教室容量不足" in error_type:
            capacity_room_counter[(clean(row.get("教室")), clean(row.get("本系统教室名称")))] += 1

    lines = [
        "# ERP 导入未成功明细归因报告",
        "",
        f"- 来源文件: {input_path}",
        f"- 标注明细: {annotated_path}",
        f"- 未成功课次数: {len(rows)}",
        "",
        "## 错误类型统计",
        "",
        "| 错误类型 | 涉及课次数 |",
        "|---|---:|",
    ]
    for row in type_rows:
        lines.append(f"| {row['错误类型']} | {row['涉及课次数']} |")
    lines.extend(["", "## 错误类型组合 Top 20", "", "| 错误类型组合 | 课次数 |", "|---|---:|"])
    for row in combo_rows[:20]:
        lines.append(f"| {row['错误类型组合']} | {row['涉及课次数']} |")
    lines.extend(["", "## 重点对象", "", "### ERP 不存在的教室 Top 10", "", "| 教室ID | 本系统教室名称 | 课次数 |", "|---|---|---:|"])
    for (room_id, room_name), count in missing_room_counter.most_common(10):
        lines.append(f"| {room_id} | {room_name} | {count} |")
    lines.extend(["", "### 班级科目不匹配 Top 10", "", "| 班级编码 | 课节科目 | 课次数 |", "|---|---|---:|"])
    for (class_id, subject), count in subject_mismatch_counter.most_common(10):
        lines.append(f"| {class_id} | {subject} | {count} |")
    lines.extend(["", "### 班级课次时段冲突 Top 10", "", "| 班级编码 | 课次数 |", "|---|---:|"])
    for class_id, count in class_conflict_counter.most_common(10):
        lines.append(f"| {class_id} | {count} |")
    lines.extend(["", "### 容量不足教室 Top 10", "", "| 教室ID | 本系统教室名称 | 课次数 |", "|---|---|---:|"])
    for (room_id, room_name), count in capacity_room_counter.most_common(10):
        lines.append(f"| {room_id} | {room_name} | {count} |")
    lines.extend(["", "## 初步判断", ""])
    lines.append("- `教室不存在` 是最大基础数据问题，说明 ERP 不认识本次导入表中的部分教室 ID，尤其是线上虚拟教室和云谷 507/508。")
    lines.append("- `班级科目不匹配` 说明 ERP 课次或班级科目配置与本系统课表不一致，这类不能靠重导解决，需要先核对 ERP 课次归属。")
    lines.append("- `老师/教室/班级时段冲突` 属于回写校验问题，需要结合 ERP 已有课次和当前课表做调课或释放占用。")
    lines.append("- `教室容量不足` 是 ERP 强校验；本系统之前按提醒口径处理，ERP 若强制校验则需要换教室或维护容量。")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="标注 ERP 课次导入失败原因")
    parser.add_argument("--input-xlsx", type=Path, required=True)
    parser.add_argument("--rooms", type=Path, default=DEFAULT_ROOMS)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--timestamp", default="")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    stamp = args.timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    rows = read_rows(args.input_xlsx)
    rooms = room_lookup(args.rooms)
    annotated = annotate_rows(rows, rooms)
    fieldnames = list(rows[0].keys()) + ["错误类型", "错误对象", "本系统教室名称", "处理优先级", "处理建议"] if rows else []

    out_dir = args.output_dir
    xlsx_path = out_dir / f"erp_import_failures_annotated_{stamp}.xlsx"
    csv_path = out_dir / f"erp_import_failures_annotated_{stamp}.csv"
    type_summary_path = out_dir / f"erp_import_failure_type_summary_{stamp}.csv"
    combo_summary_path = out_dir / f"erp_import_failure_combo_summary_{stamp}.csv"
    detail_summary_path = out_dir / f"erp_import_failure_detail_summary_{stamp}.csv"
    report_path = out_dir / f"erp_import_failure_report_{stamp}.md"

    write_xlsx(xlsx_path, fieldnames, annotated)
    write_csv_rows(csv_path, fieldnames, annotated)
    type_rows, combo_rows, detail_rows = build_summary(annotated)
    write_csv_rows(type_summary_path, ["错误类型", "涉及课次数"], type_rows)
    write_csv_rows(combo_summary_path, ["错误类型组合", "涉及课次数"], combo_rows)
    write_csv_rows(detail_summary_path, sorted({key for row in detail_rows for key in row.keys()}), detail_rows)
    write_report(report_path, args.input_xlsx, xlsx_path, type_rows, combo_rows, annotated)

    print(f"rows={len(annotated)}")
    print(f"annotated_xlsx={xlsx_path}")
    print(f"annotated_csv={csv_path}")
    print(f"type_summary={type_summary_path}")
    print(f"combo_summary={combo_summary_path}")
    print(f"detail_summary={detail_summary_path}")
    print(f"report={report_path}")
    for row in type_rows:
        print(f"{row['错误类型']}={row['涉及课次数']}")


if __name__ == "__main__":
    main()
