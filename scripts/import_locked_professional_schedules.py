#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import shutil
import warnings
from dataclasses import dataclass, field
from datetime import datetime, time
from pathlib import Path
import sys
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from data_admin_server import (
    DATA_DIR,
    OUTPUT_DIR,
    export_scheduler_input,
    load_state,
    normalize_text,
    save_state,
    split_id_list,
)
from scripts.field_utils import parse_datetime_value, row_value
from scripts.period_utils import period_from_minutes


CLASS_ID_RE = re.compile(r"(KY[A-Z]+[0-9]+)")
BLANK_MARKERS = {"", "-", "—", "无", "暂无", "NULL", "N/A"}
ONLINE_ROOM_NAMES = {"线上虚拟网络教室01"}


@dataclass
class ImportedClassSummary:
    class_id: str
    class_name: str
    source: str
    lesson_count: int = 0
    date_min: str = ""
    date_max: str = ""
    room_fallbacks: List[str] = field(default_factory=list)
    teacher_warnings: List[str] = field(default_factory=list)
    missing_stage_module_count: int = 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="导入专业课固定课表，并标记班级不参与自动排课。")
    parser.add_argument(
        "folder",
        help="包含专业课明细课表 xlsx 的文件夹",
    )
    return parser.parse_args()


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value).strip()


def is_blank_marker(value: Any) -> bool:
    return clean_cell(value).strip() in BLANK_MARKERS


def parse_datetime_cell(value: Any, label: str) -> datetime:
    return parse_datetime_value(value, label, allow_date=True)


def parse_optional_float(value: Any) -> Optional[float]:
    text = clean_cell(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def period_for(start_dt: datetime) -> str:
    return period_from_minutes(
        start_dt.hour * 60 + start_dt.minute,
        am_end_minutes=12 * 60,
        pm_end_minutes=19 * 60,
    )


def unique(values: Iterable[str]) -> List[str]:
    seen = set()
    result: List[str] = []
    for value in values:
        text = normalize_text(value)
        if text and text not in seen:
            seen.add(text)
            result.append(text)
    return result


def build_indexes(state: Dict[str, Any]) -> Dict[str, Any]:
    rooms_by_id = {room["id"]: room for room in state["rooms"]}
    rooms_by_name: Dict[str, List[Dict[str, Any]]] = {}
    for room in state["rooms"]:
        rooms_by_name.setdefault(normalize_text(room.get("name")), []).append(room)

    teachers_by_name: Dict[str, List[Dict[str, Any]]] = {}
    for teacher in state["teachers"]:
        teachers_by_name.setdefault(normalize_text(teacher.get("name")), []).append(teacher)

    products_by_id = {product["id"]: product for product in state["products"]}
    course_meta: Dict[Tuple[str, str], Dict[str, str]] = {}
    courses_by_product_name: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    by_product_subject: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for course in state["product_courses"]:
        by_product_subject.setdefault((course.get("product_id", ""), course.get("subject", "")), []).append(course)
        course_name = normalize_text(course.get("course_name"))
        if course.get("product_id") and course_name:
            courses_by_product_name.setdefault((course.get("product_id", ""), course_name), []).append(course)
    for key, courses in by_product_subject.items():
        groups = unique(course.get("course_group", "") for course in courses)
        modules = unique(course.get("course_module", "") for course in courses)
        course_meta[key] = {
            "course_group": groups[0] if len(groups) == 1 else "",
            "course_module": modules[0] if len(modules) == 1 else "",
        }

    return {
        "rooms_by_id": rooms_by_id,
        "rooms_by_name": rooms_by_name,
        "teachers_by_name": teachers_by_name,
        "products_by_id": products_by_id,
        "course_meta": course_meta,
        "courses_by_product_name": courses_by_product_name,
    }


def stage_from_course_name(course_name: str) -> str:
    if "基础" in course_name:
        return "基础"
    if "强化" in course_name:
        return "强化"
    if "冲点" in course_name or "冲刺" in course_name:
        return "冲刺"
    if "导学" in course_name:
        return "导学1"
    return ""


def course_meta_from_product_course(course: Dict[str, Any], course_name: str) -> Dict[str, str]:
    return {
        "window_name": normalize_text(row_value(course, "window_name", "quarter")),
        "stage": normalize_text(course.get("stage")),
        "course_module": normalize_text(course.get("course_module")),
        "course_group": normalize_text(course.get("course_group")),
        "course_code": normalize_text(course.get("course_code")),
        "course_name": course_name,
    }


def resolve_course_meta(
    row: Dict[str, Any],
    cls: Dict[str, Any],
    subject: str,
    indexes: Dict[str, Any],
    summary: ImportedClassSummary,
) -> Dict[str, str]:
    raw_course_name = normalize_text(row.get("课程名称") or row.get("课程内容"))
    product_id = normalize_text(cls.get("product_id"))
    if raw_course_name:
        matches = indexes["courses_by_product_name"].get((product_id, raw_course_name), [])
        stage_hint = stage_from_course_name(raw_course_name)
        if stage_hint:
            staged_matches = [course for course in matches if normalize_text(course.get("stage")) == stage_hint]
            if staged_matches:
                matches = staged_matches
        if len(matches) == 1:
            return course_meta_from_product_course(matches[0], raw_course_name)
        unique_without_stage = {
            (
                normalize_text(row_value(course, "window_name", "quarter")),
                normalize_text(course.get("course_module")),
                normalize_text(course.get("course_group")),
                normalize_text(course.get("course_code")),
            )
            for course in matches
        }
        if len(unique_without_stage) == 1 and matches:
            return course_meta_from_product_course(matches[0], raw_course_name)
        if len(matches) > 1:
            summary.teacher_warnings.append(
                f"课程名称 {raw_course_name} 在产品 {product_id} 下命中多条产品课程，保留课程名称但不自动写模块。"
            )

    meta = indexes["course_meta"].get((product_id, subject), {})
    return {
        "window_name": "",
        "stage": "",
        "course_module": meta.get("course_module", ""),
        "course_group": meta.get("course_group", ""),
        "course_code": "",
        "course_name": raw_course_name,
    }


def resolve_teacher(
    row: Dict[str, Any],
    indexes: Dict[str, Any],
    summary: ImportedClassSummary,
) -> Tuple[str, str]:
    raw_name = clean_cell(row.get("教师名称"))
    raw_code = clean_cell(row.get("教师编码"))
    if raw_name in BLANK_MARKERS:
        return "", ""
    candidates = indexes["teachers_by_name"].get(raw_name, [])
    numeric_candidates = [
        teacher for teacher in candidates
        if re.fullmatch(r"\d{6}", normalize_text(teacher.get("employee_id")))
    ]
    if len(numeric_candidates) == 1:
        return normalize_text(numeric_candidates[0].get("employee_id")), raw_name
    if re.fullmatch(r"\d{6}", raw_code):
        return raw_code, raw_name
    summary.teacher_warnings.append(f"{raw_name} 无法唯一匹配 6 位员工 ID，保留教师姓名，不写教师编码。")
    return "", raw_name


def preferred_room_for_missing(
    cls: Dict[str, Any],
    lesson_date: str,
    indexes: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    room_ids = split_id_list(cls.get("preferred_room_ids"))
    rooms = [indexes["rooms_by_id"].get(room_id) for room_id in room_ids]
    rooms = [room for room in rooms if room]
    if not rooms:
        return None
    if cls.get("id") == "KYJSJ2757" and "2026-07-01" <= lesson_date <= "2026-08-31":
        for room in rooms:
            if room.get("teaching_area_id") == "ARHFWY216":
                return room
    if len(rooms) == 1:
        return rooms[0]
    return None


def resolve_room(
    row: Dict[str, Any],
    cls: Dict[str, Any],
    lesson_date: str,
    indexes: Dict[str, Any],
    summary: ImportedClassSummary,
) -> Dict[str, Any]:
    room_name = clean_cell(row.get("教室名称"))
    if room_name and room_name not in BLANK_MARKERS:
        candidates = indexes["rooms_by_name"].get(room_name, [])
        if len(candidates) == 1:
            return candidates[0]
        if len(candidates) > 1:
            teaching_area_name = clean_cell(row.get("教学区"))
            for room in candidates:
                if teaching_area_name and teaching_area_name in clean_cell(room.get("teaching_area_name")):
                    return room
            raise ValueError(f"{summary.source} {summary.class_id} 教室 {room_name} 命中多个房间，无法确定。")
        if room_name in ONLINE_ROOM_NAMES:
            online = indexes["rooms_by_name"].get(room_name, [])
            if online:
                return online[0]
        raise ValueError(f"{summary.source} {summary.class_id} 使用了未登记教室: {room_name}")

    fallback = preferred_room_for_missing(cls, lesson_date, indexes)
    if fallback:
        note = f"{lesson_date} 原文件教室为空，使用班级优先教室 {fallback.get('name')}({fallback.get('id')})。"
        if note not in summary.room_fallbacks:
            summary.room_fallbacks.append(note)
        return fallback
    raise ValueError(f"{summary.source} {summary.class_id} {lesson_date} 原文件教室为空，且班级优先教室无法唯一确定。")


def infer_stage(cls: Dict[str, Any], lesson_date: str) -> str:
    sub_product = normalize_text(cls.get("sub_product"))
    month = int(lesson_date[5:7])
    if sub_product in {"寒暑营", "无忧寒"}:
        if month <= 2:
            return "寒假"
        if month <= 6:
            return "春季"
        if month <= 8:
            return "暑假"
        return "秋季"
    return ""


def rows_from_workbook(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(path, read_only=True, data_only=True)
    for worksheet in workbook.worksheets:
        worksheet.reset_dimensions()
        iterator = worksheet.iter_rows(values_only=True)
        try:
            headers = [clean_cell(cell) for cell in next(iterator)]
        except StopIteration:
            continue
        header_indexes = {header: index for index, header in enumerate(headers) if header}
        if "班级名称" not in header_indexes or "起始时间" not in header_indexes or "结束时间" not in header_indexes:
            continue
        for row in iterator:
            if not any(row):
                continue
            rows.append({header: row[index] if index < len(row) else "" for header, index in header_indexes.items()})
    return rows


def import_workbook(
    path: Path,
    state: Dict[str, Any],
    indexes: Dict[str, Any],
) -> Tuple[str, List[Dict[str, Any]], ImportedClassSummary]:
    match = CLASS_ID_RE.search(path.name)
    if not match:
        raise ValueError(f"{path.name} 文件名未识别到班级编码。")
    class_id = match.group(1)
    classes_by_id = {cls["id"]: cls for cls in state["classes"]}
    if class_id not in classes_by_id:
        raise ValueError(f"{path.name} 对应班级 {class_id} 不存在于 data/classes。")
    cls = classes_by_id[class_id]
    class_name = normalize_text(cls.get("name"))
    product = indexes["products_by_id"].get(cls.get("product_id"), {})
    source_rows = rows_from_workbook(path)
    if not source_rows:
        raise ValueError(f"{path.name} 未读取到课表明细行。")
    source_rows.sort(key=lambda item: parse_datetime_cell(item.get("起始时间"), "起始时间"))

    summary = ImportedClassSummary(class_id=class_id, class_name=class_name, source=path.name)
    lessons: List[Dict[str, Any]] = []
    dates: List[str] = []
    for index, row in enumerate(source_rows, start=1):
        start_dt = parse_datetime_cell(row.get("起始时间"), "起始时间")
        end_dt = parse_datetime_cell(row.get("结束时间"), "结束时间")
        lesson_date = start_dt.strftime("%Y-%m-%d")
        dates.append(lesson_date)
        duration_hours = parse_optional_float(row.get("小时数"))
        if duration_hours is None:
            duration_hours = parse_optional_float(row.get("分钟数"))
            if duration_hours is not None:
                duration_hours = duration_hours / 60
        if duration_hours is None:
            duration_hours = round((end_dt - start_dt).total_seconds() / 3600, 2)

        room = resolve_room(row, cls, lesson_date, indexes, summary)
        teacher_id, teacher_name = resolve_teacher(row, indexes, summary)
        subject = clean_cell(row.get("科目内")) or normalize_text(cls.get("subject"))
        meta = resolve_course_meta(row, cls, subject, indexes, summary)
        window_name = meta.get("window_name", "")
        stage = meta.get("stage", "") or infer_stage(cls, lesson_date)
        course_module = meta.get("course_module", "")
        course_group = meta.get("course_group", "")
        course_code = meta.get("course_code", "")
        course_name = meta.get("course_name", "")
        if not stage or not course_module:
            summary.missing_stage_module_count += 1
        lessons.append(
            {
                "id": f"LOCKED_PRO_{class_id}_{index:03d}",
                "class_id": class_id,
                "class_name": class_name,
                "date": lesson_date,
                "period": period_for(start_dt),
                "start_time": start_dt.strftime("%H:%M"),
                "end_time": end_dt.strftime("%H:%M"),
                "duration_hours": f"{duration_hours:g}",
                "teacher_id": teacher_id,
                "teacher_name": teacher_name,
                "room_id": normalize_text(room.get("id")),
                "room_name": normalize_text(room.get("name")),
                "teaching_area_id": normalize_text(room.get("teaching_area_id")),
                "business_product_id": normalize_text(cls.get("product_id")),
                "business_product_name": normalize_text(product.get("name")) or normalize_text(cls.get("product_id")),
                "subject": subject,
                "window_name": window_name,
                "stage": stage,
                "course_module": course_module,
                "course_group": course_group,
                "course_code": course_code,
                "course_name": course_name,
                "source": str(path),
                "is_locked": True,
                "notes": "专业课固定课表：不参与自动排课，仅作为已定课表和资源占用留存。",
            }
        )
    summary.lesson_count = len(lessons)
    summary.date_min = min(dates)
    summary.date_max = max(dates)
    return class_id, lessons, summary


def backup_data_dir(timestamp: str) -> Path:
    backup_dir = OUTPUT_DIR / "backups" / f"data_{timestamp}_locked_professional_schedules"
    shutil.copytree(DATA_DIR, backup_dir)
    return backup_dir


def update_class_lock_flags(state: Dict[str, Any], class_ids: Iterable[str], source_folder: Path) -> None:
    target_ids = set(class_ids)
    stamp = datetime.now().strftime("%Y-%m-%d")
    for cls in state["classes"]:
        if cls.get("id") not in target_ids:
            continue
        cls["is_manual_schedule_locked"] = True
        note = normalize_text(cls.get("notes"))
        addition = f"{stamp} 已导入专业课固定课表：{source_folder}；不参与自动排课。"
        if addition not in note:
            cls["notes"] = f"{note}; {addition}" if note else addition


def write_report(
    timestamp: str,
    folder: Path,
    backup_dir: Path,
    summaries: List[ImportedClassSummary],
    old_removed: int,
    scheduler_counts: Dict[str, int],
) -> Path:
    report_path = OUTPUT_DIR / f"locked_professional_schedule_import_{timestamp}.md"
    lines = [
        "# 专业课固定课表导入报告",
        "",
        f"- 源文件夹：`{folder}`",
        f"- 数据备份：`{backup_dir}`",
        f"- 替换旧锁定课节：{old_removed} 条",
        f"- 本次导入锁定课节：{sum(item.lesson_count for item in summaries)} 条",
        f"- 已标记不参与自动排课班级：{len(summaries)} 个",
        "",
        "## 班级汇总",
        "",
        "| 班级编码 | 班级名称 | 文件 | 课节数 | 日期范围 |",
        "| --- | --- | --- | ---: | --- |",
    ]
    for item in summaries:
        lines.append(
            f"| {item.class_id} | {item.class_name} | {item.source} | {item.lesson_count} | {item.date_min} 至 {item.date_max} |"
        )

    room_notes = [note for item in summaries for note in item.room_fallbacks]
    teacher_notes = [f"{item.class_id}: {note}" for item in summaries for note in item.teacher_warnings]
    missing_stage_items = [item for item in summaries if item.missing_stage_module_count]
    if room_notes:
        lines.extend(["", "## 教室补充说明", ""])
        lines.extend(f"- {note}" for note in room_notes)
    if teacher_notes:
        lines.extend(["", "## 教师编码提醒", ""])
        lines.extend(f"- {note}" for note in teacher_notes)
    if missing_stage_items:
        lines.extend(["", "## 阶段/模块说明", ""])
        for item in missing_stage_items:
            lines.append(
                f"- {item.class_id} 有 {item.missing_stage_module_count} 条课节的原始表未提供明确阶段或模块，系统仅按能安全推断的信息留存。"
            )

    lines.extend(
        [
            "",
            "## scheduler_input_draft 计数",
            "",
        ]
    )
    for key, value in scheduler_counts.items():
        lines.append(f"- {key}: {value}")
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return report_path


def main() -> None:
    args = parse_args()
    folder = Path(args.folder).expanduser().resolve()
    if not folder.exists():
        raise SystemExit(f"源文件夹不存在：{folder}")
    files = sorted(path for path in folder.glob("*.xlsx") if not path.name.startswith("~$"))
    if not files:
        raise SystemExit(f"源文件夹没有 xlsx：{folder}")

    state = load_state()
    indexes = build_indexes(state)
    imported_lessons: List[Dict[str, Any]] = []
    summaries: List[ImportedClassSummary] = []
    imported_class_ids: List[str] = []

    for path in files:
        class_id, lessons, summary = import_workbook(path, state, indexes)
        imported_class_ids.append(class_id)
        imported_lessons.extend(lessons)
        summaries.append(summary)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = backup_data_dir(timestamp)

    imported_class_id_set = set(imported_class_ids)
    old_lessons = state.get("locked_scheduled_lessons", [])
    remaining_lessons = [
        lesson for lesson in old_lessons
        if normalize_text(lesson.get("class_id")) not in imported_class_id_set
    ]
    old_removed = len(old_lessons) - len(remaining_lessons)
    state["locked_scheduled_lessons"] = [*remaining_lessons, *imported_lessons]
    update_class_lock_flags(state, imported_class_ids, folder)
    save_state(state)
    scheduler_result = export_scheduler_input()
    report_path = write_report(
        timestamp,
        folder,
        backup_dir,
        summaries,
        old_removed,
        scheduler_result.get("counts", {}),
    )
    print(f"导入完成：{len(imported_class_ids)} 个班级，{len(imported_lessons)} 条锁定课节")
    print(f"备份：{backup_dir}")
    print(f"报告：{report_path}")
    print(f"scheduler_input：{scheduler_result.get('path')}")


if __name__ == "__main__":
    main()
