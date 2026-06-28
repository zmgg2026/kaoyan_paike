#!/usr/bin/env python3
"""Build a reversed ERP import workbook from an ERP import result file."""

from __future__ import annotations

import argparse
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Sequence

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.csv_utils import read_csv_rows, write_csv_rows
from scripts.field_utils import normalize_excel_text as clean


DEFAULT_OUTPUT_DIR = Path("outputs")
DEFAULT_CLASSES = Path("data/classes.csv")
ERP_HEADERS = [
    "*课次ID",
    "日期",
    "时间",
    "班级编码",
    "教师1（实际授课教师）",
    "教师2",
    "教室",
    "事件",
    "备注",
    "授课方式标识",
    "课程",
    "课节科目",
]
INPUT_TO_OUTPUT = {
    "*课次ID": "课次ID",
    "日期": "日期",
    "时间": "时间",
    "班级编码": "班级编码",
    "教师1（实际授课教师）": "教师1（实际授课教师）",
    "教师2": "教师2",
    "教室": "教室",
    "事件": "事件",
    "备注": "备注",
    "授课方式标识": "授课方式标识",
    "课程": "课程",
    "课节科目": "课节科目",
}
HEADER_ROW = 2
DATA_START_ROW = 3


def read_result_rows(path: Path) -> List[Dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    # ERP result files can incorrectly declare dimension=A1 while the XML has full rows.
    worksheet.reset_dimensions()
    raw_rows = worksheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(raw_rows)]
    rows: List[Dict[str, str]] = []
    for raw in raw_rows:
        row = {headers[index]: clean(raw[index]) if index < len(raw) else "" for index in range(len(headers))}
        if clean(row.get("课次ID")):
            rows.append(row)
    return rows


def output_row(source: Dict[str, str]) -> Dict[str, str]:
    row = {header: clean(source.get(INPUT_TO_OUTPUT[header])) for header in ERP_HEADERS}
    if not row["授课方式标识"]:
        row["授课方式标识"] = "直播课"
    return row


def validate_template_headers(worksheet) -> Dict[str, int]:
    headers = {clean(worksheet.cell(HEADER_ROW, col).value): col for col in range(1, worksheet.max_column + 1)}
    missing = [header for header in ERP_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"ERP模板缺少列: {', '.join(missing)}")
    return headers


def clear_template_rows(worksheet) -> None:
    if worksheet.max_row >= DATA_START_ROW:
        worksheet.delete_rows(DATA_START_ROW, worksheet.max_row - DATA_START_ROW + 1)


def write_workbook(template: Path, output_path: Path, rows: Sequence[Dict[str, str]]) -> None:
    workbook = load_workbook(template)
    worksheet = workbook.worksheets[0]
    headers = validate_template_headers(worksheet)
    clear_template_rows(worksheet)
    for offset, row in enumerate(rows, start=DATA_START_ROW):
        for header in ERP_HEADERS:
            worksheet.cell(offset, headers[header]).value = clean(row.get(header))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def load_suite_codes(path: Path) -> Dict[str, str]:
    if not path.exists():
        return {}
    return {clean(row.get("id")): clean(row.get("suite_code")) for row in read_csv_rows(path) if clean(row.get("id"))}


def write_report(
    path: Path,
    source: Path,
    output_xlsx: Path,
    output_csv: Path,
    order_map: Path,
    rows: Sequence[Dict[str, str]],
    source_rows: Sequence[Dict[str, str]],
) -> None:
    class_counter = Counter(row["班级编码"] for row in rows)
    subject_counter = Counter(row["课节科目"] for row in rows)
    lines = [
        f"# ERP 导入失败课次倒序重导 {path.stem.rsplit('_', 1)[-1]}",
        "",
        f"- 来源导入结果: `{source}`",
        f"- 输出课次数: {len(rows)}",
        "- 导入顺序: 按导入结果明细原顺序整体倒序",
        f"- 导入表: `{output_xlsx}`",
        f"- CSV备份: `{output_csv}`",
        f"- 顺序映射: `{order_map}`",
        "",
        "## 倒序校验",
        "",
    ]
    if rows:
        lines.append(f"- 新第1条课次ID: {rows[0]['*课次ID']}，原所在行: {source_rows[-1].get('所在行', '')}")
        lines.append(f"- 新第{len(rows)}条课次ID: {rows[-1]['*课次ID']}，原所在行: {source_rows[0].get('所在行', '')}")
    lines.extend(["", "## 失败较多班级 TOP 20", "", "| 班级编码 | 课次数 |", "|---|---:|"])
    for class_id, count in class_counter.most_common(20):
        lines.append(f"| {class_id} | {count} |")
    lines.extend(["", "## 科目汇总", "", "| 科目 | 课次数 |", "|---|---:|"])
    for subject, count in subject_counter.most_common():
        lines.append(f"| {subject} | {count} |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="将ERP导入结果里的失败课次倒序写入导入模板。")
    parser.add_argument("--result-xlsx", type=Path, required=True)
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--classes", type=Path, default=DEFAULT_CLASSES)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--timestamp", default="")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    stamp = args.timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    source_rows = read_result_rows(args.result_xlsx)
    output_rows = [output_row(row) for row in reversed(source_rows)]
    output_xlsx = args.output_dir / f"erp_lesson_import_reverse_failed_{stamp}.xlsx"
    output_csv = args.output_dir / f"erp_lesson_import_reverse_failed_{stamp}.csv"
    order_map = args.output_dir / f"erp_lesson_import_reverse_failed_order_map_{stamp}.csv"
    report = args.output_dir / f"erp_lesson_import_reverse_failed_report_{stamp}.md"

    suite_codes = load_suite_codes(args.classes)
    order_rows: List[Dict[str, str]] = []
    for new_index, (out_row, source_row) in enumerate(zip(output_rows, reversed(source_rows)), start=1):
        class_id = out_row["班级编码"]
        order_rows.append(
            {
                "新导入顺序": str(new_index),
                "原失败所在行": clean(source_row.get("所在行")),
                "课次ID": out_row["*课次ID"],
                "日期": out_row["日期"],
                "时间": out_row["时间"],
                "班级编码": class_id,
                "套班编码": suite_codes.get(class_id, ""),
                "科目": out_row["课节科目"],
                "教师1": out_row["教师1（实际授课教师）"],
                "教室": out_row["教室"],
                "课程": out_row["课程"],
                "错误原因": clean(source_row.get("错误原因")),
            }
        )

    write_workbook(args.template, output_xlsx, output_rows)
    write_csv_rows(output_csv, ERP_HEADERS, output_rows, extrasaction="ignore")
    write_csv_rows(
        order_map,
        ["新导入顺序", "原失败所在行", "课次ID", "日期", "时间", "班级编码", "套班编码", "科目", "教师1", "教室", "课程", "错误原因"],
        order_rows,
        extrasaction="ignore",
    )
    write_report(report, args.result_xlsx, output_xlsx, output_csv, order_map, output_rows, source_rows)

    required_missing = Counter()
    for row in output_rows:
        for field in ("*课次ID", "日期", "时间", "班级编码", "授课方式标识"):
            if not row.get(field):
                required_missing[field] += 1
    print(f"source_failed_rows={len(source_rows)}")
    print(f"output_rows={len(output_rows)}")
    print(f"output_xlsx={output_xlsx}")
    print(f"output_csv={output_csv}")
    print(f"order_map={order_map}")
    print(f"report={report}")
    for field, count in required_missing.items():
        print(f"missing_{field}={count}")


if __name__ == "__main__":
    main()
