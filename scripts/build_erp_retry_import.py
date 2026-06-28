#!/usr/bin/env python3
"""Build a retry ERP import workbook from annotated failed rows."""

from __future__ import annotations

import argparse
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.schedule_data import load_class_metadata  # noqa: E402
from scripts.schedule_modes import assignment_reference_class_id, assignment_schedule_mode  # noqa: E402
from scripts.csv_utils import read_csv_rows, write_csv_rows  # noqa: E402
from scripts.field_utils import normalize_excel_text as clean  # noqa: E402


DEFAULT_FAILURES = Path("outputs/erp_import_failures_annotated_20260521_161645.csv")
DEFAULT_CLASSES = Path("data/classes.csv")
DEFAULT_TEACHER_ASSIGNMENTS = Path("data/class_teacher_assignments.csv")
DEFAULT_PRODUCT_COURSES = Path("data/product_courses.csv")
DEFAULT_OUTPUT_DIR = Path("outputs")
ONLINE_ROOM_ID = "RMHFWY97001"
ROOM_OVERRIDES = [
    {
        "suite_code": "2726",
        "start_date": "2026-07-01",
        "room_id": "RMHFWY01004",
        "room_name": "汇金403",
    },
    {
        "suite_code": "2727",
        "start_date": "2026-01-01",
        "room_id": "RMHFWY03022",
        "room_name": "环球金融209",
    },
]
HEADER_ROW = 2
DATA_START_ROW = 3
ERP_HEADERS = [
    "*课次ID",
    "日期",
    "时间",
    "班级编码",
    "教师1（实际授课教师）",
    "教师2",
    "教室",
    "事件",
    "备注",
    "授课方式标识",
    "课程",
    "课节科目",
]
POLITICS_SPLIT_CODES = {
    "CSHFWY000300285|CSHFWY000300288": {
        "PM1": ("CSHFWY000300285", "马原"),
        "PM2": ("CSHFWY000300288", "思修"),
        "14:00~16:00": ("CSHFWY000300285", "马原"),
        "16:20~18:20": ("CSHFWY000300288", "思修"),
    },
    "CSHFWY000300286|CSHFWY000300454": {
        "PM1": ("CSHFWY000300286", "毛中特"),
        "PM2": ("CSHFWY000300454", "新思想"),
        "14:00~16:00": ("CSHFWY000300286", "毛中特"),
        "16:20~18:20": ("CSHFWY000300454", "新思想"),
    },
}


def date_key(value: str) -> str:
    text = clean(value).replace("/", "-").replace(".", "-")
    parts = text.split("-")
    if len(parts) != 3:
        return text
    try:
        year, month, day = (int(part) for part in parts)
    except ValueError:
        return text
    return f"{year:04d}-{month:02d}-{day:02d}"


def split_remark(remark: str) -> Tuple[str, str, str]:
    parts = [part.strip() for part in remark.split("/") if part.strip()]
    if len(parts) >= 3:
        return parts[0], parts[1], parts[2]
    if len(parts) == 2:
        return "", parts[0], parts[1]
    if len(parts) == 1:
        return "", "", parts[0]
    return "", "", ""


def load_course_names(path: Path) -> Dict[str, str]:
    names: Dict[str, str] = {}
    for row in read_csv_rows(path):
        code = clean(row.get("course_code"))
        if code:
            names[code] = clean(row.get("course_name"))
    return names


def group_matches(row_group: str, assignment_group: str) -> bool:
    return not assignment_group or not row_group or row_group == assignment_group


def load_merge_modes(path: Path) -> Tuple[Dict[Tuple[str, str, str, str], str], set[Tuple[str, str, str, str]]]:
    modes: Dict[Tuple[str, str, str, str], str] = {}
    inherited_main_keys: set[Tuple[str, str, str, str]] = set()
    for row in read_csv_rows(path):
        class_id = clean(row.get("class_id"))
        subject = clean(row.get("subject"))
        stage = clean(row.get("stage"))
        group = clean(row.get("course_group"))
        mode = assignment_schedule_mode(row)
        inherited = assignment_reference_class_id(row)
        if class_id and subject and stage and mode:
            modes[(class_id, subject, stage, group)] = mode
        if inherited and subject and stage:
            inherited_main_keys.add((inherited, subject, stage, group))
    return modes, inherited_main_keys


def merge_mode_for_row(
    row: Dict[str, str],
    modes: Dict[Tuple[str, str, str, str], str],
    inherited_main_keys: set[Tuple[str, str, str, str]],
) -> str:
    class_id = clean(row.get("班级编码"))
    subject = clean(row.get("课节科目"))
    quarter, stage, _module = split_remark(clean(row.get("备注")))
    group = ""
    stage_candidates = [stage, quarter]
    for candidate_stage in stage_candidates:
        if not candidate_stage:
            continue
        for (mode_class_id, mode_subject, mode_stage, mode_group), mode in modes.items():
            if mode_class_id == class_id and mode_subject == subject and mode_stage == candidate_stage and group_matches(group, mode_group):
                return mode
        for main_class_id, main_subject, main_stage, main_group in inherited_main_keys:
            if main_class_id == class_id and main_subject == subject and main_stage == candidate_stage and group_matches(group, main_group):
                return "合班主班"
    return ""


def is_wuyou_math_class(row: Dict[str, str], class_meta: Dict[str, Dict[str, str]]) -> bool:
    class_id = clean(row.get("班级编码"))
    meta = class_meta.get(class_id, {})
    return (
        clean(meta.get("product_line")) == "考研无忧"
        and clean(meta.get("subject")) == "数学"
    )


def split_politics_course(row: Dict[str, str], course_names: Dict[str, str], actions: List[str]) -> None:
    code = clean(row.get("课程"))
    if code not in POLITICS_SPLIT_CODES:
        return
    slot_key = clean(row.get("时间"))
    if slot_key not in POLITICS_SPLIT_CODES[code]:
        # Fall back to row order by start time when lesson-slot is not available in ERP failure output.
        return
    new_code, module = POLITICS_SPLIT_CODES[code][slot_key]
    quarter, stage, _old_module = split_remark(clean(row.get("备注")))
    row["课程"] = new_code
    row["备注"] = "/".join(part for part in [quarter, stage, module] if part)
    actions.append(f"组合课程拆分为 {module}({new_code})")
    if course_names.get(new_code):
        actions.append(f"课程名称={course_names[new_code]}")


def apply_room_overrides(
    row: Dict[str, str],
    class_meta: Dict[str, Dict[str, str]],
    merge_mode: str,
    actions: List[str],
) -> None:
    if merge_mode == "共享课表":
        return
    class_id = clean(row.get("班级编码"))
    meta = class_meta.get(class_id, {})
    suite_code = clean(meta.get("suite_code"))
    row_date = date_key(clean(row.get("日期")))
    for override in ROOM_OVERRIDES:
        if suite_code != override["suite_code"]:
            continue
        if row_date < override["start_date"]:
            continue
        if clean(row.get("教室")) != override["room_id"]:
            row["教室"] = override["room_id"]
            actions.append(
                f"{suite_code}班{override['start_date']}起教室改为"
                f"{override['room_name']}({override['room_id']})"
            )
        return


def apply_rules(
    source_row: Dict[str, str],
    class_meta: Dict[str, Dict[str, str]],
    merge_modes: Dict[Tuple[str, str, str, str], str],
    inherited_main_keys: set[Tuple[str, str, str, str]],
    course_names: Dict[str, str],
) -> Tuple[Dict[str, str], List[str]]:
    row = dict(source_row)
    error_type = clean(row.get("错误类型"))
    actions: List[str] = []
    merge_mode = merge_mode_for_row(row, merge_modes, inherited_main_keys)
    original_room = clean(row.get("教室"))

    if merge_mode == "共享课表":
        if clean(row.get("教师1（实际授课教师）")) or clean(row.get("教室")) or clean(row.get("课程")):
            actions.append("共享课表从班清空教师/教室/课程")
        row["教师1（实际授课教师）"] = ""
        row["教室"] = ""
        row["课程"] = ""
    else:
        if "教室不存在" in error_type:
            if original_room.startswith("RMONLINE") and merge_mode in {"合班主班", ""}:
                row["教室"] = ONLINE_ROOM_ID
                actions.append(f"线上课教室改为 {ONLINE_ROOM_ID}")
            else:
                row["教室"] = ""
                actions.append("线下不存在教室先清空教室")
        if "教室时间冲突" in error_type:
            if clean(row.get("教室")):
                row["教室"] = ""
                actions.append("教室时间冲突清空教室")
        if "老师时间冲突" in error_type:
            if clean(row.get("教师1（实际授课教师）")):
                row["教师1（实际授课教师）"] = ""
                actions.append("老师时间冲突清空教师1")
        split_politics_course(row, course_names, actions)

    if "班级科目不匹配" in error_type and is_wuyou_math_class(row, class_meta):
        if clean(row.get("课节科目")) != "数学一":
            row["课节科目"] = "数学一"
            actions.append("无忧数学课节科目改为数学一")

    apply_room_overrides(row, class_meta, merge_mode, actions)

    # Explicit no-op rules kept for traceability.
    if "班级课次时段冲突" in error_type:
        actions.append("班级课次时段冲突按要求不调整")
    if "教室容量不足" in error_type:
        actions.append("教室容量不足按要求不调整")

    return row, actions


def validate_template_headers(ws) -> Dict[str, int]:
    headers = {clean(ws.cell(HEADER_ROW, col).value): col for col in range(1, ws.max_column + 1)}
    missing = [header for header in ERP_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"ERP模板缺少列：{', '.join(missing)}")
    return headers


def clear_template_rows(ws) -> None:
    if ws.max_row >= DATA_START_ROW:
        ws.delete_rows(DATA_START_ROW, ws.max_row - DATA_START_ROW + 1)


def output_values(row: Dict[str, str]) -> List[str]:
    return [
        clean(row.get("课次ID")),
        clean(row.get("日期")),
        clean(row.get("时间")),
        clean(row.get("班级编码")),
        clean(row.get("教师1（实际授课教师）")),
        clean(row.get("教师2")),
        clean(row.get("教室")),
        clean(row.get("事件")),
        clean(row.get("备注")),
        clean(row.get("授课方式标识")) or "直播课",
        clean(row.get("课程")),
        clean(row.get("课节科目")),
    ]


def write_retry_workbook(template: Path, output_path: Path, rows: Sequence[Dict[str, str]]) -> None:
    workbook = load_workbook(template)
    worksheet = workbook.worksheets[0]
    headers = validate_template_headers(worksheet)
    clear_template_rows(worksheet)
    for offset, row in enumerate(rows, start=DATA_START_ROW):
        for header, value in zip(ERP_HEADERS, output_values(row)):
            worksheet.cell(offset, headers[header]).value = value
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_report(path: Path, source: Path, output_xlsx: Path, rows: Sequence[Dict[str, str]], log_rows: Sequence[Dict[str, str]]) -> None:
    action_counter: Counter[str] = Counter()
    for item in log_rows:
        for action in clean(item.get("修正动作")).split("；"):
            if action:
                action_counter[action] += 1
    lines = [
        "# ERP 二次导入表修正报告",
        "",
        f"- 来源失败标注: {source}",
        f"- 二次导入表: {output_xlsx}",
        f"- 输出课次数: {len(rows)}",
        "",
        "## 修正动作统计",
        "",
        "| 修正动作 | 课次数 |",
        "|---|---:|",
    ]
    for action, count in action_counter.most_common():
        lines.append(f"| {action} | {count} |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="按 ERP 失败原因生成二次导入表")
    parser.add_argument("--failures", type=Path, default=DEFAULT_FAILURES)
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--classes", type=Path, default=DEFAULT_CLASSES)
    parser.add_argument("--teacher-assignments", type=Path, default=DEFAULT_TEACHER_ASSIGNMENTS)
    parser.add_argument("--product-courses", type=Path, default=DEFAULT_PRODUCT_COURSES)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--timestamp", default="")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    stamp = args.timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    failures = read_csv_rows(args.failures)
    class_meta = load_class_metadata(args.classes)
    merge_modes, inherited_main_keys = load_merge_modes(args.teacher_assignments)
    course_names = load_course_names(args.product_courses)

    output_rows: List[Dict[str, str]] = []
    log_rows: List[Dict[str, str]] = []
    for source in failures:
        row, actions = apply_rules(source, class_meta, merge_modes, inherited_main_keys, course_names)
        output_rows.append(row)
        log_rows.append(
            {
                "课次ID": clean(row.get("课次ID")),
                "班级编码": clean(row.get("班级编码")),
                "日期": clean(row.get("日期")),
                "时间": clean(row.get("时间")),
                "错误类型": clean(source.get("错误类型")),
                "原教师1": clean(source.get("教师1（实际授课教师）")),
                "新教师1": clean(row.get("教师1（实际授课教师）")),
                "原教室": clean(source.get("教室")),
                "新教室": clean(row.get("教室")),
                "原课程": clean(source.get("课程")),
                "新课程": clean(row.get("课程")),
                "原课节科目": clean(source.get("课节科目")),
                "新课节科目": clean(row.get("课节科目")),
                "修正动作": "；".join(actions) if actions else "未调整，按要求重导",
            }
        )

    output_xlsx = args.output_dir / f"erp_lesson_import_retry_{stamp}.xlsx"
    output_csv = args.output_dir / f"erp_lesson_import_retry_{stamp}.csv"
    log_csv = args.output_dir / f"erp_lesson_import_retry_adjustments_{stamp}.csv"
    report = args.output_dir / f"erp_lesson_import_retry_report_{stamp}.md"

    write_retry_workbook(args.template, output_xlsx, output_rows)
    write_csv_rows(output_csv, ERP_HEADERS, [dict(zip(ERP_HEADERS, output_values(row))) for row in output_rows])
    write_csv_rows(log_csv, list(log_rows[0].keys()) if log_rows else [], log_rows)
    write_report(report, args.failures, output_xlsx, output_rows, log_rows)

    action_counter: Counter[str] = Counter()
    for item in log_rows:
        for action in clean(item.get("修正动作")).split("；"):
            if action:
                action_counter[action] += 1
    print(f"rows={len(output_rows)}")
    print(f"output_xlsx={output_xlsx}")
    print(f"output_csv={output_csv}")
    print(f"adjustments={log_csv}")
    print(f"report={report}")
    for action, count in action_counter.most_common():
        print(f"{action}={count}")


if __name__ == "__main__":
    main()
