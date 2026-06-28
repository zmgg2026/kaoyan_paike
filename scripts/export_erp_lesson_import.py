#!/usr/bin/env python3
"""Export the maintenance schedule into the ERP lesson-import workbook shape."""

from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.schedule_modes import assignment_reference_class_id, assignment_schedule_mode  # noqa: E402


DEFAULT_SCHEDULE = Path("outputs/batch_schedule_maintenance.csv")
DEFAULT_OUTPUT_DIR = Path("outputs")
DEFAULT_TEACHER_ASSIGNMENTS = Path("data/class_teacher_assignments.csv")
DEFAULT_PRODUCT_COURSES = Path("data/product_courses.csv")
HEADER_ROW = 2
DATA_START_ROW = 3
PUBLIC_SUBJECTS = {"英语", "政治", "数学", "语文"}

ERP_HEADERS = [
    "*课次ID",
    "日期",
    "时间",
    "班级编码",
    "教师1（实际授课教师）",
    "教师2",
    "教室",
    "事件",
    "备注",
    "授课方式标识",
    "课程",
    "课节科目",
]


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_csv(path: Path) -> List[Dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert generated lessons into the ERP import template.")
    parser.add_argument("--template", type=Path, required=True, help="ERP import template xlsx.")
    parser.add_argument("--schedule-csv", type=Path, default=DEFAULT_SCHEDULE, help="Generated schedule CSV.")
    parser.add_argument("--lesson-id-map", type=Path, help="Optional CSV: class_id,date,start_time,end_time,erp_lesson_id.")
    parser.add_argument("--teacher-assignments", type=Path, default=DEFAULT_TEACHER_ASSIGNMENTS, help="Class teacher assignment CSV used to identify shared merge classes.")
    parser.add_argument("--product-courses", type=Path, default=DEFAULT_PRODUCT_COURSES, help="Product course CSV used to backfill missing course codes.")
    parser.add_argument("--start-date", default="2026-07-01")
    parser.add_argument("--end-date", default="2026-12-31")
    parser.add_argument("--include-subjects", default="|".join(sorted(PUBLIC_SUBJECTS)), help="Subjects to export, separated by | or comma. Default excludes fixed professional classes.")
    parser.add_argument(
        "--room-field",
        choices=["room_name", "room_id"],
        default="room_id",
        help="Field written to ERP 教室 column. Default uses room_id because ERP import requires room codes.",
    )
    parser.add_argument("--output-xlsx", type=Path)
    parser.add_argument("--gap-report", type=Path)
    return parser.parse_args()


def load_lesson_id_map(path: Optional[Path]) -> Dict[Tuple[str, str, str, str], str]:
    if not path:
        return {}
    rows = read_csv(path)
    mapping: Dict[Tuple[str, str, str, str], str] = {}
    for row in rows:
        class_id = clean(row.get("class_id") or row.get("班级编码"))
        date = normalize_date(clean(row.get("date") or row.get("日期")))
        start_time = normalize_time(clean(row.get("start_time") or row.get("开始时间")))
        end_time = normalize_time(clean(row.get("end_time") or row.get("结束时间")))
        lesson_id = clean(row.get("erp_lesson_id") or row.get("课次ID") or row.get("*课次ID"))
        if class_id and date and start_time and end_time and lesson_id:
            mapping[(class_id, date, start_time, end_time)] = lesson_id
    return mapping


def normalize_date(value: str) -> str:
    if not value:
        return ""
    value = value.replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y%m%d"):
        try:
            return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return value


def display_date(value: str) -> str:
    normalized = normalize_date(value)
    try:
        return datetime.strptime(normalized, "%Y-%m-%d").strftime("%Y/%m/%d")
    except ValueError:
        return value.replace("-", "/")


def normalize_time(value: str) -> str:
    if not value:
        return ""
    value = value.strip()
    if "~" in value:
        value = value.split("~", 1)[0]
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(value, fmt).strftime("%H:%M")
        except ValueError:
            pass
    return value


def split_values(value: str) -> List[str]:
    return [item.strip() for item in value.replace(",", "|").replace("，", "|").split("|") if item.strip()]


def load_course_code_lookup(path: Path) -> Dict[Tuple[str, ...], str]:
    if not path.exists():
        return {}
    rows = read_csv(path)
    candidates: Dict[Tuple[str, ...], set[str]] = {}
    for row in rows:
        code = clean(row.get("course_code") or row.get("课程编码"))
        if not code:
            continue
        course_name = clean(row.get("course_name") or row.get("课程名称"))
        subject = clean(row.get("subject") or row.get("科目"))
        stage = clean(row.get("stage") or row.get("阶段"))
        module = clean(row.get("course_module") or row.get("课程模块"))
        group = clean(row.get("course_group") or row.get("课程类别"))
        for key in (
            ("name", course_name),
            ("subject_stage_module_name", subject, stage, module, course_name),
            ("subject_stage_module_group", subject, stage, module, group),
        ):
            if all(key[1:]):
                candidates.setdefault(key, set()).add(code)
    return {key: next(iter(codes)) for key, codes in candidates.items() if len(codes) == 1}


def backfill_course_code(row: Dict[str, str], lookup: Dict[Tuple[str, ...], str]) -> str:
    existing = clean(row.get("course_code"))
    if existing:
        return existing
    course_name = clean(row.get("course_name"))
    subject = clean(row.get("subject"))
    stage = clean(row.get("stage"))
    module = clean(row.get("course_module"))
    group = clean(row.get("course_group"))
    for key in (
        ("subject_stage_module_name", subject, stage, module, course_name),
        ("name", course_name),
        ("subject_stage_module_group", subject, stage, module, group),
    ):
        if all(key[1:]) and key in lookup:
            return lookup[key]
    return ""


def load_shared_class_keys(path: Path) -> Dict[Tuple[str, str, str, str], str]:
    if not path.exists():
        return {}
    shared: Dict[Tuple[str, str, str, str], str] = {}
    for row in read_csv(path):
        if assignment_schedule_mode(row) != "共享课表":
            continue
        class_id = clean(row.get("class_id") or row.get("班级编码"))
        subject = clean(row.get("subject") or row.get("科目"))
        stage = clean(row.get("stage") or row.get("阶段"))
        group = clean(row.get("course_group") or row.get("课程类别"))
        inherited = assignment_reference_class_id(row)
        if class_id and subject and stage:
            shared[(class_id, subject, stage, group)] = inherited
    return shared


def is_shared_merge_row(row: Dict[str, str], shared_keys: Dict[Tuple[str, str, str, str], str]) -> bool:
    class_id = clean(row.get("class_id"))
    subject = clean(row.get("subject"))
    quarter = clean(row.get("quarter"))
    stage = clean(row.get("stage"))
    group = clean(row.get("course_group"))
    stage_candidates = [stage]
    if quarter and quarter not in stage_candidates:
        stage_candidates.append(quarter)
    return (
        any((class_id, subject, candidate, group) in shared_keys for candidate in stage_candidates)
        or any((class_id, subject, candidate, "") in shared_keys for candidate in stage_candidates)
    )


def remark(row: Dict[str, str]) -> str:
    parts = [
        clean(row.get("quarter")),
        clean(row.get("stage")),
        clean(row.get("course_module")),
    ]
    text = "/".join(part for part in parts if part)
    return text[:80]


def schedule_rows(rows: Iterable[Dict[str, str]], start_date: str, end_date: str, include_subjects: set[str]) -> List[Dict[str, str]]:
    selected = []
    for row in rows:
        date = normalize_date(clean(row.get("date")))
        subject = clean(row.get("subject"))
        if start_date <= date <= end_date and (not include_subjects or subject in include_subjects):
            selected.append(row)
    selected.sort(key=lambda r: (normalize_date(clean(r.get("date"))), clean(r.get("start_time")), clean(r.get("class_id")), clean(r.get("lesson_slot"))))
    return selected


def build_export_row(
    row: Dict[str, str],
    lesson_id_map: Dict[Tuple[str, str, str, str], str],
    room_field: str,
    shared_keys: Dict[Tuple[str, str, str, str], str],
    course_code_lookup: Dict[Tuple[str, ...], str],
) -> Tuple[List[str], List[Dict[str, str]]]:
    date = normalize_date(clean(row.get("date")))
    start_time = normalize_time(clean(row.get("start_time")))
    end_time = normalize_time(clean(row.get("end_time")))
    class_id = clean(row.get("class_id"))
    lesson_id = lesson_id_map.get((class_id, date, start_time, end_time), "")
    shared_merge_row = is_shared_merge_row(row, shared_keys)
    teacher_id = "" if shared_merge_row else clean(row.get("teacher_id"))
    room_value = "" if shared_merge_row else clean(row.get(room_field))
    course_code = "" if shared_merge_row else backfill_course_code(row, course_code_lookup)

    values = [
        lesson_id,
        display_date(date),
        f"{start_time}~{end_time}" if start_time and end_time else "",
        class_id,
        teacher_id,
        "",
        room_value,
        "",
        remark(row),
        "直播课",
        course_code,
        clean(row.get("subject")),
    ]

    gaps: List[Dict[str, str]] = []
    gap_context = {
        "class_id": class_id,
        "date": date,
        "start_time": start_time,
        "end_time": end_time,
        "subject": clean(row.get("subject")),
        "stage": clean(row.get("stage")),
        "course_module": clean(row.get("course_module")),
        "teacher_id": teacher_id,
        "teacher_name": clean(row.get("teacher_name")),
        "course_code": course_code,
        "course_name": clean(row.get("course_name")),
        "is_shared_merge_row": "是" if shared_merge_row else "否",
    }
    if not lesson_id:
        gaps.append({**gap_context, "gap_type": "missing_erp_lesson_id", "message": "模板要求课次ID必填，当前课表没有ERP课次ID。"})
    if not class_id:
        gaps.append({**gap_context, "gap_type": "missing_class_id", "message": "班级编码为空。"})
    if not teacher_id and not shared_merge_row:
        gaps.append({**gap_context, "gap_type": "missing_teacher_id", "message": "教师1为空，ERP可能校验失败。"})
    if not course_code and not shared_merge_row:
        gaps.append({**gap_context, "gap_type": "missing_course_code", "message": "课程编码为空，模板建议填课程编码。"})
    return values, gaps


def validate_template_headers(ws) -> Dict[str, int]:
    headers = {clean(ws.cell(HEADER_ROW, col).value): col for col in range(1, ws.max_column + 1)}
    missing = [header for header in ERP_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"ERP模板缺少列：{', '.join(missing)}")
    return headers


def clear_template_rows(ws) -> None:
    if ws.max_row >= DATA_START_ROW:
        ws.delete_rows(DATA_START_ROW, ws.max_row - DATA_START_ROW + 1)


def write_gap_report(path: Path, rows: List[Dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "gap_type",
        "message",
        "class_id",
        "date",
        "start_time",
        "end_time",
        "subject",
        "stage",
        "course_module",
        "teacher_id",
        "teacher_name",
        "course_code",
        "course_name",
        "is_shared_merge_row",
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_xlsx = args.output_xlsx or DEFAULT_OUTPUT_DIR / f"erp_lesson_import_draft_{timestamp}.xlsx"
    gap_report = args.gap_report or DEFAULT_OUTPUT_DIR / f"erp_lesson_import_gap_report_{timestamp}.csv"

    if not args.template.exists():
        raise FileNotFoundError(args.template)
    if not args.schedule_csv.exists():
        raise FileNotFoundError(args.schedule_csv)

    include_subjects = set(split_values(args.include_subjects))
    rows = schedule_rows(read_csv(args.schedule_csv), args.start_date, args.end_date, include_subjects)
    lesson_id_map = load_lesson_id_map(args.lesson_id_map)
    shared_keys = load_shared_class_keys(args.teacher_assignments)
    course_code_lookup = load_course_code_lookup(args.product_courses)

    workbook = load_workbook(args.template)
    ws = workbook.worksheets[0]
    headers = validate_template_headers(ws)
    clear_template_rows(ws)

    all_gaps: List[Dict[str, str]] = []
    shared_blank_count = 0
    for offset, source_row in enumerate(rows, start=DATA_START_ROW):
        values, gaps = build_export_row(source_row, lesson_id_map, args.room_field, shared_keys, course_code_lookup)
        if values[4] == "" and values[6] == "" and values[10] == "" and is_shared_merge_row(source_row, shared_keys):
            shared_blank_count += 1
        all_gaps.extend(gaps)
        for header, value in zip(ERP_HEADERS, values):
            ws.cell(offset, headers[header]).value = value

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)
    write_gap_report(gap_report, all_gaps)

    counts = Counter(gap["gap_type"] for gap in all_gaps)
    print(f"exported_rows={len(rows)}")
    print(f"shared_merge_rows_blanked={shared_blank_count}")
    print(f"output_xlsx={output_xlsx}")
    print(f"gap_report={gap_report}")
    for key in sorted(counts):
        print(f"{key}={counts[key]}")


if __name__ == "__main__":
    main()
