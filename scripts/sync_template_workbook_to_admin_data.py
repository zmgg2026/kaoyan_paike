#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

import data_admin_server
from scripts.csv_utils import serialize_csv_value, write_csv_rows
from scripts.field_utils import (
    normalize_date_text,
    normalize_text,
    normalize_time_text,
    parse_bool as normalize_bool,
    split_delimited_values,
)
from scripts.schedule_modes import (
    assignment_actual_scheduled_class_id,
    assignment_inherited_class_id,
    assignment_schedule_mode_value,
    class_schedule_mode_display_name,
    normalize_class_schedule_mode,
)
from scripts.table_schema import CLASS_JSON_EXTRA_FIELDNAMES, STANDARD_TABLE_FIELDNAMES
from scripts.template_tables import TEMPLATE_SHEET_ALIASES, TEMPLATE_SHEETS


DEFAULT_WORKBOOK = ROOT / "outputs" / "ai_scheduling_sop_20260625" / "AI排课基础数据模板.xlsx"
DATA_DIR = ROOT / "data"


SHEETS = TEMPLATE_SHEETS
SHEET_ALIASES = TEMPLATE_SHEET_ALIASES

LIST_FIELDS = {
    "season_window_ids",
    "applicable_stages",
    "selected_stages",
    "actual_schedule_window_ids",
    "preferred_teaching_area_ids",
    "preferred_room_ids",
    "allowed_weekdays",
    "allowed_periods",
    "weekdays",
    "periods",
    "schedule_window_ids",
    "class_ids",
    "product_ids",
    "product_name_keywords",
    "excluded_weekdays",
    "exception_weekdays",
    "teaching_area_ids",
    "stages",
    "class_name_keywords",
}

BOOL_FIELDS = {
    "is_active",
    "is_usable",
    "preferred_room_is_required",
    "is_manual_schedule_locked",
    "is_schedule_locked",
    "is_class_window_included",
    "is_conflict_group_active",
    "is_locked",
    "effective_after_class_start",
    "same_half_day_block_required",
    "same_half_day_4h_same_teacher_required",
    "is_enabled",
    "is_deleted",
}

CURRENT_TEACHER_ASSIGNMENT_FIELDS = [
    "class_id",
    "class_name",
    "product_id",
    "product_name",
    "subject",
    "stage",
    "course_group",
    "class_schedule_mode",
    "actual_scheduled_class_id",
    "teacher_id",
    "teacher_name",
    "assignment_extra_time_requirement",
    "notes",
]

INT_FIELDS = {
    "calendar_year",
    "window_year",
    "window_order",
    "order",
    "window_sequence",
    "stage_priority",
    "module_priority",
    "module_priority_in_group",
    "lessons_per_block",
    "max_blocks_per_class_per_day",
    "standard_capacity",
    "scheduling_capacity",
    "room_count",
    "active_room_count",
    "capacity",
    "size",
    "erp_duration_minutes",
    "erp_lesson_count",
    "erp_single_lesson_minutes",
    "duration_minutes",
    "lesson_count",
    "single_lesson_minutes",
}

FLOAT_FIELDS = {
    "duration_hours",
    "total_hours",
    "block_hours",
    "block_hours_override",
    "max_hours_per_class_per_day",
    "min_weekly_hours",
    "max_weekly_hours",
    "driving_distance_km",
    "travel_minutes",
}

DATE_FIELDS = {
    "date",
    "start_date",
    "end_date",
    "first_lesson_date",
    "earliest_date",
    "latest_date",
}
TIME_FIELDS = {"start_time", "end_time"}


def normalize_number(value: Any, *, integer: bool) -> int | float | str:
    if value in (None, ""):
        return 0 if integer else ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return normalize_text(value)
    return int(number) if integer else number


def normalize_list(value: Any) -> List[str]:
    return split_delimited_values(value)


def cell_value(field: str, value: Any) -> Any:
    if field in LIST_FIELDS:
        return normalize_list(value)
    if field in BOOL_FIELDS:
        return normalize_bool(value)
    if field in INT_FIELDS:
        return normalize_number(value, integer=True)
    if field in FLOAT_FIELDS:
        return normalize_number(value, integer=False)
    if field in DATE_FIELDS:
        return normalize_date_text(value)
    if field in TIME_FIELDS:
        return normalize_time_text(value)
    return normalize_text(value)


def read_sheet(workbook_path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    actual_sheet_name = sheet_name
    if actual_sheet_name not in wb.sheetnames:
        for alias in SHEET_ALIASES.get(sheet_name, []):
            if alias in wb.sheetnames:
                actual_sheet_name = alias
                break
    if actual_sheet_name not in wb.sheetnames:
        return []
    ws = wb[actual_sheet_name]
    headers = [normalize_text(value) for value in next(ws.iter_rows(min_row=6, max_row=6, values_only=True))]
    rows: List[Dict[str, Any]] = []
    for raw in ws.iter_rows(min_row=7, values_only=True):
        if not any(value not in (None, "") for value in raw):
            continue
        row = {
            field: cell_value(field, raw[index] if index < len(raw) else "")
            for index, field in enumerate(headers)
            if field
        }
        rows.append(row)
    return rows


def current_teacher_assignment_template_row(row: Dict[str, Any]) -> Dict[str, Any]:
    class_id = normalize_text(row.get("class_id"))
    inherited_class_id = assignment_inherited_class_id(row)
    actual_class_id = assignment_actual_scheduled_class_id(row) or inherited_class_id
    mode = normalize_class_schedule_mode(
        assignment_schedule_mode_value(row),
        inherit_from_class_id=inherited_class_id,
        actual_scheduled_class_id=actual_class_id,
        class_id=class_id,
    )
    is_shared = mode == "共享课表"
    actual_scheduled_class_id = actual_class_id if is_shared else class_id or actual_class_id
    result = {
        "class_id": class_id,
        "class_name": normalize_text(row.get("class_name")),
        "product_id": normalize_text(row.get("product_id") or row.get("canonical_product_id")),
        "product_name": normalize_text(row.get("product_name")),
        "subject": normalize_text(row.get("subject")),
        "stage": normalize_text(row.get("stage")),
        "course_group": normalize_text(row.get("course_group")),
        "class_schedule_mode": class_schedule_mode_display_name(mode),
        "actual_scheduled_class_id": actual_scheduled_class_id,
        "teacher_id": "" if is_shared else normalize_text(row.get("teacher_id")),
        "teacher_name": "" if is_shared else normalize_text(row.get("teacher_name")),
        "assignment_extra_time_requirement": normalize_text(row.get("assignment_extra_time_requirement")),
        "notes": normalize_text(row.get("notes")),
    }
    return {field: result.get(field, "") for field in CURRENT_TEACHER_ASSIGNMENT_FIELDS}


def enrich_rows(key: str, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    enriched: List[Dict[str, Any]] = []
    for row in rows:
        if key == "teachers":
            row.pop("id", None)
            row.pop("identity", None)
            row.pop("teacher_type", None)
        elif key == "product_courses":
            if not normalize_text(row.get("window_name")):
                row["window_name"] = normalize_text(row.get("quarter"))
            if not row.get("module_priority_in_group"):
                row["module_priority_in_group"] = row.get("module_priority", 0)
            row.pop("quarter", None)
            row.pop("module_priority", None)
            row.pop("block_hours", None)
            row.pop("teaching_area_ids", None)
        elif key in {"locked_scheduled_lessons", "historical_scheduled_lessons", "scheduled_lessons"}:
            if not normalize_text(row.get("window_name")):
                row["window_name"] = normalize_text(row.get("quarter"))
            row.pop("quarter", None)
        elif key == "product_schedule_rules":
            product_id = normalize_text(row.get("product_id"))
            product_ids = normalize_list(row.get("product_ids"))
            if not product_id and len(product_ids) == 1:
                row["product_id"] = product_ids[0]
            if not row.get("block_hours"):
                row["block_hours"] = row.get("block_hours_override", "")
            for old_field in (
                "rule_name",
                "scope_type",
                "product_ids",
                "product_name_keywords",
                "subject",
                "stage",
                "course_module",
                "course_group",
                "start_date",
                "end_date",
                "excluded_weekdays",
                "exception_weekdays",
                "block_hours_override",
            ):
                row.pop(old_field, None)
        elif key == "products":
            row.setdefault("season_window_ids", [])
            row.setdefault("applicable_stages", [])
        elif key == "classes":
            if not normalize_list(row.get("selected_stages")):
                row["selected_stages"] = normalize_list(row.get("stages"))
            if row.get("is_manual_schedule_locked") in ("", None):
                row["is_manual_schedule_locked"] = row.get("is_schedule_locked", False)
            row.pop("actual_schedule_window_ids", None)
            row.pop("stages", None)
            row.pop("is_schedule_locked", None)
        elif key == "business_product_mappings":
            if not normalize_text(row.get("local_product_id")):
                row["local_product_id"] = normalize_text(row.get("canonical_product_id"))
            row.pop("canonical_product_id", None)
        elif key == "class_conflict_groups":
            if row.get("is_conflict_group_active") in ("", None):
                row["is_conflict_group_active"] = row.get("is_active", True)
            if not normalize_text(row.get("conflict_source")):
                row["conflict_source"] = normalize_text(row.get("source")) or "手动"
            row.pop("is_active", None)
            row.pop("source", None)
        elif key == "class_teacher_assignments":
            enriched.append(current_teacher_assignment_template_row(row))
            continue
        enriched.append(row)
    return enriched


def merge_assignments_into_classes(classes: List[Dict[str, Any]], assignments: List[Dict[str, Any]]) -> None:
    by_class: Dict[str, List[Dict[str, Any]]] = {}
    for assignment in assignments:
        class_id = normalize_text(assignment.get("class_id"))
        if not class_id:
            continue
        payload = {
            key: value
            for key, value in assignment.items()
            if key not in {"class_id", "class_name"}
        }
        by_class.setdefault(class_id, []).append(payload)
    for cls in classes:
        cls["teacher_assignments"] = by_class.get(normalize_text(cls.get("id")), [])
        cls.setdefault("requirements", [])


def add_missing_preferred_room_placeholders(tables: Dict[str, List[Dict[str, Any]]]) -> None:
    """Keep required class room references visible when ERP room export is missing them."""
    rooms = tables.get("rooms", [])
    room_ids = {normalize_text(room.get("id")) for room in rooms if normalize_text(room.get("id"))}
    areas = {
        normalize_text(area.get("id")): area
        for area in tables.get("teaching_areas", [])
        if normalize_text(area.get("id"))
    }
    missing: Dict[str, Dict[str, Any]] = {}

    def remember(room_id: str, area_ids: List[str], source: str) -> None:
        room_id = normalize_text(room_id)
        if not room_id or room_id in room_ids:
            return
        area_id = area_ids[0] if area_ids else ""
        missing.setdefault(
            room_id,
            {
                "room_id": room_id,
                "teaching_area_id": area_id,
                "sources": set(),
            },
        )
        if area_id and not missing[room_id]["teaching_area_id"]:
            missing[room_id]["teaching_area_id"] = area_id
        missing[room_id]["sources"].add(source)

    for cls in tables.get("classes", []):
        area_ids = normalize_list(cls.get("preferred_teaching_area_ids"))
        for room_id in normalize_list(cls.get("preferred_room_ids")):
            remember(room_id, area_ids, f"班级 {normalize_text(cls.get('id'))}")

    for boundary in tables.get("class_window_boundaries", []):
        area_ids = normalize_list(boundary.get("preferred_teaching_area_ids"))
        for room_id in normalize_list(boundary.get("preferred_room_ids")):
            remember(room_id, area_ids, f"班级窗口 {normalize_text(boundary.get('class_window_id'))}")

    for room_id in sorted(missing):
        item = missing[room_id]
        area = areas.get(item["teaching_area_id"], {})
        area_name = normalize_text(area.get("short_name") or area.get("name") or item["teaching_area_id"])
        rooms.append(
            {
                "id": room_id,
                "name": f"{area_name} 待核对教室 {room_id}",
                "teaching_area_id": item["teaching_area_id"],
                "teaching_area_name": area_name,
                "campus": normalize_text(area.get("campus")),
                "capacity": 0,
                "room_type": "待核对",
                "is_active": False,
                "data_source": "班级/窗口指定教室补齐占位",
                "notes": "该教室ID出现在班级或班级窗口指定教室中，但04_教室表未找到；需人工核对ERP教室导出。引用来源："
                + "；".join(sorted(item["sources"])),
            }
        )
        room_ids.add(room_id)


def json_doc(key: str, rows: List[Dict[str, Any]], workbook_path: Path) -> Dict[str, Any]:
    return {
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source": str(workbook_path),
        "record_count": len(rows),
        key: rows,
    }


def write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    fields = output_fields_for_key("", csv_export=True, rows=rows)
    write_csv_rows(path, fields, rows, value_formatter=serialize_csv_value)


def output_fields_for_key(
    key: str,
    *,
    csv_export: bool,
    rows: List[Dict[str, Any]] | None = None,
) -> List[str]:
    if key in STANDARD_TABLE_FIELDNAMES:
        fields = list(STANDARD_TABLE_FIELDNAMES[key])
        if key == "classes" and not csv_export:
            fields.extend(CLASS_JSON_EXTRA_FIELDNAMES)
        return fields
    fields: List[str] = []
    seen = set()
    for row in rows or []:
        for field in row:
            if field not in seen:
                seen.add(field)
                fields.append(field)
    return fields


def standard_output_rows(key: str, rows: List[Dict[str, Any]], *, csv_export: bool = False) -> List[Dict[str, Any]]:
    if key not in STANDARD_TABLE_FIELDNAMES:
        return rows
    extra_fields = CLASS_JSON_EXTRA_FIELDNAMES if key == "classes" and not csv_export else ()
    return data_admin_server.standard_rows(rows, STANDARD_TABLE_FIELDNAMES[key], extra_fields)


def sync(workbook_path: Path) -> Dict[str, int]:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    tables: Dict[str, List[Dict[str, Any]]] = {}
    for sheet_name, key in SHEETS.items():
        rows = enrich_rows(key, read_sheet(workbook_path, sheet_name))
        tables[key] = rows

    merge_assignments_into_classes(tables["classes"], tables["class_teacher_assignments"])
    add_missing_preferred_room_placeholders(tables)

    for key, rows in tables.items():
        json_rows = standard_output_rows(key, rows, csv_export=False)
        csv_rows = standard_output_rows(key, rows, csv_export=True)
        (DATA_DIR / f"{key}.json").write_text(
            json.dumps(json_doc(key, json_rows, workbook_path), ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        write_csv_rows(
            DATA_DIR / f"{key}.csv",
            output_fields_for_key(key, csv_export=True),
            csv_rows,
            value_formatter=serialize_csv_value,
        )

    return {key: len(rows) for key, rows in tables.items()}


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync AI scheduling template workbook into web admin data files.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    args = parser.parse_args()
    counts = sync(args.workbook)
    for key in sorted(counts):
        print(f"{key}: {counts[key]}")


if __name__ == "__main__":
    main()
