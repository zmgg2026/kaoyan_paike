#!/usr/bin/env python3
"""Build a class schedule review workbook for ERP import failures."""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter, defaultdict
from datetime import date as Date, datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.csv_utils import read_csv_rows
from scripts.field_utils import (
    display_date_text as display_date,
    normalize_date_text as normalize_date,
    normalize_excel_text as clean,
    row_value,
)
from scripts.schedule_display import weekday_label


DEFAULT_SCHEDULE = Path("outputs/batch_schedule_maintenance.csv")
DEFAULT_CLASSES = Path("data/classes.csv")
DEFAULT_OUTPUT_DIR = Path("outputs")

SLOTS = [
    ("AM1", "上午一 08:00-10:00"),
    ("AM2", "上午二 10:20-12:20"),
    ("PM1", "下午一 14:00-16:00"),
    ("PM2", "下午二 16:20-18:20"),
    ("EVENING1", "晚上一 19:00-21:00"),
]
SUBJECT_FILL = {
    "英语": "DDEBFF",
    "政治": "FFE3DC",
    "数学": "E1F3DE",
    "语文": "F3E6FF",
}
HEADER_FILL = "D9EAF7"
FAILED_FILL = "FFF2CC"
RELATED_FILL = "E2F0D9"
THIN = Side(style="thin", color="C8D0D8")


def schedule_window_name(row: Dict[str, str]) -> str:
    return clean(row_value(row, "window_name", "quarter"))


def read_result_rows(path: Path) -> List[Dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    worksheet.reset_dimensions()
    raw_rows = worksheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(raw_rows)]
    rows: List[Dict[str, str]] = []
    for raw in raw_rows:
        row = {headers[index]: clean(raw[index]) if index < len(raw) else "" for index in range(len(headers))}
        if clean(row.get("课次ID")):
            rows.append(row)
    return rows


def load_classes(path: Path) -> Dict[str, Dict[str, str]]:
    return {clean(row.get("id")): row for row in read_csv_rows(path) if clean(row.get("id"))}


def row_slot(row: Dict[str, str]) -> str:
    slot = clean(row.get("lesson_slot"))
    if slot:
        return slot
    start = clean(row.get("start_time"))
    mapping = {
        "08:00": "AM1",
        "10:20": "AM2",
        "14:00": "PM1",
        "16:20": "PM2",
        "19:00": "EVENING1",
    }
    return mapping.get(start, clean(row.get("period")))


def result_slot(row: Dict[str, str]) -> str:
    text = clean(row.get("时间"))
    start = text.split("~", 1)[0] if "~" in text else text
    mapping = {
        "08:00": "AM1",
        "10:20": "AM2",
        "14:00": "PM1",
        "16:20": "PM2",
        "19:00": "EVENING1",
    }
    return mapping.get(start, start)


def suite_for_class(class_id: str, classes: Dict[str, Dict[str, str]]) -> str:
    return clean(classes.get(class_id, {}).get("suite_code"))


def extract_class_codes(rows: Sequence[Dict[str, str]]) -> Counter[str]:
    counter: Counter[str] = Counter()
    for row in rows:
        for code in re.findall(r"班级([A-Z0-9]+)", clean(row.get("错误原因"))):
            counter[code] += 1
    return counter


def failed_keys(rows: Sequence[Dict[str, str]]) -> set[Tuple[str, str, str, str]]:
    keys: set[Tuple[str, str, str, str]] = set()
    for row in rows:
        keys.add(
            (
                clean(row.get("班级编码")),
                normalize_date(row.get("日期")),
                clean(row.get("时间")),
                clean(row.get("课程")),
            )
        )
    return keys


def schedule_key(row: Dict[str, str]) -> Tuple[str, str, str, str]:
    time_text = f"{clean(row.get('start_time'))}~{clean(row.get('end_time'))}"
    return (clean(row.get("class_id")), normalize_date(row.get("date")), time_text, clean(row.get("course_code")))


def all_dates(start: str, end: str) -> List[str]:
    current = Date.fromisoformat(start)
    finish = Date.fromisoformat(end)
    dates: List[str] = []
    while current <= finish:
        dates.append(current.isoformat())
        current += timedelta(days=1)
    return dates


def append_sheet(ws, headers: Sequence[str], rows: Sequence[Sequence[object]]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    style_sheet(ws)


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    if ws.max_row and ws.max_column:
        ws.auto_filter.ref = ws.dimensions
    border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        values = [clean(ws.cell(row=row_idx, column=col_idx).value) for row_idx in range(1, min(ws.max_row, 200) + 1)]
        max_len = max((len(value) for value in values), default=8)
        ws.column_dimensions[letter].width = min(42, max(10, max_len + 2))


def format_matrix_sheet(ws) -> None:
    border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "C2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    for col in range(3, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24
    ws.row_dimensions[1].height = 34
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 120
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def lesson_text(row: Dict[str, str], is_failed: bool) -> str:
    prefix = "【失败】" if is_failed else ""
    subject = clean(row.get("subject"))
    class_id = clean(row.get("class_id"))
    stage = clean(row.get("stage"))
    module = clean(row.get("course_module"))
    teacher = clean(row.get("teacher_name"))
    room = clean(row.get("room_name"))
    course_name = clean(row.get("course_name"))
    return "\n".join(
        part
        for part in [
            f"{prefix}{class_id} {subject}",
            "/".join(part for part in [schedule_window_name(row), stage, module] if part),
            f"{teacher} {room}".strip(),
            course_name,
        ]
        if part
    )


def build_matrix_rows(
    schedule_rows: Sequence[Dict[str, str]],
    suite_code: str,
    failed_schedule_keys: set[Tuple[str, str, str, str]],
    start_date: str,
    end_date: str,
) -> Tuple[List[str], List[List[str]]]:
    by_date_slot: Dict[Tuple[str, str], List[Dict[str, str]]] = defaultdict(list)
    for row in schedule_rows:
        if clean(row.get("date")) < start_date or clean(row.get("date")) > end_date:
            continue
        by_date_slot[(normalize_date(row.get("date")), row_slot(row))].append(row)

    dates = all_dates(start_date, end_date)
    headers = [
        "时段",
        "起止时间",
        *[f"{display_date(date_text)}\n{weekday_label(date_text)}" for date_text in dates],
    ]
    rows: List[List[str]] = []
    for slot, label in SLOTS:
        if " " in label:
            slot_label, time_label = label.split(" ", 1)
        else:
            slot_label, time_label = label, ""
        values = [slot_label, time_label]
        for date_text in dates:
            lessons = sorted(by_date_slot.get((date_text, slot), []), key=lambda item: clean(item.get("class_id")))
            values.append("\n\n".join(lesson_text(row, schedule_key(row) in failed_schedule_keys) for row in lessons))
        rows.append(values)
    return headers, rows


def main() -> None:
    parser = argparse.ArgumentParser(description="生成ERP失败课次对应班级课表核对Excel。")
    parser.add_argument("--result-xlsx", type=Path, required=True)
    parser.add_argument("--schedule-csv", type=Path, default=DEFAULT_SCHEDULE)
    parser.add_argument("--classes", type=Path, default=DEFAULT_CLASSES)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--timestamp", default="")
    parser.add_argument("--start-date", default="2026-07-01")
    parser.add_argument("--end-date", default="2026-12-13")
    args = parser.parse_args()

    stamp = args.timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    result_rows = read_result_rows(args.result_xlsx)
    classes = load_classes(args.classes)
    schedule_rows_all = read_csv_rows(args.schedule_csv)
    failed_class_ids = {clean(row.get("班级编码")) for row in result_rows}
    mentioned_counter = extract_class_codes(result_rows)
    local_schedule_class_ids = {clean(row.get("class_id")) for row in schedule_rows_all}
    local_related_class_ids = {
        class_id for class_id in mentioned_counter if class_id in local_schedule_class_ids or class_id in classes
    }
    review_class_ids = failed_class_ids | local_related_class_ids
    review_suite_codes = {suite_for_class(class_id, classes) for class_id in review_class_ids}
    review_suite_codes.discard("")
    failed_key_set = failed_keys(result_rows)

    review_rows = [
        row
        for row in schedule_rows_all
        if clean(row.get("class_id")) in review_class_ids
        and args.start_date <= normalize_date(row.get("date")) <= args.end_date
    ]
    review_rows.sort(
        key=lambda row: (
            suite_for_class(clean(row.get("class_id")), classes),
            normalize_date(row.get("date")),
            clean(row.get("start_time")),
            clean(row.get("class_id")),
            clean(row.get("subject")),
        )
    )

    workbook = Workbook()
    summary = workbook.active
    summary.title = "汇总"
    summary_rows = [
        ["来源导入结果", str(args.result_xlsx)],
        ["失败课次数", len(result_rows)],
        ["失败班级数", len(failed_class_ids)],
        ["本地课表纳入班级数", len(review_class_ids)],
        ["本地课表纳入套班", "、".join(sorted(review_suite_codes))],
        ["课表范围", f"{args.start_date} 至 {args.end_date}"],
    ]
    append_sheet(summary, ["项目", "内容"], summary_rows)

    class_summary_rows = []
    for class_id in sorted(review_class_ids, key=lambda cid: (suite_for_class(cid, classes), cid)):
        meta = classes.get(class_id, {})
        class_summary_rows.append(
            [
                suite_for_class(class_id, classes),
                class_id,
                clean(meta.get("name")),
                clean(meta.get("sub_product")),
                clean(meta.get("subject")),
                "是" if class_id in failed_class_ids else "否",
                mentioned_counter.get(class_id, 0),
            ]
        )
    append_sheet(
        workbook.create_sheet("涉及班级"),
        ["套班编码", "班级编码", "班级名称", "子产品", "科目", "是否失败课次班级", "错误原因中出现次数"],
        class_summary_rows,
    )

    external_rows = []
    for class_id, count in mentioned_counter.most_common():
        if class_id in review_class_ids:
            continue
        external_rows.append([class_id, count, "本系统课表未纳入，可能是ERP中其他班级/课次"])
    append_sheet(workbook.create_sheet("外部冲突对象"), ["对象编码", "出现次数", "说明"], external_rows)

    failure_headers = [
        "课次ID",
        "日期",
        "时间",
        "班级编码",
        "套班编码",
        "教师1",
        "教室",
        "备注",
        "课程",
        "课节科目",
        "所在行",
        "错误原因",
    ]
    failure_rows = [
        [
            clean(row.get("课次ID")),
            clean(row.get("日期")),
            clean(row.get("时间")),
            clean(row.get("班级编码")),
            suite_for_class(clean(row.get("班级编码")), classes),
            clean(row.get("教师1（实际授课教师）")),
            clean(row.get("教室")),
            clean(row.get("备注")),
            clean(row.get("课程")),
            clean(row.get("课节科目")),
            clean(row.get("所在行")),
            clean(row.get("错误原因")),
        ]
        for row in result_rows
    ]
    append_sheet(workbook.create_sheet("失败课次142"), failure_headers, failure_rows)

    detail_headers = [
        "是否失败课次",
        "套班编码",
        "日期",
        "星期",
        "时间",
        "班级编码",
        "班级名称",
        "科目",
        "季度",
        "阶段",
        "模块",
        "课程类别",
        "课程编码",
        "课程名称",
        "教师ID",
        "教师",
        "教室ID",
        "教室",
    ]
    detail_rows = []
    failed_schedule_key_set = set()
    for row in review_rows:
        is_failed = schedule_key(row) in failed_key_set
        if is_failed:
            failed_schedule_key_set.add(schedule_key(row))
        detail_rows.append(
            [
                "是" if is_failed else "",
                suite_for_class(clean(row.get("class_id")), classes),
                display_date(clean(row.get("date"))),
                clean(row.get("weekday")),
                f"{clean(row.get('start_time'))}~{clean(row.get('end_time'))}",
                clean(row.get("class_id")),
                clean(row.get("class_name")),
                clean(row.get("subject")),
                schedule_window_name(row),
                clean(row.get("stage")),
                clean(row.get("course_module")),
                clean(row.get("course_group")),
                clean(row.get("course_code")),
                clean(row.get("course_name")),
                clean(row.get("teacher_id")),
                clean(row.get("teacher_name")),
                clean(row.get("room_id")),
                clean(row.get("room_name")),
            ]
        )
    detail_sheet = workbook.create_sheet("班级课表明细")
    append_sheet(detail_sheet, detail_headers, detail_rows)
    failed_col = 1
    for row_idx in range(2, detail_sheet.max_row + 1):
        if clean(detail_sheet.cell(row_idx, failed_col).value) == "是":
            for col_idx in range(1, detail_sheet.max_column + 1):
                detail_sheet.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=FAILED_FILL)

    failed_slots = {
        (normalize_date(row.get("日期")), result_slot(row), clean(row.get("班级编码")))
        for row in result_rows
    }
    same_slot_rows = []
    for row in review_rows:
        key = (normalize_date(row.get("date")), row_slot(row), clean(row.get("class_id")))
        date_slot = (key[0], key[1])
        if not any(date_slot == (failed_date, failed_slot) for failed_date, failed_slot, _class_id in failed_slots):
            continue
        same_slot_rows.append(
            [
                "是" if schedule_key(row) in failed_key_set else "",
                suite_for_class(clean(row.get("class_id")), classes),
                display_date(clean(row.get("date"))),
                clean(row.get("weekday")),
                row_slot(row),
                f"{clean(row.get('start_time'))}~{clean(row.get('end_time'))}",
                clean(row.get("class_id")),
                clean(row.get("subject")),
                clean(row.get("stage")),
                clean(row.get("course_module")),
                clean(row.get("teacher_name")),
                clean(row.get("room_name")),
                clean(row.get("course_name")),
            ]
        )
    same_slot = workbook.create_sheet("失败时段同刻课表")
    append_sheet(
        same_slot,
        ["是否失败课次", "套班编码", "日期", "星期", "时段", "时间", "班级编码", "科目", "阶段", "模块", "教师", "教室", "课程名称"],
        same_slot_rows,
    )
    for row_idx in range(2, same_slot.max_row + 1):
        if clean(same_slot.cell(row_idx, 1).value) == "是":
            for col_idx in range(1, same_slot.max_column + 1):
                same_slot.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=FAILED_FILL)

    for suite_code in sorted(review_suite_codes):
        suite_rows = [row for row in review_rows if suite_for_class(clean(row.get("class_id")), classes) == suite_code]
        if not suite_rows:
            continue
        matrix_sheet = workbook.create_sheet(f"{suite_code}套班课表")
        matrix_headers, matrix_rows = build_matrix_rows(suite_rows, suite_code, failed_key_set, args.start_date, args.end_date)
        matrix_sheet.append(matrix_headers)
        for matrix_row in matrix_rows:
            matrix_sheet.append(matrix_row)
        format_matrix_sheet(matrix_sheet)
        for row_idx in range(2, matrix_sheet.max_row + 1):
            for col_idx in range(3, matrix_sheet.max_column + 1):
                value = clean(matrix_sheet.cell(row_idx, col_idx).value)
                if "【失败】" in value:
                    matrix_sheet.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=FAILED_FILL)
                else:
                    for subject, color in SUBJECT_FILL.items():
                        if subject in value:
                            matrix_sheet.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=color)
                            break

    for ws in workbook.worksheets:
        ws.sheet_view.showGridLines = False

    output_path = args.output_dir / f"erp_failed_class_schedule_review_{stamp}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print(f"failed_rows={len(result_rows)}")
    print(f"review_classes={len(review_class_ids)}")
    print(f"review_suites={','.join(sorted(review_suite_codes))}")
    print(f"schedule_rows={len(review_rows)}")
    print(f"failed_rows_matched_in_schedule={len(failed_schedule_key_set)}")
    print(f"output_xlsx={output_path}")


if __name__ == "__main__":
    main()
