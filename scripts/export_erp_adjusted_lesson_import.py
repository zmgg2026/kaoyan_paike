#!/usr/bin/env python3
"""Export only lessons changed from the latest ERP schedule into the import template."""

from __future__ import annotations

import argparse
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.build_erp_lesson_id_map import (  # noqa: E402
    PUBLIC_SUBJECTS,
    erp_sort_key,
    normalize_date,
    normalize_one_time,
    normalize_time,
    schedule_sort_key,
)
from scripts.export_erp_lesson_import import (  # noqa: E402
    DATA_START_ROW,
    DEFAULT_PRODUCT_COURSES,
    DEFAULT_SCHEDULE,
    DEFAULT_TEACHER_ASSIGNMENTS,
    ERP_HEADERS,
    backfill_course_code,
    build_export_row,
    clean,
    clear_template_rows,
    display_date,
    load_course_code_lookup,
    load_shared_class_keys,
    schedule_window_name,
    split_values,
    validate_template_headers,
)
from scripts.csv_utils import read_csv_rows, write_csv_rows  # noqa: E402
from scripts.schedule_data import load_class_metadata, load_room_name_to_id  # noqa: E402
from scripts.sync_erp_adjusted_schedule import read_erp_rows  # noqa: E402


DEFAULT_CLASSES = Path("data/classes.csv")
DEFAULT_ROOMS = Path("data/rooms.csv")
DEFAULT_OUTPUT_DIR = Path("outputs")
ONLINE_ROOM_ID = "RMHFWY97001"


def group_by_class(rows: Iterable[Dict[str, str]], class_field: str) -> Dict[str, List[Dict[str, str]]]:
    grouped: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for row in rows:
        class_id = clean(row.get(class_field))
        if class_id:
            grouped[class_id].append(row)
    return grouped


def should_include_schedule_row(row: Dict[str, str], start_date: str, end_date: str, include_subjects: set[str]) -> bool:
    date = normalize_date(row.get("date"))
    subject = clean(row.get("subject"))
    return start_date <= date <= end_date and (not include_subjects or subject in include_subjects)


def should_include_erp_row(row: Dict[str, str], schedule_class_ids: set[str], start_date: str, end_date: str) -> bool:
    class_id = clean(row.get("班级编码"))
    if class_id not in schedule_class_ids:
        return False
    if clean(row.get("授课内容类型")) != "正课":
        return False
    try:
        minutes = int(float(clean(row.get("分钟数")) or 0))
    except ValueError:
        return False
    if minutes != 120:
        return False
    date = normalize_date(row.get("日期"))
    if not (start_date <= date <= end_date):
        return False
    subject = clean(row.get("课程科目"))
    course_code = clean(row.get("实际课程编码")) or clean(row.get("课程编码"))
    if subject == "不区分" or course_code == "CSHFWY000300158":
        return False
    return True


def is_shared_merge_row(row: Dict[str, str], shared_keys: Dict[Tuple[str, str, str, str], str]) -> bool:
    class_id = clean(row.get("class_id"))
    subject = clean(row.get("subject"))
    quarter = schedule_window_name(row)
    stage = clean(row.get("stage"))
    group = clean(row.get("course_group"))
    stage_candidates = [stage]
    if quarter and quarter not in stage_candidates:
        stage_candidates.append(quarter)
    return any((class_id, subject, candidate, group) in shared_keys for candidate in stage_candidates) or any(
        (class_id, subject, candidate, "") in shared_keys for candidate in stage_candidates
    )


def is_wuyou_math_class(class_id: str, class_meta: Dict[str, Dict[str, str]]) -> bool:
    meta = class_meta.get(class_id, {})
    return clean(meta.get("product_line")) == "考研无忧" and clean(meta.get("subject")) == "数学"


def import_subject_for_row(row: Dict[str, str], class_meta: Dict[str, Dict[str, str]]) -> str:
    class_id = clean(row.get("class_id"))
    subject = clean(row.get("subject"))
    if is_wuyou_math_class(class_id, class_meta):
        return "数学一"
    return subject


def normalize_import_values(values: Sequence[str], source_row: Dict[str, str], class_meta: Dict[str, Dict[str, str]]) -> List[str]:
    normalized = [clean(value) for value in values]
    class_id = normalized[3]
    room = normalized[6]
    if room.startswith("RMONLINE"):
        normalized[6] = ONLINE_ROOM_ID
    if is_wuyou_math_class(class_id, class_meta):
        normalized[11] = "数学一"
    return normalized


def semantic_current(
    row: Dict[str, str],
    values: Sequence[str],
    class_meta: Dict[str, Dict[str, str]],
    shared_keys: Dict[Tuple[str, str, str, str], str],
) -> Dict[str, str]:
    date = normalize_date(row.get("date"))
    start_time = normalize_one_time(clean(row.get("start_time")))
    end_time = normalize_one_time(clean(row.get("end_time")))
    shared = is_shared_merge_row(row, shared_keys)
    return {
        "date": date,
        "time": f"{start_time}~{end_time}",
        "class_id": clean(row.get("class_id")),
        "teacher_name": "" if shared else clean(row.get("teacher_name")),
        "room_id": "" if shared else clean(values[6]),
        "course_code": "" if shared else clean(values[10]),
        "subject": import_subject_for_row(row, class_meta),
    }


def semantic_erp(row: Dict[str, str], room_name_to_id: Dict[str, str]) -> Dict[str, str]:
    date = normalize_date(row.get("日期"))
    start_time, end_time = normalize_time(row.get("时间"))
    room_name = clean(row.get("实际教室名称")) or clean(row.get("教室名称"))
    return {
        "date": date,
        "time": f"{start_time}~{end_time}",
        "class_id": clean(row.get("班级编码")),
        "teacher_name": clean(row.get("实际教师1")) or clean(row.get("教师1")),
        "room_id": room_name_to_id.get(room_name, "") if room_name else "",
        "course_code": clean(row.get("实际课程编码")) or clean(row.get("课程编码")),
        "subject": clean(row.get("课程科目")),
    }


def subject_equivalent(current_subject: str, erp_subject: str) -> bool:
    if current_subject == erp_subject:
        return True
    return current_subject == "数学一" and erp_subject == "数学"


def diff_reasons(current: Dict[str, str], erp: Dict[str, str]) -> List[str]:
    labels = {
        "date": "日期",
        "time": "时间",
        "teacher_name": "教师",
        "room_id": "教室",
        "course_code": "课程",
    }
    reasons: List[str] = []
    for field, label in labels.items():
        if clean(current.get(field)) != clean(erp.get(field)):
            reasons.append(f"{label}: ERP={erp.get(field, '')} -> 当前={current.get(field, '')}")
    return reasons


def pair_schedule_to_erp(
    schedule_rows: Sequence[Dict[str, str]],
    erp_rows: Sequence[Dict[str, str]],
) -> Tuple[List[Tuple[Dict[str, str], Dict[str, str], int]], List[Dict[str, str]]]:
    schedule_by_class = group_by_class(schedule_rows, "class_id")
    erp_by_class = group_by_class(erp_rows, "班级编码")
    for rows in schedule_by_class.values():
        rows.sort(key=schedule_sort_key)
    for rows in erp_by_class.values():
        rows.sort(key=erp_sort_key)

    pairs: List[Tuple[Dict[str, str], Dict[str, str], int]] = []
    issues: List[Dict[str, str]] = []
    for class_id in sorted(set(schedule_by_class) | set(erp_by_class)):
        schedule_items = schedule_by_class.get(class_id, [])
        erp_items = erp_by_class.get(class_id, [])
        pair_count = min(len(schedule_items), len(erp_items))
        for index in range(pair_count):
            pairs.append((schedule_items[index], erp_items[index], index + 1))
        if len(schedule_items) != len(erp_items):
            issues.append(
                {
                    "issue_type": "count_mismatch",
                    "class_id": class_id,
                    "schedule_count": str(len(schedule_items)),
                    "erp_count": str(len(erp_items)),
                    "message": "当前课表与ERP课次数量不一致，超出部分不会进入导入表。",
                }
            )
        for index, row in enumerate(schedule_items[pair_count:], start=pair_count + 1):
            issues.append(
                {
                    "issue_type": "missing_erp_lesson_id",
                    "class_id": class_id,
                    "schedule_count": str(len(schedule_items)),
                    "erp_count": str(len(erp_items)),
                    "message": f"第{index}个当前课表缺ERP课次ID: {normalize_date(row.get('date'))} {clean(row.get('start_time'))}~{clean(row.get('end_time'))}",
                }
            )
        for index, row in enumerate(erp_items[pair_count:], start=pair_count + 1):
            issues.append(
                {
                    "issue_type": "unused_erp_lesson",
                    "class_id": class_id,
                    "schedule_count": str(len(schedule_items)),
                    "erp_count": str(len(erp_items)),
                    "message": f"第{index}个ERP课次未使用: {clean(row.get('课次ID'))} {normalize_date(row.get('日期'))} {clean(row.get('时间'))}",
                }
            )
    return pairs, issues


def write_workbook(template: Path, output_path: Path, rows: Sequence[Sequence[str]]) -> None:
    workbook = load_workbook(template)
    worksheet = workbook.worksheets[0]
    headers = validate_template_headers(worksheet)
    clear_template_rows(worksheet)
    for offset, values in enumerate(rows, start=DATA_START_ROW):
        for header, value in zip(ERP_HEADERS, values):
            worksheet.cell(offset, headers[header]).value = clean(value)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_report(
    path: Path,
    erp_export: Path,
    output_xlsx: Path,
    changed_rows: Sequence[Dict[str, str]],
    issues: Sequence[Dict[str, str]],
    gaps: Sequence[Dict[str, str]],
    start_date: str,
    end_date: str,
) -> None:
    reason_counter: Counter[str] = Counter()
    for row in changed_rows:
        for reason in split_values(row.get("change_reasons", "")):
            reason_counter[reason.split(":", 1)[0]] += 1
    issue_counter = Counter(row["issue_type"] for row in issues)
    gap_counter = Counter(row["gap_type"] for row in gaps)

    lines = [
        "# ERP 差异课次导入表报告",
        "",
        f"- ERP 对比文件：`{erp_export}`",
        f"- 对比范围：{start_date} 至 {end_date}",
        f"- 导入表：`{output_xlsx}`",
        f"- 本次导出调整课次：{len(changed_rows)}",
        f"- 映射/数量异常：{len(issues)}",
        f"- 导入字段缺口：{len(gaps)}",
        "",
        "## 变化字段统计",
        "",
    ]
    if reason_counter:
        for reason, count in reason_counter.most_common():
            lines.append(f"- {reason}: {count}")
    else:
        lines.append("- 无")
    if issue_counter:
        lines.extend(["", "## 映射/数量异常"])
        for issue_type, count in issue_counter.most_common():
            lines.append(f"- {issue_type}: {count}")
    if gap_counter:
        lines.extend(["", "## 导入字段缺口"])
        for gap_type, count in gap_counter.most_common():
            lines.append(f"- {gap_type}: {count}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="对比ERP现有课表，只导出当前已调整课次到导入模板。")
    parser.add_argument("--erp-export", type=Path, required=True)
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--schedule-csv", type=Path, default=DEFAULT_SCHEDULE)
    parser.add_argument("--teacher-assignments", type=Path, default=DEFAULT_TEACHER_ASSIGNMENTS)
    parser.add_argument("--product-courses", type=Path, default=DEFAULT_PRODUCT_COURSES)
    parser.add_argument("--classes", type=Path, default=DEFAULT_CLASSES)
    parser.add_argument("--rooms", type=Path, default=DEFAULT_ROOMS)
    parser.add_argument("--start-date", default="2026-07-01")
    parser.add_argument("--end-date", default="2026-12-13")
    parser.add_argument("--include-subjects", default="|".join(sorted(PUBLIC_SUBJECTS)))
    parser.add_argument("--output-xlsx", type=Path)
    parser.add_argument("--output-csv", type=Path)
    parser.add_argument("--diff-csv", type=Path)
    parser.add_argument("--issue-csv", type=Path)
    parser.add_argument("--gap-report", type=Path)
    parser.add_argument("--report", type=Path)
    parser.add_argument("--timestamp", default="")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    stamp = args.timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    output_xlsx = args.output_xlsx or DEFAULT_OUTPUT_DIR / f"erp_lesson_import_adjusted_{stamp}.xlsx"
    output_csv = args.output_csv or DEFAULT_OUTPUT_DIR / f"erp_lesson_import_adjusted_{stamp}.csv"
    diff_csv = args.diff_csv or DEFAULT_OUTPUT_DIR / f"erp_lesson_import_adjusted_diff_{stamp}.csv"
    issue_csv = args.issue_csv or DEFAULT_OUTPUT_DIR / f"erp_lesson_import_adjusted_mapping_issues_{stamp}.csv"
    gap_report = args.gap_report or DEFAULT_OUTPUT_DIR / f"erp_lesson_import_adjusted_gap_report_{stamp}.csv"
    report = args.report or DEFAULT_OUTPUT_DIR / f"erp_lesson_import_adjusted_report_{stamp}.md"

    include_subjects = set(split_values(args.include_subjects))
    schedule_source = [
        row
        for row in read_csv_rows(args.schedule_csv)
        if should_include_schedule_row(row, args.start_date, args.end_date, include_subjects)
    ]
    schedule_class_ids = {clean(row.get("class_id")) for row in schedule_source}
    erp_source = [
        row
        for row in read_erp_rows(args.erp_export)
        if should_include_erp_row(row, schedule_class_ids, args.start_date, args.end_date)
    ]
    pairs, mapping_issues = pair_schedule_to_erp(schedule_source, erp_source)

    class_meta = load_class_metadata(args.classes)
    room_name_to_id = load_room_name_to_id(args.rooms)
    shared_keys = load_shared_class_keys(args.teacher_assignments)
    course_code_lookup = load_course_code_lookup(args.product_courses)

    export_rows: List[List[str]] = []
    export_csv_rows: List[Dict[str, str]] = []
    diff_rows: List[Dict[str, str]] = []
    all_gaps: List[Dict[str, str]] = []

    for schedule_row, erp_row, sequence in pairs:
        values, gaps = build_export_row(schedule_row, {}, "room_id", shared_keys, course_code_lookup)
        values[0] = clean(erp_row.get("课次ID"))
        values = normalize_import_values(values, schedule_row, class_meta)
        # Make sure the course code fallback uses the same lookup even if a row was imported before the tag existed.
        if not values[10] and not is_shared_merge_row(schedule_row, shared_keys):
            values[10] = backfill_course_code(schedule_row, course_code_lookup)
        current_semantic = semantic_current(schedule_row, values, class_meta, shared_keys)
        erp_semantic = semantic_erp(erp_row, room_name_to_id)
        reasons = diff_reasons(current_semantic, erp_semantic)
        if not reasons:
            continue

        export_rows.append(values)
        export_csv_rows.append(dict(zip(ERP_HEADERS, values)))
        diff_rows.append(
            {
                "class_id": clean(schedule_row.get("class_id")),
                "sequence": str(sequence),
                "erp_lesson_id": clean(erp_row.get("课次ID")),
                "erp_date": normalize_date(erp_row.get("日期")),
                "erp_time": clean(erp_row.get("时间")),
                "current_date": normalize_date(schedule_row.get("date")),
                "current_time": f"{normalize_one_time(clean(schedule_row.get('start_time')))}~{normalize_one_time(clean(schedule_row.get('end_time')))}",
                "subject": clean(schedule_row.get("subject")),
                "stage": clean(schedule_row.get("stage")),
                "course_module": clean(schedule_row.get("course_module")),
                "current_teacher": clean(schedule_row.get("teacher_name")),
                "erp_teacher": clean(erp_row.get("实际教师1")) or clean(erp_row.get("教师1")),
                "current_room_id": values[6],
                "erp_room_name": clean(erp_row.get("实际教室名称")) or clean(erp_row.get("教室名称")),
                "current_course_code": values[10],
                "erp_course_code": clean(erp_row.get("实际课程编码")) or clean(erp_row.get("课程编码")),
                "change_reasons": "|".join(reasons),
            }
        )
        for gap in gaps:
            if gap.get("gap_type") == "missing_erp_lesson_id":
                continue
            all_gaps.append({**gap, "erp_lesson_id": clean(erp_row.get("课次ID"))})

    write_workbook(args.template, output_xlsx, export_rows)
    write_csv_rows(output_csv, ERP_HEADERS, export_csv_rows, extrasaction="ignore")
    write_csv_rows(
        diff_csv,
        [
            "class_id",
            "sequence",
            "erp_lesson_id",
            "erp_date",
            "erp_time",
            "current_date",
            "current_time",
            "subject",
            "stage",
            "course_module",
            "current_teacher",
            "erp_teacher",
            "current_room_id",
            "erp_room_name",
            "current_course_code",
            "erp_course_code",
            "change_reasons",
        ],
        diff_rows,
        extrasaction="ignore",
    )
    write_csv_rows(issue_csv, ["issue_type", "class_id", "schedule_count", "erp_count", "message"], mapping_issues)
    gap_fields = [
        "erp_lesson_id",
        "gap_type",
        "message",
        "class_id",
        "date",
        "start_time",
        "end_time",
        "subject",
        "stage",
        "course_module",
        "teacher_id",
        "teacher_name",
        "course_code",
        "course_name",
        "is_shared_merge_row",
    ]
    write_csv_rows(gap_report, gap_fields, all_gaps, extrasaction="ignore")
    write_report(report, args.erp_export, output_xlsx, diff_rows, mapping_issues, all_gaps, args.start_date, args.end_date)

    print(f"schedule_rows={len(schedule_source)}")
    print(f"erp_rows={len(erp_source)}")
    print(f"paired_rows={len(pairs)}")
    print(f"adjusted_rows={len(export_rows)}")
    print(f"mapping_issues={len(mapping_issues)}")
    print(f"gaps={len(all_gaps)}")
    print(f"output_xlsx={output_xlsx}")
    print(f"output_csv={output_csv}")
    print(f"diff_csv={diff_csv}")
    print(f"issue_csv={issue_csv}")
    print(f"gap_report={gap_report}")
    print(f"report={report}")


if __name__ == "__main__":
    main()
