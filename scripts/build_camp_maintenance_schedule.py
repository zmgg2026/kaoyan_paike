#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import re
import signal
import shutil
import sys
import time
from collections import Counter, defaultdict
from dataclasses import replace
from datetime import date as Date, datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Set, Tuple

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import scheduler
from scripts.csv_utils import read_csv_rows
from scripts.field_utils import (
    normalize_blank_marker,
    parse_bool,
    parse_enabled,
    parse_datetime_value,
    row_value,
    split_delimited_values as split_pipe_values,
)
from scripts.schedule_class_windows import (
    ClassWindowConstraint,
    load_class_window_constraint_items,
    load_class_window_constraints,
    suite_window_constraints_from_class_windows,
)
from scripts.schedule_constraints import (
    assignment_date_period_keys,
    assignments_conflicting_with_candidate,
    candidate_conflicts_for_repair,
)
from scripts.schedule_conflicts import (
    assignments_are_same_shared_merge_event,
    assignments_overlap,
    minutes,
    teacher_time_conflict_groups,
    teacher_time_conflict_lines,
    write_teacher_time_conflicts_csv,
)
from scripts.schedule_display import assignment_standard_lesson_count
from scripts.schedule_modes import assignment_reference_class_id, assignment_schedule_mode
from scripts.schedule_data import (
    infer_class_subject,
    load_active_blackout_dates,
    load_class_metadata as raw_load_class_metadata,
    load_room_maps as room_maps,
    load_room_names as raw_load_room_names,
    load_teacher_maps as teacher_maps,
)
from scripts.schedule_first_lesson import (
    first_lesson_module_violations,
    replacement_preserves_first_lesson_module,
)
from scripts.schedule_run_rules import run_dates_over_limit
from scripts.schedule_outputs import write_batch_csv, write_day_table_html
from scripts.schedule_scope import (
    class_ids_for_suite_codes as raw_class_ids_for_suite_codes,
    filtered_schedule_input,
)
from scripts.schedule_week_balance import (
    average_subject_week_bounds_from_counts,
    balanced_week_quotas,
    week_key,
)
from scripts.schedule_batch import (
    alternate_teacher_tasks,
    candidate_hours_by_date,
    candidate_is_valid,
    candidate_teacher_key,
    candidate_avoids_same_class_teacher_day_limit,
    improve_total_week_balance,
    rebalance_subject_teacher_alternation,
    schedule_balanced_camp,
    schedule_round_robin,
    schedule_balanced_camp_by_suite,
    LONG_CAMP_MATH_MAX_CONSECUTIVE_DAYS,
    SUMMER_PREFERRED_WEEKLY_HALFDAY_MAX,
)

_raw_class_ids_for_suite_codes = raw_class_ids_for_suite_codes
_raw_load_class_metadata = raw_load_class_metadata
_raw_load_room_names = raw_load_room_names


SUMMER_CAMP_SUITES = ("2790", "2791", "2792", "2793", "2794")
HALF_YEAR_BATCH_SUITES = tuple(str(code) for code in range(2770, 2780))
HALF_YEAR_BATCH_PRIORITY = (
    "2777",
    "2779",
    "2776",
    "2778",
    "2775",
    "2773",
    "2771",
    "2770",
    "2772",
    "2774",
)
MAINTENANCE_SUITES = (
    "2750",
    "2751",
    "2752",
    "2753",
    "2754",
    "2755",
    "2756",
    "2757",
    "2759",
    "2711",
    "2712",
    *SUMMER_CAMP_SUITES,
)
PREVIOUS_PILOT_SUITES = ("2751", "2753", "2755", "2757", "2759")
ONLINE_MERGE_GROUPS = {
    "2751": ("2750", "2751", "2752", "2753", "2754"),
    "2755": ("2755", "2756", "2757", "2759", "2711", "2712"),
    "2791": SUMMER_CAMP_SUITES,
}
ONLINE_PRIMARY_SUITES = tuple(ONLINE_MERGE_GROUPS)
PUBLIC_PRODUCT_ORDER = (
    "寒暑营",
    "无忧寒",
    "暑假营",
    "半年营",
    "全年营",
    "无忧暑",
    "无忧秋",
    "无忧春",
    "冲刺营",
)
ADDITIONAL_PUBLIC_PRODUCTS = ("全年营", "无忧暑", "无忧秋", "无忧春", "冲刺营")
PUBLIC_SUBJECT_PRIORITY = {"数学": 0, "政治": 1, "英语": 2, "语文": 3}
PUBLIC_SAME_CLASS_SUBJECT_DAY_HOURS = 8
WYQC_PRODUCTS = {"无忧秋", "无忧春"}
WYQC_SUITE_PRIORITY = {
    # These two suites have the tightest remaining 27考研无忧春 resource fit in
    # the current data set, so solve them before the lighter suites consume the
    # same teachers/rooms.
    "2724": 0,
    "2723": 1,
}
WUYOU_PRODUCTS = {"无忧寒", "无忧暑", "无忧秋", "无忧春"}
WUYOU_PRODUCT_BLACKOUT_DATES = {"2026-10-10"}
HISTORY_DEDUCT_PRODUCTS = WYQC_PRODUCTS | {"全年营"}
WYQC_SUMMER_START = "2026-07-04"
WYQC_FOUNDATION_END = "2026-08-31"
WYQC_FOUNDATION_END_BY_SUITE = {
    "2701": "2026-07-26",
    "2702": "2026-08-02",
    "2703": "2026-08-16",
    "2704": "2026-07-26",
    "2706": "2026-07-26",
    "2721": "2026-08-02",
    "2720": "2026-08-16",
}
WUYOU_SUMMER_WEEK_BOUNDS_WITH_MATH = {
    "数学": (3, 4),
    "政治": (3, 4),
    "英语": (3, 4),
}
WUYOU_SUMMER_WEEK_BOUNDS_WITHOUT_MATH = {
    "政治": (3, 4),
    "英语": (3, 4),
}
WUYOU_SUMMER_TOTAL_MAX_WITH_MATH = 10
WUYOU_SUMMER_TOTAL_MAX_WITHOUT_MATH = 8
WYQC_SUMMER_WEEK_BOUNDS = {
    "数学": (2, 3),
    "政治": (1, 2),
    "英语": (2, 4),
}
WYQC_AUTUMN_START = "2026-09-05"
WYQC_AUTUMN_END = "2026-12-06"
WYQC_AUTUMN_MAKEUP_DATES = {"2026-10-06", "2026-10-07"}
WYS_PRODUCT = "无忧暑"
WYS_SUMMER_START = "2026-07-01"
WYS_SUMMER_END = "2026-08-31"
WYS_AUTUMN_START = "2026-09-01"
WYS_FOUNDATION_END = "2026-10-28"
WYS_SPRINT_START = "2026-10-28"
WYS_SUMMER_WEEK_BOUNDS = WUYOU_SUMMER_WEEK_BOUNDS_WITH_MATH
WYS_SUMMER_WEEK_BOUNDS_WITHOUT_MATH = WUYOU_SUMMER_WEEK_BOUNDS_WITHOUT_MATH
WYS_SUMMER_TOTAL_MAX_WITH_MATH = WUYOU_SUMMER_TOTAL_MAX_WITH_MATH
WYS_SUMMER_TOTAL_MAX_WITHOUT_MATH = WUYOU_SUMMER_TOTAL_MAX_WITHOUT_MATH
WYS_SPRINT_WEEK_BOUNDS_WITH_MATH = {
    "数学": (2, 2),
    "政治": (1, 2),
    "英语": (1, 2),
}
WYS_SPRINT_WEEKLY_TOTAL_WITH_MATH = 5
WYS_STAGE_PRIORITY_SUITE = "2731"
WYS_STAGE_PRIORITY_ORDER = ("基础", "强化", "冲刺")
WYS_STAGELESS_WEEKLY_SUITE = WYS_STAGE_PRIORITY_SUITE
WYS_STAGELESS_AUTUMN_WEEKLY_MIN = 5
WYS_STAGELESS_AUTUMN_WEEKLY_MAX = 6
WYS_STAGELESS_SUMMER_WEEKLY_MAX = 10
WYS_AUTUMN_ALLOWED_WEEKDAYS = {2, 5, 6}
WYS_NO_MATH_BALANCED_SUITE = "2727"
WYS_NO_MATH_SUMMER_WEEKLY_MAX = 8
WYS_NO_MATH_SUMMER_SUBJECT_WEEKLY_BOUNDS = {
    "英语": (3, 4),
    "政治": (3, 4),
}
WYS_NO_MATH_AUTUMN_WEEKLY_MIN = 2
WYS_NO_MATH_AUTUMN_WEEKLY_MAX = 3
WYS_NO_MATH_SUBJECT_WEEKLY_BOUNDS = {
    "英语": (1, 2),
    "政治": (0, 1),
}
WYQC_SPRINT_START_BY_SUBJECT = {
    "英语": WYQC_AUTUMN_START,
    "数学": WYQC_AUTUMN_START,
    "政治": WYQC_AUTUMN_START,
}
SPRINT_CAMP_PRODUCT = "冲刺营"
SPRINT_CAMP_WEEK_BOUNDS = {
    "数学": (2, 3),
    "政治": (1, 2),
    "英语": (2, 3),
}
SUMMER_PREPLAN_STAGES = {"暑假", "基础", "强化"}
SUMMER_CLASS_WINDOW_IDS = {"WINDOW_SUMMER", "暑假"}
SUMMER_PUBLIC_SUBJECTS = {"英语", "政治", "数学"}
SUMMER_STAGE_ORDER = {"基础": 0, "强化": 1, "冲刺": 2}
AUTUMN_PREPLAN_STAGES = {"秋季", "冲刺"}
LONG_CAMP_PREFERRED_PERIODS = {"数学": "AM", "英语": "PM", "政治": "PM"}
LONG_CAMP_BALANCE_START = "2026-07-01"
LONG_CAMP_SUBJECT_WEEKLY_MIN = {"数学": 1, "英语": 1, "政治": 1}
LONG_CAMP_SUBJECT_WEEKLY_MAX = {"数学": 4, "英语": 4, "政治": 3}
HISTORY_PATH = Path(
    os.environ.get(
        "HISTORY_PATH",
        ROOT / "outputs" / "manual_sources" / "current_history" / "20240301-20260630历史课表【20260520】.xlsx",
    )
)
OUTPUT_CSV = Path("outputs/batch_schedule_maintenance.csv")
OUTPUT_HTML = Path("outputs/batch_schedule_maintenance.html")
OUTPUT_REPORT = Path("outputs/batch_schedule_maintenance_report.md")
TEACHER_CONFLICT_CSV = Path("outputs/teacher_time_conflicts.csv")
LEGACY_OUTPUT_CSV = Path("outputs/summer_camp_schedule.csv")
LEGACY_OUTPUT_HTML = Path("outputs/summer_camp_schedule.html")
LEGACY_OUTPUT_REPORT = Path("outputs/summer_camp_schedule_build_report.md")
SUMMER_PAGE_CSV = OUTPUT_CSV
SUMMER_FALLBACK_CSVS = (
    LEGACY_OUTPUT_CSV,
    Path("outputs/schedule_2751_summer_batch.csv"),
    Path("outputs/schedule_2753_2755_2757_2759_summer_batch.csv"),
)
SUMMER_FAST_PRODUCTS = {"寒暑营", "无忧寒", "暑假营"}
ONLINE_ROOM_NAME = "线上虚拟网络教室01"
ONLINE_ROOM_ID = "RMHFWY97001"
RENJIE_TEACHER_IDS = {"351339"}
RENJIE_TEACHER_NAMES = {"任洁"}
RENJIE_OFFLINE_WINDOWS = (
    ("2026-07-25", "AM", "2026-08-18", "AM"),
    ("2026-08-21", "PM", "2026-08-27", "AM"),
)
AUTUMN_START = "2026-09-01"
AUTUMN_END = "2026-12-13"
AUTUMN_ENGLISH_START = "2026-09-08"
AUTUMN_MATH_START = "2026-09-16"
AUTUMN_POLITICS_START = "2026-09-19"
AUTUMN_SPECIAL_SUBJECT_RULES = {
    ("2791", "英语"): {
        "start_date": "2026-09-15",
        "allowed_weekdays": {2, 4},
        "allowed_periods": {"EVENING"},
        "preferred_period": "EVENING",
        "block_hours": 2,
        "label": "2791 英语周三/周五晚课",
    },
    ("2791", "数学"): {
        "start_date": "2026-09-17",
        "allowed_weekdays": {1, 3},
        "allowed_periods": {"EVENING"},
        "preferred_period": "EVENING",
        "block_hours": 2,
        "label": "2791 数学周二/周四晚课",
    },
    ("2791", "政治"): {
        "start_date": "2026-11-08",
        "allowed_weekdays": {5, 6},
        "allowed_periods": {"AM", "PM"},
        "preferred_period": "PM",
        "block_hours": 4,
        "one_per_week": True,
        "label": "2791 政治周末半天课",
    },
}


def row_matches_phase(row: dict, phases: Set[str]) -> bool:
    return (row.get("quarter") or "") in phases or (row.get("stage") or "") in phases


def task_matches_phase(task: scheduler.CourseBlock, phases: Set[str]) -> bool:
    return (task.quarter or "") in phases or (task.stage or "") in phases


def assignment_matches_phase(assignment: scheduler.Assignment, phases: Set[str]) -> bool:
    return task_matches_phase(assignment.task, phases)


def online_merge_phase_for_row(row: dict) -> str:
    quarter = row.get("quarter") or ""
    stage = row.get("stage") or ""
    if quarter in {"春季", "秋季"}:
        return quarter
    if stage in {"春季", "秋季"}:
        return stage
    if stage == "冲刺":
        return "秋季"
    return ""


def online_merge_phase_for_assignment(assignment: scheduler.Assignment) -> str:
    task = assignment.task
    quarter = task.quarter or ""
    stage = task.stage or ""
    if quarter in {"春季", "秋季"}:
        return quarter
    if stage in {"春季", "秋季"}:
        return stage
    if stage == "冲刺":
        return "秋季"
    return ""
AUTUMN_HOLIDAYS = {
    "2026-10-01",
    "2026-10-02",
    "2026-10-03",
    "2026-10-04",
    "2026-10-05",
    "2026-10-06",
    "2026-10-07",
    "2026-10-10",
    "2026-09-25",
}
_SCHEDULER_INPUT_DATA_CACHE: Dict[Path, Tuple[int, Dict[str, Any]]] = {}
_SCHEDULE_INPUT_CACHE: Dict[Tuple[Path, int, Tuple[str, ...]], scheduler.ScheduleInput] = {}
_CLASS_METADATA_CACHE: Dict[Tuple[Path, int], Dict[str, Dict[str, str]]] = {}
_ROOM_NAMES_CACHE: Dict[Tuple[Path, int], Dict[str, str]] = {}
_CLASS_WINDOW_CONSTRAINT_CACHE: Dict[Tuple[Path, int, int, int, Optional[Tuple[str, ...]]], Dict[str, ClassWindowConstraint]] = {}
_CLASS_IDS_CACHE: Dict[Tuple[Path, int, Tuple[str, ...], Optional[Tuple[str, ...]]], List[str]] = {}
_SHARED_ASSIGNMENT_CACHE: Dict[Tuple[Path, int], Dict[Tuple[str, str, str, str], str]] = {}
FAST_ATTEMPT_TIMEOUT_SECONDS = 12
FULL_SPACING_IMPROVEMENT_PASSES = int(os.environ.get("FULL_SPACING_IMPROVEMENT_PASSES", "8"))
GLOBAL_REPAIR_MAX_SECONDS = int(os.environ.get("GLOBAL_REPAIR_MAX_SECONDS", "180"))
ALLOW_PUBLISH_WITH_TEACHER_CONFLICTS = os.environ.get(
    "ALLOW_PUBLISH_WITH_TEACHER_CONFLICTS",
    "",
)
ALLOW_PUBLISH_WITH_TEACHER_CONFLICTS = parse_bool(ALLOW_PUBLISH_WITH_TEACHER_CONFLICTS)
ALLOW_PUBLISH_WITH_COVERAGE_GAPS = os.environ.get(
    "ALLOW_PUBLISH_WITH_COVERAGE_GAPS",
    "",
)
ALLOW_PUBLISH_WITH_COVERAGE_GAPS = parse_bool(ALLOW_PUBLISH_WITH_COVERAGE_GAPS)


class FastAttemptTimeout(ValueError):
    pass


class PerfLog:
    def __init__(self) -> None:
        self.started_at = time.perf_counter()
        self.last_mark = self.started_at
        self.entries: List[Tuple[str, float, float]] = []

    def mark(self, label: str) -> None:
        now = time.perf_counter()
        step_seconds = now - self.last_mark
        total_seconds = now - self.started_at
        self.entries.append((label, step_seconds, total_seconds))
        self.last_mark = now
        print(f"[perf] {label}: {step_seconds:.2f}s / total {total_seconds:.2f}s", flush=True)

    def total_seconds(self) -> float:
        return time.perf_counter() - self.started_at

    def markdown_lines(self) -> List[str]:
        lines = [f"- 总耗时: {self.total_seconds():.2f}s"]
        lines.extend(
            f"- {label}: {step_seconds:.2f}s，累计 {total_seconds:.2f}s"
            for label, step_seconds, total_seconds in self.entries
        )
        return lines


def file_mtime_ns(path: Path) -> int:
    return path.stat().st_mtime_ns if path.exists() else 0


def load_class_metadata(path: Path) -> Dict[str, Dict[str, str]]:
    resolved = path.resolve()
    classes_path = resolved / "classes.csv"
    cache_key = (resolved, file_mtime_ns(classes_path))
    cached = _CLASS_METADATA_CACHE.get(cache_key)
    if cached is not None:
        return cached
    data = _raw_load_class_metadata(path)
    _CLASS_METADATA_CACHE.clear()
    _CLASS_METADATA_CACHE[cache_key] = data
    return data


def load_room_names(path: Path) -> Dict[str, str]:
    resolved = path.resolve()
    if resolved.is_file():
        cache_key = (resolved, file_mtime_ns(resolved))
    else:
        rooms_json_path = resolved / "rooms.json"
        rooms_csv_path = resolved / "rooms.csv"
        cache_key = (resolved, file_mtime_ns(rooms_json_path), file_mtime_ns(rooms_csv_path))
    cached = _ROOM_NAMES_CACHE.get(cache_key)
    if cached is not None:
        return cached
    data = _raw_load_room_names(path)
    _ROOM_NAMES_CACHE.clear()
    _ROOM_NAMES_CACHE[cache_key] = data
    return data


def load_class_window_constraints_for_data_dir(
    path: Path,
    season_window_ids: Optional[Set[str]] = None,
) -> Dict[str, ClassWindowConstraint]:
    resolved = path.resolve()
    class_window_path = resolved / "class_window_boundaries.csv"
    season_key = tuple(sorted(season_window_ids)) if season_window_ids else None
    cache_key = (
        resolved,
        file_mtime_ns(class_window_path),
        file_mtime_ns(resolved / "rooms.json"),
        file_mtime_ns(resolved / "rooms.csv"),
        season_key,
    )
    cached = _CLASS_WINDOW_CONSTRAINT_CACHE.get(cache_key)
    if cached is not None:
        return cached
    data = load_class_window_constraints(
        class_window_path,
        season_window_ids=season_window_ids,
    )
    _CLASS_WINDOW_CONSTRAINT_CACHE.clear()
    _CLASS_WINDOW_CONSTRAINT_CACHE[cache_key] = data
    return data


def load_summer_class_window_constraints(path: Path) -> Dict[str, ClassWindowConstraint]:
    return load_class_window_constraints_for_data_dir(path, SUMMER_CLASS_WINDOW_IDS)


def load_all_class_window_constraints(path: Path) -> Dict[str, ClassWindowConstraint]:
    return load_class_window_constraints_for_data_dir(path)


def load_all_class_window_constraint_items(path: Path) -> Dict[str, List[ClassWindowConstraint]]:
    return load_class_window_constraint_items(path / "class_window_boundaries.csv")


def load_summer_suite_window_constraints(
    path: Path,
    class_metadata: Optional[Dict[str, Dict[str, str]]] = None,
) -> Dict[str, ClassWindowConstraint]:
    metadata = class_metadata if class_metadata is not None else load_class_metadata(path)
    return suite_window_constraints_from_class_windows(
        load_summer_class_window_constraints(path),
        metadata,
        set(MAINTENANCE_SUITES),
    )


def class_ids_for_suite_codes(
    path: Path,
    suite_codes: Sequence[str],
    subjects: Optional[Set[str]],
) -> List[str]:
    resolved = path.resolve()
    classes_path = resolved / "classes.csv"
    normalized_subjects = tuple(sorted(subjects)) if subjects else None
    cache_key = (
        resolved,
        file_mtime_ns(classes_path),
        tuple(suite_codes),
        normalized_subjects,
    )
    cached = _CLASS_IDS_CACHE.get(cache_key)
    if cached is not None:
        return list(cached)
    data = _raw_class_ids_for_suite_codes(path, suite_codes, subjects)
    _CLASS_IDS_CACHE[cache_key] = list(data)
    return list(data)
IGNORED_HISTORY_GAP_DATES = {"2026-06-10", "2026-06-17", "2026-06-23", "2026-06-24"}
LOCKED_SCHEDULE_PATH = Path("data/locked_scheduled_lessons.csv")


def clean(value: object) -> str:
    return normalize_blank_marker(value)


def parse_dt(value: object) -> datetime:
    return parse_datetime_value(value, "历史课表时间")


def date_text(value: object) -> str:
    return parse_dt(value).date().isoformat()


def time_text(value: object) -> str:
    return parse_dt(value).strftime("%H:%M")


def period_for_time(start: str) -> str:
    hour = int(start.split(":", 1)[0])
    if hour < 13:
        return "AM"
    if hour < 18:
        return "PM"
    return "EVENING"


def stage_for_display_date(value: str) -> str:
    month = Date.fromisoformat(value).month
    if month in {1, 2}:
        return "寒假"
    if 3 <= month <= 6:
        return "春季"
    if month in {7, 8}:
        return "暑假"
    return "秋季"


def normalize_subject(value: object) -> str:
    subject = clean(value)
    if subject.startswith("数学"):
        return "数学"
    return subject


def without_blackout_dates(
    schedule_input: scheduler.ScheduleInput,
    blackout_dates: Set[str],
) -> scheduler.ScheduleInput:
    if not blackout_dates:
        return schedule_input
    return replace(
        schedule_input,
        time_slots=[
            slot
            for slot in schedule_input.time_slots
            if slot.date not in blackout_dates
        ],
    )


def without_dates(
    schedule_input: scheduler.ScheduleInput,
    dates: Set[str],
) -> scheduler.ScheduleInput:
    if not dates:
        return schedule_input
    return replace(
        schedule_input,
        time_slots=[slot for slot in schedule_input.time_slots if slot.date not in dates],
    )


def with_monday_exception(schedule_input: scheduler.ScheduleInput) -> scheduler.ScheduleInput:
    classes: Dict[str, scheduler.SchoolClass] = {}
    for class_id, cls in schedule_input.classes.items():
        requirements = []
        for requirement in cls.requirements:
            allowed_weekdays = (
                set(requirement.allowed_weekdays) | {0}
                if requirement.allowed_weekdays
                else requirement.allowed_weekdays
            )
            rules = []
            for rule in requirement.schedule_rules:
                rule_weekdays = (
                    set(rule.allowed_weekdays) | {0}
                    if rule.allowed_weekdays
                    else rule.allowed_weekdays
                )
                rules.append(replace(rule, allowed_weekdays=rule_weekdays))
            requirements.append(
                replace(
                    requirement,
                    allowed_weekdays=allowed_weekdays,
                    schedule_rules=tuple(rules),
                )
            )
        classes[class_id] = replace(cls, requirements=requirements)
    return replace(schedule_input, classes=classes)


def improve_long_camp_week_balance(
    schedule_input: scheduler.ScheduleInput,
    assignments: Sequence[scheduler.Assignment],
) -> List[scheduler.Assignment]:
    assignments = scheduler.sorted_assignments(assignments)
    if not assignments:
        return []
    max_block_hours = max((assignment.task.block_hours for assignment in assignments), default=4)
    slot_blocks = scheduler.build_contiguous_slot_blocks(schedule_input.time_slots, max_block_hours)
    week_quotas = balanced_week_quotas(slot_blocks, len(assignments))
    if not week_quotas:
        return assignments
    domains = scheduler.candidate_domains(
        scheduler.build_course_blocks(schedule_input.classes),
        schedule_input,
    )
    return improve_total_week_balance(
        schedule_input,
        assignments,
        domains,
        week_quotas,
        LONG_CAMP_PREFERRED_PERIODS,
    )


def load_rooms_for_capacity(path: Path) -> Dict[str, scheduler.Room]:
    data = load_scheduler_input_data(path.parent)
    return {
        room["id"]: scheduler.Room(
            id=room["id"],
            capacity=room.get("capacity"),
            capacity_unlimited=scheduler.parse_bool(room.get("capacity_unlimited")),
            teaching_area_id=clean(room.get("teaching_area_id")),
            teaching_area_name=clean(room.get("teaching_area_name")),
            region_tag=clean(room.get("region_tag")),
        )
        for room in data.get("rooms", [])
        if room.get("id")
    }


def load_area_travel_minutes(data_dir: Path) -> Dict[Tuple[str, str], int]:
    data = load_scheduler_input_data(data_dir)
    return scheduler.parse_area_travel_minutes(data.get("teaching_area_links", []))


def load_scheduler_input_data(data_dir: Path) -> Dict[str, Any]:
    input_path = data_dir / "scheduler_input_draft.json"
    resolved = input_path.resolve()
    mtime_ns = resolved.stat().st_mtime_ns
    cached = _SCHEDULER_INPUT_DATA_CACHE.get(resolved)
    if cached and cached[0] == mtime_ns:
        return cached[1]
    data = json.loads(resolved.read_text(encoding="utf-8"))
    _SCHEDULER_INPUT_DATA_CACHE[resolved] = (mtime_ns, data)
    return data


def load_schedule_input_for_classes(data_dir: Path, class_ids: Iterable[str]) -> scheduler.ScheduleInput:
    data = load_scheduler_input_data(data_dir)
    class_id_set = set(class_ids)
    input_path = (data_dir / "scheduler_input_draft.json").resolve()
    mtime_ns = input_path.stat().st_mtime_ns
    cache_key = (input_path, mtime_ns, tuple(sorted(class_id_set)))
    cached = _SCHEDULE_INPUT_CACHE.get(cache_key)
    if cached is not None:
        return cached
    selected_classes = [
        cls for cls in data.get("classes", [])
        if cls.get("id") in class_id_set
    ]
    product_ids = {
        cls.get("product_id")
        for cls in selected_classes
        if cls.get("product_id")
    }
    filtered = dict(data)
    filtered["classes"] = selected_classes
    filtered["products"] = [
        product for product in data.get("products", [])
        if product.get("id") in product_ids
    ]
    locked_class_ids = {
        lesson.get("class_id")
        for lesson in data.get("locked_lessons", [])
        if lesson.get("class_id")
    }
    allowed_class_ids = class_id_set | locked_class_ids
    conflict_groups = []
    for group in data.get("conflict_groups", []):
        ids = [class_id for class_id in group.get("class_ids", []) if class_id in allowed_class_ids]
        if len(ids) < 2:
            continue
        normalized_group = dict(group)
        normalized_group["class_ids"] = ids
        conflict_groups.append(normalized_group)
    filtered["conflict_groups"] = conflict_groups
    schedule_input = scheduler.load_input_data(filtered)
    _SCHEDULE_INPUT_CACHE[cache_key] = schedule_input
    return schedule_input


def suite_code_from_class_id(class_id: str) -> str:
    match = re.search(r"(\d{4})$", class_id or "")
    return match.group(1) if match else ""


def suite_subject_classes(data_dir: Path) -> Dict[str, Dict[str, Dict[str, str]]]:
    result: Dict[str, Dict[str, Dict[str, str]]] = defaultdict(dict)
    for row in read_csv_rows(data_dir / "classes.csv"):
        suite_code = clean(row.get("suite_code")) or suite_code_from_class_id(clean(row.get("id")))
        subject = normalize_subject(infer_class_subject(row))
        if suite_code not in MAINTENANCE_SUITES or subject not in {"英语", "数学", "政治"}:
            continue
        result[suite_code][subject] = {
            "id": clean(row.get("id")),
            "name": clean(row.get("name")) or clean(row.get("id")),
        }
    return result


def maintenance_class_ids(data_dir: Path) -> Set[str]:
    result: Set[str] = set()
    for subject_classes in suite_subject_classes(data_dir).values():
        for class_info in subject_classes.values():
            if class_info.get("id"):
                result.add(class_info["id"])
    return result


def halfyear_batch_class_ids(data_dir: Path) -> Set[str]:
    result: Set[str] = set()
    for suite_code in HALF_YEAR_BATCH_SUITES:
        try:
            result.update(class_ids_for_suite_codes(data_dir, [suite_code], None))
        except ValueError:
            continue
    return result


def wuyou_spring_autumn_class_ids(data_dir: Path) -> Set[str]:
    classes_path = data_dir / "classes.csv"
    if not classes_path.exists():
        return set()
    result: Set[str] = set()
    for row in read_csv_rows(classes_path):
        if clean(row.get("subject_category")) != "公共课":
            continue
        if clean(row.get("sub_product")) not in WYQC_PRODUCTS:
            continue
        if clean(row.get("is_schedule_locked")) in {"是", "1", "true", "True", "yes", "Y", "y"}:
            continue
        class_id = clean(row.get("id"))
        if class_id:
            result.add(class_id)
    return result


def full_year_public_class_ids(data_dir: Path) -> Set[str]:
    classes_path = data_dir / "classes.csv"
    if not classes_path.exists():
        return set()
    result: Set[str] = set()
    for row in read_csv_rows(classes_path):
        if clean(row.get("subject_category")) != "公共课":
            continue
        if clean(row.get("sub_product")) != "全年营":
            continue
        if clean(row.get("is_schedule_locked")) in {"是", "1", "true", "True", "yes", "Y", "y"}:
            continue
        class_id = clean(row.get("id"))
        if class_id:
            result.add(class_id)
    return result


def history_class_ids(data_dir: Path) -> Set[str]:
    return (
        maintenance_class_ids(data_dir)
        | halfyear_batch_class_ids(data_dir)
        | wuyou_spring_autumn_class_ids(data_dir)
        | full_year_public_class_ids(data_dir)
    )


def online_primary_for_suite(suite_code: str) -> Optional[str]:
    for primary_suite, suite_codes in ONLINE_MERGE_GROUPS.items():
        if suite_code in suite_codes:
            return primary_suite
    return None


def clone_row_for_suite(
    row: dict,
    suite_code: str,
    primary_suite: str,
    class_lookup: Dict[str, Dict[str, Dict[str, str]]],
) -> Optional[dict]:
    class_info = class_lookup.get(suite_code, {}).get(row.get("subject") or "")
    if not class_info:
        return None
    cloned = dict(row)
    cloned["class_id"] = class_info["id"]
    cloned["class_name"] = class_info["name"] if suite_code == primary_suite else f"{class_info['name']}（合班到{primary_suite}）"
    return cloned


def expand_online_merge_rows(rows: Sequence[dict], data_dir: Path) -> Tuple[List[dict], List[str]]:
    class_lookup = suite_subject_classes(data_dir)
    result: List[dict] = []
    merge_lines: List[str] = []

    for stage in {"春季", "秋季"}:
        for primary_suite, suite_codes in ONLINE_MERGE_GROUPS.items():
            source_rows = [
                row
                for row in rows
                if online_merge_phase_for_row(row) == stage
                and suite_code_from_class_id(row.get("class_id", "")) == primary_suite
            ]
            if not source_rows:
                continue
            for suite_code in suite_codes:
                added = 0
                for row in source_rows:
                    cloned = clone_row_for_suite(row, suite_code, primary_suite, class_lookup)
                    if cloned:
                        result.append(cloned)
                        added += 1
                if suite_code != primary_suite and added:
                    merge_lines.append(f"{stage} {suite_code} 合班到 {primary_suite}: {added} 条")

    online_suites = {suite_code for suite_codes in ONLINE_MERGE_GROUPS.values() for suite_code in suite_codes}
    for row in rows:
        stage = online_merge_phase_for_row(row)
        suite_code = suite_code_from_class_id(row.get("class_id", ""))
        if stage and suite_code in online_suites:
            continue
        result.append(row)

    return sorted(
        result,
        key=lambda item: (
            item["date"],
            scheduler.period_sort_value(item["period"]),
            item.get("start_time", ""),
            item.get("class_id", ""),
        ),
    ), merge_lines


def infer_module_group(subject: str, course_name: str) -> Tuple[str, str]:
    if subject == "数学":
        if "概率" in course_name:
            return "概率论", "数学类"
        if "线性" in course_name or "线代" in course_name:
            return "线代", "数学类"
        if "真题" in course_name:
            return "真题精讲", "数学类"
        if "拔高" in course_name:
            return "拔高专题", "数学类"
        return "高数" if course_name else "", "数学类"
    if subject == "英语":
        if "词汇" in course_name:
            return "词汇", "阅读类"
        if "语法" in course_name:
            return "语法", "写作类"
        if "完形" in course_name:
            return "完形", "阅读类"
        if "新题型" in course_name:
            return "新题型", "阅读类"
        if "阅读" in course_name:
            return "阅读", "阅读类"
        if "翻译" in course_name:
            return "翻译", "写作类"
        if "写作" in course_name:
            return "写作", "写作类"
        return "", ""
    if subject == "政治":
        for module in ("马原", "思修", "史纲", "毛中特", "新思想", "时政"):
            if module in course_name:
                return module, "马原类" if module in {"马原", "思修", "时政"} else "毛史类"
        return "", ""
    return "", ""


def product_course_code_lookup(data_dir: Path) -> Dict[Tuple[str, ...], str]:
    path = data_dir / "product_courses.csv"
    if not path.exists():
        return {}
    candidates: Dict[Tuple[str, ...], Set[str]] = defaultdict(set)
    for row in read_csv_rows(path):
        code = clean(row.get("course_code"))
        if not code:
            continue
        course_name = clean(row.get("course_name"))
        subject = clean(row.get("subject"))
        stage = clean(row.get("stage"))
        module = clean(row.get("course_module"))
        group = clean(row.get("course_group"))
        for key in (
            ("name", course_name),
            ("subject_stage_module_name", subject, stage, module, course_name),
            ("subject_stage_module_group", subject, stage, module, group),
        ):
            if all(key[1:]):
                candidates[key].add(code)
    return {key: next(iter(codes)) for key, codes in candidates.items() if len(codes) == 1}


def infer_course_code_from_lookup(row: dict, lookup: Dict[Tuple[str, ...], str]) -> str:
    existing = clean(row.get("course_code"))
    if existing:
        return existing
    course_name = clean(row.get("course_name"))
    subject = clean(row.get("subject"))
    stage = clean(row.get("stage"))
    module = clean(row.get("course_module"))
    group = clean(row.get("course_group"))
    for key in (
        ("subject_stage_module_name", subject, stage, module, course_name),
        ("name", course_name),
        ("subject_stage_module_group", subject, stage, module, group),
    ):
        if all(key[1:]) and key in lookup:
            return lookup[key]
    return ""


def infer_history_stage(
    start_date: str,
    course_name: str = "",
    class_id: str = "",
    sub_product: str = "",
) -> str:
    if sub_product == "全年营":
        for stage in ("导学1", "一轮", "二轮", "三轮", "四轮"):
            if stage in course_name:
                return stage
        if "专项突破" in course_name:
            return "二轮"
        return "一轮"
    if suite_code_from_class_id(class_id) in HALF_YEAR_BATCH_SUITES or sub_product in WYQC_PRODUCTS:
        for stage in ("基础", "强化", "冲刺"):
            if stage in course_name:
                return stage
    return "寒假" if start_date < "2026-04-01" else "春季"


def should_ignore_history_gap(row: dict) -> bool:
    return (
        row.get("class_id") == "KYJXZ2759"
        and row.get("date") in IGNORED_HISTORY_GAP_DATES
        and (not row.get("course_module") or not row.get("teacher_name"))
    )


def normalize_history_rows(path: Path, data_dir: Path) -> Tuple[List[dict], List[str], List[str]]:
    from openpyxl import load_workbook

    teacher_by_id, teacher_by_name = teacher_maps(data_dir)
    room_by_id, room_by_name = room_maps(data_dir)
    class_metadata = load_class_metadata(data_dir)
    course_code_lookup = product_course_code_lookup(data_dir)
    included_class_ids = history_class_ids(data_dir)
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    worksheet.reset_dimensions()
    raw_rows = worksheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(raw_rows)]

    rows: List[dict] = []
    for raw in raw_rows:
        row = {headers[index]: raw[index] if index < len(raw) else None for index in range(len(headers))}
        class_id = clean(row.get("班级编码"))
        if class_id not in included_class_ids:
            continue
        start_time = time_text(row.get("起始时间"))
        end_time = time_text(row.get("结束时间"))
        lesson_date = date_text(row.get("起始时间"))
        course_code = clean(row_value(row, "课程编码", "课程编号", "course_code"))
        course_name = clean(row_value(row, "课程名称", "课程内容"))
        teacher_id = clean(row_value(row, "教师ID", "合班教师ID", "teacher_id"))
        teacher_name = clean(row_value(row, "教师姓名", "教师名称", "配课老师姓名", "teacher_name"))
        if teacher_name and not teacher_id:
            teacher_id = teacher_by_name.get(teacher_name, "")
        if teacher_id and not teacher_name:
            teacher_name = teacher_by_id.get(teacher_id, "")
        room_name = clean(row_value(row, "教室名称", "上课教室", "room_name"))
        room_id = clean(row_value(row, "教室编码", "上课教室编码", "room_id"))
        if room_name and not room_id:
            room_id = room_by_name.get(room_name, "")
        if room_id and not room_name:
            room_name = room_by_id.get(room_id, "")
        rows.append(
            {
                "date": lesson_date,
                "period": period_for_time(start_time),
                "start_time": start_time,
                "end_time": end_time,
                "class_id": class_id,
                "class_name": clean(row_value(row, "班级名称", "班级名称（外）")) or class_id,
                "subject": normalize_subject(row_value(row, "科目", "科目内")),
                "stage": infer_history_stage(
                    lesson_date,
                    course_name,
                    class_id,
                    class_metadata.get(class_id, {}).get("sub_product", ""),
                ),
                "course_code": course_code,
                "course_name": course_name,
                "teacher_id": teacher_id,
                "teacher_name": teacher_name,
                "merge_teacher_id": clean(row_value(row, "合班教师ID", "合班教师编码")),
                "room_id": room_id,
                "room_name": room_name,
                "duration_hours": int(float(row.get("小时数") or 0)),
            }
        )

    donors_by_slot: Dict[Tuple[str, str, str, str], List[dict]] = defaultdict(list)
    for row in rows:
        if row["course_name"] and row["teacher_name"]:
            donors_by_slot[(row["date"], row["start_time"], row["end_time"], row["subject"])].append(row)

    warnings: List[str] = []
    ignored: List[str] = []
    normalized_rows: List[dict] = []
    for row in rows:
        donors = donors_by_slot.get((row["date"], row["start_time"], row["end_time"], row["subject"]), [])
        donor = next(
            (
                item
                for item in donors
                if row["merge_teacher_id"] and item["teacher_id"] == row["merge_teacher_id"]
            ),
            donors[0] if len(donors) == 1 else None,
        )
        if donor:
            for field in ("course_name", "room_id", "room_name"):
                if not row[field]:
                    row[field] = donor[field]
            if not row["teacher_id"]:
                row["teacher_id"] = donor["teacher_id"]
            if not row["teacher_name"]:
                row["teacher_name"] = donor["teacher_name"]
        if row["teacher_id"] and not row["teacher_name"]:
            row["teacher_name"] = teacher_by_id.get(row["teacher_id"], "")
        if row["teacher_name"] and not row["teacher_id"]:
            row["teacher_id"] = teacher_by_name.get(row["teacher_name"], "")
        if not row["room_id"] and row["room_name"]:
            warnings.append(f"{row['class_id']} {row['date']} {row['start_time']} 教室 {row['room_name']} 未匹配到教室编码")
        if not row["room_id"]:
            row["room_id"] = "RMHFWY97001"
        if not row["room_name"]:
            row["room_name"] = ONLINE_ROOM_NAME
        row["course_module"], row["course_group"] = infer_module_group(row["subject"], row["course_name"])
        if not row["course_code"]:
            row["course_code"] = infer_course_code_from_lookup(row, course_code_lookup)
        if should_ignore_history_gap(row):
            ignored.append(f"{row['class_id']} {row['date']} {row['start_time']} 历史错位缺课记录")
            continue
        if not row["course_module"]:
            warnings.append(f"{row['class_id']} {row['date']} {row['start_time']} 缺课程名称/模块")
        if not row["teacher_name"]:
            warnings.append(f"{row['class_id']} {row['date']} {row['start_time']} 缺老师")
        normalized_rows.append(row)

    return merge_history_rows(normalized_rows), warnings, ignored


def merge_history_rows(rows: Sequence[dict]) -> List[dict]:
    grouped: Dict[Tuple[str, str, str, str, str, str, str, str, str], List[dict]] = defaultdict(list)
    for row in rows:
        key = (
            row["date"],
            row["period"],
            row["class_id"],
            row["subject"],
            row["stage"],
            row["course_module"],
            row["teacher_id"],
            row["room_id"],
            row["course_name"],
        )
        grouped[key].append(row)

    merged: List[dict] = []
    for group in grouped.values():
        group = sorted(group, key=lambda item: item["start_time"])
        base = dict(group[0])
        base["start_time"] = min(item["start_time"] for item in group)
        base["end_time"] = max(item["end_time"] for item in group)
        base["duration_hours"] = sum(int(item["duration_hours"]) for item in group)
        merged.append(base)
    return sorted(merged, key=lambda item: (item["date"], scheduler.period_sort_value(item["period"]), item["start_time"], item["class_id"]))


STANDARD_SLOT_SPECS: Dict[str, Tuple[Tuple[int, str, str, str, int], ...]] = {
    "AM": (
        (1, "上午一", "08:00", "10:00", 2),
        (2, "上午二", "10:20", "12:20", 2),
    ),
    "PM": (
        (1, "下午一", "14:00", "16:00", 2),
        (2, "下午二", "16:20", "18:20", 2),
    ),
    "EVENING": (
        (1, "晚上", "19:00", "21:00", 2),
    ),
}


def standard_slots_for_row(row: dict, task_id: str) -> Tuple[scheduler.TimeSlot, ...]:
    date_text = row["date"]
    period = row["period"]
    duration = int(float(row.get("duration_hours") or 0))
    start_time = row.get("start_time") or ""
    end_time = row.get("end_time") or ""
    specs = STANDARD_SLOT_SPECS.get(period, ())
    selected_specs: List[Tuple[int, str, str, str, int]] = []

    start_minute = minutes(start_time) if start_time else None
    end_minute = minutes(end_time) if end_time else None
    if start_minute is not None and end_minute is not None:
        selected_specs = [
            spec
            for spec in specs
            if (minutes(spec[2]) or 0) >= start_minute and (minutes(spec[3]) or 0) <= end_minute
        ]
    elif duration and specs:
        remaining = duration
        for spec in specs:
            if remaining <= 0:
                break
            selected_specs.append(spec)
            remaining -= spec[4]

    if selected_specs and sum(spec[4] for spec in selected_specs) == duration:
        return tuple(
            scheduler.TimeSlot(
                id=f"{date_text}-{period}-{order}",
                date=date_text,
                period=period,
                name=name,
                order=order,
                start_time=slot_start,
                end_time=slot_end,
                duration_hours=slot_duration,
            )
            for order, name, slot_start, slot_end, slot_duration in selected_specs
        )

    slot = scheduler.TimeSlot(
        id=f"{date_text}-{period}-{row['class_id']}-{task_id}",
        date=date_text,
        period=period,
        name=period,
        order=1,
        start_time=start_time,
        end_time=end_time,
        duration_hours=duration,
    )
    return (slot,)


def assignment_from_row(row: dict, task_id: str) -> scheduler.Assignment:
    duration = int(float(row.get("duration_hours") or 0))
    slots = standard_slots_for_row(row, task_id)
    task = scheduler.CourseBlock(
        task_id=task_id,
        class_id=row["class_id"],
        class_name=row.get("class_name") or row["class_id"],
        product_id=None,
        product_name=shared_assignment_note_for_row(row) or None,
        class_size=None,
        subject_category=row.get("subject_category") or "公共课",
        subject=row.get("subject") or "",
        quarter=row.get("quarter") or None,
        stage=row.get("stage") or None,
        course_module=row.get("course_module") or None,
        course_group=row.get("course_group") or None,
        teacher_id=row.get("teacher_id") or "",
        teacher_name=row.get("teacher_name") or "",
        block_hours=duration,
        room_ids={row.get("room_id") or ""},
        start_date=row["date"],
        end_date=row["date"],
        allowed_periods={row["period"]},
        allowed_weekdays=None,
        excluded_weekdays=None,
        schedule_rules=(),
        is_locked=True,
        course_code=row.get("course_code") or "",
        course_name=row.get("course_name") or "",
    )
    return scheduler.Assignment(
        task=task,
        candidate=scheduler.Candidate(
            slots=slots,
            teacher_id=row.get("teacher_id") or "",
            teacher_name=row.get("teacher_name") or "",
            room_id=row.get("room_id") or "",
        ),
    )


def assignments_from_rows(rows: Sequence[dict], prefix: str) -> List[scheduler.Assignment]:
    return [assignment_from_row(row, f"{prefix}:{index}") for index, row in enumerate(rows, start=1)]


def load_locked_schedule_rows(data_dir: Path) -> List[dict]:
    path = data_dir / LOCKED_SCHEDULE_PATH.name
    if not path.exists():
        return []

    class_metadata = load_class_metadata(data_dir)
    course_code_lookup = product_course_code_lookup(data_dir)
    rows: List[dict] = []
    for row in read_csv_rows(path):
        if clean(row.get("is_locked")) not in {"是", "1", "true", "True", "yes", "Y", "y"}:
            continue
        class_id = clean(row.get("class_id"))
        lesson_date = clean(row.get("date"))
        room_id = clean(row.get("room_id"))
        if not class_id or not lesson_date or not room_id:
            continue
        class_meta = class_metadata.get(class_id, {})
        normalized = {
            "date": lesson_date,
            "period": clean(row.get("period")) or period_for_time(clean(row.get("start_time"))),
            "start_time": clean(row.get("start_time")),
            "end_time": clean(row.get("end_time")),
            "class_id": class_id,
            "class_name": clean(row.get("class_name")) or class_meta.get("name") or class_id,
            "subject_category": clean(row.get("subject_category")) or class_meta.get("subject_category") or "专业课",
            "subject": normalize_subject(row.get("subject")) or class_meta.get("subject") or "已定课程",
            "stage": clean(row.get("stage")) or stage_for_display_date(lesson_date),
            "course_code": clean(row.get("course_code")),
            "course_name": clean(row.get("course_name")),
            "course_module": clean(row.get("course_module")),
            "course_group": clean(row.get("course_group")),
            "teacher_id": clean(row.get("teacher_id")),
            "teacher_name": clean(row.get("teacher_name")),
            "room_id": room_id,
            "room_name": clean(row.get("room_name")) or room_id,
            "duration_hours": int(float(row.get("duration_hours") or 0)),
        }
        if not normalized["course_code"]:
            normalized["course_code"] = infer_course_code_from_lookup(normalized, course_code_lookup)
        rows.append(normalized)
    return sorted(
        rows,
        key=lambda item: (
            item["date"],
            scheduler.period_sort_value(item["period"]),
            item.get("start_time", ""),
            item.get("class_id", ""),
        ),
    )


def load_locked_schedule_assignments(data_dir: Path) -> Tuple[List[scheduler.Assignment], List[dict]]:
    rows = load_locked_schedule_rows(data_dir)
    return assignments_from_rows(rows, "LOCKED"), rows


def clone_assignment_for_suite(
    assignment: scheduler.Assignment,
    suite_code: str,
    primary_suite: str,
    class_lookup: Dict[str, Dict[str, Dict[str, str]]],
) -> Optional[scheduler.Assignment]:
    class_info = class_lookup.get(suite_code, {}).get(assignment.task.subject)
    if not class_info:
        return None
    task = replace(
        assignment.task,
        task_id=f"{assignment.task.task_id}:MERGED:{suite_code}",
        class_id=class_info["id"],
        class_name=class_info["name"] if suite_code == primary_suite else f"{class_info['name']}（合班到{primary_suite}）",
        product_name="" if suite_code == primary_suite else f"合班到 {primary_suite}",
    )
    return scheduler.Assignment(task=task, candidate=assignment.candidate)


def expand_online_merge_assignments(
    assignments: Sequence[scheduler.Assignment],
    data_dir: Path,
) -> Tuple[List[scheduler.Assignment], List[str]]:
    class_lookup = suite_subject_classes(data_dir)
    result: List[scheduler.Assignment] = []
    merge_lines: List[str] = []
    by_primary_stage: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)

    for assignment in assignments:
        primary_suite = suite_code_from_class_id(assignment.task.class_id)
        phase = online_merge_phase_for_assignment(assignment)
        if phase and primary_suite in ONLINE_MERGE_GROUPS:
            by_primary_stage[(primary_suite, phase)].append(assignment)
        else:
            result.append(assignment)

    for (primary_suite, stage), source_assignments in sorted(by_primary_stage.items()):
        for suite_code in ONLINE_MERGE_GROUPS[primary_suite]:
            added = 0
            for assignment in source_assignments:
                cloned = clone_assignment_for_suite(assignment, suite_code, primary_suite, class_lookup)
                if cloned:
                    result.append(cloned)
                    added += 1
            if suite_code != primary_suite and added:
                merge_lines.append(f"{stage} {suite_code} 合班到 {primary_suite}: {added} 条")

    return scheduler.sorted_assignments(result), merge_lines


def load_existing_summer(path: Path) -> List[scheduler.Assignment]:
    rows: List[dict] = []
    seen: Set[Tuple[str, str, str, str, str, str, str, str]] = set()

    def append_row(row: dict) -> None:
        if not row_matches_phase(row, SUMMER_PREPLAN_STAGES):
            return
        if row.get("subject") not in SUMMER_PUBLIC_SUBJECTS:
            return
        suite_code = (row.get("class_id") or "")[-4:]
        if suite_code not in MAINTENANCE_SUITES:
            return
        key = (
            row.get("date") or "",
            row.get("period") or "",
            row.get("class_id") or "",
            row.get("subject") or "",
            row.get("quarter") or "",
            row.get("stage") or "",
            row.get("course_module") or "",
            row.get("teacher_id") or "",
            row.get("room_id") or "",
        )
        if key in seen:
            return
        seen.add(key)
        rows.append(row)

    source_paths = [path] if path.exists() else []
    for source_path in source_paths:
        for row in read_csv_rows(source_path):
            append_row(row)
    if not rows:
        for source_path in SUMMER_FALLBACK_CSVS:
            if not source_path.exists():
                continue
            for row in read_csv_rows(source_path):
                append_row(row)
    return assignments_from_rows(rows, "SUMMER")


def load_shared_assignment_notes(data_dir: Path = ROOT / "data") -> Dict[Tuple[str, str, str, str], str]:
    path = data_dir / "class_teacher_assignments.csv"
    cache_key = (path.resolve(), file_mtime_ns(path))
    cached = _SHARED_ASSIGNMENT_CACHE.get(cache_key)
    if cached is not None:
        return cached
    notes: Dict[Tuple[str, str, str, str], str] = {}
    if not path.exists():
        _SHARED_ASSIGNMENT_CACHE[cache_key] = notes
        return notes
    for row in read_csv_rows(path):
        mode = assignment_schedule_mode(row)
        inherit_from = assignment_reference_class_id(row)
        if mode != "共享课表" and not inherit_from:
            continue
        class_id = clean(row.get("class_id"))
        subject = clean(row.get("subject"))
        phase = clean(row.get("stage"))
        course_group = clean(row.get("course_group"))
        if not class_id or not phase:
            continue
        note = f"合班到 {inherit_from}" if inherit_from else "共享课表"
        for key in (
            (class_id, subject, phase, course_group),
            (class_id, subject, phase, ""),
            (class_id, "", phase, course_group),
            (class_id, "", phase, ""),
        ):
            notes.setdefault(key, note)
    _SHARED_ASSIGNMENT_CACHE.clear()
    _SHARED_ASSIGNMENT_CACHE[cache_key] = notes
    return notes


def shared_assignment_note_for_row(row: dict) -> str:
    class_id = clean(row.get("class_id"))
    if not class_id:
        return ""
    notes = load_shared_assignment_notes()
    subject = clean(row.get("subject"))
    course_group = clean(row.get("course_group"))
    phases = [
        clean(row.get("quarter")),
        clean(row.get("stage")),
        stage_for_display_date(clean(row.get("date"))) if clean(row.get("date")) else "",
    ]
    for phase in [item for item in phases if item]:
        for key in (
            (class_id, subject, phase, course_group),
            (class_id, subject, phase, ""),
            (class_id, "", phase, course_group),
            (class_id, "", phase, ""),
        ):
            note = notes.get(key)
            if note:
                return note
    return ""


def summer_conflict_suite_codes(assignments: Sequence[scheduler.Assignment]) -> List[str]:
    grouped: Dict[Tuple[str, str, str, str], List[scheduler.Assignment]] = defaultdict(list)
    teacher_grouped: Dict[Tuple[str, str, str], List[scheduler.Assignment]] = defaultdict(list)
    for assignment in assignments:
        if not assignment_matches_phase(assignment, SUMMER_PREPLAN_STAGES):
            continue
        suite_code = suite_code_from_class_id(assignment.task.class_id)
        if suite_code not in MAINTENANCE_SUITES:
            continue
        first = assignment.candidate.slots[0]
        grouped[(suite_code, assignment.task.class_id, first.date, first.period)].append(assignment)
        teacher_key = candidate_teacher_key(assignment.candidate)
        if teacher_key:
            teacher_grouped[(teacher_key, first.date, first.period)].append(assignment)

    conflict_suites: Set[str] = set()
    for (suite_code, _class_id, _date, _period), group in grouped.items():
        for left_index, left in enumerate(group):
            if any(assignments_overlap(left, right) for right in group[left_index + 1 :]):
                conflict_suites.add(suite_code)
                break
    def suite_rank(suite_code: str) -> int:
        if suite_code in SUMMER_CAMP_SUITES:
            return 2
        if suite_code in {"2711", "2712"}:
            return 1
        return 0

    def conflict_rebuild_suite(left: scheduler.Assignment, right: scheduler.Assignment) -> str:
        left_suite = suite_code_from_class_id(left.task.class_id)
        right_suite = suite_code_from_class_id(right.task.class_id)
        return max(
            (left_suite, right_suite),
            key=lambda suite_code: (
                suite_rank(suite_code),
                suite_code,
            ),
        )

    for group in teacher_grouped.values():
        for left_index, left in enumerate(group):
            for right in group[left_index + 1 :]:
                if left.task.class_id == right.task.class_id:
                    continue
                if assignments_are_same_shared_merge_event(left, right):
                    continue
                if not assignments_overlap(left, right):
                    continue
                suite_code = conflict_rebuild_suite(left, right)
                if suite_rank(suite_code) > 0:
                    conflict_suites.add(suite_code)
    return sorted(conflict_suites)


def max_consecutive_dates(date_values: Iterable[str]) -> int:
    parsed = sorted({Date.fromisoformat(value) for value in date_values if value})
    best = 0
    current_run = 0
    previous: Optional[Date] = None
    for item in parsed:
        if previous is not None and (item - previous).days == 1:
            current_run += 1
        else:
            current_run = 1
        best = max(best, current_run)
        previous = item
    return best


def summer_distribution_rebuild_suite_codes(assignments: Sequence[scheduler.Assignment]) -> List[str]:
    teacher_dates: Dict[Tuple[str, str, str], List[str]] = defaultdict(list)
    subject_week_loads: Dict[Tuple[str, str], Dict[Tuple[int, int], int]] = defaultdict(lambda: defaultdict(int))
    suite_week_loads: Dict[str, Dict[Tuple[int, int], int]] = defaultdict(lambda: defaultdict(int))
    class_stage_assignments: Dict[Tuple[str, str, str], Dict[int, List[scheduler.Assignment]]] = defaultdict(lambda: defaultdict(list))
    class_teacher_day_hours: Dict[Tuple[str, str, str, str], float] = defaultdict(float)
    rebuild_suites: Set[str] = set()
    for assignment in assignments:
        if not assignment_matches_phase(assignment, SUMMER_PREPLAN_STAGES) or assignment.task.subject not in SUMMER_PUBLIC_SUBJECTS:
            continue
        suite_code = suite_code_from_class_id(assignment.task.class_id)
        if suite_code not in MAINTENANCE_SUITES:
            continue
        first = assignment.candidate.slots[0]
        teacher_key = assignment.candidate.teacher_id or assignment.candidate.teacher_name or assignment.task.teacher_id
        if teacher_key:
            teacher_dates[(suite_code, assignment.task.subject, teacher_key)].append(first.date)
        subject_week_loads[(suite_code, assignment.task.subject)][week_key(assignment.candidate.slots)] += 1
        suite_week_loads[suite_code][week_key(assignment.candidate.slots)] += 1
        stage_rank = SUMMER_STAGE_ORDER.get(assignment.task.stage or "")
        if stage_rank is not None:
            class_stage_assignments[(suite_code, assignment.task.class_id, assignment.task.subject)][stage_rank].append(assignment)
        teacher_key = candidate_teacher_key(assignment.candidate)
        if teacher_key:
            for date_text, hours in candidate_hours_by_date(assignment.candidate).items():
                class_teacher_day_hours[(suite_code, assignment.task.class_id, teacher_key, date_text)] += hours
    for _key, _required_module, first_assignment, _anchor_assignment in first_lesson_module_violations(assignments):
        if not assignment_matches_phase(first_assignment, SUMMER_PREPLAN_STAGES):
            continue
        suite_code = suite_code_from_class_id(first_assignment.task.class_id)
        if suite_code in MAINTENANCE_SUITES:
            rebuild_suites.add(suite_code)

    for (suite_code, _class_id, _teacher_key, _date_text), hours in class_teacher_day_hours.items():
        if hours >= 8:
            rebuild_suites.add(suite_code)

    for suite_code, _subject, _week, _count, _weekly_min, _weekly_max in summer_subject_week_rule_violations(assignments):
        rebuild_suites.add(suite_code)

    for suite_code, loads in suite_week_loads.items():
        if any(count >= 12 for count in loads.values()):
            rebuild_suites.add(suite_code)

    for (suite_code, _class_id, _subject), by_rank in class_stage_assignments.items():
        ranks = sorted(by_rank)
        for lower_rank, higher_rank in zip(ranks, ranks[1:]):
            latest_lower = max(by_rank[lower_rank], key=lambda item: scheduler.slot_sort_key(item.candidate.slots[-1]))
            earliest_higher = min(by_rank[higher_rank], key=lambda item: scheduler.slot_sort_key(item.candidate.slots[0]))
            if scheduler.slot_sort_key(latest_lower.candidate.slots[-1]) >= scheduler.slot_sort_key(earliest_higher.candidate.slots[0]):
                rebuild_suites.add(suite_code)

    for (suite_code, subject, _teacher_key), dates in teacher_dates.items():
        run_length = max_consecutive_dates(dates)
        if subject in {"英语", "政治"} and run_length > 3:
            rebuild_suites.add(suite_code)
        elif subject == "数学" and run_length >= 4:
            rebuild_suites.add(suite_code)

    for (suite_code, subject), loads in subject_week_loads.items():
        if not loads or sum(loads.values()) < 8:
            continue
        values = list(loads.values())
        if max(values) - min(values) >= (4 if subject == "数学" else 5):
            rebuild_suites.add(suite_code)
    return sorted(rebuild_suites)


def class_window_venue_rebuild_suite_codes(
    assignments: Sequence[scheduler.Assignment],
    class_window_constraints: Dict[str, ClassWindowConstraint],
) -> List[str]:
    rebuild_suites: Set[str] = set()
    for assignment in assignments:
        if not assignment_matches_phase(assignment, SUMMER_PREPLAN_STAGES) or assignment.task.subject not in SUMMER_PUBLIC_SUBJECTS:
            continue
        constraint = class_window_constraints.get(assignment.task.class_id)
        if not constraint:
            continue
        first = assignment.candidate.slots[0]
        if constraint.earliest_date and first.date < constraint.earliest_date:
            continue
        if constraint.latest_date and first.date > constraint.latest_date:
            continue
        if constraint.room_ids and assignment.candidate.room_id not in constraint.room_ids:
            rebuild_suites.add(suite_code_from_class_id(assignment.task.class_id))
    return sorted(rebuild_suites)


def summer_subject_week_rule_violations(
    assignments: Sequence[scheduler.Assignment],
) -> List[Tuple[str, str, Tuple[int, int], int, int, int]]:
    subject_weeks: Dict[Tuple[str, str], Dict[Tuple[int, int], int]] = defaultdict(lambda: defaultdict(int))
    suite_subjects: Dict[str, Set[str]] = defaultdict(set)
    suite_subject_blocks: Dict[Tuple[str, str], List[Tuple[scheduler.TimeSlot, ...]]] = defaultdict(list)
    for assignment in assignments:
        if not assignment_matches_phase(assignment, SUMMER_PREPLAN_STAGES) or assignment.task.subject not in SUMMER_PUBLIC_SUBJECTS:
            continue
        suite_code = suite_code_from_class_id(assignment.task.class_id)
        if suite_code not in MAINTENANCE_SUITES:
            continue
        suite_subjects[suite_code].add(assignment.task.subject)
        subject_weeks[(suite_code, assignment.task.subject)][week_key(assignment.candidate.slots)] += 1
        suite_subject_blocks[(suite_code, assignment.task.subject)].append(assignment.candidate.slots)

    violations: List[Tuple[str, str, Tuple[int, int], int, int, int]] = []
    for suite_code, subjects in sorted(suite_subjects.items()):
        subject_counts = {
            subject: sum(subject_weeks.get((suite_code, subject), {}).values())
            for subject in subjects
        }
        bounds = average_subject_week_bounds_from_counts(
            {
                subject: suite_subject_blocks.get((suite_code, subject), [])
                for subject in subjects
            },
            subject_counts,
        )
        for subject in sorted(subjects):
            if subject not in bounds:
                continue
            weekly_min, weekly_max = bounds[subject]
            subject_loads = subject_weeks.get((suite_code, subject), {})
            edge_weeks = {min(subject_loads), max(subject_loads)} if subject_loads else set()
            for key, count in sorted(subject_loads.items()):
                if weekly_min and count < weekly_min:
                    if key in edge_weeks:
                        continue
                    violations.append((suite_code, subject, key, count, weekly_min, weekly_max or 999))
                elif weekly_max and count > weekly_max:
                    violations.append((suite_code, subject, key, count, weekly_min or 0, weekly_max))
    return violations


def summer_window_bounds(
    constraints: Sequence[ClassWindowConstraint],
) -> Tuple[ClassWindowConstraint, ClassWindowConstraint]:
    if not constraints:
        raise ValueError("缺少班级暑假排课窗口")
    start_constraint = min(
        constraints,
        key=lambda constraint: (
            constraint.earliest_date,
            scheduler.period_sort_value(constraint.earliest_period),
        ),
    )
    end_constraint = max(
        constraints,
        key=lambda constraint: (
            constraint.latest_date,
            scheduler.period_sort_value(constraint.latest_period),
        ),
    )
    return start_constraint, end_constraint


def summer_schedule_input_for_suites(
    data_dir: Path,
    suite_codes: Sequence[str],
    suite_window_constraints: Dict[str, ClassWindowConstraint],
    protected_assignments: Sequence[scheduler.Assignment],
) -> Tuple[scheduler.ScheduleInput, List[str]]:
    suite_code_list = list(suite_codes)
    class_ids = class_ids_for_suite_codes(data_dir, suite_code_list, None)
    source = load_schedule_input_for_classes(data_dir, class_ids)
    if protected_assignments:
        source = replace(source, locked_assignments=[*source.locked_assignments, *protected_assignments])
    all_class_window_constraints = load_summer_class_window_constraints(data_dir)
    class_window_constraints = {
        class_id: all_class_window_constraints[class_id]
        for class_id in class_ids
        if class_id in all_class_window_constraints
    }
    selected_constraints = [
        suite_window_constraints[suite_code]
        for suite_code in suite_code_list
        if suite_code in suite_window_constraints
    ]
    if not selected_constraints:
        suite_label = ", ".join(suite_code_list) or "未指定套班"
        raise ValueError(f"{suite_label} 缺少班级暑假排课窗口")
    start_constraint, end_constraint = summer_window_bounds(selected_constraints)
    summer_input = filtered_schedule_input(
        source,
        class_ids=class_ids,
        stages=SUMMER_PREPLAN_STAGES,
        subjects=None,
        start=start_constraint.earliest_date,
        end=end_constraint.latest_date,
        start_period=start_constraint.earliest_period,
        end_period=end_constraint.latest_period,
        periods={"AM", "PM"},
        room_ids=None,
        quarters={"暑假"},
        class_window_constraints=class_window_constraints,
    )
    return summer_input, class_ids


SummerRebuildAttempt = Tuple[str, Callable[[], List[scheduler.Assignment]], Optional[int]]


def run_summer_rebuild_attempts(
    attempts: Sequence[SummerRebuildAttempt],
) -> Tuple[List[scheduler.Assignment], str, List[str]]:
    errors: List[str] = []
    for label, callback, timeout_seconds in attempts:
        try:
            result = (
                run_fast_callback_with_timeout(callback, timeout_seconds)
                if timeout_seconds is not None
                else callback()
            )
            return result, label, errors
        except ValueError as exc:
            errors.append(f"{label}: {exc}")
    raise ValueError("；".join(errors))


def rebuild_summer_suite_with_core_scheduler(
    data_dir: Path,
    suite_code: str,
    class_metadata: Dict[str, Dict[str, str]],
    suite_window_constraints: Dict[str, ClassWindowConstraint],
    protected_assignments: Sequence[scheduler.Assignment],
) -> List[scheduler.Assignment]:
    summer_input, class_ids = summer_schedule_input_for_suites(
        data_dir,
        [suite_code],
        suite_window_constraints,
        protected_assignments,
    )
    assignments = scheduler.schedule(summer_input)
    selected_class_ids = set(class_ids)
    return scheduler.sorted_assignments([
        assignment for assignment in assignments if assignment.task.class_id in selected_class_ids
    ])


def rebuild_summer_suites_with_core_scheduler(
    data_dir: Path,
    suite_codes: Sequence[str],
    class_metadata: Dict[str, Dict[str, str]],
    suite_window_constraints: Dict[str, ClassWindowConstraint],
    protected_assignments: Sequence[scheduler.Assignment],
) -> List[scheduler.Assignment]:
    summer_input, class_ids = summer_schedule_input_for_suites(
        data_dir,
        suite_codes,
        suite_window_constraints,
        protected_assignments,
    )
    assignments = scheduler.schedule(summer_input)
    selected_class_ids = set(class_ids)
    return scheduler.sorted_assignments([
        assignment for assignment in assignments if assignment.task.class_id in selected_class_ids
    ])


def rebuild_summer_suite_balanced(
    data_dir: Path,
    suite_code: str,
    class_metadata: Dict[str, Dict[str, str]],
    suite_window_constraints: Dict[str, ClassWindowConstraint],
    protected_assignments: Sequence[scheduler.Assignment],
) -> List[scheduler.Assignment]:
    summer_input, class_ids = summer_schedule_input_for_suites(
        data_dir,
        [suite_code],
        suite_window_constraints,
        protected_assignments,
    )
    return schedule_balanced_camp_by_suite(
        summer_input,
        class_ids,
        class_metadata,
        compact_english_politics_without_math=True,
    )


def rebuild_summer_suite_round_robin(
    data_dir: Path,
    suite_code: str,
    class_metadata: Dict[str, Dict[str, str]],
    suite_window_constraints: Dict[str, ClassWindowConstraint],
    protected_assignments: Sequence[scheduler.Assignment],
    spacing_improvement_passes: Optional[int] = None,
) -> List[scheduler.Assignment]:
    summer_input, class_ids = summer_schedule_input_for_suites(
        data_dir,
        [suite_code],
        suite_window_constraints,
        protected_assignments,
    )
    suite_subjects = {
        requirement.subject
        for cls in summer_input.classes.values()
        for requirement in cls.requirements
    }
    sub_products = {
        class_metadata.get(class_id, {}).get("sub_product", "")
        for class_id in class_ids
    }
    is_wuyou_product = bool(sub_products & WUYOU_PRODUCTS)
    is_wuyou_han = "无忧寒" in sub_products
    subject_week_bounds = (
        wuyou_summer_week_bounds(suite_subjects)
        if is_wuyou_han
        else None
    )
    weekly_total_max = (
        wuyou_summer_weekly_total_max(suite_subjects)
        if is_wuyou_han
        else SUMMER_PREFERRED_WEEKLY_HALFDAY_MAX
    )
    return schedule_round_robin(
        summer_input,
        class_ids,
        subject_week_bounds,
        weekly_total_max,
        hard_weekly_total_max=weekly_total_max if is_wuyou_han else None,
        avoid_public_subject_consecutive_days=True,
        prefer_public_teacher_alternation=True,
        teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
        spacing_improvement_passes=spacing_improvement_passes if spacing_improvement_passes is not None else (0 if is_wuyou_product else 300),
        require_all_subject_weeks=True,
        strict_subject_week_quotas=is_wuyou_han,
        use_average_subject_week_bounds=not is_wuyou_han,
    )


def rebuild_summer_suite_conflict_first(
    data_dir: Path,
    suite_code: str,
    class_metadata: Dict[str, Dict[str, str]],
    suite_window_constraints: Dict[str, ClassWindowConstraint],
    protected_assignments: Sequence[scheduler.Assignment],
) -> List[scheduler.Assignment]:
    summer_input, class_ids = summer_schedule_input_for_suites(
        data_dir,
        [suite_code],
        suite_window_constraints,
        protected_assignments,
    )
    return schedule_round_robin(
        summer_input,
        class_ids,
        avoid_public_subject_consecutive_days=True,
        prefer_public_teacher_alternation=True,
        teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
        spacing_improvement_passes=0,
        require_all_subject_weeks=False,
    )


def summer_suite_codes_with_class_windows(data_dir: Path) -> Tuple[List[str], List[str]]:
    constraints = load_summer_suite_window_constraints(data_dir)
    with_constraints = [suite_code for suite_code in MAINTENANCE_SUITES if suite_code in constraints]
    missing = [
        suite_code
        for suite_code in MAINTENANCE_SUITES
        if suite_code not in constraints
        and any(
            class_info.get("id")
            for class_info in suite_subject_classes(data_dir).get(suite_code, {}).values()
        )
    ]
    return with_constraints, missing


def build_summer_assignments(
    data_dir: Path,
    initial_protected_assignments: Sequence[scheduler.Assignment] = (),
) -> Tuple[List[scheduler.Assignment], List[str], List[str], List[str]]:
    class_metadata = load_class_metadata(data_dir)
    class_window_constraints = load_summer_class_window_constraints(data_dir)
    suite_window_constraints = load_summer_suite_window_constraints(data_dir, class_metadata)
    suite_codes, missing_suite_codes = summer_suite_codes_with_class_windows(data_dir)
    if not suite_codes:
        return [], missing_suite_codes, [], []

    rebuild_summer = parse_bool(os.environ.get("REBUILD_SUMMER", ""))
    existing: List[scheduler.Assignment] = []
    suites_to_schedule = list(suite_codes)
    conflict_rebuild_suites: List[str] = []
    distribution_rebuild_suites: List[str] = []
    existing_by_suite: Dict[str, List[scheduler.Assignment]] = defaultdict(list)
    if not rebuild_summer:
        existing = load_existing_summer(SUMMER_PAGE_CSV)
        for assignment in existing:
            existing_by_suite[suite_code_from_class_id(assignment.task.class_id)].append(assignment)
        conflict_rebuild_suites = summer_conflict_suite_codes(existing)
        distribution_rebuild_suites = summer_distribution_rebuild_suite_codes(existing)
        class_window_venue_rebuild_suites = class_window_venue_rebuild_suite_codes(existing, class_window_constraints)
        distribution_rebuild_suites = sorted(set(distribution_rebuild_suites) | set(class_window_venue_rebuild_suites))
        rebuild_suite_set = set(conflict_rebuild_suites) | set(distribution_rebuild_suites)
        if rebuild_suite_set:
            existing = [
                assignment
                for assignment in existing
                if suite_code_from_class_id(assignment.task.class_id) not in rebuild_suite_set
            ]
        existing_suites = {
            suite_code_from_class_id(assignment.task.class_id)
            for assignment in existing
        }
        suites_to_schedule = [
            suite_code
            for suite_code in suite_codes
            if suite_code not in existing_suites or suite_code in rebuild_suite_set
        ]
        if not suites_to_schedule:
            return existing, missing_suite_codes, [], []

    fallback_suites = [
        suite_code
        for suite_code in suites_to_schedule
        if suite_code in set(conflict_rebuild_suites) | set(distribution_rebuild_suites)
    ]
    fallback_assignments: List[scheduler.Assignment] = []
    protected_assignments = [*initial_protected_assignments, *existing]
    for suite_code in fallback_suites:
        print(f"预排 暑假 {suite_code}", flush=True)
        suite_class_ids = class_ids_for_suite_codes(data_dir, [suite_code], None)
        constraint = suite_window_constraints[suite_code]
        suite_protected_assignments = relevant_locked_assignments_for_replan(
            data_dir,
            suite_class_ids,
            protected_assignments,
            constraint.earliest_date,
            constraint.latest_date,
        )
        try:
            rebuilt, _used_label, _errors = run_summer_rebuild_attempts(
                (
                    (
                        "快速轮排",
                        lambda suite=suite_code: rebuild_summer_suite_round_robin(
                            data_dir,
                            suite,
                            class_metadata,
                            suite_window_constraints,
                            suite_protected_assignments,
                            0,
                        ),
                        20,
                    ),
                    (
                        "均衡",
                        lambda suite=suite_code: rebuild_summer_suite_balanced(
                            data_dir,
                            suite,
                            class_metadata,
                            suite_window_constraints,
                            suite_protected_assignments,
                        ),
                        20,
                    ),
                    (
                        "基础调度器",
                        lambda suite=suite_code: rebuild_summer_suite_with_core_scheduler(
                            data_dir,
                            suite,
                            class_metadata,
                            suite_window_constraints,
                            suite_protected_assignments,
                        ),
                        20,
                    ),
                    (
                        "冲突优先",
                        lambda suite=suite_code: rebuild_summer_suite_conflict_first(
                            data_dir,
                            suite,
                            class_metadata,
                            suite_window_constraints,
                            suite_protected_assignments,
                        ),
                        None,
                    ),
                )
            )
        except ValueError:
            relaxed_protected = relevant_locked_assignments_for_replan(
                data_dir,
                suite_class_ids,
                initial_protected_assignments,
                constraint.earliest_date,
                constraint.latest_date,
            )
            try:
                rebuilt = rebuild_summer_suite_conflict_first(
                    data_dir,
                    suite_code,
                    class_metadata,
                    suite_window_constraints,
                    relaxed_protected,
                )
            except ValueError:
                rebuilt = existing_by_suite.get(suite_code, [])
                if not rebuilt:
                    raise
                conflict_rebuild_suites = [
                    item for item in conflict_rebuild_suites if item != suite_code
                ]
                distribution_rebuild_suites = [
                    item for item in distribution_rebuild_suites if item != suite_code
                ]
        fallback_assignments.extend(rebuilt)
        protected_assignments.extend(rebuilt)

    balanced_suite_codes = [
        suite_code for suite_code in suites_to_schedule if suite_code not in set(fallback_suites)
    ]
    if not balanced_suite_codes:
        return (
            scheduler.sorted_assignments([*existing, *fallback_assignments]),
            missing_suite_codes,
            conflict_rebuild_suites,
            distribution_rebuild_suites,
        )

    assignments: List[scheduler.Assignment] = []
    for suite_code in balanced_suite_codes:
        print(f"预排 暑假 {suite_code}", flush=True)
        suite_class_ids = class_ids_for_suite_codes(data_dir, [suite_code], None)
        constraint = suite_window_constraints[suite_code]
        suite_protected_assignments = relevant_locked_assignments_for_replan(
            data_dir,
            suite_class_ids,
            protected_assignments,
            constraint.earliest_date,
            constraint.latest_date,
        )
        try:
            rebuilt, _used_label, _errors = run_summer_rebuild_attempts(
                (
                    (
                        "轮排",
                        lambda suite=suite_code: rebuild_summer_suite_round_robin(
                            data_dir,
                            suite,
                            class_metadata,
                            suite_window_constraints,
                            suite_protected_assignments,
                        ),
                        None,
                    ),
                    (
                        "均衡",
                        lambda suite=suite_code: rebuild_summer_suite_balanced(
                            data_dir,
                            suite,
                            class_metadata,
                            suite_window_constraints,
                            suite_protected_assignments,
                        ),
                        None,
                    ),
                    (
                        "冲突优先",
                        lambda suite=suite_code: rebuild_summer_suite_conflict_first(
                            data_dir,
                            suite,
                            class_metadata,
                            suite_window_constraints,
                            suite_protected_assignments,
                        ),
                        None,
                    ),
                )
            )
        except ValueError:
            relaxed_protected = relevant_locked_assignments_for_replan(
                data_dir,
                suite_class_ids,
                initial_protected_assignments,
                constraint.earliest_date,
                constraint.latest_date,
            )
            rebuilt = rebuild_summer_suite_conflict_first(
                data_dir,
                suite_code,
                class_metadata,
                suite_window_constraints,
                relaxed_protected,
            )
        assignments.extend(rebuilt)
        protected_assignments.extend(rebuilt)
    return (
        scheduler.sorted_assignments([*existing, *fallback_assignments, *assignments]),
        missing_suite_codes,
        conflict_rebuild_suites,
        distribution_rebuild_suites,
    )


def autumn_sunday_slots(start: str, end: str, holidays: Set[str]) -> List[scheduler.TimeSlot]:
    slot_specs = (
        ("AM", "上午一", 1, "08:00", "10:00"),
        ("AM", "上午二", 2, "10:20", "12:20"),
        ("PM", "下午一", 1, "14:00", "16:00"),
        ("PM", "下午二", 2, "16:20", "18:20"),
    )
    current = Date.fromisoformat(start)
    last = Date.fromisoformat(end)
    slots: List[scheduler.TimeSlot] = []
    while current <= last:
        date_text = current.isoformat()
        if current.weekday() == 6 and date_text not in holidays:
            for period, name, order, start_time, end_time in slot_specs:
                slots.append(
                    scheduler.TimeSlot(
                        id=f"{date_text}-{period}-{order}",
                        date=date_text,
                        period=period,
                        name=name,
                        order=order,
                        start_time=start_time,
                        end_time=end_time,
                        duration_hours=2,
                    )
                )
        current += timedelta(days=1)
    return slots


def autumn_time_slots(source_slots: Sequence[scheduler.TimeSlot]) -> List[scheduler.TimeSlot]:
    by_id = {
        slot.id: slot
        for slot in source_slots
        if AUTUMN_START <= slot.date <= AUTUMN_END and slot.date not in AUTUMN_HOLIDAYS
    }
    for slot in autumn_sunday_slots(AUTUMN_POLITICS_START, AUTUMN_END, AUTUMN_HOLIDAYS):
        by_id.setdefault(slot.id, slot)
    return sorted(by_id.values(), key=scheduler.slot_sort_key)


def default_autumn_subject_rule(subject: str) -> Dict[str, object]:
    if subject == "英语":
        return {
            "start_date": AUTUMN_ENGLISH_START,
            "allowed_weekdays": {1, 3},
            "allowed_periods": {"EVENING"},
            "preferred_period": "EVENING",
            "block_hours": 2,
            "label": "英语周二/周四晚课",
        }
    if subject == "数学":
        return {
            "start_date": AUTUMN_MATH_START,
            "allowed_weekdays": {2, 4},
            "allowed_periods": {"EVENING"},
            "preferred_period": "EVENING",
            "block_hours": 2,
            "label": "数学周三/周五晚课",
        }
    if subject == "政治":
        return {
            "start_date": AUTUMN_POLITICS_START,
            "allowed_weekdays": {5, 6},
            "allowed_periods": {"AM", "PM"},
            "preferred_period": "PM",
            "block_hours": 4,
            "one_per_week": True,
            "label": "政治周末半天课",
        }
    return {
        "start_date": AUTUMN_START,
        "allowed_weekdays": None,
        "allowed_periods": None,
        "preferred_period": None,
        "block_hours": None,
        "label": subject,
    }


def autumn_subject_rule(suite_code: str, subject: str) -> Dict[str, object]:
    return AUTUMN_SPECIAL_SUBJECT_RULES.get((suite_code, subject), default_autumn_subject_rule(subject))


def autumn_requirement(requirement: scheduler.Requirement, room_id: str, suite_code: str) -> scheduler.Requirement:
    rule = autumn_subject_rule(suite_code, requirement.subject)
    if requirement.subject == "英语":
        return replace(
            requirement,
            room_ids={room_id},
            block_hours=int(rule["block_hours"]),
            start_date=str(rule["start_date"]),
            end_date=AUTUMN_END,
            allowed_periods=set(rule["allowed_periods"]),
            allowed_weekdays=set(rule["allowed_weekdays"]),
            excluded_weekdays=None,
            schedule_rules=(),
        )
    if requirement.subject == "数学":
        return replace(
            requirement,
            room_ids={room_id},
            block_hours=int(rule["block_hours"]),
            start_date=str(rule["start_date"]),
            end_date=AUTUMN_END,
            allowed_periods=set(rule["allowed_periods"]),
            allowed_weekdays=set(rule["allowed_weekdays"]),
            excluded_weekdays=None,
            schedule_rules=(),
        )
    if requirement.subject == "政治":
        return replace(
            requirement,
            room_ids={room_id},
            block_hours=int(rule["block_hours"]),
            start_date=str(rule["start_date"]),
            end_date=AUTUMN_END,
            allowed_periods=set(rule["allowed_periods"]),
            allowed_weekdays=set(rule["allowed_weekdays"]),
            excluded_weekdays=None,
            schedule_rules=(),
        )
    return replace(
        requirement,
        room_ids={room_id},
        start_date=AUTUMN_START,
        end_date=AUTUMN_END,
        schedule_rules=(),
    )


def autumn_rule_dates(rule: Dict[str, object], count: int) -> List[str]:
    dates: List[str] = []
    used_weeks: Set[Tuple[int, int]] = set()
    current = Date.fromisoformat(str(rule["start_date"]))
    last = Date.fromisoformat(AUTUMN_END)
    allowed_weekdays = set(rule.get("allowed_weekdays") or set())
    one_per_week = bool(rule.get("one_per_week"))
    while current <= last and len(dates) < count:
        date_text = current.isoformat()
        week = current.isocalendar()[:2]
        if (
            current.weekday() in allowed_weekdays
            and date_text not in AUTUMN_HOLIDAYS
            and (not one_per_week or week not in used_weeks)
        ):
            dates.append(date_text)
            used_weeks.add(week)
        current += timedelta(days=1)
    if len(dates) < count:
        label = str(rule.get("label") or "秋季课节")
        raise ValueError(f"{label}课节不足，需要 {count} 次，仅找到 {len(dates)} 次")
    return dates


def fixed_subject_assignments(
    autumn_input: scheduler.ScheduleInput,
    class_ids: Sequence[str],
    subject: str,
    fixed_suite_codes: Optional[Set[str]] = None,
) -> List[scheduler.Assignment]:
    result: List[scheduler.Assignment] = []
    for class_id in class_ids:
        suite_code = suite_code_from_class_id(class_id)
        if fixed_suite_codes is not None and suite_code not in fixed_suite_codes:
            continue
        cls = autumn_input.classes.get(class_id)
        if not cls:
            continue
        tasks = [
            task
            for task in scheduler.build_course_blocks({class_id: cls})
            if task.subject == subject
        ]
        if not tasks:
            continue
        if subject == "英语":
            tasks = alternate_teacher_tasks(tasks)
        rule = autumn_subject_rule(suite_code, subject)
        dates = autumn_rule_dates(rule, len(tasks))
        preferred_period = str(rule.get("preferred_period") or "")
        for task, date_text in zip(tasks, dates):
            candidates = [
                candidate
                for candidate in scheduler.candidate_assignments(task, autumn_input)
                if candidate.slots[0].date == date_text
                and (not preferred_period or candidate.slots[0].period == preferred_period)
            ]
            if not candidates:
                label = str(rule.get("label") or subject)
                raise ValueError(f"{label}任务 {task.task_id} 没有 {date_text} 可用课节")
            candidates.sort(
                key=lambda candidate: (
                    scheduler.candidate_same_day_teacher_travel_penalty(
                        autumn_input,
                        [*autumn_input.locked_assignments, *result],
                        task,
                        candidate,
                    ),
                    candidate.room_id,
                )
            )
            result.append(scheduler.Assignment(task=task, candidate=candidates[0]))
    return scheduler.sorted_assignments(result)


def apply_fixed_autumn_sequences(
    autumn_input: scheduler.ScheduleInput,
    class_ids: Sequence[str],
    assignments: Sequence[scheduler.Assignment],
) -> List[scheduler.Assignment]:
    english = fixed_subject_assignments(autumn_input, class_ids, "英语")
    politics = fixed_subject_assignments(autumn_input, class_ids, "政治")
    special_math = fixed_subject_assignments(autumn_input, class_ids, "数学", {"2791"})
    fixed = [*english, *politics, *special_math]
    fixed_task_ids = {assignment.task.task_id for assignment in fixed}
    return scheduler.sorted_assignments([
        *[assignment for assignment in assignments if assignment.task.task_id not in fixed_task_ids],
        *fixed,
    ])


def autumn_teacher_overlap_lines(assignments: Sequence[scheduler.Assignment]) -> List[str]:
    grouped: Dict[Tuple[str, str, str, str, str], List[scheduler.Assignment]] = defaultdict(list)
    for assignment in assignments:
        if not assignment.candidate.teacher_id:
            continue
        first = assignment.candidate.slots[0]
        last = assignment.candidate.slots[-1]
        grouped[
            (
                assignment.candidate.teacher_id,
                first.date,
                first.period,
                first.start_time,
                last.end_time,
            )
        ].append(assignment)

    lines: List[str] = []
    for (teacher_id, date_text, period, start_time, end_time), group in sorted(grouped.items()):
        if len(group) <= 1:
            continue
        teacher_name = group[0].candidate.teacher_name or teacher_id
        details = "；".join(
            f"{item.task.class_id} {item.task.subject}/{item.task.course_module or ''}"
            for item in sorted(group, key=lambda assignment: assignment.task.class_id)
        )
        lines.append(f"{date_text} {period} {start_time}-{end_time} {teacher_name}: {details}")
    return lines


def build_autumn_assignments(
    data_dir: Path,
    protected_assignments: Sequence[scheduler.Assignment] = (),
) -> Tuple[List[scheduler.Assignment], Dict[str, str]]:
    class_metadata = load_class_metadata(data_dir)
    class_ids = class_ids_for_suite_codes(data_dir, ONLINE_PRIMARY_SUITES, None)
    source = load_schedule_input_for_classes(data_dir, class_ids)
    if protected_assignments:
        source = replace(source, locked_assignments=[*source.locked_assignments, *protected_assignments])
    autumn_input = filtered_schedule_input(
        source,
        class_ids=class_ids,
        stages=AUTUMN_PREPLAN_STAGES,
        subjects=None,
        start=AUTUMN_START,
        end=AUTUMN_END,
        start_period="AM",
        end_period="EVENING",
        periods={"AM", "PM", "EVENING"},
        room_ids=None,
    )

    rooms = dict(autumn_input.rooms)
    room_names: Dict[str, str] = {}
    classes = {}
    for class_id, cls in autumn_input.classes.items():
        suite_code = class_metadata.get(class_id, {}).get("suite_code") or class_id[-4:]
        room_id = f"RMONLINE{suite_code}"
        rooms[room_id] = scheduler.Room(id=room_id, capacity=None, capacity_unlimited=True)
        room_names[room_id] = f"线上教室{suite_code}"
        classes[class_id] = replace(
            cls,
            room_ids={room_id},
            start_date=AUTUMN_START,
            start_period="AM",
            end_date=AUTUMN_END,
            end_period="EVENING",
            requirements=[autumn_requirement(requirement, room_id, suite_code) for requirement in cls.requirements],
        )

    autumn_input = replace(autumn_input, rooms=rooms, classes=classes)
    autumn_input = replace(autumn_input, time_slots=autumn_time_slots(autumn_input.time_slots))
    assignments = schedule_balanced_camp_by_suite(
        autumn_input,
        class_ids,
        class_metadata,
        {"英语": "EVENING", "数学": "EVENING", "政治": "PM"},
        lock_previous_assignments=False,
    )
    assignments = apply_fixed_autumn_sequences(autumn_input, class_ids, assignments)
    return assignments, room_names


def history_requirement_key(row: dict) -> Tuple[str, str, str, str]:
    return scheduler.requirement_mapping_key(row)


def apply_history_deductions(
    schedule_input: scheduler.ScheduleInput,
    history_rows: Sequence[dict],
    before_date: str,
) -> scheduler.ScheduleInput:
    consumed_hours: Dict[Tuple[str, Tuple[str, str, str, str]], int] = defaultdict(int)
    for row in history_rows:
        class_id = row.get("class_id") or ""
        if class_id not in schedule_input.classes:
            continue
        if (row.get("date") or "") >= before_date:
            continue
        key = history_requirement_key(row)
        if not all(key[:3]):
            continue
        consumed_hours[(class_id, key)] += int(float(row.get("duration_hours") or 0))

    if not consumed_hours:
        return schedule_input

    classes: Dict[str, scheduler.SchoolClass] = {}
    for class_id, cls in schedule_input.classes.items():
        adjusted_requirements: List[scheduler.Requirement] = []
        for requirement in cls.requirements:
            consumed = consumed_hours.get((class_id, scheduler.requirement_object_key(requirement)), 0)
            if consumed <= 0:
                adjusted_requirements.append(requirement)
                continue
            remaining = max(0, requirement.total_hours - consumed)
            if remaining <= 0:
                continue
            block_hours = requirement.block_hours
            if block_hours > remaining or remaining % block_hours != 0:
                block_hours = remaining
            adjusted_requirements.append(
                replace(requirement, total_hours=remaining, block_hours=block_hours)
            )
        if adjusted_requirements:
            classes[class_id] = replace(cls, requirements=adjusted_requirements)
    return replace(schedule_input, classes=classes)


def remove_history_replaced_assignments(
    assignments: Sequence[scheduler.Assignment],
    history_rows: Sequence[dict],
    before_date: str,
) -> Tuple[List[scheduler.Assignment], int]:
    consumed_hours: Dict[Tuple[str, Tuple[str, str, str, str]], int] = defaultdict(int)
    for row in history_rows:
        if (row.get("date") or "") >= before_date:
            continue
        key = history_requirement_key(row)
        if not all(key[:3]):
            continue
        consumed_hours[(row.get("class_id") or "", key)] += int(float(row.get("duration_hours") or 0))

    if not consumed_hours:
        return list(assignments), 0

    result: List[scheduler.Assignment] = []
    removed = 0
    for assignment in scheduler.sorted_assignments(assignments):
        key = (assignment.task.class_id, scheduler.requirement_object_key(assignment.task))
        remaining_consumed = consumed_hours.get(key, 0)
        if remaining_consumed >= assignment.task.block_hours:
            consumed_hours[key] = remaining_consumed - assignment.task.block_hours
            removed += 1
            continue
        result.append(assignment)
    return result, removed


def build_halfyear_batch_assignments(
    data_dir: Path,
    protected_assignments: Sequence[scheduler.Assignment],
    history_rows: Sequence[dict],
    target_suite_codes: Optional[Set[str]] = None,
    fast_scope_locked_filter: bool = False,
) -> Tuple[List[scheduler.Assignment], List[str]]:
    class_metadata = load_class_metadata(data_dir)
    blackout_dates = load_active_blackout_dates(data_dir)
    result: List[scheduler.Assignment] = []
    lock_result: List[scheduler.Assignment] = []
    strategy_lines: List[str] = []

    def suite_start_key(suite_code: str) -> Tuple[str, str]:
        class_ids = class_ids_for_suite_codes(data_dir, [suite_code], None)
        source = load_schedule_input_for_classes(data_dir, class_ids)
        start_dates = [
            cls.start_date
            for class_id, cls in source.classes.items()
            if class_id in class_ids and cls.start_date
        ]
        return (max(start_dates) if start_dates else "9999-12-31", suite_code)

    priority_suites = [
        suite_code
        for suite_code in HALF_YEAR_BATCH_PRIORITY
        if suite_code in HALF_YEAR_BATCH_SUITES
    ]
    remaining_suites = [
        suite_code
        for suite_code in HALF_YEAR_BATCH_SUITES
        if suite_code not in set(priority_suites)
    ]
    ordered_suites = [*priority_suites, *sorted(remaining_suites, key=suite_start_key)]
    if target_suite_codes is not None:
        ordered_suites = [
            suite_code for suite_code in ordered_suites
            if suite_code in target_suite_codes
        ]

    for suite_code in ordered_suites:
        print(f"预排 半年营 {suite_code}", flush=True)
        class_ids = class_ids_for_suite_codes(data_dir, [suite_code], None)
        source = load_schedule_input_for_classes(data_dir, class_ids)
        suite_protected_assignments = (
            relevant_locked_assignments_for_replan(data_dir, class_ids, protected_assignments, "2026-07-01", AUTUMN_END)
            if fast_scope_locked_filter
            else protected_assignments
        )
        source = replace(
            source,
            locked_assignments=[*source.locked_assignments, *suite_protected_assignments, *lock_result],
        )
        batch_input = filtered_schedule_input(
            source,
            class_ids=class_ids,
            stages=SUMMER_STAGE_ORDER.keys(),
            subjects=None,
            start="2026-07-01",
            end=AUTUMN_END,
            start_period="AM",
            end_period="EVENING",
            periods={"AM", "PM", "EVENING"},
            room_ids=None,
        )
        batch_input = without_blackout_dates(batch_input, blackout_dates)
        def attempt_schedule(
            current_input: scheduler.ScheduleInput,
        ) -> Tuple[Optional[List[scheduler.Assignment]], str, List[str]]:
            attempts = (
                ("轮排兜底", lambda: schedule_round_robin(
                    current_input,
                    class_ids,
                    balance_public_subject_weeks=True,
                    avoid_public_subject_consecutive_days=True,
                    prefer_public_teacher_alternation=True,
                    teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
                    spacing_improvement_passes=FULL_SPACING_IMPROVEMENT_PASSES,
                    require_all_subject_weeks=False,
                    strict_subject_week_quotas=True,
                )),
                (
                    "均衡",
                    lambda: schedule_balanced_camp_by_suite(
                        current_input,
                        class_ids,
                        class_metadata,
                        lock_previous_assignments=True,
                        balance_public_subject_weeks=True,
                        require_all_subject_weeks=False,
                    ),
                ),
                ("基础均衡", lambda: schedule_balanced_camp(
                    current_input,
                    class_ids,
                    balance_public_subject_weeks=True,
                    avoid_public_subject_consecutive_days=True,
                    prefer_public_teacher_alternation=True,
                    teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
                    require_all_subject_weeks=False,
                )),
                ("消冲突兜底", lambda: schedule_round_robin(
                    current_input,
                    class_ids,
                    balance_public_subject_weeks=True,
                    avoid_public_subject_consecutive_days=True,
                    prefer_public_teacher_alternation=True,
                    teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
                    spacing_improvement_passes=0,
                    require_all_subject_weeks=False,
                )),
            )
            attempt_errors: List[str] = []
            for name, callback in attempts:
                try:
                    return callback(), name, attempt_errors
                except ValueError as exc:
                    attempt_errors.append(f"{name}: {exc}")
            return None, "", attempt_errors

        suite_assignments, used_strategy, errors = attempt_schedule(batch_input)
        monday_exception_used = False
        if suite_assignments is None:
            monday_input = with_monday_exception(batch_input)
            suite_assignments, used_strategy, monday_errors = attempt_schedule(monday_input)
            if suite_assignments is not None:
                monday_exception_used = True
            else:
                errors.extend([f"周一例外 {item}" for item in monday_errors])
        if suite_assignments is None:
            details = "；".join(errors[:3])
            strategy_lines.append(f"{suite_code}: 阻塞，未生成课表；{details}")
            continue
        lock_result.extend(suite_assignments)
        visible_assignments, removed_count = remove_history_replaced_assignments(
            suite_assignments,
            history_rows,
            "2026-07-01",
        )
        result.extend(visible_assignments)
        start_date = visible_assignments[0].candidate.slots[0].date if visible_assignments else ""
        end_date = visible_assignments[-1].candidate.slots[0].date if visible_assignments else ""
        strategy_lines.append(
            f"{suite_code}: {used_strategy}，按剩余课量周平均配额，{len(visible_assignments)} 节，{start_date} 至 {end_date}"
            + ("，启用周一例外" if monday_exception_used else "")
            + (f"，已由历史课表抵扣 {removed_count} 节" if removed_count else "")
        )
    return scheduler.sorted_assignments(result), strategy_lines


def suite_public_class_groups(
    class_metadata: Dict[str, Dict[str, str]],
    sub_product: str,
    excluded_class_ids: Set[str],
) -> List[Tuple[str, List[str]]]:
    grouped: Dict[str, List[str]] = defaultdict(list)
    for class_id, meta in class_metadata.items():
        if class_id in excluded_class_ids:
            continue
        if meta.get("subject_category") != "公共课":
            continue
        if meta.get("sub_product") != sub_product:
            continue
        if meta.get("is_schedule_locked") in {"是", "1", "true", "True", "yes", "Y", "y"}:
            continue
        suite_code = meta.get("suite_code") or class_id
        grouped[suite_code].append(class_id)

    return sorted(
        grouped.items(),
        key=lambda item: suite_planning_key(item[0], item[1], class_metadata),
    )


def suite_planning_key(
    suite_code: str,
    class_ids: Sequence[str],
    class_metadata: Dict[str, Dict[str, str]],
) -> Tuple[int, int, str, str]:
    has_math = any(class_metadata.get(class_id, {}).get("subject") == "数学" for class_id in class_ids)
    spans: List[int] = []
    start_dates: List[str] = []
    for class_id in class_ids:
        meta = class_metadata.get(class_id, {})
        start_date = meta.get("start_date") or "9999-12-31"
        end_date = meta.get("end_date") or "9999-12-31"
        start_dates.append(start_date)
        try:
            spans.append((Date.fromisoformat(end_date) - Date.fromisoformat(start_date)).days)
        except ValueError:
            spans.append(9999)
    return (
        0 if has_math else 1,
        min(spans) if spans else 9999,
        min(start_dates) if start_dates else "9999-12-31",
        suite_code,
    )


def ordered_suite_class_ids(
    suite_class_ids: Sequence[str],
    class_metadata: Dict[str, Dict[str, str]],
) -> List[str]:
    return sorted(
        suite_class_ids,
        key=lambda class_id: (
            PUBLIC_SUBJECT_PRIORITY.get(class_metadata.get(class_id, {}).get("subject"), 99),
            class_id,
        ),
    )


def with_preferred_class_rooms(
    schedule_input: scheduler.ScheduleInput,
    class_metadata: Dict[str, Dict[str, str]],
) -> scheduler.ScheduleInput:
    classes: Dict[str, scheduler.SchoolClass] = {}
    for class_id, cls in schedule_input.classes.items():
        preferred_room_ids = set(split_pipe_values(class_metadata.get(class_id, {}).get("preferred_room_ids")))
        if not preferred_room_ids:
            classes[class_id] = cls
            continue
        requirements = [
            replace(
                requirement,
                room_ids=(set(requirement.room_ids) & preferred_room_ids) or preferred_room_ids,
            )
            for requirement in cls.requirements
        ]
        classes[class_id] = replace(cls, room_ids=preferred_room_ids, requirements=requirements)
    return replace(schedule_input, classes=classes)


def min_date_text(*values: Optional[str]) -> Optional[str]:
    dates = [value for value in values if value]
    return min(dates) if dates else None


def max_date_text(*values: Optional[str]) -> Optional[str]:
    dates = [value for value in values if value]
    return max(dates) if dates else None


def wuyou_qc_block_hours(requirement: scheduler.Requirement) -> int:
    if requirement.total_hours >= 4 and requirement.total_hours % 4 == 0:
        return 4
    return requirement.block_hours


def wuyou_summer_week_bounds(subjects: Set[str]) -> Dict[str, Tuple[Optional[int], Optional[int]]]:
    return (
        WUYOU_SUMMER_WEEK_BOUNDS_WITH_MATH
        if "数学" in subjects
        else WUYOU_SUMMER_WEEK_BOUNDS_WITHOUT_MATH
    )


def wuyou_summer_weekly_total_max(subjects: Set[str]) -> int:
    return (
        WUYOU_SUMMER_TOTAL_MAX_WITH_MATH
        if "数学" in subjects
        else WUYOU_SUMMER_TOTAL_MAX_WITHOUT_MATH
    )


def max_only_week_bounds(
    week_bounds: Dict[str, Tuple[Optional[int], Optional[int]]],
) -> Dict[str, Tuple[Optional[int], Optional[int]]]:
    return {
        subject: (None, weekly_max)
        for subject, (_weekly_min, weekly_max) in week_bounds.items()
    }


def with_wuyou_qc_stage_windows(schedule_input: scheduler.ScheduleInput) -> scheduler.ScheduleInput:
    classes: Dict[str, scheduler.SchoolClass] = {}
    for class_id, cls in schedule_input.classes.items():
        suite_code = suite_code_from_class_id(class_id)
        foundation_end = WYQC_FOUNDATION_END_BY_SUITE.get(suite_code, WYQC_FOUNDATION_END)
        requirements: List[scheduler.Requirement] = []
        has_sprint = False
        for requirement in cls.requirements:
            stage = requirement.stage or ""
            if stage in {"基础", "强化"}:
                requirements.append(
                    replace(
                        requirement,
                        start_date=max_date_text(requirement.start_date, WYQC_SUMMER_START),
                        end_date=min_date_text(requirement.end_date, foundation_end),
                        allowed_periods={"AM", "PM"},
                        allowed_weekdays={0, 1, 2, 3, 4, 5},
                        block_hours=wuyou_qc_block_hours(requirement),
                    )
                )
            elif stage == "冲刺":
                has_sprint = True
                start_date = WYQC_SPRINT_START_BY_SUBJECT.get(requirement.subject, WYQC_AUTUMN_START)
                schedule_rules = (
                    scheduler.ScheduleRule(
                        subject=requirement.subject,
                        stage=requirement.stage,
                        course_module=requirement.course_module,
                        course_group=requirement.course_group,
                        start_date=start_date,
                        end_date=WYQC_AUTUMN_END,
                        allowed_periods={"AM", "PM"},
                        allowed_weekdays={5, 6},
                        excluded_weekdays=None,
                        block_hours=wuyou_qc_block_hours(requirement),
                    ),
                    scheduler.ScheduleRule(
                        subject=requirement.subject,
                        stage=requirement.stage,
                        course_module=requirement.course_module,
                        course_group=requirement.course_group,
                        start_date=min(WYQC_AUTUMN_MAKEUP_DATES),
                        end_date=max(WYQC_AUTUMN_MAKEUP_DATES),
                        allowed_periods={"AM", "PM"},
                        allowed_weekdays={1, 2},
                        excluded_weekdays=None,
                        block_hours=wuyou_qc_block_hours(requirement),
                    ),
                )
                requirements.append(
                    replace(
                        requirement,
                        start_date=max_date_text(requirement.start_date, start_date),
                        end_date=min_date_text(requirement.end_date, WYQC_AUTUMN_END) or WYQC_AUTUMN_END,
                        allowed_periods=None,
                        allowed_weekdays=None,
                        schedule_rules=schedule_rules,
                        block_hours=wuyou_qc_block_hours(requirement),
                    )
                )
            else:
                requirements.append(requirement)
        class_updates = {"requirements": requirements}
        if has_sprint:
            class_updates["end_date"] = min_date_text(cls.end_date, WYQC_AUTUMN_END) or WYQC_AUTUMN_END
        classes[class_id] = replace(cls, **class_updates)
    return replace(schedule_input, classes=classes)


def with_requirement_weekdays(
    schedule_input: scheduler.ScheduleInput,
    weekdays: Set[int],
) -> scheduler.ScheduleInput:
    classes: Dict[str, scheduler.SchoolClass] = {}
    for class_id, cls in schedule_input.classes.items():
        requirements = [
            replace(
                requirement,
                allowed_weekdays=set(weekdays),
                excluded_weekdays=None,
                schedule_rules=tuple(
                    replace(rule, allowed_weekdays=set(weekdays), excluded_weekdays=None)
                    for rule in requirement.schedule_rules
                ),
            )
            for requirement in cls.requirements
        ]
        classes[class_id] = replace(cls, requirements=requirements)
    return replace(schedule_input, classes=classes)


def wuyou_summer_autumn_slot_allowed(
    slot: scheduler.TimeSlot,
    include_wed_pm: bool,
    include_wed_am: bool,
) -> bool:
    if slot.date < WYS_AUTUMN_START or slot.date > AUTUMN_END:
        return True
    weekday = Date.fromisoformat(slot.date).weekday()
    if weekday in {5, 6} and slot.period in {"AM", "PM"}:
        return True
    if include_wed_pm and weekday == 2 and slot.period == "PM":
        return True
    if include_wed_am and weekday == 2 and slot.period == "AM":
        return True
    return False


def with_wuyou_summer_autumn_slots(
    schedule_input: scheduler.ScheduleInput,
    include_wed_pm: bool = True,
    include_wed_am: bool = False,
) -> scheduler.ScheduleInput:
    return replace(
        schedule_input,
        time_slots=[
            slot
            for slot in schedule_input.time_slots
            if wuyou_summer_autumn_slot_allowed(slot, include_wed_pm, include_wed_am)
        ],
    )


def with_wuyou_summer_stage_windows(schedule_input: scheduler.ScheduleInput) -> scheduler.ScheduleInput:
    classes: Dict[str, scheduler.SchoolClass] = {}
    for class_id, cls in schedule_input.classes.items():
        requirements: List[scheduler.Requirement] = []
        has_sprint = False
        for requirement in cls.requirements:
            stage = requirement.stage or ""
            if stage in {"基础", "强化"}:
                schedule_rules = (
                    scheduler.ScheduleRule(
                        subject=requirement.subject,
                        stage=requirement.stage,
                        course_module=requirement.course_module,
                        course_group=requirement.course_group,
                        start_date=WYS_SUMMER_START,
                        end_date=WYS_SUMMER_END,
                        allowed_periods={"AM", "PM"},
                        allowed_weekdays={0, 1, 2, 3, 4, 5},
                        excluded_weekdays=None,
                        block_hours=wuyou_qc_block_hours(requirement),
                    ),
                    scheduler.ScheduleRule(
                        subject=requirement.subject,
                        stage=requirement.stage,
                        course_module=requirement.course_module,
                        course_group=requirement.course_group,
                        start_date=WYS_AUTUMN_START,
                        end_date=WYS_FOUNDATION_END,
                        allowed_periods={"AM", "PM"},
                        allowed_weekdays={2, 5, 6},
                        excluded_weekdays=None,
                        block_hours=wuyou_qc_block_hours(requirement),
                    ),
                )
                requirements.append(
                    replace(
                        requirement,
                        start_date=max_date_text(requirement.start_date, cls.start_date, WYS_SUMMER_START),
                        end_date=max_date_text(requirement.end_date, WYS_FOUNDATION_END),
                        allowed_periods=None,
                        allowed_weekdays=None,
                        schedule_rules=schedule_rules,
                        block_hours=wuyou_qc_block_hours(requirement),
                    )
                )
            elif stage == "冲刺":
                has_sprint = True
                requirements.append(
                    replace(
                        requirement,
                        start_date=max_date_text(requirement.start_date, WYS_SPRINT_START),
                        end_date=max_date_text(requirement.end_date, AUTUMN_END),
                        allowed_periods={"AM", "PM"},
                        allowed_weekdays={2, 5, 6},
                        block_hours=wuyou_qc_block_hours(requirement),
                    )
                )
            else:
                requirements.append(requirement)
        class_updates = {"requirements": requirements}
        if has_sprint:
            class_updates["end_date"] = max_date_text(cls.end_date, AUTUMN_END)
        classes[class_id] = replace(cls, **class_updates)
    return replace(schedule_input, classes=classes)


def with_stage_subset(
    schedule_input: scheduler.ScheduleInput,
    stages: Set[str],
) -> scheduler.ScheduleInput:
    classes: Dict[str, scheduler.SchoolClass] = {}
    for class_id, cls in schedule_input.classes.items():
        requirements = [
            requirement
            for requirement in cls.requirements
            if (requirement.stage or "") in stages
        ]
        if requirements:
            classes[class_id] = replace(cls, requirements=requirements)
    return replace(schedule_input, classes=classes)


def with_subject_subset(
    schedule_input: scheduler.ScheduleInput,
    subjects: Set[str],
) -> scheduler.ScheduleInput:
    classes: Dict[str, scheduler.SchoolClass] = {}
    for class_id, cls in schedule_input.classes.items():
        requirements = [
            requirement
            for requirement in cls.requirements
            if requirement.subject in subjects
        ]
        if requirements:
            classes[class_id] = replace(cls, requirements=requirements)
    return replace(schedule_input, classes=classes)


def rebalance_public_teacher_sequences(
    schedule_input: scheduler.ScheduleInput,
    class_ids: Sequence[str],
    assignments: Sequence[scheduler.Assignment],
) -> List[scheduler.Assignment]:
    public_tasks = [
        task
        for task in scheduler.build_course_blocks(schedule_input.classes)
        if task.class_id in set(class_ids) and task.subject in SUMMER_PUBLIC_SUBJECTS
    ]
    if not public_tasks:
        return scheduler.sorted_assignments(list(assignments))
    domains = scheduler.candidate_domains(public_tasks, schedule_input)
    return rebalance_subject_teacher_alternation(
        schedule_input,
        assignments,
        domains,
        SUMMER_PUBLIC_SUBJECTS,
    )


def schedule_autumn_weekly_four_hours_by_subject(
    schedule_input: scheduler.ScheduleInput,
    class_ids: Sequence[str],
    subject_order: Sequence[str] = ("数学", "政治", "英语"),
) -> List[scheduler.Assignment]:
    assignments: List[scheduler.Assignment] = []
    for subject in subject_order:
        subject_input = with_subject_subset(
            replace(
                schedule_input,
                locked_assignments=[*schedule_input.locked_assignments, *assignments],
            ),
            {subject},
        )
        subject_class_ids = [
            class_id
            for class_id in class_ids
            if class_id in subject_input.classes
        ]
        if not subject_class_ids:
            continue
        scheduled: Optional[List[scheduler.Assignment]] = None
        errors: List[str] = []
        for label, week_bounds, require_all_weeks, strict_quotas, max_consecutive_days in (
            ("每周1个半天", {subject: (1, 1)}, True, True, 2),
            ("每周最多1个半天软兜底", {subject: (None, 1)}, False, True, 2),
            ("每周最多2个半天软兜底", {subject: (None, 2)}, False, False, 2),
            ("日期窗口内软兜底", None, False, False, 2),
            ("日期窗口内同科连续兜底", None, False, False, None),
        ):
            try:
                scheduled = schedule_round_robin(
                    subject_input,
                    subject_class_ids,
                    subject_week_bounds=week_bounds,
                    avoid_public_subject_consecutive_days=False,
                    prefer_public_teacher_alternation=True,
                    subject_max_consecutive_days=max_consecutive_days,
                    allow_same_subject_day_fallback=max_consecutive_days is None,
                    teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
                    spacing_improvement_passes=0,
                    require_all_subject_weeks=require_all_weeks,
                    strict_subject_week_quotas=strict_quotas,
                )
                break
            except ValueError as exc:
                errors.append(f"{label}: {exc}")
        if scheduled is None:
            raise ValueError("；".join(errors[:3]))
        assignments.extend(scheduled)
    return scheduler.sorted_assignments(assignments)


def fast_additional_phase() -> str:
    return clean(os.environ.get("FAST_ADDITIONAL_PHASE")).lower()


def build_wyqc_autumn_assignments(
    data_dir: Path,
    protected_assignments: Sequence[scheduler.Assignment],
    history_rows: Sequence[dict],
    target_suite_codes: Set[str],
    target_sub_products: Set[str],
    allow_existing_public_adjustment: bool = False,
    fast_scope_locked_filter: bool = True,
) -> Tuple[List[scheduler.Assignment], List[str]]:
    class_metadata = load_class_metadata(data_dir)
    blackout_dates = load_active_blackout_dates(data_dir)
    result: List[scheduler.Assignment] = []
    strategy_lines: List[str] = []
    target_products = target_sub_products & WYQC_PRODUCTS
    if not target_products:
        return [], []

    for sub_product in ADDITIONAL_PUBLIC_PRODUCTS:
        if sub_product not in target_products:
            continue
        groups = suite_public_class_groups(class_metadata, sub_product, set())
        if target_suite_codes:
            groups = [
                (suite_code, suite_class_ids)
                for suite_code, suite_class_ids in groups
                if suite_code in target_suite_codes
            ]
        groups = sorted(
            groups,
            key=lambda item: (
                len({
                    class_metadata.get(class_id, {}).get("subject", "")
                    for class_id in item[1]
                    if class_metadata.get(class_id, {}).get("subject")
                }),
                "数学"
                in {
                    class_metadata.get(class_id, {}).get("subject", "")
                    for class_id in item[1]
                },
                WYQC_SUITE_PRIORITY.get(item[0], 100),
                suite_planning_key(item[0], item[1], class_metadata),
            ),
        )
        for suite_code, suite_class_ids in groups:
            class_ids = ordered_suite_class_ids(suite_class_ids, class_metadata)
            print(f"快速重排 {sub_product} {suite_code} 秋季: {', '.join(class_ids)}", flush=True)
            scheduled: Optional[List[scheduler.Assignment]] = None
            scheduled_input: Optional[scheduler.ScheduleInput] = None
            errors: List[str] = []
            used_label = ""
            protect_modes = (True, False) if allow_existing_public_adjustment else (True,)
            for protect_existing_public in protect_modes:
                locked_source = [*protected_assignments, *result]
                if fast_scope_locked_filter or not protect_existing_public:
                    locked_public = relevant_locked_assignments_for_replan(
                        data_dir,
                        class_ids,
                        locked_source,
                        WYQC_AUTUMN_START,
                        WYQC_AUTUMN_END,
                        include_teacher=True,
                        include_room=True,
                    )
                else:
                    locked_public = locked_source
                for use_preferred_rooms in (True, False):
                    source = load_schedule_input_for_classes(data_dir, class_ids)
                    source = apply_history_deductions(source, history_rows, "2026-07-01")
                    source = with_wuyou_qc_stage_windows(source)
                    source = filtered_schedule_input(
                        source,
                        class_ids,
                        {"冲刺"},
                        None,
                        WYQC_AUTUMN_START,
                        WYQC_AUTUMN_END,
                        "",
                        "",
                        {"AM", "PM"},
                        None,
                    )
                    if use_preferred_rooms:
                        source = with_preferred_class_rooms(source, class_metadata)
                    source = replace(
                        source,
                        locked_assignments=[*source.locked_assignments, *locked_public],
                    )
                    source = with_conflict_groups_for_locked(data_dir, source, source.locked_assignments)
                    source = without_blackout_dates(source, blackout_dates)
                    source = without_dates(source, WUYOU_PRODUCT_BLACKOUT_DATES)
                    label = (
                        ("锁定既有公共课" if protect_existing_public else "允许调整既有公共课")
                        + "+"
                        + ("优先教室" if use_preferred_rooms else "同校区教室")
                    )
                    try:
                        scheduled = run_fast_callback_with_timeout(
                            lambda current=source: schedule_autumn_weekly_four_hours_by_subject(
                                current,
                                class_ids,
                            ),
                            45,
                        )
                        scheduled_input = source
                        used_label = label
                        break
                    except ValueError as exc:
                        errors.append(f"{label}: {exc}")
                if scheduled is not None:
                    break
            if scheduled is None:
                raise ValueError(
                    f"{sub_product} {suite_code} 秋季重排失败: {'；'.join(errors[:6])}"
                )
            if scheduled_input is None:
                raise ValueError(f"{sub_product} {suite_code} 秋季重排失败: 未记录排课输入")
            result.extend(scheduled)
            start_date = scheduled[0].candidate.slots[0].date if scheduled else ""
            end_date = scheduled[-1].candidate.slots[0].date if scheduled else ""
            subject_counts: Dict[str, int] = defaultdict(int)
            for assignment in scheduled:
                subject_counts[assignment.task.subject] += 1
            subject_summary = "，".join(
                f"{subject}{subject_counts[subject]}个半天"
                for subject in sorted(subject_counts, key=lambda item: PUBLIC_SUBJECT_PRIORITY.get(item, 99))
            )
            strategy_lines.append(
                f"{sub_product} {suite_code} 秋季: {used_label}，{len(scheduled)} 节，"
                f"{start_date} 至 {end_date}，{subject_summary}，按每科周均衡分布"
            )
    return scheduler.sorted_assignments(result), strategy_lines


def conflict_groups_for_classes(data_dir: Path, class_ids: Set[str]) -> Tuple[Dict[str, Set[str]], Dict[str, Set[str]]]:
    input_path = data_dir / "scheduler_input_draft.json"
    data = json.loads(input_path.read_text(encoding="utf-8"))
    conflict_groups: Dict[str, Set[str]] = {}
    class_conflict_groups: Dict[str, Set[str]] = defaultdict(set)
    for index, group in enumerate(data.get("conflict_groups", []), start=1):
        group_id = clean(group.get("id")) or f"CONFLICT_{index}"
        group_class_ids = {clean(class_id) for class_id in group.get("class_ids", []) if clean(class_id)}
        selected = group_class_ids & class_ids
        if len(selected) < 2:
            continue
        conflict_groups[group_id] = selected
        for class_id in selected:
            class_conflict_groups[class_id].add(group_id)
    for class_id in class_ids:
        class_conflict_groups.setdefault(class_id, set())
    return conflict_groups, dict(class_conflict_groups)


def with_conflict_groups_for_locked(
    data_dir: Path,
    schedule_input: scheduler.ScheduleInput,
    locked_assignments: Sequence[scheduler.Assignment],
) -> scheduler.ScheduleInput:
    known_class_ids = set(schedule_input.classes) | {
        assignment.task.class_id for assignment in locked_assignments
    }
    conflict_groups, class_conflict_groups = conflict_groups_for_classes(data_dir, known_class_ids)
    return replace(
        schedule_input,
        conflict_groups=conflict_groups,
        class_conflict_groups=class_conflict_groups,
    )


def conflict_related_locked_assignments(
    data_dir: Path,
    class_ids: Sequence[str],
    locked_assignments: Sequence[scheduler.Assignment],
) -> List[scheduler.Assignment]:
    data = load_scheduler_input_data(data_dir)
    target_ids = set(class_ids)
    related_ids: Set[str] = set()
    for group in data.get("conflict_groups", []):
        group_class_ids = {clean(class_id) for class_id in group.get("class_ids", []) if clean(class_id)}
        if group_class_ids & target_ids:
            related_ids.update(group_class_ids)
    return [
        assignment
        for assignment in locked_assignments
        if assignment.task.class_id in related_ids
    ]


def relevant_locked_assignments_for_replan(
    data_dir: Path,
    class_ids: Sequence[str],
    locked_assignments: Sequence[scheduler.Assignment],
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    include_teacher: bool = True,
    include_room: bool = True,
) -> List[scheduler.Assignment]:
    if not locked_assignments:
        return []
    source = load_schedule_input_for_classes(data_dir, class_ids)
    target_teacher_keys: Set[str] = set()
    target_room_ids: Set[str] = set()
    for cls in source.classes.values():
        if cls.room_ids:
            target_room_ids.update(cls.room_ids)
        for requirement in cls.requirements:
            if requirement.teacher_id:
                target_teacher_keys.add(requirement.teacher_id)
            if requirement.teacher_name:
                target_teacher_keys.add(requirement.teacher_name)
            if requirement.room_ids:
                target_room_ids.update(requirement.room_ids)
    conflict_related = conflict_related_locked_assignments(data_dir, class_ids, locked_assignments)
    conflict_task_ids = {assignment.task.class_id for assignment in conflict_related}

    result: List[scheduler.Assignment] = []
    seen: Set[Tuple[str, str, str, str, str]] = set()
    for assignment in locked_assignments:
        first = assignment.candidate.slots[0]
        if start_date and first.date < start_date:
            continue
        if end_date and first.date > end_date:
            continue
        teacher_key = candidate_teacher_key(assignment.candidate)
        related = (
            assignment.task.class_id in conflict_task_ids
            or (include_teacher and teacher_key and teacher_key in target_teacher_keys)
            or (include_room and assignment.candidate.room_id and assignment.candidate.room_id in target_room_ids)
        )
        if not related:
            continue
        key = (
            assignment.task.class_id,
            first.date,
            first.period,
            teacher_key,
            assignment.candidate.room_id,
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(assignment)
    return scheduler.sorted_assignments(result)


def schedule_public_suite_attempt(
    schedule_input: scheduler.ScheduleInput,
    class_ids: Sequence[str],
    sub_product: str,
    allow_basic_fallback: bool = True,
) -> List[scheduler.Assignment]:
    if sub_product in WYQC_PRODUCTS:
        return schedule_wuyou_qc_suite_attempt(schedule_input, class_ids, allow_basic_fallback)
    if sub_product == WYS_PRODUCT:
        return schedule_wuyou_summer_suite_attempt(schedule_input, class_ids, allow_basic_fallback)
    if sub_product == SPRINT_CAMP_PRODUCT:
        return schedule_sprint_camp_suite_attempt(schedule_input, class_ids)
    is_long_camp = sub_product in {"全年营", "半年营"}
    is_public_product = sub_product in PUBLIC_PRODUCT_ORDER
    round_robin_attempts = (True, False) if is_long_camp else (False,)
    errors: List[str] = []
    for strict_average in round_robin_attempts:
        try:
            return schedule_round_robin(
                schedule_input,
                class_ids,
                balance_public_subject_weeks=is_long_camp,
                avoid_public_subject_consecutive_days=is_public_product,
                prefer_public_teacher_alternation=is_public_product,
                teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS if is_public_product else None,
                spacing_improvement_passes=0 if sub_product in WUYOU_PRODUCTS and not is_long_camp else FULL_SPACING_IMPROVEMENT_PASSES,
                require_all_subject_weeks=is_long_camp,
                strict_subject_week_quotas=strict_average,
            )
        except ValueError as exc:
            errors.append(("严格平均" if strict_average else "软平均") + f": {exc}")
    if errors:
        if is_long_camp or not allow_basic_fallback:
            raise ValueError("；".join(errors[:2]))
    try:
        return schedule_round_robin(
            schedule_input,
            class_ids,
            avoid_public_subject_consecutive_days=is_public_product,
            prefer_public_teacher_alternation=is_public_product,
            teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS if is_public_product else None,
            spacing_improvement_passes=0 if sub_product in WUYOU_PRODUCTS and not is_long_camp else FULL_SPACING_IMPROVEMENT_PASSES,
        )
    except ValueError:
        if not allow_basic_fallback:
            raise
        return [
            assignment
            for assignment in scheduler.schedule(schedule_input)
            if assignment.task.class_id in schedule_input.classes
        ]


def schedule_public_suite_core_fallback(
    schedule_input: scheduler.ScheduleInput,
    class_ids: Sequence[str],
) -> List[scheduler.Assignment]:
    assignments = scheduler.schedule(schedule_input)
    selected_class_ids = set(class_ids)
    return scheduler.sorted_assignments([
        assignment for assignment in assignments if assignment.task.class_id in selected_class_ids
    ])


def schedule_public_suite_conflict_first(
    schedule_input: scheduler.ScheduleInput,
    class_ids: Sequence[str],
) -> List[scheduler.Assignment]:
    return schedule_round_robin(
        schedule_input,
        class_ids,
        avoid_public_subject_consecutive_days=True,
        prefer_public_teacher_alternation=True,
        teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
        spacing_improvement_passes=0,
        require_all_subject_weeks=False,
    )


def schedule_wuyou_summer_suite_attempt(
    schedule_input: scheduler.ScheduleInput,
    class_ids: Sequence[str],
    allow_basic_fallback: bool,
) -> List[scheduler.Assignment]:
    windowed_input = with_wuyou_summer_stage_windows(schedule_input)
    foundation_source = with_stage_subset(windowed_input, {"基础", "强化"})
    foundation_input_attempts = (
        ("周末", with_wuyou_summer_autumn_slots(foundation_source, include_wed_pm=False, include_wed_am=False)),
        ("周末+周三下午", with_wuyou_summer_autumn_slots(foundation_source, include_wed_pm=True, include_wed_am=False)),
        ("周末+周三下午+周三上午", with_wuyou_summer_autumn_slots(foundation_source, include_wed_pm=True, include_wed_am=True)),
    )
    foundation_assignments: List[scheduler.Assignment] = []
    foundation_class_ids = [class_id for class_id in class_ids if class_id in foundation_source.classes]
    suite_subjects = {
        requirement.subject
        for cls in foundation_source.classes.values()
        for requirement in cls.requirements
    }
    foundation_week_bounds = (
        WYS_SUMMER_WEEK_BOUNDS
        if "数学" in suite_subjects
        else WYS_SUMMER_WEEK_BOUNDS_WITHOUT_MATH
    )
    foundation_weekly_total_max = (
        WYS_SUMMER_TOTAL_MAX_WITH_MATH
        if "数学" in suite_subjects
        else WYS_SUMMER_TOTAL_MAX_WITHOUT_MATH
    )
    last_error: Optional[ValueError] = None
    if foundation_class_ids:
        for _label, foundation_input in foundation_input_attempts:
            try:
                foundation_assignments = schedule_round_robin(
                    foundation_input,
                    foundation_class_ids,
                    subject_week_bounds=foundation_week_bounds,
                    preferred_weekly_total_max=foundation_weekly_total_max,
                    hard_weekly_total_max=foundation_weekly_total_max,
                    avoid_public_subject_consecutive_days=True,
                    prefer_public_teacher_alternation=True,
                    subject_max_consecutive_days=2,
                    allow_same_subject_day_fallback=False,
                    teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
                    spacing_improvement_passes=0,
                    require_all_subject_weeks=True,
                    strict_subject_week_quotas=True,
                )
                break
            except ValueError as exc:
                last_error = exc
        if not foundation_assignments and not allow_basic_fallback and last_error:
            raise last_error

    if last_error and allow_basic_fallback and not foundation_assignments:
        fallback_errors: List[ValueError] = [last_error]
        foundation_attempts: List[Tuple[Optional[Dict[str, Tuple[Optional[int], Optional[int]]]], Optional[int], bool]] = [
            (foundation_week_bounds, None, True),
            (max_only_week_bounds(foundation_week_bounds), 2, True),
            (max_only_week_bounds(foundation_week_bounds), None, False),
        ]
        for _label, foundation_input in foundation_input_attempts:
            for week_bounds, max_consecutive_days, avoid_same_subject in foundation_attempts:
                try:
                    foundation_assignments = schedule_round_robin(
                        foundation_input,
                        foundation_class_ids,
                        subject_week_bounds=week_bounds,
                        preferred_weekly_total_max=foundation_weekly_total_max,
                        hard_weekly_total_max=foundation_weekly_total_max,
                        avoid_public_subject_consecutive_days=avoid_same_subject,
                        prefer_public_teacher_alternation=avoid_same_subject,
                        subject_max_consecutive_days=max_consecutive_days,
                        allow_same_subject_day_fallback=not avoid_same_subject,
                        teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
                        spacing_improvement_passes=0,
                        require_all_subject_weeks=bool(week_bounds),
                        strict_subject_week_quotas=False,
                    )
                    break
                except ValueError as exc:
                    fallback_errors.append(exc)
            if foundation_assignments:
                break
        if not foundation_assignments:
            raise fallback_errors[-1]

    sprint_input = with_stage_subset(windowed_input, {"冲刺"})
    sprint_class_ids = [class_id for class_id in class_ids if class_id in sprint_input.classes]
    if not sprint_class_ids:
        return rebalance_public_teacher_sequences(
            schedule_input,
            class_ids,
            scheduler.sorted_assignments(foundation_assignments),
        )

    sprint_input = replace(
        sprint_input,
        locked_assignments=[*sprint_input.locked_assignments, *foundation_assignments],
    )
    sprint_errors: List[str] = []
    sprint_assignments: Optional[List[scheduler.Assignment]] = None
    sprint_subjects = {
        requirement.subject
        for cls in sprint_input.classes.values()
        for requirement in cls.requirements
    }
    if "数学" in sprint_subjects:
        for label, candidate_input, strict_quotas, allow_same_subject_day in (
            (
                "秋季冲刺每周5半天",
                with_wuyou_summer_autumn_slots(sprint_input, include_wed_pm=True, include_wed_am=False),
                True,
                False,
            ),
            (
                "秋季冲刺每周5半天+周三上午",
                with_wuyou_summer_autumn_slots(sprint_input, include_wed_pm=True, include_wed_am=True),
                True,
                False,
            ),
            (
                "秋季冲刺每周5半天软配额",
                with_wuyou_summer_autumn_slots(sprint_input, include_wed_pm=True, include_wed_am=True),
                False,
                False,
            ),
            (
                "秋季冲刺每周5半天同科同日兜底",
                with_wuyou_summer_autumn_slots(sprint_input, include_wed_pm=True, include_wed_am=True),
                False,
                True,
            ),
        ):
            try:
                sprint_assignments = schedule_round_robin(
                    candidate_input,
                    sprint_class_ids,
                    subject_week_bounds=WYS_SPRINT_WEEK_BOUNDS_WITH_MATH,
                    preferred_weekly_total_max=WYS_SPRINT_WEEKLY_TOTAL_WITH_MATH,
                    hard_weekly_total_max=WYS_SPRINT_WEEKLY_TOTAL_WITH_MATH,
                    avoid_public_subject_consecutive_days=True,
                    prefer_public_teacher_alternation=True,
                    subject_max_consecutive_days=2,
                    allow_same_subject_day_fallback=allow_same_subject_day,
                    teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
                    spacing_improvement_passes=80,
                    require_all_subject_weeks=True,
                    strict_subject_week_quotas=strict_quotas,
                )
                break
            except ValueError as exc:
                sprint_errors.append(f"{label}: {exc}")
    if sprint_assignments is not None:
        return rebalance_public_teacher_sequences(
            schedule_input,
            class_ids,
            scheduler.sorted_assignments([*foundation_assignments, *sprint_assignments]),
        )
    for label, candidate_input in (
        ("周末", with_wuyou_summer_autumn_slots(sprint_input, include_wed_pm=False, include_wed_am=False)),
        ("周末+周三下午", with_wuyou_summer_autumn_slots(sprint_input, include_wed_pm=True, include_wed_am=False)),
        ("周末+周三下午+周三上午", with_wuyou_summer_autumn_slots(sprint_input, include_wed_pm=True, include_wed_am=True)),
        ("完整秋季窗口", sprint_input),
    ):
        try:
            sprint_assignments = schedule_round_robin(
                candidate_input,
                sprint_class_ids,
                avoid_public_subject_consecutive_days=True,
                prefer_public_teacher_alternation=True,
                subject_max_consecutive_days=2,
                allow_same_subject_day_fallback=False,
                teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
                spacing_improvement_passes=80,
            )
            break
        except ValueError as exc:
            sprint_errors.append(f"{label}: {exc}")
    if sprint_assignments is None:
        for label, candidate_input in (
            ("周末同科目同日兜底", with_wuyou_summer_autumn_slots(sprint_input, include_wed_pm=False, include_wed_am=False)),
            ("周末+周三下午同科目同日兜底", with_wuyou_summer_autumn_slots(sprint_input, include_wed_pm=True, include_wed_am=False)),
            ("周末+周三下午+周三上午同科目同日兜底", with_wuyou_summer_autumn_slots(sprint_input, include_wed_pm=True, include_wed_am=True)),
            ("完整秋季窗口同科目同日兜底", sprint_input),
        ):
            try:
                sprint_assignments = schedule_round_robin(
                    candidate_input,
                    sprint_class_ids,
                    avoid_public_subject_consecutive_days=True,
                    prefer_public_teacher_alternation=True,
                    subject_max_consecutive_days=2,
                    allow_same_subject_day_fallback=True,
                    teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
                    spacing_improvement_passes=80,
                )
                break
            except ValueError as exc:
                sprint_errors.append(f"{label}: {exc}")
    if sprint_assignments is None:
        raise ValueError("无忧暑秋季冲刺阶段预排失败: " + "；".join(sprint_errors[:3]))
    return rebalance_public_teacher_sequences(
        schedule_input,
        class_ids,
        scheduler.sorted_assignments([*foundation_assignments, *sprint_assignments]),
    )


def schedule_wuyou_qc_suite_attempt(
    schedule_input: scheduler.ScheduleInput,
    class_ids: Sequence[str],
    allow_basic_fallback: bool,
) -> List[scheduler.Assignment]:
    foundation_input = with_stage_subset(schedule_input, {"基础", "强化"})
    foundation_assignments: List[scheduler.Assignment] = []
    foundation_class_ids = [class_id for class_id in class_ids if class_id in foundation_input.classes]
    suite_subjects = {
        requirement.subject
        for cls in foundation_input.classes.values()
        for requirement in cls.requirements
    }
    # 无忧秋/无忧春会先扣除历史课时；扣除后部分科目剩余课时可能
    # 少于“每周 3 个半天”的下限，因此这里把周上限和总量作为硬约束，
    # 周下限留给报告核对，避免再次用基础调度把课程堆到局部几周。
    foundation_week_bounds = max_only_week_bounds(wuyou_summer_week_bounds(suite_subjects))
    foundation_weekly_total_max = wuyou_summer_weekly_total_max(suite_subjects)
    if foundation_class_ids:
        try:
            foundation_assignments = schedule_balanced_camp(
                foundation_input,
                foundation_class_ids,
                subject_week_bounds=foundation_week_bounds,
                preferred_weekly_total_max=foundation_weekly_total_max,
                avoid_public_subject_consecutive_days=True,
                prefer_public_teacher_alternation=True,
                teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
                require_all_subject_weeks=True,
            )
        except ValueError as exc:
            if not allow_basic_fallback:
                raise
            last_error: Optional[ValueError] = exc
        else:
            last_error = None

        foundation_attempts: List[Tuple[
            Optional[Dict[str, Tuple[Optional[int], Optional[int]]]],
            Optional[int],
            bool,
            bool,
            Optional[int],
        ]] = [
            (foundation_week_bounds, 2, True, True, foundation_weekly_total_max),
        ]
        if allow_basic_fallback:
            foundation_attempts.extend([
                # 无忧秋/无忧春会先扣历史课表，剩余课时常常不是整齐的周配额。
                # 第一层仍保留每周上限，后两层只保留硬冲突/阶段顺序/同科连续限制，
                # 避免因为“必须每周都有每科”而整套班完全卡死。
                (foundation_week_bounds, None, False, False, foundation_weekly_total_max),
                (foundation_week_bounds, 2, False, False, None),
                (None, 2, False, False, None),
                (None, None, False, False, None),
            ])
        if not foundation_assignments:
            for (
                week_bounds,
                max_consecutive_days,
                require_all_weeks,
                strict_quotas,
                hard_total_max,
            ) in foundation_attempts:
                try:
                    foundation_assignments = schedule_round_robin(
                        foundation_input,
                        foundation_class_ids,
                        subject_week_bounds=week_bounds,
                        preferred_weekly_total_max=foundation_weekly_total_max,
                        hard_weekly_total_max=hard_total_max,
                        avoid_public_subject_consecutive_days=True,
                        prefer_public_teacher_alternation=True,
                        subject_max_consecutive_days=max_consecutive_days,
                        allow_same_subject_day_fallback=max_consecutive_days is None,
                        teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
                        spacing_improvement_passes=80,
                        require_all_subject_weeks=require_all_weeks,
                        strict_subject_week_quotas=strict_quotas,
                    )
                    break
                except ValueError as exc:
                    last_error = exc
        if not foundation_assignments and last_error:
            raise last_error

    sprint_input = with_stage_subset(schedule_input, {"冲刺"})
    sprint_class_ids = [class_id for class_id in class_ids if class_id in sprint_input.classes]
    if not sprint_class_ids:
        return scheduler.sorted_assignments(foundation_assignments)

    sprint_input = replace(
        sprint_input,
        locked_assignments=[*sprint_input.locked_assignments, *foundation_assignments],
    )
    sprint_assignments = schedule_autumn_weekly_four_hours_by_subject(
        sprint_input,
        sprint_class_ids,
    )
    return scheduler.sorted_assignments([*foundation_assignments, *sprint_assignments])


def schedule_sprint_camp_suite_attempt(
    schedule_input: scheduler.ScheduleInput,
    class_ids: Sequence[str],
) -> List[scheduler.Assignment]:
    errors: List[str] = []
    average_week_bounds = max_only_week_bounds(SPRINT_CAMP_WEEK_BOUNDS)
    for label, candidate_input in (
        ("周二到周六", with_requirement_weekdays(schedule_input, {1, 2, 3, 4, 5})),
        ("周一到周六", with_requirement_weekdays(schedule_input, {0, 1, 2, 3, 4, 5})),
    ):
        for strict_average in (True, False):
            try:
                return schedule_round_robin(
                    candidate_input,
                    class_ids,
                    subject_week_bounds=average_week_bounds,
                    avoid_public_subject_consecutive_days=True,
                    prefer_public_teacher_alternation=True,
                    subject_max_consecutive_days=2,
                    teacher_alternation_subjects=SUMMER_PUBLIC_SUBJECTS,
                    spacing_improvement_passes=80,
                    strict_subject_week_quotas=strict_average,
                )
            except ValueError as exc:
                errors.append(f"{label}+{'严格平均' if strict_average else '软平均'}: {exc}")
    raise ValueError("冲刺营预排失败: " + "；".join(errors[:2]))


def candidate_on_makeup_date(
    schedule_input: scheduler.ScheduleInput,
    assignment: scheduler.Assignment,
    date_text: str,
    occupied_assignments: Sequence[scheduler.Assignment],
) -> Optional[scheduler.Candidate]:
    original_period = assignment.candidate.slots[0].period
    relaxed_task = replace(
        assignment.task,
        start_date=date_text,
        end_date=date_text,
        allowed_periods=None,
        allowed_weekdays=None,
        excluded_weekdays=None,
        schedule_rules=(),
    )
    candidate_input = replace(
        schedule_input,
        locked_assignments=[*schedule_input.locked_assignments, *occupied_assignments],
    )
    class_slot_used, teacher_slot_used, room_slot_used, conflict_group_slot_used = scheduler.locked_constraint_sets(candidate_input)
    candidates = [
        candidate
        for candidate in scheduler.candidate_assignments(relaxed_task, candidate_input)
        if candidate.slots[0].date == date_text
        and candidate_avoids_original_shape(assignment.candidate, candidate)
        and candidate_is_valid(
            candidate_input,
            class_slot_used,
            teacher_slot_used,
            room_slot_used,
            conflict_group_slot_used,
            relaxed_task,
            candidate,
        )
    ]
    candidates.sort(
        key=lambda candidate: (
            candidate.slots[0].period != original_period,
            scheduler.period_sort_value(candidate.slots[0].period),
            candidate.room_id != assignment.candidate.room_id,
            candidate.room_id,
        )
    )
    return candidates[0] if candidates else None


def candidate_avoids_original_shape(
    original: scheduler.Candidate,
    candidate: scheduler.Candidate,
) -> bool:
    return len(original.slots) == len(candidate.slots)


def shift_wuyou_makeup_day_assignments(
    schedule_input: scheduler.ScheduleInput,
    assignments: Sequence[scheduler.Assignment],
    sub_product: str,
) -> List[scheduler.Assignment]:
    if sub_product not in {"无忧秋", "无忧春", "无忧暑"}:
        return list(assignments)
    result = scheduler.sorted_assignments(list(assignments))
    task_ids_to_shift = [
        assignment.task.task_id
        for assignment in result
        if assignment.candidate.slots[0].date == "2026-10-10"
    ]
    for task_id in task_ids_to_shift:
        index = next(
            item_index
            for item_index, item in enumerate(result)
            if item.task.task_id == task_id
        )
        assignment = result[index]
        if assignment.candidate.slots[0].date != "2026-10-10":
            continue
        occupied = [item for item_index, item in enumerate(result) if item_index != index]
        replacement: Optional[scheduler.Candidate] = None
        for target_date in ("2026-10-07", "2026-10-06"):
            replacement = candidate_on_makeup_date(schedule_input, assignment, target_date, occupied)
            if replacement:
                break
        if not replacement:
            raise ValueError(
                f"{sub_product} {suite_code_from_class_id(assignment.task.class_id)} "
                f"{assignment.task.class_id} {assignment.task.subject} 10/10 调班无法挪到 10/7 或 10/6"
            )
        result[index] = scheduler.Assignment(
            task=assignment.task,
            candidate=replacement,
        )
        result = scheduler.sorted_assignments(result)
    return result


def load_base_time_slots(data_dir: Path) -> List[scheduler.TimeSlot]:
    data = json.loads((data_dir / "scheduler_input_draft.json").read_text(encoding="utf-8"))
    return [
        scheduler.TimeSlot(
            id=slot["id"],
            date=slot["date"],
            period=slot["period"],
            name=slot.get("name") or slot["period"],
            order=int(slot.get("order") or 1),
            start_time=slot.get("start_time"),
            end_time=slot.get("end_time"),
            duration_hours=int(slot.get("duration_hours") or 2),
        )
        for slot in data.get("time_slots", [])
        if slot.get("id") and slot.get("date") and slot.get("period")
    ]


def standard_period_slots(date_text: str, period: str) -> Tuple[scheduler.TimeSlot, ...]:
    return tuple(
        scheduler.TimeSlot(
            id=f"{date_text}-{period}-{order}",
            date=date_text,
            period=period,
            name=name,
            order=order,
            start_time=start_time,
            end_time=end_time,
            duration_hours=duration,
        )
        for order, name, start_time, end_time, duration in STANDARD_SLOT_SPECS.get(period, ())
    )


def time_slots_with_standard_days(
    time_slots: Sequence[scheduler.TimeSlot],
    start_date: str,
    end_date: str,
    periods: Set[str],
    weekdays: Optional[Set[int]] = None,
    excluded_dates: Optional[Set[str]] = None,
) -> List[scheduler.TimeSlot]:
    result = list(time_slots)
    seen_ids = {slot.id for slot in result}
    excluded_dates = excluded_dates or set()
    current = Date.fromisoformat(start_date)
    end = Date.fromisoformat(end_date)
    while current <= end:
        date_text = current.isoformat()
        if date_text in excluded_dates or (weekdays is not None and current.weekday() not in weekdays):
            current += timedelta(days=1)
            continue
        for period in periods:
            for slot in standard_period_slots(date_text, period):
                if slot.id in seen_ids:
                    continue
                seen_ids.add(slot.id)
                result.append(slot)
        current += timedelta(days=1)
    return sorted(
        result,
        key=lambda slot: (
            slot.date,
            scheduler.period_sort_value(slot.period),
            slot.order,
            slot.id,
        ),
    )


def relaxed_candidates_on_date(
    assignment: scheduler.Assignment,
    time_slots: Sequence[scheduler.TimeSlot],
    date_text: str,
    preferred_period: Optional[str] = None,
) -> List[scheduler.Candidate]:
    task = assignment.task
    if task.start_date and date_text < task.start_date:
        return []
    if task.end_date and date_text > task.end_date:
        return []
    room_id = assignment.candidate.room_id
    if task.room_ids and room_id not in task.room_ids:
        return []
    candidates: List[scheduler.Candidate] = []
    day_slots = [slot for slot in time_slots if slot.date == date_text]
    for slot_block in scheduler.build_contiguous_slot_blocks(day_slots, task.block_hours):
        if preferred_period and slot_block[0].period != preferred_period:
            continue
        if task.allowed_periods and slot_block[0].period not in task.allowed_periods:
            continue
        candidates.append(
            scheduler.Candidate(
                slots=slot_block,
                teacher_id=assignment.candidate.teacher_id,
                teacher_name=assignment.candidate.teacher_name,
                room_id=room_id,
            )
        )
    original_period = assignment.candidate.slots[0].period
    candidates.sort(
        key=lambda candidate: (
            candidate.slots[0].period != original_period,
            scheduler.period_sort_value(candidate.slots[0].period),
            candidate.slots[0].order,
        )
    )
    return candidates


def apply_2731_stageless_weekly_rule(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
) -> Tuple[List[scheduler.Assignment], List[str]]:
    suite_code = WYS_STAGELESS_WEEKLY_SUITE
    suite_assignments = [
        assignment
        for assignment in assignments
        if suite_code_for_class(assignment.task.class_id, class_metadata) == suite_code
        and assignment.task.subject in SUMMER_PUBLIC_SUBJECTS
    ]
    if not suite_assignments:
        return scheduler.sorted_assignments(list(assignments)), []

    long_camp_only_rest_dates = {
        (Date.fromisoformat("2026-08-29") + timedelta(days=offset)).isoformat()
        for offset in range(7)
    }
    excluded_dates = (set(blackout_dates) - long_camp_only_rest_dates) | WUYOU_PRODUCT_BLACKOUT_DATES
    time_slots = time_slots_with_standard_days(
        load_base_time_slots(data_dir),
        WYS_AUTUMN_START,
        AUTUMN_END,
        {"AM", "PM"},
        weekdays=WYS_AUTUMN_ALLOWED_WEEKDAYS,
        excluded_dates=excluded_dates,
    )
    slots_by_date_period: Dict[Tuple[str, str], List[scheduler.TimeSlot]] = defaultdict(list)
    for slot in time_slots:
        slots_by_date_period[(slot.date, slot.period)].append(slot)
    for key in list(slots_by_date_period):
        slots_by_date_period[key] = sorted(
            slots_by_date_period[key],
            key=lambda slot: (slot.order, slot.start_time or "", slot.id),
        )

    autumn_weeks: List[Date] = []
    current = Date.fromisoformat(WYS_AUTUMN_START)
    end = Date.fromisoformat(AUTUMN_END)
    while current <= end:
        week = week_monday(current.isoformat())
        if week not in autumn_weeks:
            autumn_weeks.append(week)
        current += timedelta(days=1)

    result = scheduler.sorted_assignments(list(assignments))
    lines: List[str] = []

    def suite_week_counts(items: Sequence[scheduler.Assignment]) -> Dict[Date, int]:
        counts: Dict[Date, int] = defaultdict(int)
        for item in items:
            if suite_code_for_class(item.task.class_id, class_metadata) != suite_code:
                continue
            first = item.candidate.slots[0]
            if WYS_SUMMER_START <= first.date <= AUTUMN_END:
                counts[week_monday(first.date)] += 1
        return counts

    def suite_week_subjects(items: Sequence[scheduler.Assignment]) -> Dict[Date, Set[str]]:
        subjects: Dict[Date, Set[str]] = defaultdict(set)
        for item in items:
            if suite_code_for_class(item.task.class_id, class_metadata) != suite_code:
                continue
            first = item.candidate.slots[0]
            if WYS_SUMMER_START <= first.date <= AUTUMN_END:
                subjects[week_monday(first.date)].add(item.task.subject)
        return subjects

    def target_dates_for_week(week: Date) -> List[str]:
        dates: List[str] = []
        for offset in range(7):
            day = week + timedelta(days=offset)
            date_text = day.isoformat()
            if (
                date_text < WYS_AUTUMN_START
                or date_text > AUTUMN_END
                or date_text in excluded_dates
                or day.weekday() not in WYS_AUTUMN_ALLOWED_WEEKDAYS
            ):
                continue
            dates.append(date_text)
        return dates

    def effective_week_bounds(week: Date) -> Tuple[int, int]:
        available_half_days = len(target_dates_for_week(week)) * 2
        if available_half_days <= 0:
            return 0, 0
        return (
            min(WYS_STAGELESS_AUTUMN_WEEKLY_MIN, available_half_days),
            min(WYS_STAGELESS_AUTUMN_WEEKLY_MAX, available_half_days),
        )

    def candidate_on_target(
        assignment: scheduler.Assignment,
        date_text: str,
        period: str,
        validation_pool: Sequence[scheduler.Assignment],
        avoid_same_teacher_day: bool,
        avoid_same_class_day: bool = True,
    ) -> Optional[scheduler.Assignment]:
        slots = slots_by_date_period.get((date_text, period), [])
        required_hours = sum(slot.duration_hours for slot in assignment.candidate.slots)
        selected: List[scheduler.TimeSlot] = []
        remaining = required_hours
        for slot in slots:
            if remaining <= 0:
                break
            selected.append(slot)
            remaining -= slot.duration_hours
        if not selected or remaining > 0:
            return None
        teacher_key = candidate_teacher_key(assignment.candidate)
        for item in validation_pool:
            item_first = item.candidate.slots[0]
            if (
                avoid_same_class_day
                and item.task.class_id == assignment.task.class_id
                and item_first.date == date_text
            ):
                return None
            if (
                avoid_same_teacher_day
                and teacher_key
                and candidate_teacher_key(item.candidate) == teacher_key
                and item_first.date == date_text
            ):
                return None
        moved = scheduler.Assignment(
            task=assignment.task,
            candidate=scheduler.Candidate(
                slots=tuple(selected),
                teacher_id=assignment.candidate.teacher_id,
                teacher_name=assignment.candidate.teacher_name,
                room_id=assignment.candidate.room_id,
            ),
        )
        if assignments_conflicting_with_candidate(moved, validation_pool, class_conflict_groups):
            return None
        return moved

    for current_assignment in list(result):
        first = current_assignment.candidate.slots[0]
        if (
            suite_code_for_class(current_assignment.task.class_id, class_metadata) != suite_code
            or first.date < WYS_AUTUMN_START
            or first.date > AUTUMN_END
            or Date.fromisoformat(first.date).weekday() in WYS_AUTUMN_ALLOWED_WEEKDAYS
        ):
            continue
        original_week = week_monday(first.date)
        target_weeks = [
            *(
                [original_week]
                if suite_week_counts(result).get(original_week, 0) <= effective_week_bounds(original_week)[1]
                else []
            ),
            *[
                week
                for week in autumn_weeks
                if week != original_week
                and suite_week_counts(result).get(week, 0) < effective_week_bounds(week)[1]
            ],
        ]
        replacement: Optional[scheduler.Assignment] = None
        for avoid_same_class_day in (True, False):
            for avoid_same_teacher_day in (True, False):
                for target_week in target_weeks:
                    validation_pool = [
                        item for item in result if item.task.task_id != current_assignment.task.task_id
                    ]
                    for date_text in target_dates_for_week(target_week):
                        for period in ("AM", "PM"):
                            replacement = candidate_on_target(
                                current_assignment,
                                date_text,
                                period,
                                validation_pool,
                                avoid_same_teacher_day=avoid_same_teacher_day,
                                avoid_same_class_day=avoid_same_class_day,
                            )
                            if replacement is not None:
                                break
                        if replacement is not None:
                            break
                    if replacement is not None:
                        break
                if replacement is not None:
                    break
            if replacement is not None:
                break
        if replacement is None:
            lines.append(
                f"2731 无阶段周均衡: {current_assignment.task.class_id} "
                f"{current_assignment.task.subject}/{current_assignment.task.stage or ''}/{current_assignment.task.course_module or ''} "
                f"{first.date} {first.period} 未能挪入周三/周末"
            )
            continue
        new_first = replacement.candidate.slots[0]
        result = scheduler.sorted_assignments(
            [
                item
                for item in result
                if item.task.task_id != current_assignment.task.task_id
            ]
            + [replacement]
        )
        lines.append(
            f"2731 无阶段周均衡: {current_assignment.task.class_id} "
            f"{current_assignment.task.subject}/{current_assignment.task.stage or ''}/{current_assignment.task.course_module or ''} "
            f"{first.date} {first.period} -> {new_first.date} {new_first.period}，修正为周三/周末"
        )

    moved_count = 0
    for target_week in autumn_weeks:
        target_min, target_max = effective_week_bounds(target_week)
        while suite_week_counts(result).get(target_week, 0) < target_min:
            counts = suite_week_counts(result)
            subjects_by_week = suite_week_subjects(result)
            missing_subjects = SUMMER_PUBLIC_SUBJECTS - subjects_by_week.get(target_week, set())
            target_dates = target_dates_for_week(target_week)
            if not target_dates:
                break
            donors = [
                assignment
                for assignment in result
                if suite_code_for_class(assignment.task.class_id, class_metadata) == suite_code
                and WYS_SUMMER_START <= assignment.candidate.slots[0].date < WYS_AUTUMN_START
            ]
            donors.sort(
                key=lambda assignment: (
                    assignment.task.subject not in missing_subjects,
                    -counts.get(week_monday(assignment.candidate.slots[0].date), 0),
                    assignment.candidate.slots[0].date,
                    PUBLIC_SUBJECT_PRIORITY.get(assignment.task.subject, 99),
                )
            )
            moved_assignment: Optional[scheduler.Assignment] = None
            moved_source: Optional[scheduler.Assignment] = None
            for avoid_same_teacher_day in (True, False):
                for donor in donors:
                    source_week = week_monday(donor.candidate.slots[0].date)
                    if counts.get(source_week, 0) <= 1:
                        continue
                    validation_pool = [
                        item for item in result if item.task.task_id != donor.task.task_id
                    ]
                    for date_text in target_dates:
                        for period in ("AM", "PM"):
                            if counts.get(target_week, 0) >= target_max:
                                continue
                            candidate = candidate_on_target(
                                donor,
                                date_text,
                                period,
                                validation_pool,
                                avoid_same_teacher_day,
                            )
                            if candidate is None:
                                continue
                            moved_assignment = candidate
                            moved_source = donor
                            break
                        if moved_assignment is not None:
                            break
                    if moved_assignment is not None:
                        break
                if moved_assignment is not None:
                    break
            if moved_assignment is None or moved_source is None:
                break
            result = scheduler.sorted_assignments(
                [
                    item
                    for item in result
                    if item.task.task_id != moved_source.task.task_id
                ]
                + [moved_assignment]
            )
            moved_count += 1
            old_first = moved_source.candidate.slots[0]
            new_first = moved_assignment.candidate.slots[0]
            lines.append(
                f"2731 无阶段周均衡: {moved_source.task.class_id} "
                f"{moved_source.task.subject}/{moved_source.task.stage or ''}/{moved_source.task.course_module or ''} "
                f"{old_first.date} {old_first.period} -> {new_first.date} {new_first.period}"
            )

    final_counts = suite_week_counts(result)
    summer_issues = [
        f"{week_display_label(week)} {count} 个半天"
        for week, count in sorted(final_counts.items())
        if WYS_SUMMER_START <= week.isoformat() < WYS_AUTUMN_START
        and count > WYS_STAGELESS_SUMMER_WEEKLY_MAX
    ]
    autumn_issues = [
        f"{week_display_label(week)} {final_counts.get(week, 0)} 个半天"
        for week in autumn_weeks
        if not (effective_week_bounds(week)[0] <= final_counts.get(week, 0) <= effective_week_bounds(week)[1])
    ]
    limited_weeks = [
        f"{week_display_label(week)} 可用 {effective_week_bounds(week)[1]} 个半天"
        for week in autumn_weeks
        if 0 < effective_week_bounds(week)[1] < WYS_STAGELESS_AUTUMN_WEEKLY_MIN
    ]
    if moved_count:
        lines.insert(0, f"2731 无阶段周均衡: 移动 {moved_count} 节，目标为暑假每周不超过 10、秋季周三/周末每周 5-6")
    if summer_issues:
        lines.append("2731 暑假周课量仍超上限: " + "；".join(summer_issues))
    if autumn_issues:
        lines.append("2731 秋季周课量仍未达 5-6: " + "；".join(autumn_issues))
    if limited_weeks:
        lines.append("2731 秋季周三/周末受停课影响的例外周: " + "；".join(limited_weeks))
    if not summer_issues and not autumn_issues and limited_weeks:
        lines.append("2731 无阶段周均衡: 当前已满足暑假每周不超过 10；秋季非例外周每周 5-6，例外周按周三/周末可用容量排课")
    elif not moved_count and not summer_issues and not autumn_issues:
        lines.append("2731 无阶段周均衡: 当前已满足暑假每周不超过 10、秋季周三/周末每周 5-6")
    return scheduler.sorted_assignments(result), lines


def load_class_teacher_assignment_lookup(data_dir: Path) -> Dict[Tuple[str, str, str], Tuple[str, str]]:
    path = data_dir / "class_teacher_assignments.csv"
    lookup: Dict[Tuple[str, str, str], Tuple[str, str]] = {}
    if not path.exists():
        return lookup
    for row in read_csv_rows(path):
        class_id = clean(row.get("class_id"))
        stage = clean(row.get("stage"))
        course_group = clean(row.get("course_group"))
        teacher_id = clean(row.get("teacher_id"))
        teacher_name = clean(row.get("teacher_name"))
        if not class_id or not teacher_id and not teacher_name:
            continue
        lookup.setdefault((class_id, stage, course_group), (teacher_id, teacher_name))
        if stage:
            lookup.setdefault((class_id, stage, ""), (teacher_id, teacher_name))
        if course_group:
            lookup.setdefault((class_id, "", course_group), (teacher_id, teacher_name))
    return lookup


def stage_priority_week_total_max(date_text: str) -> int:
    return WYS_STAGELESS_SUMMER_WEEKLY_MAX if date_text <= WYS_SUMMER_END else WYS_STAGELESS_AUTUMN_WEEKLY_MAX


def stage_priority_subject_week_max(date_text: str, subject: str) -> int:
    if date_text <= WYS_SUMMER_END:
        return 4
    return 4


def stage_priority_date_allowed(date_text: str, blackout_dates: Set[str]) -> bool:
    value = Date.fromisoformat(date_text)
    if date_text in blackout_dates or date_text in WUYOU_PRODUCT_BLACKOUT_DATES:
        return False
    if WYS_SUMMER_START <= date_text <= WYS_SUMMER_END:
        return value.weekday() != 6
    if WYS_AUTUMN_START <= date_text <= AUTUMN_END:
        return value.weekday() in WYS_AUTUMN_ALLOWED_WEEKDAYS
    return False


def stage_priority_slots(
    data_dir: Path,
    start_date: str,
    blackout_dates: Set[str],
) -> List[Tuple[str, str, Tuple[scheduler.TimeSlot, ...]]]:
    long_camp_only_rest_dates = {
        (Date.fromisoformat("2026-08-29") + timedelta(days=offset)).isoformat()
        for offset in range(7)
    }
    excluded_dates = (set(blackout_dates) - long_camp_only_rest_dates) | WUYOU_PRODUCT_BLACKOUT_DATES
    time_slots = time_slots_with_standard_days(
        load_base_time_slots(data_dir),
        max(start_date, WYS_SUMMER_START),
        AUTUMN_END,
        {"AM", "PM"},
        excluded_dates=excluded_dates,
    )
    grouped: Dict[Tuple[str, str], List[scheduler.TimeSlot]] = defaultdict(list)
    for slot in time_slots:
        if slot.period not in {"AM", "PM"}:
            continue
        if slot.date < start_date or not stage_priority_date_allowed(slot.date, excluded_dates):
            continue
        grouped[(slot.date, slot.period)].append(slot)
    result: List[Tuple[str, str, Tuple[scheduler.TimeSlot, ...]]] = []
    for (date_text, period), slots in grouped.items():
        block = tuple(sorted(slots, key=lambda slot: (slot.order, slot.start_time or "", slot.id)))
        if sum(slot.duration_hours for slot in block) < 4:
            block = standard_period_slots(date_text, period)
        result.append((date_text, period, block))
    return sorted(
        result,
        key=lambda item: (
            item[0],
            scheduler.period_sort_value(item[1]),
        ),
    )


def build_2731_stage_priority_blocks(
    data_dir: Path,
    class_metadata: Dict[str, Dict[str, str]],
) -> Tuple[Dict[str, Dict[str, List[scheduler.CourseBlock]]], List[str]]:
    teacher_lookup = load_class_teacher_assignment_lookup(data_dir)
    suite_class_ids = {
        class_id
        for class_id, meta in class_metadata.items()
        if suite_code_for_class(class_id, class_metadata) == WYS_STAGE_PRIORITY_SUITE
        and meta.get("sub_product") == WYS_PRODUCT
        and meta.get("subject") in SUMMER_PUBLIC_SUBJECTS
    }
    products_by_class = {
        class_id: clean(class_metadata[class_id].get("product_id"))
        for class_id in suite_class_ids
        if clean(class_metadata[class_id].get("product_id"))
    }
    product_to_classes: Dict[str, List[str]] = defaultdict(list)
    for class_id, product_id in products_by_class.items():
        product_to_classes[product_id].append(class_id)

    blocks: Dict[str, Dict[str, List[scheduler.CourseBlock]]] = {
        stage: defaultdict(list) for stage in WYS_STAGE_PRIORITY_ORDER
    }
    warnings: List[str] = []
    product_path = data_dir / "product_courses.csv"
    if not product_path.exists():
        return blocks, ["2731 阶段优先重建: 缺少 product_courses.csv，无法按产品课时重建"]

    task_seq = 1
    for row in read_csv_rows(product_path):
        product_id = clean(row.get("product_id"))
        if product_id not in product_to_classes:
            continue
        subject = normalize_subject(row.get("subject"))
        if subject not in SUMMER_PUBLIC_SUBJECTS:
            continue
        stage = clean(row.get("stage"))
        if stage not in WYS_STAGE_PRIORITY_ORDER:
            continue
        course_module = clean(row.get("course_module"))
        course_group = clean(row.get("course_group"))
        try:
            total_hours = int(float(clean(row.get("total_hours")) or 0))
            block_hours = int(float(clean(row.get("block_hours")) or 4))
        except ValueError:
            warnings.append(f"2731 阶段优先重建: {product_id} {stage} {course_module} 课时不是数字，已跳过")
            continue
        if total_hours <= 0 or block_hours <= 0:
            continue
        block_count = max(1, (total_hours + block_hours - 1) // block_hours)
        for class_id in sorted(product_to_classes[product_id]):
            meta = class_metadata.get(class_id, {})
            teacher_id, teacher_name = teacher_lookup.get((class_id, stage, course_group), ("", ""))
            if not teacher_id and not teacher_name:
                teacher_id, teacher_name = teacher_lookup.get((class_id, stage, ""), ("", ""))
            if not teacher_id and not teacher_name:
                teacher_id, teacher_name = teacher_lookup.get((class_id, "", course_group), ("", ""))
            if not teacher_id and not teacher_name:
                warnings.append(
                    f"2731 阶段优先重建: {class_id} {stage} {course_group or course_module} 缺课程老师，已跳过"
                )
                continue
            room_ids = split_arg_values([meta.get("preferred_room_ids", "")])
            for index in range(block_count):
                task = scheduler.CourseBlock(
                    task_id=f"2731_STAGE:{task_seq}",
                    class_id=class_id,
                    class_name=meta.get("name") or class_id,
                    product_id=product_id,
                    product_name=clean(row.get("product_name")) or meta.get("product_name"),
                    class_size=None,
                    subject_category=clean(row.get("subject_category")) or "公共课",
                    subject=subject,
                    quarter=clean(row.get("quarter")) or None,
                    stage=stage,
                    course_module=course_module or None,
                    course_group=course_group or None,
                    teacher_id=teacher_id,
                    teacher_name=teacher_name,
                    block_hours=block_hours,
                    room_ids=room_ids or None,
                    start_date=meta.get("start_date") or None,
                    end_date=AUTUMN_END,
                    allowed_periods={"AM", "PM"},
                    allowed_weekdays=None,
                    excluded_weekdays=None,
                    schedule_rules=(),
                    is_locked=False,
                    course_code=clean(row.get("course_code")),
                    course_name=clean(row.get("course_name")),
                )
                blocks[stage][subject].append(task)
                task_seq += 1
    return blocks, warnings


def stage_priority_candidate_rooms(task: scheduler.CourseBlock) -> List[str]:
    room_ids = sorted(task.room_ids or set())
    if room_ids:
        return room_ids
    return ["RMHFWY03023"]


def stage_priority_candidate_assignment(
    task: scheduler.CourseBlock,
    slots: Tuple[scheduler.TimeSlot, ...],
    room_id: str,
) -> scheduler.Assignment:
    return scheduler.Assignment(
        task=replace(
            task,
            start_date=slots[0].date,
            end_date=slots[0].date,
            allowed_periods={slots[0].period},
            allowed_weekdays={Date.fromisoformat(slots[0].date).weekday()},
        ),
        candidate=scheduler.Candidate(
            slots=slots,
            teacher_id=task.teacher_id,
            teacher_name=task.teacher_name,
            room_id=room_id,
        ),
    )


def assignment_with_replaced_slots(
    assignment: scheduler.Assignment,
    slots: Tuple[scheduler.TimeSlot, ...],
) -> scheduler.Assignment:
    return scheduler.Assignment(
        task=replace(
            assignment.task,
            start_date=slots[0].date,
            end_date=slots[0].date,
            allowed_periods={slots[0].period},
            allowed_weekdays={Date.fromisoformat(slots[0].date).weekday()},
        ),
        candidate=scheduler.Candidate(
            slots=slots,
            teacher_id=assignment.candidate.teacher_id,
            teacher_name=assignment.candidate.teacher_name,
            room_id=assignment.candidate.room_id,
        ),
    )


def stage_order_ok_for_suite(
    assignments: Sequence[scheduler.Assignment],
    suite_code: str,
    class_metadata: Dict[str, Dict[str, str]],
    stage_order: Sequence[str],
) -> bool:
    ranges: Dict[Tuple[str, str], List[str]] = defaultdict(list)
    for assignment in assignments:
        if suite_code_for_class(assignment.task.class_id, class_metadata) != suite_code:
            continue
        stage = clean(assignment.task.stage)
        if stage not in stage_order:
            continue
        ranges[(assignment.task.class_id, stage)].append(assignment.candidate.slots[0].date)
    class_ids = {
        assignment.task.class_id
        for assignment in assignments
        if suite_code_for_class(assignment.task.class_id, class_metadata) == suite_code
    }
    for class_id in class_ids:
        for before, after in zip(stage_order, stage_order[1:]):
            before_dates = ranges.get((class_id, before), [])
            after_dates = ranges.get((class_id, after), [])
            if before_dates and after_dates and max(before_dates) > min(after_dates):
                return False
    return True


def suite_same_subject_day_overloads(
    assignments: Sequence[scheduler.Assignment],
    suite_code: str,
    class_metadata: Dict[str, Dict[str, str]],
) -> Dict[Tuple[str, str, str], float]:
    loads: Dict[Tuple[str, str, str], float] = defaultdict(float)
    for assignment in assignments:
        if suite_code_for_class(assignment.task.class_id, class_metadata) != suite_code:
            continue
        if assignment.task.subject not in SUMMER_PUBLIC_SUBJECTS:
            continue
        first = assignment.candidate.slots[0]
        loads[(assignment.task.class_id, assignment.task.subject, first.date)] += sum(
            slot.duration_hours for slot in assignment.candidate.slots
        )
    return {key: value for key, value in loads.items() if value >= PUBLIC_SAME_CLASS_SUBJECT_DAY_HOURS}


def repair_2731_same_subject_day_overloads(
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
) -> Tuple[List[scheduler.Assignment], List[str]]:
    result = scheduler.sorted_assignments(list(assignments))
    lines: List[str] = []

    def swap_score(target: scheduler.Assignment, candidate: scheduler.Assignment) -> Tuple[int, int, int, str, int]:
        target_date = Date.fromisoformat(target.candidate.slots[0].date)
        candidate_date = Date.fromisoformat(candidate.candidate.slots[0].date)
        return (
            abs((candidate_date - target_date).days),
            clean(candidate.task.stage) != clean(target.task.stage),
            scheduler.period_sort_value(candidate.candidate.slots[0].period),
            candidate.candidate.slots[0].date,
            PUBLIC_SUBJECT_PRIORITY.get(candidate.task.subject, 99),
        )

    for _attempt in range(8):
        overloads = suite_same_subject_day_overloads(result, WYS_STAGE_PRIORITY_SUITE, class_metadata)
        if not overloads:
            break
        repaired = False
        for (class_id, subject, date_text), _hours in sorted(overloads.items()):
            overloaded_assignments = [
                assignment
                for assignment in result
                if assignment.task.class_id == class_id
                and assignment.task.subject == subject
                and assignment.candidate.slots[0].date == date_text
            ]
            overloaded_assignments.sort(
                key=lambda assignment: (
                    renjie_teacher_assignment(assignment),
                    scheduler.period_sort_value(assignment.candidate.slots[0].period) != scheduler.period_sort_value("PM"),
                )
            )
            for left in overloaded_assignments:
                candidates = [
                    assignment
                    for assignment in result
                    if assignment.task.task_id != left.task.task_id
                    and suite_code_for_class(assignment.task.class_id, class_metadata) == WYS_STAGE_PRIORITY_SUITE
                    and assignment.task.subject != left.task.subject
                    and assignment.candidate.slots[0].date != date_text
                ]
                candidates.sort(key=lambda assignment: swap_score(left, assignment))
                for right in candidates:
                    left_swapped = assignment_with_replaced_slots(left, right.candidate.slots)
                    right_swapped = assignment_with_replaced_slots(right, left.candidate.slots)
                    if renjie_offline_assignment_violation(left_swapped) or renjie_offline_assignment_violation(right_swapped):
                        continue
                    others = [
                        assignment
                        for assignment in result
                        if assignment.task.task_id not in {left.task.task_id, right.task.task_id}
                    ]
                    if assignments_conflicting_with_candidate(left_swapped, others, class_conflict_groups):
                        continue
                    if assignments_conflicting_with_candidate(right_swapped, others, class_conflict_groups):
                        continue
                    proposed = scheduler.sorted_assignments([*others, left_swapped, right_swapped])
                    proposed_overloads = suite_same_subject_day_overloads(
                        proposed,
                        WYS_STAGE_PRIORITY_SUITE,
                        class_metadata,
                    )
                    target_key = (class_id, subject, date_text)
                    if target_key in proposed_overloads:
                        continue
                    if set(proposed_overloads) - (set(overloads) - {target_key}):
                        continue
                    if not stage_order_ok_for_suite(proposed, WYS_STAGE_PRIORITY_SUITE, class_metadata, WYS_STAGE_PRIORITY_ORDER):
                        continue
                    result = proposed
                    lines.append(
                        "2731 同科同日 8 小时修正: "
                        f"{left.task.class_id} {left.task.subject}/{left.task.stage}/{left.task.course_module} "
                        f"{left.candidate.slots[0].date} {left.candidate.slots[0].period} "
                        f"<-> {right.task.class_id} {right.task.subject}/{right.task.stage}/{right.task.course_module} "
                        f"{right.candidate.slots[0].date} {right.candidate.slots[0].period}"
                    )
                    repaired = True
                    break
                if repaired:
                    break
            if repaired:
                break
        if not repaired:
            unresolved = "；".join(
                f"{class_id} {date_text} {subject} {hours:g} 小时"
                for (class_id, subject, date_text), hours in sorted(overloads.items())
            )
            lines.append(f"2731 同科同日 8 小时修正: 暂未找到无冲突换位，残留 {unresolved}")
            break
    return scheduler.sorted_assignments(result), lines


def rebuild_2731_stage_priority_schedule(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
) -> Tuple[List[scheduler.Assignment], List[str]]:
    suite_class_ids = {
        class_id
        for class_id, meta in class_metadata.items()
        if suite_code_for_class(class_id, class_metadata) == WYS_STAGE_PRIORITY_SUITE
        and meta.get("sub_product") == WYS_PRODUCT
        and meta.get("subject") in SUMMER_PUBLIC_SUBJECTS
    }
    if not suite_class_ids:
        return scheduler.sorted_assignments(list(assignments)), []

    protected_assignments = [
        assignment for assignment in assignments if assignment.task.class_id not in suite_class_ids
    ]
    stage_blocks, lines = build_2731_stage_priority_blocks(data_dir, class_metadata)
    total_required = sum(
        len(tasks)
        for stage_subjects in stage_blocks.values()
        for tasks in stage_subjects.values()
    )
    if total_required <= 0:
        return scheduler.sorted_assignments(list(assignments)), lines
    stage_subject_total_counts: Dict[Tuple[str, str], int] = {}
    stage_subject_teacher_total_counts: Dict[Tuple[str, str, str], int] = defaultdict(int)
    stage_subject_teacher_counts: Dict[Tuple[str, str], Set[str]] = defaultdict(set)
    for stage, subjects in stage_blocks.items():
        for subject, tasks in subjects.items():
            stage_subject_total_counts[(stage, subject)] = len(tasks)
            for task in tasks:
                teacher_key = task.teacher_id or task.teacher_name
                stage_subject_teacher_total_counts[(stage, subject, teacher_key)] += 1
                stage_subject_teacher_counts[(stage, subject)].add(teacher_key)

    starts = [
        clean(class_metadata.get(class_id, {}).get("start_date"))
        for class_id in suite_class_ids
        if clean(class_metadata.get(class_id, {}).get("start_date"))
    ]
    start_date = max(WYS_SUMMER_START, min(starts) if starts else WYS_SUMMER_START)
    slot_pool = stage_priority_slots(data_dir, start_date, blackout_dates)
    placed: List[scheduler.Assignment] = list(protected_assignments)
    new_assignments: List[scheduler.Assignment] = []
    week_total_counts: Dict[Date, int] = defaultdict(int)
    week_subject_counts: Dict[Tuple[Date, str], int] = defaultdict(int)
    date_subjects: Dict[str, Set[str]] = defaultdict(set)
    scheduled_stage_subject_counts: Dict[Tuple[str, str], int] = defaultdict(int)
    scheduled_stage_subject_teacher_counts: Dict[Tuple[str, str, str], int] = defaultdict(int)
    last_english_politics_subject: Optional[str] = None
    last_english_politics_run_count = 0
    last_teacher_by_stage_subject: Dict[Tuple[str, str], str] = {}
    teacher_run_by_stage_subject: Dict[Tuple[str, str], Tuple[str, int]] = {}
    skipped_conflicts: List[str] = []

    def remaining_for_stage(stage: str) -> int:
        return sum(len(tasks) for tasks in stage_blocks.get(stage, {}).values())

    def task_teacher_key(task: scheduler.CourseBlock) -> str:
        return task.teacher_id or task.teacher_name

    def subject_progress(stage: str, subject: str) -> float:
        total = stage_subject_total_counts.get((stage, subject)) or 1
        return scheduled_stage_subject_counts[(stage, subject)] / total

    def teacher_progress(stage: str, subject: str, teacher_key: str) -> float:
        total = stage_subject_teacher_total_counts.get((stage, subject, teacher_key)) or 1
        return scheduled_stage_subject_teacher_counts[(stage, subject, teacher_key)] / total

    def subject_has_multiple_teachers(stage: str, subject: str) -> bool:
        return len(stage_subject_teacher_counts.get((stage, subject), set())) > 1

    def task_indexes_for_subject(stage: str, subject: str) -> List[int]:
        tasks = stage_blocks.get(stage, {}).get(subject, [])
        if not tasks:
            return []
        stage_subject_key = (stage, subject)
        scheduled_count = scheduled_stage_subject_counts[stage_subject_key]
        if scheduled_count == 0:
            if stage == "基础" and subject == "英语":
                anchor_indexes = [
                    index for index, task in enumerate(tasks)
                    if clean(task.course_module) == "词汇"
                ]
                if anchor_indexes:
                    return anchor_indexes
            if stage == "基础" and subject == "政治":
                anchor_indexes = [
                    index for index, task in enumerate(tasks)
                    if clean(task.course_module) == "马原"
                ]
                if anchor_indexes:
                    return anchor_indexes
        last_teacher = last_teacher_by_stage_subject.get(stage_subject_key)
        indexes = list(range(len(tasks)))
        indexes.sort(
            key=lambda index: (
                subject_has_multiple_teachers(stage, subject)
                and bool(last_teacher and task_teacher_key(tasks[index]) == last_teacher),
                teacher_progress(stage, subject, task_teacher_key(tasks[index])),
                index,
            )
        )
        return indexes

    def candidate_tasks_for_slot(date_text: str, stage: str) -> List[Tuple[str, int, scheduler.CourseBlock]]:
        week = week_monday(date_text)
        if week_total_counts[week] >= stage_priority_week_total_max(date_text):
            return []
        subjects = [
            subject
            for subject, tasks in stage_blocks.get(stage, {}).items()
            if tasks and week_subject_counts[(week, subject)] < stage_priority_subject_week_max(date_text, subject)
        ]
        if not subjects:
            return []
        candidates: List[Tuple[str, int, scheduler.CourseBlock]] = []
        for subject in subjects:
            for index in task_indexes_for_subject(stage, subject):
                candidates.append((subject, index, stage_blocks[stage][subject][index]))
        candidates.sort(
            key=lambda item: (
                item[0] in date_subjects[date_text],
                item[0] in {"英语", "政治"}
                and item[0] == last_english_politics_subject
                and last_english_politics_run_count >= 2,
                subject_has_multiple_teachers(stage, item[0])
                and teacher_run_by_stage_subject.get((stage, item[0]), ("", 0))[0] == task_teacher_key(item[2])
                and teacher_run_by_stage_subject.get((stage, item[0]), ("", 0))[1] >= 2,
                subject_has_multiple_teachers(stage, item[0])
                and teacher_run_by_stage_subject.get((stage, item[0]), ("", 0))[0] == task_teacher_key(item[2])
                and teacher_run_by_stage_subject.get((stage, item[0]), ("", 0))[1] >= 4,
                subject_progress(stage, item[0]),
                item[0] in {"英语", "政治"} and item[0] == last_english_politics_subject,
                week_subject_counts[(week, item[0])],
                PUBLIC_SUBJECT_PRIORITY.get(item[0], 99),
                subject_has_multiple_teachers(stage, item[0])
                and bool(last_teacher_by_stage_subject.get((stage, item[0])) == task_teacher_key(item[2])),
                teacher_progress(stage, item[0], task_teacher_key(item[2])),
                item[1],
            )
        )
        return candidates

    stage_index = 0
    for date_text, _period, slots in slot_pool:
        while stage_index < len(WYS_STAGE_PRIORITY_ORDER) and remaining_for_stage(WYS_STAGE_PRIORITY_ORDER[stage_index]) == 0:
            stage_index += 1
        if stage_index >= len(WYS_STAGE_PRIORITY_ORDER):
            break
        stage = WYS_STAGE_PRIORITY_ORDER[stage_index]
        candidates_for_slot = candidate_tasks_for_slot(date_text, stage)
        if not candidates_for_slot:
            continue
        unused_subject_candidates = [
            item for item in candidates_for_slot if item[0] not in date_subjects[date_text]
        ]
        if unused_subject_candidates:
            candidates_for_slot = unused_subject_candidates
        placed_assignment: Optional[Tuple[str, int, scheduler.Assignment]] = None
        blocked_task: Optional[scheduler.CourseBlock] = None
        blocked_subject = ""
        for subject, task_index, task in candidates_for_slot:
            blocked_task = task
            blocked_subject = subject
            for room_id in stage_priority_candidate_rooms(task):
                candidate = stage_priority_candidate_assignment(task, slots, room_id)
                if renjie_offline_assignment_violation(candidate):
                    continue
                if assignments_conflicting_with_candidate(candidate, placed, class_conflict_groups):
                    continue
                placed_assignment = (subject, task_index, candidate)
                break
            if placed_assignment:
                break
        if not placed_assignment:
            task = blocked_task or candidates_for_slot[0][2]
            subject = blocked_subject or candidates_for_slot[0][0]
            skipped_conflicts.append(
                f"{date_text} {slots[0].period} {task.class_id} {stage} {task.course_module or task.course_group or subject}"
            )
            continue
        subject, task_index, assignment = placed_assignment
        stage_blocks[stage][subject].pop(task_index)
        placed.append(assignment)
        new_assignments.append(assignment)
        week = week_monday(date_text)
        week_total_counts[week] += 1
        week_subject_counts[(week, subject)] += 1
        date_subjects[date_text].add(subject)
        scheduled_stage_subject_counts[(stage, subject)] += 1
        teacher_key = task_teacher_key(assignment.task)
        scheduled_stage_subject_teacher_counts[(stage, subject, teacher_key)] += 1
        last_teacher_by_stage_subject[(stage, subject)] = teacher_key
        previous_teacher, previous_teacher_run = teacher_run_by_stage_subject.get((stage, subject), ("", 0))
        teacher_run_by_stage_subject[(stage, subject)] = (
            teacher_key,
            previous_teacher_run + 1 if previous_teacher == teacher_key else 1,
        )
        if subject in {"英语", "政治"}:
            last_english_politics_run_count = (
                last_english_politics_run_count + 1
                if last_english_politics_subject == subject
                else 1
            )
            last_english_politics_subject = subject

    remaining_lines: List[str] = []
    scheduled_by_stage: Dict[str, int] = defaultdict(int)
    for assignment in new_assignments:
        scheduled_by_stage[assignment.task.stage or ""] += 1
    for stage in WYS_STAGE_PRIORITY_ORDER:
        remaining = remaining_for_stage(stage)
        scheduled = scheduled_by_stage.get(stage, 0)
        required = scheduled + remaining
        if remaining:
            for subject, tasks in sorted(stage_blocks[stage].items(), key=lambda item: PUBLIC_SUBJECT_PRIORITY.get(item[0], 99)):
                if not tasks:
                    continue
                module_counts: Dict[str, int] = defaultdict(int)
                for task in tasks:
                    module_counts[task.course_module or task.course_group or subject] += 1
                module_text = "、".join(f"{module} {count} 个半天" for module, count in sorted(module_counts.items()))
                remaining_lines.append(f"2731 {stage} {subject}: 剩余 {len(tasks)} 个半天未排（{module_text}）")
        lines.append(f"2731 阶段优先重建: {stage} 已排 {scheduled}/{required} 个半天")

    lines.insert(
        0,
        f"2731 阶段优先重建: 移除旧无阶段课表，按 基础 -> 强化 -> 冲刺 重新生成 {len(new_assignments)}/{total_required} 个半天",
    )
    if skipped_conflicts:
        lines.append("2731 阶段优先重建: 以下候选因可排窗口或老师/教室/互斥冲突跳过，已尝试后续时段: " + "；".join(skipped_conflicts[:20]))
    lines.extend(remaining_lines)
    rebuilt = scheduler.sorted_assignments([*protected_assignments, *new_assignments])
    rebuilt, same_day_lines = repair_2731_same_subject_day_overloads(
        rebuilt,
        class_metadata,
        class_conflict_groups,
    )
    lines.extend(same_day_lines)
    return scheduler.sorted_assignments(rebuilt), lines


def apply_2727_no_math_summer_rule(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
) -> Tuple[List[scheduler.Assignment], List[str]]:
    suite_code = WYS_NO_MATH_BALANCED_SUITE
    suite_class_ids = {
        class_id
        for class_id, meta in class_metadata.items()
        if suite_code_for_class(class_id, class_metadata) == suite_code
        and meta.get("sub_product") == WYS_PRODUCT
        and meta.get("subject") in {"英语", "政治"}
    }
    if not suite_class_ids:
        return scheduler.sorted_assignments(list(assignments)), []
    suite_assignments = [
        assignment
        for assignment in assignments
        if assignment.task.class_id in suite_class_ids
        and assignment.task.subject in {"英语", "政治"}
    ]
    if not suite_assignments:
        return scheduler.sorted_assignments(list(assignments)), []

    starts = [
        clean(class_metadata.get(class_id, {}).get("start_date"))
        for class_id in suite_class_ids
        if clean(class_metadata.get(class_id, {}).get("start_date"))
    ]
    summer_start = max(WYS_SUMMER_START, min(starts) if starts else WYS_SUMMER_START)
    excluded_dates = set(blackout_dates) | WUYOU_PRODUCT_BLACKOUT_DATES
    time_slots = time_slots_with_standard_days(
        load_base_time_slots(data_dir),
        summer_start,
        WYS_SUMMER_END,
        {"AM", "PM"},
        weekdays={0, 1, 2, 3, 4, 5},
        excluded_dates=excluded_dates,
    )
    slots_by_date_period: Dict[Tuple[str, str], List[scheduler.TimeSlot]] = defaultdict(list)
    for slot in time_slots:
        slots_by_date_period[(slot.date, slot.period)].append(slot)
    for key in list(slots_by_date_period):
        slots_by_date_period[key] = sorted(
            slots_by_date_period[key],
            key=lambda slot: (slot.order, slot.start_time or "", slot.id),
        )

    summer_weeks: List[Date] = []
    current_week = week_monday(summer_start)
    last_week = week_monday(WYS_SUMMER_END)

    def target_dates_for_week(week: Date) -> List[str]:
        dates: List[str] = []
        for offset in range(7):
            day = week + timedelta(days=offset)
            date_text = day.isoformat()
            if (
                date_text < summer_start
                or date_text > WYS_SUMMER_END
                or date_text in excluded_dates
                or day.weekday() == 6
            ):
                continue
            dates.append(date_text)
        return dates

    while current_week <= last_week:
        if target_dates_for_week(current_week):
            summer_weeks.append(current_week)
        current_week += timedelta(days=7)

    result = scheduler.sorted_assignments(list(assignments))
    lines: List[str] = []

    def weekly_state(
        items: Sequence[scheduler.Assignment],
    ) -> Tuple[Dict[Date, int], Dict[Date, Dict[str, int]]]:
        totals: Dict[Date, int] = defaultdict(int)
        subjects: Dict[Date, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        for item in items:
            if item.task.class_id not in suite_class_ids or item.task.subject not in {"英语", "政治"}:
                continue
            first = item.candidate.slots[0]
            if summer_start <= first.date <= WYS_SUMMER_END:
                week = week_monday(first.date)
                totals[week] += 1
                subjects[week][item.task.subject] += 1
        return totals, subjects

    def candidate_on_target(
        assignment: scheduler.Assignment,
        date_text: str,
        period: str,
        validation_pool: Sequence[scheduler.Assignment],
        avoid_same_teacher_day: bool,
    ) -> Optional[scheduler.Assignment]:
        slots = slots_by_date_period.get((date_text, period), [])
        required_hours = sum(slot.duration_hours for slot in assignment.candidate.slots)
        selected: List[scheduler.TimeSlot] = []
        remaining = required_hours
        for slot in slots:
            if remaining <= 0:
                break
            selected.append(slot)
            remaining -= slot.duration_hours
        if not selected or remaining > 0:
            return None
        teacher_key = candidate_teacher_key(assignment.candidate)
        for item in validation_pool:
            item_first = item.candidate.slots[0]
            if item.task.class_id == assignment.task.class_id and item_first.date == date_text:
                return None
            if (
                avoid_same_teacher_day
                and teacher_key
                and candidate_teacher_key(item.candidate) == teacher_key
                and item_first.date == date_text
            ):
                return None
        moved = scheduler.Assignment(
            task=assignment.task,
            candidate=scheduler.Candidate(
                slots=tuple(selected),
                teacher_id=assignment.candidate.teacher_id,
                teacher_name=assignment.candidate.teacher_name,
                room_id=assignment.candidate.room_id,
            ),
        )
        if assignments_conflicting_with_candidate(moved, validation_pool, class_conflict_groups):
            return None
        return moved

    def can_spare(
        assignment: scheduler.Assignment,
        totals: Dict[Date, int],
        subject_counts: Dict[Date, Dict[str, int]],
    ) -> bool:
        first = assignment.candidate.slots[0]
        if not (summer_start <= first.date <= WYS_SUMMER_END):
            return True
        source_week = week_monday(first.date)
        subject = assignment.task.subject
        subject_min = WYS_NO_MATH_SUMMER_SUBJECT_WEEKLY_BOUNDS.get(subject, (0, 999))[0]
        return (
            totals.get(source_week, 0) - 1 >= 6
            and subject_counts.get(source_week, {}).get(subject, 0) - 1 >= subject_min
        )

    def try_move_to_week(target_week: Date, target_subject: str) -> bool:
        totals, subject_counts = weekly_state(result)
        if totals.get(target_week, 0) >= WYS_NO_MATH_SUMMER_WEEKLY_MAX:
            return False
        if (
            subject_counts.get(target_week, {}).get(target_subject, 0)
            >= WYS_NO_MATH_SUMMER_SUBJECT_WEEKLY_BOUNDS[target_subject][1]
        ):
            return False
        donors = [
            assignment
            for assignment in result
            if assignment.task.class_id in suite_class_ids
            and assignment.task.subject == target_subject
            and week_monday(assignment.candidate.slots[0].date) != target_week
            and can_spare(assignment, totals, subject_counts)
        ]
        donors.sort(
            key=lambda assignment: (
                summer_start <= assignment.candidate.slots[0].date <= WYS_SUMMER_END,
                -totals.get(week_monday(assignment.candidate.slots[0].date), 0),
                assignment.candidate.slots[0].date,
                scheduler.period_sort_value(assignment.candidate.slots[0].period),
            )
        )
        for avoid_same_teacher_day in (True, False):
            for donor in donors:
                validation_pool = [
                    item for item in result if item.task.task_id != donor.task.task_id
                ]
                for date_text in target_dates_for_week(target_week):
                    for period in ("AM", "PM"):
                        candidate = candidate_on_target(
                            donor,
                            date_text,
                            period,
                            validation_pool,
                            avoid_same_teacher_day,
                        )
                        if candidate is None:
                            continue
                        old_first = donor.candidate.slots[0]
                        new_first = candidate.candidate.slots[0]
                        result[:] = scheduler.sorted_assignments(
                            [
                                item
                                for item in result
                                if item.task.task_id != donor.task.task_id
                            ]
                            + [candidate]
                        )
                        lines.append(
                            f"2727 暑假英政周均衡: {donor.task.class_id} "
                            f"{donor.task.subject}/{donor.task.stage or ''}/{donor.task.course_module or ''} "
                            f"{old_first.date} {old_first.period} -> {new_first.date} {new_first.period}"
                        )
                        return True
        return False

    def scheduled_requirement_counts() -> Dict[Tuple[str, str, str, str, str], int]:
        counts: Dict[Tuple[str, str, str, str, str], int] = defaultdict(int)
        for item in result:
            if item.task.class_id not in suite_class_ids:
                continue
            key = (
                item.task.class_id,
                item.task.subject,
                item.task.stage or "",
                item.task.course_module or "",
                item.task.course_group or "",
            )
            counts[key] += 1
        return counts

    def missing_requirement_templates(target_subject: str) -> List[Tuple[scheduler.SchoolClass, scheduler.Requirement, int]]:
        source = load_schedule_input_for_classes(data_dir, sorted(suite_class_ids))
        counts = scheduled_requirement_counts()
        templates: List[Tuple[scheduler.SchoolClass, scheduler.Requirement, int]] = []
        for class_id in sorted(source.classes):
            cls = source.classes[class_id]
            for requirement in cls.requirements:
                if requirement.subject != target_subject:
                    continue
                key = (
                    class_id,
                    requirement.subject,
                    requirement.stage or "",
                    requirement.course_module or "",
                    requirement.course_group or "",
                )
                required_count = requirement.total_hours // requirement.block_hours
                missing_count = max(0, required_count - counts.get(key, 0))
                for index in range(missing_count):
                    templates.append((cls, requirement, index + 1))
        templates.sort(
            key=lambda item: (
                item[1].stage or "",
                item[1].course_group or "",
                item[1].course_module or "",
                item[2],
            )
        )
        return templates

    def room_for_class(class_id: str) -> str:
        for item in result:
            if item.task.class_id == class_id and item.candidate.room_id:
                return item.candidate.room_id
        meta_room = clean(class_metadata.get(class_id, {}).get("preferred_room_ids"))
        if meta_room:
            meta_rooms = split_pipe_values(meta_room)
            return meta_rooms[0] if meta_rooms else ""
        return ""

    def add_missing_requirement_to_week(target_week: Date, target_subject: str) -> bool:
        totals, subject_counts = weekly_state(result)
        if totals.get(target_week, 0) >= WYS_NO_MATH_SUMMER_WEEKLY_MAX:
            return False
        if (
            subject_counts.get(target_week, {}).get(target_subject, 0)
            >= WYS_NO_MATH_SUMMER_SUBJECT_WEEKLY_BOUNDS[target_subject][1]
        ):
            return False
        for cls, requirement, missing_index in missing_requirement_templates(target_subject):
            room_id = room_for_class(cls.id)
            if not room_id:
                continue
            task = scheduler.CourseBlock(
                task_id=(
                    f"ADD:{suite_code}:{cls.id}:{requirement.subject}:"
                    f"{requirement.stage or ''}:{requirement.course_module or ''}:"
                    f"{requirement.course_group or ''}:{missing_index}"
                ),
                class_id=cls.id,
                class_name=cls.name,
                product_id=cls.product_id,
                product_name=cls.product_name,
                class_size=cls.size,
                subject_category=requirement.subject_category,
                subject=requirement.subject,
                quarter=requirement.quarter,
                stage=requirement.stage,
                course_module=requirement.course_module,
                course_group=requirement.course_group,
                teacher_id=requirement.teacher_id,
                teacher_name=requirement.teacher_name,
                block_hours=requirement.block_hours,
                course_code=requirement.course_code,
                course_name=requirement.course_name,
                room_ids={room_id},
                start_date=summer_start,
                end_date=WYS_SUMMER_END,
                allowed_periods={"AM", "PM"},
                allowed_weekdays={0, 1, 2, 3, 4, 5},
                excluded_weekdays={6},
                schedule_rules=(),
            )
            for avoid_same_teacher_day in (True, False):
                for date_text in target_dates_for_week(target_week):
                    for period in ("AM", "PM"):
                        slots = slots_by_date_period.get((date_text, period), [])
                        selected: List[scheduler.TimeSlot] = []
                        remaining = requirement.block_hours
                        for slot in slots:
                            if remaining <= 0:
                                break
                            selected.append(slot)
                            remaining -= slot.duration_hours
                        if not selected or remaining > 0:
                            continue
                        teacher_key = clean(requirement.teacher_id) or clean(requirement.teacher_name)
                        blocked = False
                        for item in result:
                            first = item.candidate.slots[0]
                            if item.task.class_id == cls.id and first.date == date_text:
                                blocked = True
                                break
                            if (
                                avoid_same_teacher_day
                                and teacher_key
                                and candidate_teacher_key(item.candidate) == teacher_key
                                and first.date == date_text
                            ):
                                blocked = True
                                break
                        if blocked:
                            continue
                        candidate = scheduler.Assignment(
                            task=task,
                            candidate=scheduler.Candidate(
                                slots=tuple(selected),
                                teacher_id=requirement.teacher_id,
                                teacher_name=requirement.teacher_name,
                                room_id=room_id,
                            ),
                        )
                        if assignments_conflicting_with_candidate(candidate, result, class_conflict_groups):
                            continue
                        result.append(candidate)
                        result.sort(
                            key=lambda item: (
                                item.candidate.slots[0].date,
                                scheduler.period_sort_value(item.candidate.slots[0].period),
                                item.candidate.slots[0].start_time or "",
                                item.task.class_id,
                            )
                        )
                        first = candidate.candidate.slots[0]
                        lines.append(
                            f"2727 暑假英政周均衡: 补入缺失课程 {candidate.task.class_id} "
                            f"{candidate.task.subject}/{candidate.task.stage or ''}/{candidate.task.course_module or ''} "
                            f"{first.date} {first.period}"
                        )
                        return True
        return False

    for _pass_index in range(500):
        totals, subject_counts = weekly_state(result)
        moved = False
        for week in summer_weeks:
            if totals.get(week, 0) >= WYS_NO_MATH_SUMMER_WEEKLY_MAX:
                continue
            for subject, (subject_min, _subject_max) in WYS_NO_MATH_SUMMER_SUBJECT_WEEKLY_BOUNDS.items():
                if subject_counts.get(week, {}).get(subject, 0) >= subject_min:
                    continue
                if try_move_to_week(week, subject):
                    moved = True
                    break
            if moved:
                break
        if not moved:
            break

    for _pass_index in range(50):
        totals, subject_counts = weekly_state(result)
        added = False
        for week in summer_weeks:
            if totals.get(week, 0) >= WYS_NO_MATH_SUMMER_WEEKLY_MAX:
                continue
            for subject, (subject_min, _subject_max) in WYS_NO_MATH_SUMMER_SUBJECT_WEEKLY_BOUNDS.items():
                if subject_counts.get(week, {}).get(subject, 0) >= subject_min:
                    continue
                if add_missing_requirement_to_week(week, subject):
                    added = True
                    break
            if added:
                break
        if not added:
            break

    totals, subject_counts = weekly_state(result)
    issues: List[str] = []
    for week in summer_weeks:
        total = totals.get(week, 0)
        if total > WYS_NO_MATH_SUMMER_WEEKLY_MAX:
            issues.append(f"{week_display_label(week)} 合计{total} 个半天")
        for subject, (subject_min, subject_max) in WYS_NO_MATH_SUMMER_SUBJECT_WEEKLY_BOUNDS.items():
            count = subject_counts.get(week, {}).get(subject, 0)
            if count < subject_min or count > subject_max:
                issues.append(f"{week_display_label(week)} {subject}{count} 个半天")
    if issues:
        lines.append("2727 暑假英政周均衡仍需核对: " + "；".join(issues))
    else:
        lines.insert(
            0,
            "2727 暑假英政周均衡: 已按每周英语 3-4、政治 3-4、合计不超过 8 个半天调整",
        )
    return scheduler.sorted_assignments(result), lines


def apply_2727_no_math_weekly_rule(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
) -> Tuple[List[scheduler.Assignment], List[str]]:
    suite_code = WYS_NO_MATH_BALANCED_SUITE
    suite_assignments = [
        assignment
        for assignment in assignments
        if suite_code_for_class(assignment.task.class_id, class_metadata) == suite_code
        and assignment.task.subject in {"英语", "政治"}
        and WYS_AUTUMN_START <= assignment.candidate.slots[0].date <= AUTUMN_END
    ]
    if not suite_assignments:
        return scheduler.sorted_assignments(list(assignments)), []

    long_camp_only_rest_dates = {
        (Date.fromisoformat("2026-08-29") + timedelta(days=offset)).isoformat()
        for offset in range(7)
    }
    excluded_dates = (set(blackout_dates) - long_camp_only_rest_dates) | WUYOU_PRODUCT_BLACKOUT_DATES
    time_slots = time_slots_with_standard_days(
        load_base_time_slots(data_dir),
        WYS_AUTUMN_START,
        AUTUMN_END,
        {"AM", "PM"},
        weekdays=WYS_AUTUMN_ALLOWED_WEEKDAYS,
        excluded_dates=excluded_dates,
    )
    slots_by_date_period: Dict[Tuple[str, str], List[scheduler.TimeSlot]] = defaultdict(list)
    for slot in time_slots:
        slots_by_date_period[(slot.date, slot.period)].append(slot)
    for key in list(slots_by_date_period):
        slots_by_date_period[key] = sorted(
            slots_by_date_period[key],
            key=lambda slot: (slot.order, slot.start_time or "", slot.id),
        )

    autumn_weeks: List[Date] = []
    current = Date.fromisoformat(WYS_AUTUMN_START)
    end = Date.fromisoformat(AUTUMN_END)
    while current <= end:
        week = week_monday(current.isoformat())
        if week not in autumn_weeks:
            autumn_weeks.append(week)
        current += timedelta(days=1)

    result = scheduler.sorted_assignments(list(assignments))
    lines: List[str] = []

    def target_dates_for_week(week: Date) -> List[str]:
        dates: List[str] = []
        for offset in range(7):
            day = week + timedelta(days=offset)
            date_text = day.isoformat()
            if (
                date_text < WYS_AUTUMN_START
                or date_text > AUTUMN_END
                or date_text in excluded_dates
                or day.weekday() not in WYS_AUTUMN_ALLOWED_WEEKDAYS
            ):
                continue
            dates.append(date_text)
        return dates

    def weekly_state(
        items: Sequence[scheduler.Assignment],
    ) -> Tuple[Dict[Date, int], Dict[Date, Dict[str, int]]]:
        totals: Dict[Date, int] = defaultdict(int)
        subjects: Dict[Date, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        for item in items:
            if suite_code_for_class(item.task.class_id, class_metadata) != suite_code:
                continue
            first = item.candidate.slots[0]
            if WYS_AUTUMN_START <= first.date <= AUTUMN_END and item.task.subject in {"英语", "政治"}:
                week = week_monday(first.date)
                totals[week] += 1
                subjects[week][item.task.subject] += 1
        return totals, subjects

    def candidate_on_target(
        assignment: scheduler.Assignment,
        date_text: str,
        period: str,
        validation_pool: Sequence[scheduler.Assignment],
        avoid_same_teacher_day: bool,
    ) -> Optional[scheduler.Assignment]:
        slots = slots_by_date_period.get((date_text, period), [])
        required_hours = sum(slot.duration_hours for slot in assignment.candidate.slots)
        selected: List[scheduler.TimeSlot] = []
        remaining = required_hours
        for slot in slots:
            if remaining <= 0:
                break
            selected.append(slot)
            remaining -= slot.duration_hours
        if not selected or remaining > 0:
            return None
        teacher_key = candidate_teacher_key(assignment.candidate)
        for item in validation_pool:
            item_first = item.candidate.slots[0]
            if item.task.class_id == assignment.task.class_id and item_first.date == date_text:
                return None
            if (
                avoid_same_teacher_day
                and teacher_key
                and candidate_teacher_key(item.candidate) == teacher_key
                and item_first.date == date_text
            ):
                return None
        moved = scheduler.Assignment(
            task=assignment.task,
            candidate=scheduler.Candidate(
                slots=tuple(selected),
                teacher_id=assignment.candidate.teacher_id,
                teacher_name=assignment.candidate.teacher_name,
                room_id=assignment.candidate.room_id,
            ),
        )
        if assignments_conflicting_with_candidate(moved, validation_pool, class_conflict_groups):
            return None
        return moved

    def can_spare(
        assignment: scheduler.Assignment,
        totals: Dict[Date, int],
        subject_counts: Dict[Date, Dict[str, int]],
    ) -> bool:
        source_week = week_monday(assignment.candidate.slots[0].date)
        subject = assignment.task.subject
        subject_min = WYS_NO_MATH_SUBJECT_WEEKLY_BOUNDS.get(subject, (0, 999))[0]
        return (
            totals.get(source_week, 0) - 1 >= WYS_NO_MATH_AUTUMN_WEEKLY_MIN
            and subject_counts.get(source_week, {}).get(subject, 0) - 1 >= subject_min
        )

    def try_move_to_week(
        target_week: Date,
        target_subjects: Set[str],
        source_weeks: Optional[Set[Date]] = None,
    ) -> bool:
        totals, subject_counts = weekly_state(result)
        target_total = totals.get(target_week, 0)
        if target_total >= WYS_NO_MATH_AUTUMN_WEEKLY_MAX:
            return False
        donors = [
            assignment
            for assignment in result
            if suite_code_for_class(assignment.task.class_id, class_metadata) == suite_code
            and assignment.task.subject in target_subjects
            and WYS_AUTUMN_START <= assignment.candidate.slots[0].date <= AUTUMN_END
            and week_monday(assignment.candidate.slots[0].date) != target_week
            and (source_weeks is None or week_monday(assignment.candidate.slots[0].date) in source_weeks)
            and can_spare(assignment, totals, subject_counts)
            and subject_counts.get(target_week, {}).get(assignment.task.subject, 0)
            < WYS_NO_MATH_SUBJECT_WEEKLY_BOUNDS.get(assignment.task.subject, (0, 999))[1]
        ]
        donors.sort(
            key=lambda assignment: (
                -totals.get(week_monday(assignment.candidate.slots[0].date), 0),
                -subject_counts.get(week_monday(assignment.candidate.slots[0].date), {}).get(assignment.task.subject, 0),
                assignment.candidate.slots[0].date,
                assignment.task.subject,
            )
        )
        for avoid_same_teacher_day in (True, False):
            for donor in donors:
                validation_pool = [
                    item for item in result if item.task.task_id != donor.task.task_id
                ]
                for date_text in target_dates_for_week(target_week):
                    for period in ("AM", "PM"):
                        candidate = candidate_on_target(
                            donor,
                            date_text,
                            period,
                            validation_pool,
                            avoid_same_teacher_day,
                        )
                        if candidate is None:
                            continue
                        old_first = donor.candidate.slots[0]
                        new_first = candidate.candidate.slots[0]
                        result[:] = scheduler.sorted_assignments(
                            [
                                item
                                for item in result
                                if item.task.task_id != donor.task.task_id
                            ]
                            + [candidate]
                        )
                        lines.append(
                            f"2727 秋季周均衡: {donor.task.class_id} "
                            f"{donor.task.subject}/{donor.task.stage or ''}/{donor.task.course_module or ''} "
                            f"{old_first.date} {old_first.period} -> {new_first.date} {new_first.period}"
                        )
                        return True
        return False

    for current_assignment in list(result):
        first = current_assignment.candidate.slots[0]
        if (
            suite_code_for_class(current_assignment.task.class_id, class_metadata) != suite_code
            or first.date < WYS_AUTUMN_START
            or first.date > AUTUMN_END
            or Date.fromisoformat(first.date).weekday() in WYS_AUTUMN_ALLOWED_WEEKDAYS
        ):
            continue
        source_week = week_monday(first.date)
        validation_pool = [
            item for item in result if item.task.task_id != current_assignment.task.task_id
        ]
        replacement: Optional[scheduler.Assignment] = None
        for date_text in target_dates_for_week(source_week):
            for period in ("AM", "PM"):
                replacement = candidate_on_target(
                    current_assignment,
                    date_text,
                    period,
                    validation_pool,
                    avoid_same_teacher_day=True,
                )
                if replacement is not None:
                    break
            if replacement is not None:
                break
        if replacement is None:
            lines.append(
                f"2727 秋季周均衡: {current_assignment.task.class_id} "
                f"{current_assignment.task.subject}/{current_assignment.task.stage or ''}/{current_assignment.task.course_module or ''} "
                f"{first.date} {first.period} 未能挪入周三/周末"
            )
            continue
        new_first = replacement.candidate.slots[0]
        result = scheduler.sorted_assignments(
            [
                item
                for item in result
                if item.task.task_id != current_assignment.task.task_id
            ]
            + [replacement]
        )
        lines.append(
            f"2727 秋季周均衡: {current_assignment.task.class_id} "
            f"{current_assignment.task.subject}/{current_assignment.task.stage or ''}/{current_assignment.task.course_module or ''} "
            f"{first.date} {first.period} -> {new_first.date} {new_first.period}，修正为周三/周末"
        )

    for _pass_index in range(200):
        totals, subject_counts = weekly_state(result)
        moved = False
        for week in autumn_weeks:
            missing_subjects = {
                subject
                for subject, (subject_min, _subject_max) in WYS_NO_MATH_SUBJECT_WEEKLY_BOUNDS.items()
                if subject_counts.get(week, {}).get(subject, 0) < subject_min
            }
            if totals.get(week, 0) < WYS_NO_MATH_AUTUMN_WEEKLY_MIN or missing_subjects:
                if try_move_to_week(week, missing_subjects or {"英语", "政治"}):
                    moved = True
                    break
        if moved:
            continue
        source_weeks = {
            week
            for week in autumn_weeks
            if totals.get(week, 0) > WYS_NO_MATH_AUTUMN_WEEKLY_MAX
            or any(
                subject_counts.get(week, {}).get(subject, 0) > subject_max
                for subject, (_subject_min, subject_max) in WYS_NO_MATH_SUBJECT_WEEKLY_BOUNDS.items()
            )
        }
        if not source_weeks:
            break
        for week in autumn_weeks:
            if totals.get(week, 0) >= WYS_NO_MATH_AUTUMN_WEEKLY_MAX:
                continue
            target_subjects = {
                subject
                for subject, (_subject_min, subject_max) in WYS_NO_MATH_SUBJECT_WEEKLY_BOUNDS.items()
                if subject_counts.get(week, {}).get(subject, 0) < subject_max
            }
            if try_move_to_week(week, target_subjects, source_weeks=source_weeks):
                moved = True
                break
        if not moved:
            break

    totals, subject_counts = weekly_state(result)
    issues: List[str] = []
    for week in autumn_weeks:
        total = totals.get(week, 0)
        if total < WYS_NO_MATH_AUTUMN_WEEKLY_MIN or total > WYS_NO_MATH_AUTUMN_WEEKLY_MAX:
            issues.append(f"{week_display_label(week)} {total} 个半天")
            continue
        for subject, (subject_min, subject_max) in WYS_NO_MATH_SUBJECT_WEEKLY_BOUNDS.items():
            count = subject_counts.get(week, {}).get(subject, 0)
            if count < subject_min or count > subject_max:
                issues.append(f"{week_display_label(week)} {subject}{count} 个半天")
    if issues:
        lines.append("2727 秋季周均衡仍需核对: " + "；".join(issues))
    else:
        lines.insert(
            0,
            "2727 秋季周均衡: 已调整为周三/周末排课，周总量 2-3，英语每周 1 个半天，政治每周 1-2 个半天",
        )
    return scheduler.sorted_assignments(result), lines


def rebuild_2727_no_math_sequence_schedule(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
) -> Tuple[List[scheduler.Assignment], List[str]]:
    suite_code = WYS_NO_MATH_BALANCED_SUITE
    suite_class_ids = {
        class_id
        for class_id, meta in class_metadata.items()
        if suite_code_for_class(class_id, class_metadata) == suite_code
        and meta.get("sub_product") == WYS_PRODUCT
        and normalize_subject(meta.get("subject")) in {"英语", "政治"}
    }
    if not suite_class_ids:
        return scheduler.sorted_assignments(list(assignments)), []

    protected_assignments = [
        assignment
        for assignment in assignments
        if assignment.task.class_id not in suite_class_ids
        or assignment.task.subject not in {"英语", "政治"}
    ]
    source = load_schedule_input_for_classes(data_dir, sorted(suite_class_ids))
    if not source.classes:
        return scheduler.sorted_assignments(list(assignments)), []

    room_by_class: Dict[str, str] = {}
    for class_id, cls in source.classes.items():
        preferred_room = clean(class_metadata.get(class_id, {}).get("preferred_room_ids"))
        if preferred_room:
            preferred_rooms = split_pipe_values(preferred_room)
            room_by_class[class_id] = preferred_rooms[0] if preferred_rooms else ""
        elif cls.room_ids:
            room_by_class[class_id] = sorted(cls.room_ids)[0]
        else:
            room_by_class[class_id] = ""

    stage_order = ["基础", "强化", "冲刺"]
    pools: Dict[Tuple[str, str], List[scheduler.CourseBlock]] = defaultdict(list)
    for cls in source.classes.values():
        for req_index, requirement in enumerate(cls.requirements, start=1):
            if requirement.subject not in {"英语", "政治"}:
                continue
            block_count = requirement.total_hours // requirement.block_hours
            room_id = room_by_class.get(cls.id, "")
            room_ids = {room_id} if room_id else requirement.room_ids
            for block_index in range(block_count):
                pools[(requirement.subject, requirement.stage or "")].append(
                    scheduler.CourseBlock(
                        task_id=(
                            f"SEQ2727:{cls.id}:{requirement.subject}:"
                            f"{req_index}:{block_index + 1}"
                        ),
                        class_id=cls.id,
                        class_name=cls.name,
                        product_id=cls.product_id,
                        product_name=cls.product_name,
                        class_size=cls.size,
                        subject_category=requirement.subject_category,
                        subject=requirement.subject,
                        quarter=requirement.quarter,
                        stage=requirement.stage,
                        course_module=requirement.course_module,
                        course_group=requirement.course_group,
                        teacher_id=requirement.teacher_id,
                        teacher_name=requirement.teacher_name,
                        block_hours=requirement.block_hours,
                        course_code=requirement.course_code,
                        course_name=requirement.course_name,
                        room_ids=room_ids,
                        start_date=requirement.start_date,
                        end_date=requirement.end_date,
                        allowed_periods=requirement.allowed_periods,
                        allowed_weekdays=requirement.allowed_weekdays,
                        excluded_weekdays=requirement.excluded_weekdays,
                        schedule_rules=requirement.schedule_rules,
                    )
                )
    for key in list(pools):
        pools[key].sort(
            key=lambda task: (
                task.course_group or "",
                task.course_module or "",
                task.teacher_name or "",
                task.task_id,
            )
        )

    def iter_week_starts(start: str, end: str) -> List[Date]:
        weeks: List[Date] = []
        current = week_monday(start)
        last = week_monday(end)
        while current <= last:
            weeks.append(current)
            current += timedelta(days=7)
        return weeks

    long_camp_only_rest_dates = {
        (Date.fromisoformat("2026-08-29") + timedelta(days=offset)).isoformat()
        for offset in range(7)
    }
    summer_excluded_dates = set(blackout_dates) | WUYOU_PRODUCT_BLACKOUT_DATES
    autumn_excluded_dates = (set(blackout_dates) - long_camp_only_rest_dates) | WUYOU_PRODUCT_BLACKOUT_DATES
    time_slots = time_slots_with_standard_days(
        load_base_time_slots(data_dir),
        "2026-07-06",
        AUTUMN_END,
        {"AM", "PM"},
        excluded_dates=set(),
    )
    slots_by_date_period: Dict[Tuple[str, str], List[scheduler.TimeSlot]] = defaultdict(list)
    for slot in time_slots:
        slots_by_date_period[(slot.date, slot.period)].append(slot)
    for key in list(slots_by_date_period):
        slots_by_date_period[key] = sorted(
            slots_by_date_period[key],
            key=lambda slot: (slot.order, slot.start_time or "", slot.id),
        )

    def target_dates_for_week(week: Date, season: str) -> List[str]:
        dates: List[str] = []
        for offset in range(7):
            day = week + timedelta(days=offset)
            day_text = day.isoformat()
            if season == "summer":
                if (
                    day_text < "2026-07-06"
                    or day_text > WYS_SUMMER_END
                    or day_text in summer_excluded_dates
                    or day.weekday() == 6
                ):
                    continue
            else:
                if (
                    day_text < WYS_AUTUMN_START
                    or day_text > AUTUMN_END
                    or day_text in autumn_excluded_dates
                    or day.weekday() not in WYS_AUTUMN_ALLOWED_WEEKDAYS
                ):
                    continue
            dates.append(day_text)
        return dates

    summer_weeks = [
        week
        for week in iter_week_starts("2026-07-06", "2026-08-24")
        if target_dates_for_week(week, "summer")
    ]
    autumn_weeks = [
        week
        for week in iter_week_starts(WYS_AUTUMN_START, AUTUMN_END)
        if target_dates_for_week(week, "autumn")
    ]
    if len(summer_weeks) != 8:
        raise ValueError(f"2727 暑假周数异常: {len(summer_weeks)}")
    if len(autumn_weeks) < 12:
        raise ValueError(f"2727 秋季可排周数异常: {len(autumn_weeks)}")

    plan: List[Tuple[Date, str, str]] = []
    for index, week in enumerate(summer_weeks):
        if index in {2, 3}:
            week_subjects = ("英语", "政治", "英语", "政治", "英语", "政治", "政治")
        else:
            week_subjects = ("英语", "政治", "英语", "政治", "英语", "政治")
        for subject in week_subjects:
            plan.append((week, "summer", subject))

    politics_week_indexes = {0, 2, 4, 6, 8, 10, 12}
    for index, week in enumerate(autumn_weeks):
        english_count = 1 if index in politics_week_indexes else 2
        politics_count = 1 if index in politics_week_indexes else 0
        if english_count == 2:
            week_subjects = ("英语", "英语")
        elif politics_count:
            week_subjects = ("英语", "政治")
        else:
            week_subjects = ("英语",)
        for subject in week_subjects:
            plan.append((week, "autumn", subject))

    planned_counts: Dict[str, int] = defaultdict(int)
    for _week, _season, subject in plan:
        planned_counts[subject] += 1
    required_counts = {
        subject: sum(len(pools.get((subject, stage), [])) for stage in stage_order)
        for subject in ("英语", "政治")
    }
    if dict(planned_counts) != required_counts:
        raise ValueError(
            "2727 顺序重排计划课时与产品课时不一致: "
            f"计划 {dict(planned_counts)}，产品 {required_counts}"
        )

    generated_assignments: List[scheduler.Assignment] = []
    last_teacher_by_subject: Dict[str, str] = defaultdict(str)
    first_lesson_rank_by_subject: Dict[str, Tuple[object, ...]] = {}
    first_lesson_done: Dict[str, bool] = defaultdict(bool)

    def current_stage(subject: str) -> Optional[str]:
        for stage in stage_order:
            if pools.get((subject, stage)):
                return stage
        return None

    def make_assignment(
        task: scheduler.CourseBlock,
        lesson_date: str,
        period: str,
    ) -> Optional[scheduler.Assignment]:
        day_slots = slots_by_date_period.get((lesson_date, period), [])
        selected_slots: List[scheduler.TimeSlot] = []
        remaining_hours = task.block_hours
        for slot in day_slots:
            if remaining_hours <= 0:
                break
            selected_slots.append(slot)
            remaining_hours -= slot.duration_hours
        if not selected_slots or remaining_hours > 0:
            return None
        room_id = sorted(task.room_ids or {""})[0]
        return scheduler.Assignment(
            task=task,
            candidate=scheduler.Candidate(
                slots=tuple(selected_slots),
                teacher_id=task.teacher_id,
                teacher_name=task.teacher_name,
                room_id=room_id,
            ),
        )

    def slot_rank(slot: scheduler.TimeSlot) -> Tuple[object, ...]:
        return (
            slot.date,
            scheduler.period_sort_value(slot.period),
            slot.start_time or "",
            slot.id,
        )

    def feasible_assignments_for_task(
        task: scheduler.CourseBlock,
        week: Date,
        season: str,
    ) -> List[scheduler.Assignment]:
        candidates: List[Tuple[Tuple[object, ...], scheduler.Assignment]] = []
        task_teacher = task.teacher_id or task.teacher_name
        for lesson_date in target_dates_for_week(week, season):
            same_subject_on_date = any(
                assignment.task.subject == task.subject
                and assignment.candidate.slots[0].date == lesson_date
                for assignment in generated_assignments
            )
            same_class_teacher_on_date = any(
                assignment.task.class_id == task.class_id
                and candidate_teacher_key(assignment.candidate) == task_teacher
                and assignment.candidate.slots[0].date == lesson_date
                for assignment in generated_assignments
            )
            for period in ("AM", "PM"):
                candidate = make_assignment(task, lesson_date, period)
                if candidate is None:
                    continue
                candidate_rank = slot_rank(candidate.candidate.slots[0])
                first_subject_rank = first_lesson_rank_by_subject.get(task.subject)
                if first_subject_rank and candidate_rank <= first_subject_rank:
                    continue
                task_stage_rank = stage_order.index(task.stage or "") if (task.stage or "") in stage_order else 0
                previous_stage_ranks = [
                    slot_rank(assignment.candidate.slots[0])
                    for assignment in generated_assignments
                    if assignment.task.subject == task.subject
                    and (assignment.task.stage or "") in stage_order
                    and stage_order.index(assignment.task.stage or "") < task_stage_rank
                ]
                if previous_stage_ranks and candidate_rank <= max(previous_stage_ranks):
                    continue
                if assignments_conflicting_with_candidate(
                    candidate,
                    [*protected_assignments, *generated_assignments],
                    class_conflict_groups,
                ):
                    continue
                same_class_on_date = sum(
                    1
                    for assignment in generated_assignments
                    if assignment.task.class_id == task.class_id
                    and assignment.candidate.slots[0].date == lesson_date
                )
                first_slot = candidate.candidate.slots[0]
                candidates.append(
                    (
                        (
                            same_subject_on_date,
                            same_class_teacher_on_date,
                            same_class_on_date,
                            scheduler.period_sort_value(first_slot.period),
                            first_slot.date,
                            first_slot.start_time or "",
                        ),
                        candidate,
                    )
                )
        candidates.sort(key=lambda item: item[0])
        return [candidate for _rank, candidate in candidates]

    def task_options_for_subject(
        subject: str,
        week: Date,
        season: str,
    ) -> List[Tuple[scheduler.CourseBlock, List[scheduler.Assignment]]]:
        stage = current_stage(subject)
        if not stage:
            return []
        tasks = list(pools.get((subject, stage), []))
        if not first_lesson_done[subject] and stage == "基础":
            first_module = "词汇" if subject == "英语" else "马原"
            first_module_tasks = [
                task for task in tasks if task.course_module == first_module
            ]
            if first_module_tasks:
                tasks = first_module_tasks
        options: List[Tuple[scheduler.CourseBlock, List[scheduler.Assignment]]] = []
        for task in tasks:
            feasible = feasible_assignments_for_task(task, week, season)
            if feasible:
                options.append((task, feasible))
        options.sort(
            key=lambda item: (
                len(item[1]),
                candidate_teacher_key(item[1][0].candidate) == last_teacher_by_subject[subject],
                item[0].teacher_name or "",
                item[0].course_group or "",
                item[0].course_module or "",
                item[0].task_id,
            )
        )
        return options

    lines: List[str] = []
    for step_index, (week, season, subject) in enumerate(plan, start=1):
        options = task_options_for_subject(subject, week, season)
        if not options:
            remaining = {
                stage: len(pools.get((subject, stage), []))
                for stage in stage_order
            }
            raise ValueError(
                "2727 无忧暑英政顺序重排失败: "
                f"第 {step_index} 步 {week_display_label(week)} {subject} "
                f"{current_stage(subject) or ''} 无可用时段，剩余 {remaining}"
            )
        task, feasible = options[0]
        assignment = feasible[0]
        pools[(subject, task.stage or "")].remove(task)
        generated_assignments.append(assignment)
        last_teacher_by_subject[subject] = candidate_teacher_key(assignment.candidate)
        if not first_lesson_done[subject]:
            first_lesson_rank_by_subject[subject] = slot_rank(assignment.candidate.slots[0])
        first_lesson_done[subject] = True

    def stage_sequence_valid(items: Sequence[scheduler.Assignment]) -> bool:
        stage_rank = {stage: index for index, stage in enumerate(stage_order)}
        first_module_by_subject = {"英语": "词汇", "政治": "马原"}
        for subject, first_module in first_module_by_subject.items():
            subject_items = sorted(
                [assignment for assignment in items if assignment.task.subject == subject],
                key=lambda assignment: slot_rank(assignment.candidate.slots[0]),
            )
            if not subject_items:
                continue
            if subject_items[0].task.course_module != first_module:
                return False
            max_seen = -1
            for assignment in subject_items:
                rank = stage_rank.get(assignment.task.stage or "", 99)
                if rank < max_seen:
                    return False
                max_seen = max(max_seen, rank)
        return True

    def duplicate_same_subject_dates() -> Dict[Tuple[str, str], List[scheduler.Assignment]]:
        grouped: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
        for assignment in generated_assignments:
            grouped[(assignment.candidate.slots[0].date, assignment.task.subject)].append(assignment)
        return {
            key: sorted(
                items,
                key=lambda assignment: (
                    scheduler.period_sort_value(assignment.candidate.slots[0].period),
                    assignment.task.class_id,
                    assignment.task.task_id,
                ),
            )
            for key, items in grouped.items()
            if len(items) > 1
        }

    def move_candidate_for_assignment(
        assignment: scheduler.Assignment,
        target_date: str,
        period: str,
    ) -> Optional[scheduler.Assignment]:
        candidate = make_assignment(assignment.task, target_date, period)
        if candidate is None:
            return None
        candidate = scheduler.Assignment(
            task=assignment.task,
            candidate=scheduler.Candidate(
                slots=candidate.candidate.slots,
                teacher_id=assignment.candidate.teacher_id,
                teacher_name=assignment.candidate.teacher_name,
                room_id=assignment.candidate.room_id,
            ),
        )
        occupied = [
            item
            for item in [*protected_assignments, *generated_assignments]
            if item.task.task_id != assignment.task.task_id
        ]
        if assignments_conflicting_with_candidate(candidate, occupied, class_conflict_groups):
            return None
        candidate_generated = [
            candidate if item.task.task_id == assignment.task.task_id else item
            for item in generated_assignments
        ]
        if not stage_sequence_valid(candidate_generated):
            return None
        return candidate

    def raw_moved_assignment(
        assignment: scheduler.Assignment,
        target_date: str,
        period: str,
    ) -> Optional[scheduler.Assignment]:
        candidate = make_assignment(assignment.task, target_date, period)
        if candidate is None:
            return None
        return scheduler.Assignment(
            task=assignment.task,
            candidate=scheduler.Candidate(
                slots=candidate.candidate.slots,
                teacher_id=assignment.candidate.teacher_id,
                teacher_name=assignment.candidate.teacher_name,
                room_id=assignment.candidate.room_id,
            ),
        )

    window_constraints = load_summer_class_window_constraints(data_dir)
    repair_input_cache: Dict[str, scheduler.ScheduleInput] = {}

    def protected_assignment_index(assignment: scheduler.Assignment) -> Optional[int]:
        for index, item in enumerate(protected_assignments):
            if item.task.task_id == assignment.task.task_id:
                return index
        return None

    def can_move_protected_assignment(assignment: scheduler.Assignment) -> bool:
        meta = class_metadata.get(assignment.task.class_id, {})
        return (
            is_public_schedulable_meta(meta)
            and assignment.task.subject in SUMMER_PUBLIC_SUBJECTS
            and assignment.task.class_id not in suite_class_ids
        )

    def move_candidate_with_one_protected_displacement(
        assignment: scheduler.Assignment,
        target_date: str,
        period: str,
    ) -> Tuple[Optional[scheduler.Assignment], Optional[Tuple[int, scheduler.Assignment, scheduler.Assignment]]]:
        candidate = make_assignment(assignment.task, target_date, period)
        if candidate is None:
            return None, None
        candidate = scheduler.Assignment(
            task=assignment.task,
            candidate=scheduler.Candidate(
                slots=candidate.candidate.slots,
                teacher_id=assignment.candidate.teacher_id,
                teacher_name=assignment.candidate.teacher_name,
                room_id=assignment.candidate.room_id,
            ),
        )
        candidate_generated = [
            candidate if item.task.task_id == assignment.task.task_id else item
            for item in generated_assignments
        ]
        if not stage_sequence_valid(candidate_generated):
            return None, None
        direct_conflicts = assignments_conflicting_with_candidate(
            candidate,
            [
                *protected_assignments,
                *[
                    item
                    for item in generated_assignments
                    if item.task.task_id != assignment.task.task_id
                ],
            ],
            class_conflict_groups,
        )
        movable_conflicts: List[Tuple[int, scheduler.Assignment]] = []
        for conflict in direct_conflicts:
            protected_index = protected_assignment_index(conflict)
            if protected_index is None or not can_move_protected_assignment(conflict):
                continue
            movable_conflicts.append((protected_index, conflict))
        if len(direct_conflicts) != 1 or len(movable_conflicts) != 1:
            return None, None
        protected_index, conflict = movable_conflicts[0]
        repaired_conflict = find_teacher_conflict_repair_assignment(
            data_dir,
            conflict,
            protected_index,
            protected_assignments,
            candidate_generated,
            class_metadata,
            window_constraints,
            class_conflict_groups,
            blackout_dates,
            repair_input_cache,
        )
        if repaired_conflict is None:
            return None, None
        trial_protected = [
            repaired_conflict if index == protected_index else item
            for index, item in enumerate(protected_assignments)
        ]
        other_generated = [
            item
            for item in candidate_generated
            if item.task.task_id != candidate.task.task_id
        ]
        if assignments_conflicting_with_candidate(
            candidate,
            [*trial_protected, *other_generated],
            class_conflict_groups,
        ):
            return None, None
        if assignments_conflicting_with_candidate(
            repaired_conflict,
            [
                *candidate_generated,
                *[
                    item
                    for index, item in enumerate(protected_assignments)
                    if index != protected_index
                ],
            ],
            class_conflict_groups,
        ):
            return None, None
        return candidate, (protected_index, conflict, repaired_conflict)

    spread_moves: List[str] = []
    for _pass_index in range(20):
        duplicates = duplicate_same_subject_dates()
        if not duplicates:
            break
        moved = False
        for (lesson_date, subject), duplicate_items in sorted(duplicates.items()):
            week = week_monday(lesson_date)
            season = "summer" if lesson_date < WYS_AUTUMN_START else "autumn"
            target_dates = target_dates_for_week(week, season)
            for duplicate in duplicate_items:
                replacement: Optional[scheduler.Assignment] = None
                displacement: Optional[Tuple[int, scheduler.Assignment, scheduler.Assignment]] = None
                for target_date in target_dates:
                    if target_date == lesson_date:
                        continue
                    if any(
                        item.task.task_id != duplicate.task.task_id
                        and item.task.subject == subject
                        and item.candidate.slots[0].date == target_date
                        for item in generated_assignments
                    ):
                        continue
                    for period in ("AM", "PM"):
                        replacement = move_candidate_for_assignment(duplicate, target_date, period)
                        if replacement is None:
                            replacement, displacement = move_candidate_with_one_protected_displacement(
                                duplicate,
                                target_date,
                                period,
                            )
                        if replacement is not None:
                            break
                    if replacement is not None:
                        break
                if replacement is None:
                    continue
                if displacement is not None:
                    protected_index, old_conflict, repaired_conflict = displacement
                    protected_assignments[protected_index] = repaired_conflict
                old_slot = duplicate.candidate.slots[0]
                new_slot = replacement.candidate.slots[0]
                generated_assignments = scheduler.sorted_assignments(
                    [
                        replacement if item.task.task_id == duplicate.task.task_id else item
                        for item in generated_assignments
                    ]
                )
                spread_moves.append(
                    f"2727 同科同日摊开: {duplicate.task.subject}/"
                    f"{duplicate.task.stage or ''}/{duplicate.task.course_module or ''} "
                    f"{old_slot.date} {old_slot.period} -> {new_slot.date} {new_slot.period}"
                )
                if displacement is not None:
                    _protected_index, old_conflict, repaired_conflict = displacement
                    old_conflict_slot = old_conflict.candidate.slots[0]
                    repaired_conflict_slot = repaired_conflict.candidate.slots[0]
                    spread_moves.append(
                        "2727 让位移动前序公共课: "
                        f"{old_conflict.task.class_id} {old_conflict.task.subject}/"
                        f"{old_conflict.task.stage or ''}/{old_conflict.task.course_module or ''} "
                        f"{old_conflict_slot.date} {old_conflict_slot.period} -> "
                        f"{repaired_conflict_slot.date} {repaired_conflict_slot.period}"
                    )
                moved = True
                break
            if moved:
                break
        if not moved:
            break

    for _pass_index in range(20):
        duplicates = duplicate_same_subject_dates()
        if not duplicates:
            break
        exchanged = False
        for (lesson_date, subject), duplicate_items in sorted(duplicates.items()):
            week = week_monday(lesson_date)
            season = "summer" if lesson_date < WYS_AUTUMN_START else "autumn"
            target_dates = [
                target_date
                for target_date in target_dates_for_week(week, season)
                if target_date != lesson_date
                and not any(
                    item.task.subject == subject
                    and item.candidate.slots[0].date == target_date
                    for item in generated_assignments
                )
            ]
            if not target_dates:
                continue
            for duplicate in duplicate_items:
                duplicate_slot = duplicate.candidate.slots[0]
                duplicate_teacher = candidate_teacher_key(duplicate.candidate)
                donor_items = [
                    item
                    for item in generated_assignments
                    if item.task.task_id != duplicate.task.task_id
                    and item.task.subject == subject
                    and (item.task.stage or "") == (duplicate.task.stage or "")
                    and item.candidate.slots[0].date not in {lesson_date, *target_dates}
                ]
                donor_items.sort(
                    key=lambda item: (
                        candidate_teacher_key(item.candidate) == duplicate_teacher,
                        abs(
                            (
                                Date.fromisoformat(item.candidate.slots[0].date)
                                - Date.fromisoformat(duplicate_slot.date)
                            ).days
                        ),
                        item.candidate.slots[0].date,
                        scheduler.period_sort_value(item.candidate.slots[0].period),
                    )
                )
                for donor in donor_items:
                    donor_slot = donor.candidate.slots[0]
                    for target_date in target_dates:
                        for donor_period in ("AM", "PM"):
                            moved_donor = raw_moved_assignment(donor, target_date, donor_period)
                            if moved_donor is None:
                                continue
                            for duplicate_period in ("AM", "PM"):
                                moved_duplicate = raw_moved_assignment(
                                    duplicate,
                                    donor_slot.date,
                                    duplicate_period,
                                )
                                if moved_duplicate is None:
                                    continue
                                other_generated = [
                                    item
                                    for item in generated_assignments
                                    if item.task.task_id
                                    not in {duplicate.task.task_id, donor.task.task_id}
                                ]
                                trial_generated = scheduler.sorted_assignments(
                                    [*other_generated, moved_duplicate, moved_donor]
                                )
                                if not stage_sequence_valid(trial_generated):
                                    continue
                                if assignments_conflicting_with_candidate(
                                    moved_donor,
                                    [*protected_assignments, *other_generated, moved_duplicate],
                                    class_conflict_groups,
                                ):
                                    continue
                                if assignments_conflicting_with_candidate(
                                    moved_duplicate,
                                    [*protected_assignments, *other_generated, moved_donor],
                                    class_conflict_groups,
                                ):
                                    continue
                                generated_assignments = trial_generated
                                spread_moves.append(
                                    f"2727 同科内部交换: {donor.task.subject}/"
                                    f"{donor.task.stage or ''}/{donor.task.course_module or ''} "
                                    f"{donor_slot.date} {donor_slot.period} -> "
                                    f"{moved_donor.candidate.slots[0].date} {moved_donor.candidate.slots[0].period}；"
                                    f"{duplicate.task.course_module or ''} "
                                    f"{duplicate_slot.date} {duplicate_slot.period} -> "
                                    f"{moved_duplicate.candidate.slots[0].date} {moved_duplicate.candidate.slots[0].period}"
                                )
                                exchanged = True
                                break
                            if exchanged:
                                break
                        if exchanged:
                            break
                    if exchanged:
                        break
                if exchanged:
                    break
            if exchanged:
                break
        if not exchanged:
            break

    remaining_pools = {
        (subject, stage): len(tasks)
        for (subject, stage), tasks in pools.items()
        if tasks
    }
    if remaining_pools:
        raise ValueError(f"2727 无忧暑英政顺序重排后仍有剩余课程: {remaining_pools}")

    summer_subject_counts: Dict[Date, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    autumn_subject_counts: Dict[Date, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for assignment in generated_assignments:
        first = assignment.candidate.slots[0]
        week = week_monday(first.date)
        if first.date < WYS_AUTUMN_START:
            summer_subject_counts[week][assignment.task.subject] += 1
        else:
            autumn_subject_counts[week][assignment.task.subject] += 1

    summer_issues: List[str] = []
    for week in summer_weeks:
        total = sum(summer_subject_counts.get(week, {}).values())
        if total > WYS_NO_MATH_SUMMER_WEEKLY_MAX:
            summer_issues.append(f"{week_display_label(week)} 合计 {total}")
        for subject, (subject_min, subject_max) in WYS_NO_MATH_SUMMER_SUBJECT_WEEKLY_BOUNDS.items():
            count = summer_subject_counts.get(week, {}).get(subject, 0)
            if count < subject_min or count > subject_max:
                summer_issues.append(f"{week_display_label(week)} {subject}{count}")
    autumn_issues: List[str] = []
    for week in autumn_weeks:
        total = sum(autumn_subject_counts.get(week, {}).values())
        if total < WYS_NO_MATH_AUTUMN_WEEKLY_MIN or total > WYS_NO_MATH_AUTUMN_WEEKLY_MAX:
            autumn_issues.append(f"{week_display_label(week)} 合计 {total}")
        for subject, (subject_min, subject_max) in WYS_NO_MATH_SUBJECT_WEEKLY_BOUNDS.items():
            count = autumn_subject_counts.get(week, {}).get(subject, 0)
            if count < subject_min or count > subject_max:
                autumn_issues.append(f"{week_display_label(week)} {subject}{count}")
    if summer_issues or autumn_issues:
        raise ValueError(
            "2727 无忧暑英政顺序重排周均衡校验失败: "
            + "；".join([*summer_issues, *autumn_issues])
        )

    combined = scheduler.sorted_assignments([*protected_assignments, *generated_assignments])
    conflict_lines = teacher_time_conflict_lines(combined)
    if conflict_lines:
        lines.append(
            "2727 无忧暑英政顺序重排后仍有老师硬冲突，交由后续全局修复: "
            + "；".join(conflict_lines[:5])
        )

    lines.append(
        "2727 无忧暑英政顺序重排: 已重建 80 节，暑假 7-8 月先排，"
        "每周英语 3、政治 3-4；剩余课时进入秋季，秋季每周 2 个半天并按周三/周末均衡分布"
    )
    lines.extend(spread_moves)
    first_items = sorted(
        generated_assignments,
        key=lambda assignment: (
            assignment.task.subject,
            assignment.candidate.slots[0].date,
            scheduler.period_sort_value(assignment.candidate.slots[0].period),
        ),
    )
    for subject in ("英语", "政治"):
        subject_first = next(
            (assignment for assignment in first_items if assignment.task.subject == subject),
            None,
        )
        if subject_first:
            slot = subject_first.candidate.slots[0]
            lines.append(
                f"2727 {subject}首课: {slot.date} {slot.period} "
                f"{subject_first.task.stage or ''}/{subject_first.task.course_module or ''} "
                f"{subject_first.candidate.teacher_name}"
            )
    return combined, lines


def replace_assignment_by_task_id(
    assignments: Sequence[scheduler.Assignment],
    replacement: scheduler.Assignment,
) -> List[scheduler.Assignment]:
    return [
        replacement if assignment.task.task_id == replacement.task.task_id else assignment
        for assignment in assignments
    ]


def assignment_dedupe_key(assignment: scheduler.Assignment) -> Tuple[object, ...]:
    first = assignment.candidate.slots[0]
    last = assignment.candidate.slots[-1]
    return (
        assignment.task.class_id,
        first.date,
        first.period,
        first.start_time,
        last.end_time,
        assignment.task.subject,
        assignment.task.quarter,
        assignment.task.stage,
        assignment.task.course_module,
        assignment.task.course_group,
        assignment.candidate.teacher_id,
        assignment.candidate.teacher_name,
        assignment.candidate.room_id,
    )


def deduplicate_assignments(assignments: Sequence[scheduler.Assignment]) -> List[scheduler.Assignment]:
    result: List[scheduler.Assignment] = []
    seen: Set[Tuple[object, ...]] = set()
    for assignment in scheduler.sorted_assignments(list(assignments)):
        key = assignment_dedupe_key(assignment)
        if key in seen:
            continue
        seen.add(key)
        result.append(assignment)
    return result


def class_teacher_day_loads_for_assignments(
    assignments: Sequence[scheduler.Assignment],
    excluded_task_id: str,
) -> Dict[Tuple[str, str, str], float]:
    loads: Dict[Tuple[str, str, str], float] = defaultdict(float)
    for assignment in assignments:
        if assignment.task.task_id == excluded_task_id:
            continue
        if assignment.task.subject not in {"英语", "政治", "数学"}:
            continue
        teacher_key = candidate_teacher_key(assignment.candidate)
        if not teacher_key:
            continue
        for date_text, hours in candidate_hours_by_date(assignment.candidate).items():
            loads[(assignment.task.class_id, teacher_key, date_text)] += hours
    return loads


def candidate_within_window_boundary(
    candidate: scheduler.Candidate,
    constraint: Any,
) -> bool:
    if not constraint.earliest_date or not constraint.latest_date:
        return True
    first = candidate.slots[0]
    last = candidate.slots[-1]
    first_key = (first.date, scheduler.period_sort_value(first.period))
    last_key = (last.date, scheduler.period_sort_value(last.period))
    start_key = (constraint.earliest_date, scheduler.period_sort_value(constraint.earliest_period or "AM"))
    end_key = (constraint.latest_date, scheduler.period_sort_value(constraint.latest_period or "EVENING"))
    return start_key <= first_key and last_key <= end_key


def assignment_window_constraint(
    assignment: scheduler.Assignment,
    class_metadata: Dict[str, Dict[str, str]],
    window_constraints: Dict[str, Any],
) -> Optional[Any]:
    class_constraint = window_constraints.get(assignment.task.class_id)
    if class_constraint:
        return class_constraint
    suite_code = suite_code_for_class(assignment.task.class_id, class_metadata)
    return window_constraints.get(suite_code)


def suite_window_constraint_items(
    suite_code: str,
    class_metadata: Dict[str, Dict[str, str]],
    window_constraints: Dict[str, Any],
) -> List[Any]:
    direct = window_constraints.get(suite_code)
    if direct:
        return [direct]
    return [
        constraint
        for class_id, constraint in window_constraints.items()
        if class_metadata.get(class_id, {}).get("suite_code") == suite_code
    ]


def assignment_sub_product(
    assignment: scheduler.Assignment,
    class_metadata: Dict[str, Dict[str, str]],
) -> str:
    return class_metadata.get(assignment.task.class_id, {}).get("sub_product", "")


def public_product_move_rank(
    assignment: scheduler.Assignment,
    class_metadata: Dict[str, Dict[str, str]],
) -> Tuple[int, int, int, str]:
    sub_product = assignment_sub_product(assignment, class_metadata)
    try:
        product_rank = PUBLIC_PRODUCT_ORDER.index(sub_product)
    except ValueError:
        product_rank = len(PUBLIC_PRODUCT_ORDER)
    suite_code = suite_code_for_class(assignment.task.class_id, class_metadata)
    try:
        suite_number = int(suite_code)
    except ValueError:
        suite_number = 0
    subject_rank = PUBLIC_SUBJECT_PRIORITY.get(assignment.task.subject, 99)
    return product_rank, suite_number, subject_rank, assignment.task.class_id


def imported_maintenance_task(task: scheduler.CourseBlock) -> bool:
    return task.task_id.startswith(("SUMMER:", "FAST_BASE:", "HISTORY:", "LOCKED:"))


def compact_course_module(value: Optional[str]) -> str:
    text = clean(value)
    for prefix in ("寒假", "春季", "暑假", "秋季", "基础", "强化", "冲刺", "一轮", "二轮", "三轮", "四轮"):
        if text.startswith(prefix) and len(text) > len(prefix):
            return text[len(prefix) :]
    return text


def inferred_stage_from_module(value: Optional[str]) -> str:
    text = clean(value)
    for prefix in ("寒假", "春季", "暑假", "秋季", "基础", "强化", "冲刺", "一轮", "二轮", "三轮", "四轮"):
        if text.startswith(prefix):
            return prefix
    return ""


def repair_task_match_score(source_task: scheduler.CourseBlock, assignment: scheduler.Assignment) -> Optional[Tuple[int, str]]:
    task = assignment.task
    if source_task.subject != task.subject:
        return None
    if source_task.block_hours != task.block_hours:
        return None
    assignment_stage = clean(task.stage)
    source_stage = clean(source_task.stage)
    if assignment_stage and source_stage and assignment_stage != source_stage:
        return None
    if candidate_teacher_key(
        scheduler.Candidate((), source_task.teacher_id, source_task.teacher_name, "")
    ) != candidate_teacher_key(assignment.candidate):
        return None
    assignment_group = clean(task.course_group)
    source_group = clean(source_task.course_group)
    if assignment_group and source_group and assignment_group != source_group:
        return None
    assignment_module = compact_course_module(task.course_module)
    source_module = compact_course_module(source_task.course_module)
    if assignment_module and source_module and assignment_module != source_module:
        return None
    inferred_stage = inferred_stage_from_module(task.course_module)
    score = 0
    if inferred_stage and clean(source_task.stage) == inferred_stage:
        score -= 5
    if assignment_module and source_module:
        score -= 3
    if assignment_group and source_group:
        score -= 1
    return score, source_task.task_id


def flexible_repair_task_for_assignment(
    assignment: scheduler.Assignment,
    repair_input: scheduler.ScheduleInput,
    class_metadata: Dict[str, Dict[str, str]],
) -> scheduler.CourseBlock:
    source_tasks = scheduler.build_course_blocks(repair_input.classes)
    scored = [
        (score, source_task)
        for source_task in source_tasks
        for score in [repair_task_match_score(source_task, assignment)]
        if score is not None
    ]
    if not scored:
        return assignment.task
    scored.sort(key=lambda item: item[0])
    matched = scored[0][1]
    room_ids = matched.room_ids
    sub_product = assignment_sub_product(assignment, class_metadata)
    if (
        sub_product in SUMMER_FAST_PRODUCTS
        and assignment_matches_phase(assignment, SUMMER_PREPLAN_STAGES)
        and assignment.candidate.room_id
    ):
        room_ids = {assignment.candidate.room_id}
    return replace(
        matched,
        task_id=assignment.task.task_id,
        class_name=assignment.task.class_name,
        product_id=assignment.task.product_id,
        product_name=assignment.task.product_name,
        subject_category=assignment.task.subject_category or matched.subject_category,
        teacher_id=assignment.candidate.teacher_id or matched.teacher_id,
        teacher_name=assignment.candidate.teacher_name or matched.teacher_name,
        room_ids=room_ids,
        is_locked=False,
    )


def repaired_display_task(
    original_task: scheduler.CourseBlock,
    repair_task: scheduler.CourseBlock,
) -> scheduler.CourseBlock:
    return replace(
        original_task,
        start_date=repair_task.start_date,
        end_date=repair_task.end_date,
        allowed_periods=repair_task.allowed_periods,
        allowed_weekdays=repair_task.allowed_weekdays,
        excluded_weekdays=repair_task.excluded_weekdays,
        schedule_rules=repair_task.schedule_rules,
        room_ids=repair_task.room_ids,
        is_locked=False,
    )


def repair_schedule_input_for_class(
    data_dir: Path,
    class_id: str,
    class_metadata: Dict[str, Dict[str, str]],
) -> scheduler.ScheduleInput:
    source = load_schedule_input_for_classes(data_dir, [class_id])
    sub_product = class_metadata.get(class_id, {}).get("sub_product", "")
    if sub_product == WYS_PRODUCT:
        source = with_wuyou_summer_stage_windows(source)
        source = with_wuyou_summer_autumn_slots(source, include_wed_pm=True, include_wed_am=True)
    elif sub_product in WYQC_PRODUCTS:
        source = with_wuyou_qc_stage_windows(source)
    if sub_product in WUYOU_PRODUCTS:
        source = without_dates(source, WUYOU_PRODUCT_BLACKOUT_DATES)
    return source


def repair_candidate_allowed_by_outer_rules(
    assignment: scheduler.Assignment,
    candidate: scheduler.Candidate,
    class_metadata: Dict[str, Dict[str, str]],
    window_constraints: Dict[str, Any],
    blackout_dates: Set[str],
) -> bool:
    candidate_dates = {slot.date for slot in candidate.slots}
    if candidate_dates & blackout_dates:
        return False
    sub_product = assignment_sub_product(assignment, class_metadata)
    if sub_product in WUYOU_PRODUCTS and candidate_dates & WUYOU_PRODUCT_BLACKOUT_DATES:
        return False
    constraint = assignment_window_constraint(assignment, class_metadata, window_constraints)
    if (
        sub_product in SUMMER_FAST_PRODUCTS
        and assignment_matches_phase(assignment, SUMMER_PREPLAN_STAGES)
        and constraint
        and not candidate_within_window_boundary(candidate, constraint)
    ):
        return False
    return True


def wyqc_autumn_subject_week_limit(
    assignment: scheduler.Assignment,
    class_metadata: Dict[str, Dict[str, str]],
) -> Optional[int]:
    sub_product = assignment_sub_product(assignment, class_metadata)
    if sub_product not in WYQC_PRODUCTS:
        return None
    first = assignment.candidate.slots[0]
    if first.date < WYQC_AUTUMN_START or first.date > WYQC_AUTUMN_END:
        return None
    if assignment.task.subject not in SUMMER_PUBLIC_SUBJECTS:
        return None
    # 初排仍优先每科每周 1 个半天；老师硬冲突修复阶段最多允许
    # 同科同周 2 个半天，避免为了消冲突退回到 3 个半天以上的集中堆积。
    return 2


def assignment_week_balance_key(
    assignment: scheduler.Assignment,
    class_metadata: Dict[str, Dict[str, str]],
) -> Tuple[str, str, str, str, str, str, str, str, str]:
    first = assignment.candidate.slots[0]
    suite_code = suite_code_for_class(assignment.task.class_id, class_metadata)
    teacher_key = candidate_teacher_key(assignment.candidate)
    return (
        suite_code,
        assignment.task.class_id,
        assignment.task.subject,
        first.date,
        first.period,
        assignment.task.stage or "",
        assignment.task.course_module or "",
        assignment.task.course_group or "",
        teacher_key,
    )


def candidate_preserves_wyqc_autumn_subject_week_limit(
    assignment: scheduler.Assignment,
    candidate: scheduler.Candidate,
    other_assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
) -> bool:
    candidate_assignment = scheduler.Assignment(task=assignment.task, candidate=candidate)
    limit = wyqc_autumn_subject_week_limit(candidate_assignment, class_metadata)
    if limit is None:
        return True
    first = candidate.slots[0]
    candidate_week = week_key(candidate.slots)
    suite_code = suite_code_for_class(assignment.task.class_id, class_metadata)
    subject = assignment.task.subject
    existing_keys = {
        assignment_week_balance_key(other, class_metadata)
        for other in other_assignments
        if suite_code_for_class(other.task.class_id, class_metadata) == suite_code
        and other.task.subject == subject
        and other.candidate.slots
        and WYQC_AUTUMN_START <= other.candidate.slots[0].date <= WYQC_AUTUMN_END
        and week_key(other.candidate.slots) == candidate_week
    }
    candidate_key = assignment_week_balance_key(candidate_assignment, class_metadata)
    if candidate_key in existing_keys:
        return True
    return len(existing_keys) < limit


def repair_preserves_observed_stage_order(
    schedule_input: scheduler.ScheduleInput,
    current_assignments: Sequence[scheduler.Assignment],
    assignment_index: int,
    task: scheduler.CourseBlock,
    candidate: scheduler.Candidate,
) -> bool:
    cls = schedule_input.classes.get(task.class_id)
    if not cls or not task.stage:
        return True
    rank = cls.stage_order.get(task.stage)
    if rank is None:
        return True

    candidate_start = scheduler.slot_sort_key(candidate.slots[0])
    candidate_end = scheduler.slot_sort_key(candidate.slots[-1])
    for index, assignment in enumerate(current_assignments):
        if index == assignment_index or assignment.task.class_id != task.class_id:
            continue
        other_stage = assignment.task.stage
        if not other_stage:
            continue
        other_rank = cls.stage_order.get(other_stage)
        if other_rank is None or other_rank == rank:
            continue
        other_start = scheduler.slot_sort_key(assignment.candidate.slots[0])
        other_end = scheduler.slot_sort_key(assignment.candidate.slots[-1])
        if other_rank < rank and candidate_start <= other_end:
            return False
        if other_rank > rank and candidate_end >= other_start:
            return False
    return True


def class_stage_order_violation_count(
    schedule_input: scheduler.ScheduleInput,
    assignments: Sequence[scheduler.Assignment],
    class_id: str,
) -> int:
    cls = schedule_input.classes.get(class_id)
    if not cls or not cls.stage_order:
        return 0
    relevant: List[Tuple[int, Tuple[str, int, str], Tuple[str, int, str]]] = []
    for assignment in assignments:
        if assignment.task.class_id != class_id or not assignment.task.stage:
            continue
        rank = cls.stage_order.get(assignment.task.stage)
        if rank is None:
            continue
        relevant.append(
            (
                rank,
                scheduler.slot_sort_key(assignment.candidate.slots[0]),
                scheduler.slot_sort_key(assignment.candidate.slots[-1]),
            )
        )
    violation_count = 0
    for left_index, (left_rank, _left_start, left_end) in enumerate(relevant):
        for right_rank, right_start, _right_end in relevant[left_index + 1 :]:
            if left_rank < right_rank and left_end >= right_start:
                violation_count += 1
            elif right_rank < left_rank and _right_end >= _left_start:
                violation_count += 1
    return violation_count


def calendar_weeks_between(start: str, end: str) -> Set[Tuple[int, int]]:
    if not start or not end:
        return set()
    current = Date.fromisoformat(start)
    final = Date.fromisoformat(end)
    weeks: Set[Tuple[int, int]] = set()
    while current <= final:
        iso = current.isocalendar()
        weeks.add((iso.year, iso.week))
        current += timedelta(days=1)
    return weeks


def summer_teacher_repair_week_balance_key(
    assignment: scheduler.Assignment,
    candidate: scheduler.Candidate,
    other_assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    window_constraints: Dict[str, Any],
) -> Tuple[int, int, int, int, int]:
    sub_product = assignment_sub_product(assignment, class_metadata)
    if (
        sub_product not in SUMMER_FAST_PRODUCTS
        or assignment.task.subject not in SUMMER_PUBLIC_SUBJECTS
        or not assignment_matches_phase(assignment, SUMMER_PREPLAN_STAGES)
    ):
        return (0, 0, 0, 0, 0)

    suite_code = suite_code_for_class(assignment.task.class_id, class_metadata)
    if not suite_code:
        return (0, 0, 0, 0, 0)
    active_weeks = set()
    for constraint in suite_window_constraint_items(suite_code, class_metadata, window_constraints):
        active_weeks.update(calendar_weeks_between(constraint.earliest_date, constraint.latest_date))

    subject_counts: Dict[str, int] = defaultdict(int)
    subject_week_counts: Dict[Tuple[str, Tuple[int, int]], int] = defaultdict(int)
    total_week_counts: Dict[Tuple[int, int], int] = defaultdict(int)
    total_count = 1
    candidate_week = week_key(candidate.slots)
    active_weeks.add(candidate_week)

    for item in other_assignments:
        if item.task.subject not in SUMMER_PUBLIC_SUBJECTS:
            continue
        if not assignment_matches_phase(item, SUMMER_PREPLAN_STAGES):
            continue
        if suite_code_for_class(item.task.class_id, class_metadata) != suite_code:
            continue
        week = week_key(item.candidate.slots)
        active_weeks.add(week)
        subject_counts[item.task.subject] += 1
        subject_week_counts[(item.task.subject, week)] += 1
        total_week_counts[week] += 1
        total_count += 1

    subject_counts[assignment.task.subject] += 1
    week_count = max(1, len(active_weeks))
    subject_target = max(1, (subject_counts[assignment.task.subject] + week_count - 1) // week_count)
    total_target = max(1, (total_count + week_count - 1) // week_count)
    subject_after = subject_week_counts[(assignment.task.subject, candidate_week)] + 1
    total_after = total_week_counts[candidate_week] + 1
    subject_over = max(0, subject_after - subject_target)
    total_over = max(0, total_after - total_target)
    return (
        subject_over,
        total_over,
        subject_after,
        total_after,
        0 if subject_after >= max(1, subject_counts[assignment.task.subject] // week_count) else 1,
    )


def find_teacher_conflict_repair_assignment(
    data_dir: Path,
    assignment: scheduler.Assignment,
    assignment_index: int,
    current_assignments: Sequence[scheduler.Assignment],
    protected_assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    window_constraints: Dict[str, Any],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
    input_cache: Dict[str, scheduler.ScheduleInput],
) -> Optional[scheduler.Assignment]:
    class_id = assignment.task.class_id
    if class_id not in input_cache:
        input_cache[class_id] = repair_schedule_input_for_class(
            data_dir,
            class_id,
            class_metadata,
        )
    repair_input = input_cache[class_id]
    repair_task = flexible_repair_task_for_assignment(assignment, repair_input, class_metadata)
    other_assignments = [
        *protected_assignments,
        *[
            item
            for index, item in enumerate(current_assignments)
            if index != assignment_index
        ],
    ]
    other_by_date_period: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
    for item in other_assignments:
        for key in assignment_date_period_keys(item):
            other_by_date_period[key].append(item)
    day_loads = class_teacher_day_loads_for_assignments(other_assignments, assignment.task.task_id)
    original_first = assignment.candidate.slots[0]
    original_date = Date.fromisoformat(original_first.date)
    original_slot_key = scheduler.slot_sort_key(original_first)
    try:
        candidates = scheduler.candidate_assignments(repair_task, repair_input)
    except (KeyError, ValueError):
        return None
    candidates.sort(
        key=lambda candidate: (
            summer_teacher_repair_week_balance_key(
                assignment,
                candidate,
                other_assignments,
                class_metadata,
                window_constraints,
            ),
            candidate.slots[0].period != original_first.period,
            abs((Date.fromisoformat(candidate.slots[0].date) - original_date).days),
            scheduler.candidate_same_day_teacher_travel_penalty(
                repair_input,
                other_assignments,
                repair_task,
                candidate,
            ),
            candidate.room_id != assignment.candidate.room_id,
            scheduler.slot_sort_key(candidate.slots[0]),
        )
    )
    same_class_teacher_day_fallback: Optional[scheduler.Assignment] = None
    for candidate in candidates:
        if scheduler.slot_sort_key(candidate.slots[0]) == original_slot_key and candidate.room_id == assignment.candidate.room_id:
            continue
        if not repair_candidate_allowed_by_outer_rules(
            assignment,
            candidate,
            class_metadata,
            window_constraints,
            blackout_dates,
        ):
            continue
        skip_sequence_checks = imported_maintenance_task(assignment.task)
        if not skip_sequence_checks:
            if not replacement_preserves_first_lesson_module(
                current_assignments,
                assignment_index,
                repair_task,
                candidate,
            ):
                continue
            if not repair_preserves_observed_stage_order(
                repair_input,
                current_assignments,
                assignment_index,
                repair_task,
                candidate,
            ):
                continue
        violates_same_class_teacher_day = not candidate_avoids_same_class_teacher_day_limit(
            day_loads,
            repair_task,
            candidate,
        )
        candidate_assignment = scheduler.Assignment(
            task=repaired_display_task(assignment.task, repair_task),
            candidate=candidate,
        )
        if not candidate_preserves_wyqc_autumn_subject_week_limit(
            candidate_assignment,
            candidate,
            other_assignments,
            class_metadata,
        ):
            continue
        if candidate_conflicts_for_repair(candidate_assignment, other_by_date_period, class_conflict_groups):
            continue
        if violates_same_class_teacher_day:
            if same_class_teacher_day_fallback is None:
                same_class_teacher_day_fallback = candidate_assignment
            continue
        return candidate_assignment
    return same_class_teacher_day_fallback


def repair_global_teacher_time_conflicts(
    data_dir: Path,
    movable_assignments: Sequence[scheduler.Assignment],
    protected_assignments: Sequence[scheduler.Assignment],
    class_conflict_groups: Dict[str, Set[str]],
    max_moves: int = 300,
    max_seconds: int = GLOBAL_REPAIR_MAX_SECONDS,
) -> Tuple[List[scheduler.Assignment], List[str]]:
    result = scheduler.sorted_assignments(list(movable_assignments))
    class_metadata = load_class_metadata(data_dir)
    window_constraints = load_summer_class_window_constraints(data_dir)
    blackout_dates = load_active_blackout_dates(data_dir)
    input_cache: Dict[str, scheduler.ScheduleInput] = {}
    initial_conflicts = teacher_time_conflict_groups([*protected_assignments, *result])
    if not initial_conflicts:
        return result, ["全局老师同时间硬冲突修复: 初始无硬冲突"]

    lines: List[str] = []
    moved_count = 0
    started_at = time.perf_counter()
    stopped_by_time = False
    for _pass_index in range(max_moves):
        if max_seconds > 0 and time.perf_counter() - started_at > max_seconds:
            stopped_by_time = True
            lines.append(f"达到全局修复时间上限 {max_seconds} 秒，停止继续搜索")
            break
        conflict_groups = teacher_time_conflict_groups([*protected_assignments, *result])
        if not conflict_groups:
            break
        index_by_task_id = {
            assignment.task.task_id: index
            for index, assignment in enumerate(result)
        }
        moved = False
        for group in conflict_groups:
            movable_group = [
                (index_by_task_id[item.task.task_id], item)
                for item in group
                if item.task.task_id in index_by_task_id
            ]
            if not movable_group:
                continue
            movable_group.sort(
                key=lambda pair: public_product_move_rank(pair[1], class_metadata),
                reverse=True,
            )
            for assignment_index, assignment in movable_group:
                repaired_assignment = find_teacher_conflict_repair_assignment(
                    data_dir,
                    assignment,
                    assignment_index,
                    result,
                    protected_assignments,
                    class_metadata,
                    window_constraints,
                    class_conflict_groups,
                    blackout_dates,
                    input_cache,
                )
                if repaired_assignment is None:
                    continue
                old_first = assignment.candidate.slots[0]
                new_first = repaired_assignment.candidate.slots[0]
                result[assignment_index] = repaired_assignment
                result = scheduler.sorted_assignments(result)
                moved_count += 1
                moved = True
                lines.append(
                    f"{assignment.task.class_id} {assignment.task.subject}/{assignment.task.course_module or ''}: "
                    f"{old_first.date} {old_first.period} -> {new_first.date} {new_first.period}"
                )
                break
            if moved:
                break
        if not moved:
            break

    final_conflicts = teacher_time_conflict_groups([*protected_assignments, *result])
    summary = (
        f"全局老师同时间硬冲突修复: {len(initial_conflicts)} -> {len(final_conflicts)} 组，"
        f"移动 {moved_count} 节"
        + ("，因时间上限停止" if stopped_by_time else "")
    )
    if final_conflicts:
        samples = []
        for group in final_conflicts[:8]:
            first = group[0].candidate.slots[0]
            teacher = group[0].candidate.teacher_name or group[0].candidate.teacher_id
            samples.append(
                f"{first.date} {first.period} {teacher}: "
                + " / ".join(item.task.class_id for item in group)
            )
        lines.append("仍有硬冲突未自动修复: " + "；".join(samples))
    return result, [summary, *lines[:200]]


def rebuild_class_subject_segment_for_conflicts(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_id: str,
    subject: str,
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
    target_phase: Optional[str] = None,
) -> Optional[List[scheduler.Assignment]]:
    target_assignments = [
        assignment
        for assignment in assignments
        if assignment.task.class_id == class_id
        and assignment.task.subject == subject
        and movable_public_experience_assignment(assignment, class_metadata)
        and (
            target_phase is None
            or student_experience_phase(
                assignment.candidate.slots[0].date,
                assignment_sub_product(assignment, class_metadata),
            )
            == target_phase
        )
    ]
    if not target_assignments:
        return None
    repair_input = repair_schedule_input_for_class(data_dir, class_id, class_metadata)
    cls = repair_input.classes.get(class_id)
    stage_rank = cls.stage_order if cls else {}

    def target_sort_key(assignment: scheduler.Assignment) -> Tuple[int, Tuple[str, int, str], str]:
        rank = stage_rank.get(assignment.task.stage or "", 99)
        return rank, scheduler.slot_sort_key(assignment.candidate.slots[0]), assignment.task.task_id

    ordered_targets = sorted(target_assignments, key=target_sort_key)
    target_task_ids = {assignment.task.task_id for assignment in target_assignments}
    protected_assignments = [
        assignment
        for assignment in assignments
        if assignment.task.task_id not in target_task_ids
    ]
    window_constraints = load_summer_class_window_constraints(data_dir)

    def build_period_index(items: Sequence[scheduler.Assignment]) -> Dict[Tuple[str, str], List[scheduler.Assignment]]:
        by_date_period: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
        for item in items:
            for key in assignment_date_period_keys(item):
                by_date_period[key].append(item)
        return by_date_period

    def attempt(
        avoid_same_subject_day: bool,
        enforce_same_class_teacher_day: bool,
        keep_chronological: bool,
    ) -> Optional[List[scheduler.Assignment]]:
        placed: List[scheduler.Assignment] = []
        other_by_date_period = build_period_index(protected_assignments)
        same_subject_dates = {
            item.candidate.slots[0].date
            for item in protected_assignments
            if item.task.class_id == class_id and item.task.subject == subject
        }
        last_slot_key: Optional[Tuple[str, int, str]] = None

        for original in ordered_targets:
            repair_task = flexible_repair_task_for_assignment(original, repair_input, class_metadata)
            try:
                candidates = scheduler.candidate_assignments(repair_task, repair_input)
            except (KeyError, ValueError):
                return None
            original_first = original.candidate.slots[0]
            suite_code = suite_code_for_class(class_id, class_metadata)

            def candidate_week_load(candidate: scheduler.Candidate) -> int:
                target_week = week_monday(candidate.slots[0].date)
                return sum(
                    1
                    for item in [*protected_assignments, *placed]
                    if suite_code_for_class(item.task.class_id, class_metadata) == suite_code
                    and item.task.subject in SUMMER_PUBLIC_SUBJECTS
                    and week_monday(item.candidate.slots[0].date) == target_week
                )

            candidates.sort(
                key=lambda candidate: (
                    candidate_week_load(candidate),
                    abs((Date.fromisoformat(candidate.slots[0].date) - Date.fromisoformat(original_first.date)).days),
                    candidate.slots[0].period != original_first.period,
                    scheduler.candidate_same_day_teacher_travel_penalty(
                        repair_input,
                        [*protected_assignments, *placed],
                        repair_task,
                        candidate,
                    ),
                    candidate.room_id != original.candidate.room_id,
                    scheduler.slot_sort_key(candidate.slots[0]),
                )
            )

            selected: Optional[scheduler.Assignment] = None
            for candidate in candidates:
                first = candidate.slots[0]
                candidate_key = scheduler.slot_sort_key(first)
                if keep_chronological and last_slot_key and candidate_key < last_slot_key:
                    continue
                if avoid_same_subject_day and first.date in same_subject_dates:
                    continue
                if not repair_candidate_allowed_by_outer_rules(
                    original,
                    candidate,
                    class_metadata,
                    window_constraints,
                    blackout_dates,
                ):
                    continue
                day_loads = class_teacher_day_loads_for_assignments(
                    [*protected_assignments, *placed],
                    repair_task.task_id,
                )
                if enforce_same_class_teacher_day and not candidate_avoids_same_class_teacher_day_limit(
                    day_loads,
                    repair_task,
                    candidate,
                ):
                    continue
                candidate_assignment = scheduler.Assignment(
                    task=repaired_display_task(original.task, repair_task),
                    candidate=candidate,
                )
                if candidate_conflicts_for_repair(candidate_assignment, other_by_date_period, class_conflict_groups):
                    continue
                selected = candidate_assignment
                break

            if selected is None:
                return None
            placed.append(selected)
            same_subject_dates.add(selected.candidate.slots[0].date)
            last_slot_key = scheduler.slot_sort_key(selected.candidate.slots[0])
            for key in assignment_date_period_keys(selected):
                other_by_date_period[key].append(selected)

        return scheduler.sorted_assignments([*protected_assignments, *placed])

    for avoid_same_subject_day, enforce_day_load, keep_chronological in (
        (True, True, True),
        (False, True, True),
        (False, False, True),
        (False, False, False),
    ):
        rebuilt = attempt(avoid_same_subject_day, enforce_day_load, keep_chronological)
        if rebuilt is not None:
            return rebuilt
    return None


def repair_politics_segments_for_teacher_conflicts(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
    max_passes: int = 12,
) -> Tuple[List[scheduler.Assignment], List[str]]:
    result = scheduler.sorted_assignments(list(assignments))
    lines: List[str] = []
    moved_segments = 0

    for _pass_index in range(max_passes):
        conflict_groups = teacher_time_conflict_groups(result)
        target_keys = {
            (
                assignment.task.class_id,
                student_experience_phase(
                    assignment.candidate.slots[0].date,
                    assignment_sub_product(assignment, class_metadata),
                )
                or "",
            )
            for group in conflict_groups
            for assignment in group
            if assignment.task.subject == "政治"
            and movable_public_experience_assignment(assignment, class_metadata)
        }
        if not target_keys:
            break
        before_count = len(conflict_groups)
        ordered_targets = sorted(
            target_keys,
            key=lambda item: (
                public_product_move_rank(
                    next(
                        assignment
                        for assignment in result
                        if assignment.task.class_id == item[0] and assignment.task.subject == "政治"
                    ),
                    class_metadata,
                ),
                suite_code_for_class(item[0], class_metadata),
                item[0],
            ),
            reverse=True,
        )
        improved = False
        for class_id, phase in ordered_targets:
            rebuilt = rebuild_class_subject_segment_for_conflicts(
                data_dir,
                result,
                class_id,
                "政治",
                class_metadata,
                class_conflict_groups,
                blackout_dates,
                target_phase=phase or None,
            )
            if rebuilt is None:
                continue
            after_count = len(teacher_time_conflict_groups(rebuilt))
            if after_count >= before_count:
                continue
            result = rebuilt
            moved_segments += 1
            improved = True
            lines.append(
                f"{class_id} 政治{phase or ''}段重排: 老师硬冲突 {before_count} -> {after_count}"
            )
            break
        if not improved:
            break

    final_count = len(teacher_time_conflict_groups(result))
    if moved_segments or final_count:
        lines.insert(0, f"政治整段重排消冲突: 重排 {moved_segments} 个班级政治段，剩余硬冲突 {final_count} 组")
    return result, lines[:160]


def movable_public_experience_assignment(
    assignment: scheduler.Assignment,
    class_metadata: Dict[str, Dict[str, str]],
) -> bool:
    if assignment.task.task_id.startswith(("HISTORY:", "LOCKED:")):
        return False
    meta = class_metadata.get(assignment.task.class_id, {})
    return is_public_schedulable_meta(meta) and assignment.task.subject in SUMMER_PUBLIC_SUBJECTS


def find_public_same_subject_day_repair_assignment(
    data_dir: Path,
    assignment: scheduler.Assignment,
    assignment_index: int,
    current_assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    window_constraints: Dict[str, Any],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
    input_cache: Dict[str, scheduler.ScheduleInput],
) -> Optional[scheduler.Assignment]:
    class_id = assignment.task.class_id
    if class_id not in input_cache:
        input_cache[class_id] = repair_schedule_input_for_class(
            data_dir,
            class_id,
            class_metadata,
        )
    repair_input = input_cache[class_id]
    repair_task = flexible_repair_task_for_assignment(assignment, repair_input, class_metadata)
    other_assignments = [
        item
        for index, item in enumerate(current_assignments)
        if index != assignment_index
    ]
    other_by_date_period: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
    for item in other_assignments:
        for key in assignment_date_period_keys(item):
            other_by_date_period[key].append(item)
    day_loads = class_teacher_day_loads_for_assignments(other_assignments, assignment.task.task_id)
    original_first = assignment.candidate.slots[0]
    original_date = Date.fromisoformat(original_first.date)
    original_week = week_monday(original_first.date)
    suite_code = suite_code_for_class(class_id, class_metadata)

    def same_class_subject_on_date(candidate: scheduler.Candidate) -> bool:
        candidate_dates = {slot.date for slot in candidate.slots}
        for item in other_assignments:
            if item.task.class_id != class_id or item.task.subject != assignment.task.subject:
                continue
            if candidate_dates & {slot.date for slot in item.candidate.slots}:
                return True
        return False

    def suite_week_load(candidate: scheduler.Candidate) -> int:
        target_week = week_monday(candidate.slots[0].date)
        count = 0
        for item in other_assignments:
            if suite_code_for_class(item.task.class_id, class_metadata) != suite_code:
                continue
            if item.task.subject not in SUMMER_PUBLIC_SUBJECTS:
                continue
            if week_monday(item.candidate.slots[0].date) == target_week:
                count += 1
        return count

    try:
        candidates = scheduler.candidate_assignments(repair_task, repair_input)
    except (KeyError, ValueError):
        return None
    candidates.sort(
        key=lambda candidate: (
            week_monday(candidate.slots[0].date) != original_week,
            suite_week_load(candidate),
            abs((Date.fromisoformat(candidate.slots[0].date) - original_date).days),
            candidate.slots[0].period != original_first.period,
            scheduler.candidate_same_day_teacher_travel_penalty(
                repair_input,
                other_assignments,
                repair_task,
                candidate,
            ),
            candidate.room_id != assignment.candidate.room_id,
            scheduler.slot_sort_key(candidate.slots[0]),
        )
    )
    for candidate in candidates:
        if candidate.slots[0].date == original_first.date:
            continue
        if same_class_subject_on_date(candidate):
            continue
        if not repair_candidate_allowed_by_outer_rules(
            assignment,
            candidate,
            class_metadata,
            window_constraints,
            blackout_dates,
        ):
            continue
        if not replacement_preserves_first_lesson_module(
            current_assignments,
            assignment_index,
            repair_task,
            candidate,
        ):
            continue
        if not repair_preserves_observed_stage_order(
            repair_input,
            current_assignments,
            assignment_index,
            repair_task,
            candidate,
        ):
            continue
        if not candidate_avoids_same_class_teacher_day_limit(day_loads, repair_task, candidate):
            continue
        candidate_assignment = scheduler.Assignment(
            task=repaired_display_task(assignment.task, repair_task),
            candidate=candidate,
        )
        if candidate_conflicts_for_repair(candidate_assignment, other_by_date_period, class_conflict_groups):
            continue
        return candidate_assignment
    return None


def repair_public_same_subject_day_overloads(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
    target_class_ids: Optional[Set[str]] = None,
    max_moves: int = 120,
    max_seconds: int = 45,
) -> Tuple[List[scheduler.Assignment], List[str]]:
    result = scheduler.sorted_assignments(list(assignments))
    window_constraints = load_summer_class_window_constraints(data_dir)
    input_cache: Dict[str, scheduler.ScheduleInput] = {}
    lines: List[str] = []
    moved_count = 0
    started_at = time.perf_counter()
    stopped_by_time = False

    def duplicate_groups() -> List[Tuple[Tuple[str, str, str], List[scheduler.Assignment], float]]:
        grouped: Dict[Tuple[str, str, str], List[scheduler.Assignment]] = defaultdict(list)
        for item in result:
            if not movable_public_experience_assignment(item, class_metadata):
                continue
            if target_class_ids is not None and item.task.class_id not in target_class_ids:
                continue
            first = item.candidate.slots[0]
            if first.date < "2026-06-25" or first.date > AUTUMN_END:
                continue
            grouped[(item.task.class_id, item.task.subject, first.date)].append(item)
        duplicates: List[Tuple[Tuple[str, str, str], List[scheduler.Assignment], float]] = []
        for key, items in grouped.items():
            total_hours = sum(
                sum(slot.duration_hours for slot in item.candidate.slots)
                for item in items
            )
            if len(items) > 1 or total_hours >= PUBLIC_SAME_CLASS_SUBJECT_DAY_HOURS:
                duplicates.append((key, items, total_hours))
        duplicates.sort(
            key=lambda item: (
                -item[2],
                item[0][2],
                item[0][0],
                PUBLIC_SUBJECT_PRIORITY.get(item[0][1], 99),
            )
        )
        return duplicates

    for _pass_index in range(max_moves):
        if max_seconds > 0 and time.perf_counter() - started_at > max_seconds:
            stopped_by_time = True
            break
        duplicates = duplicate_groups()
        if not duplicates:
            break
        moved = False
        index_by_task_id = {
            assignment.task.task_id: index
            for index, assignment in enumerate(result)
        }
        for (_class_id, _subject, _date_text), items, _hours in duplicates:
            items = sorted(
                items,
                key=lambda item: (
                    scheduler.period_sort_value(item.candidate.slots[0].period),
                    item.candidate.slots[0].start_time or "",
                    item.task.task_id,
                ),
                reverse=True,
            )
            for assignment in items:
                assignment_index = index_by_task_id.get(assignment.task.task_id)
                if assignment_index is None:
                    continue
                repaired = find_public_same_subject_day_repair_assignment(
                    data_dir,
                    assignment,
                    assignment_index,
                    result,
                    class_metadata,
                    window_constraints,
                    class_conflict_groups,
                    blackout_dates,
                    input_cache,
                )
                if repaired is None:
                    continue
                old_first = assignment.candidate.slots[0]
                new_first = repaired.candidate.slots[0]
                result[assignment_index] = repaired
                result = scheduler.sorted_assignments(result)
                moved_count += 1
                moved = True
                lines.append(
                    f"公共课同班同科同日摊开: {assignment.task.class_id} "
                    f"{assignment.task.subject}/{assignment.task.stage or ''}/{assignment.task.course_module or ''} "
                    f"{old_first.date} {old_first.period} -> {new_first.date} {new_first.period}"
                )
                break
            if moved:
                break
        if not moved:
            break

    remaining = duplicate_groups()
    summary = (
        f"公共课同班同科同日 8 小时规避: 移动 {moved_count} 节，"
        f"剩余 {len(remaining)} 组需人工核对"
        + ("，因时间上限停止" if stopped_by_time else "")
    )
    if remaining:
        samples = []
        for (class_id, subject, date_text), _items, hours in remaining[:12]:
            suite_code = suite_code_for_class(class_id, class_metadata)
            samples.append(f"{suite_code} {class_id} {date_text} {subject} {hours:g} 小时")
        lines.append("仍有同班同科同日高负荷: " + "；".join(samples))
    return scheduler.sorted_assignments(result), [summary, *lines[:200]]


def repair_public_same_subject_day_overloads_with_internal_swap(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
    target_class_ids: Optional[Set[str]] = None,
    max_moves: int = 40,
    max_seconds: int = 30,
) -> Tuple[List[scheduler.Assignment], List[str]]:
    result = scheduler.sorted_assignments(list(assignments))
    window_constraints = load_summer_class_window_constraints(data_dir)
    input_cache: Dict[str, scheduler.ScheduleInput] = {}
    lines: List[str] = []
    moved_count = 0
    started_at = time.perf_counter()
    stopped_by_time = False

    def duplicate_groups() -> List[Tuple[Tuple[str, str, str], List[scheduler.Assignment], float]]:
        grouped: Dict[Tuple[str, str, str], List[scheduler.Assignment]] = defaultdict(list)
        for item in result:
            if not movable_public_experience_assignment(item, class_metadata):
                continue
            if target_class_ids is not None and item.task.class_id not in target_class_ids:
                continue
            first = item.candidate.slots[0]
            if first.date < "2026-06-25" or first.date > AUTUMN_END:
                continue
            grouped[(item.task.class_id, item.task.subject, first.date)].append(item)
        duplicates: List[Tuple[Tuple[str, str, str], List[scheduler.Assignment], float]] = []
        for key, items in grouped.items():
            total_hours = sum(
                sum(slot.duration_hours for slot in item.candidate.slots)
                for item in items
            )
            if len(items) > 1 or total_hours >= PUBLIC_SAME_CLASS_SUBJECT_DAY_HOURS:
                duplicates.append((key, items, total_hours))
        duplicates.sort(
            key=lambda item: (
                -item[2],
                item[0][2],
                item[0][0],
                PUBLIC_SUBJECT_PRIORITY.get(item[0][1], 99),
            )
        )
        return duplicates

    def sequence_valid_for_class(items: Sequence[scheduler.Assignment], class_id: str) -> bool:
        if class_id not in input_cache:
            input_cache[class_id] = repair_schedule_input_for_class(
                data_dir,
                class_id,
                class_metadata,
            )
        cls = input_cache[class_id].classes.get(class_id)
        if not cls:
            return True
        first_module_by_subject = {"英语": "词汇", "政治": "马原"}
        class_items = [item for item in items if item.task.class_id == class_id]
        for subject, first_module in first_module_by_subject.items():
            subject_items = sorted(
                [item for item in class_items if item.task.subject == subject],
                key=lambda item: scheduler.slot_sort_key(item.candidate.slots[0]),
            )
            if subject_items and subject_items[0].task.course_module != first_module:
                return False
        for subject in SUMMER_PUBLIC_SUBJECTS:
            max_rank = -1
            for item in sorted(
                [entry for entry in class_items if entry.task.subject == subject],
                key=lambda entry: scheduler.slot_sort_key(entry.candidate.slots[0]),
            ):
                rank = cls.stage_order.get(item.task.stage or "", 99)
                if rank < max_rank:
                    return False
                max_rank = max(max_rank, rank)
        return True

    def repaired_at(
        assignment: scheduler.Assignment,
        assignment_index: int,
        target_date: str,
        target_period: str,
        excluded_task_ids: Set[str],
        extra_assignments: Sequence[scheduler.Assignment] = (),
    ) -> Optional[scheduler.Assignment]:
        class_id = assignment.task.class_id
        if class_id not in input_cache:
            input_cache[class_id] = repair_schedule_input_for_class(
                data_dir,
                class_id,
                class_metadata,
            )
        repair_input = input_cache[class_id]
        repair_task = flexible_repair_task_for_assignment(assignment, repair_input, class_metadata)
        other_assignments = [
            item
            for item in result
            if item.task.task_id not in excluded_task_ids
        ]
        other_with_extra = [*other_assignments, *extra_assignments]
        other_by_date_period: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
        for item in other_with_extra:
            for key in assignment_date_period_keys(item):
                other_by_date_period[key].append(item)
        day_loads = class_teacher_day_loads_for_assignments(other_with_extra, assignment.task.task_id)
        try:
            candidates = scheduler.candidate_assignments(repair_task, repair_input)
        except (KeyError, ValueError):
            return None
        for candidate in candidates:
            first = candidate.slots[0]
            if first.date != target_date or first.period != target_period:
                continue
            if not repair_candidate_allowed_by_outer_rules(
                assignment,
                candidate,
                class_metadata,
                window_constraints,
                blackout_dates,
            ):
                continue
            if not replacement_preserves_first_lesson_module(
                result,
                assignment_index,
                repair_task,
                candidate,
            ):
                continue
            if not repair_preserves_observed_stage_order(
                repair_input,
                result,
                assignment_index,
                repair_task,
                candidate,
            ):
                continue
            if not candidate_avoids_same_class_teacher_day_limit(day_loads, repair_task, candidate):
                continue
            candidate_assignment = scheduler.Assignment(
                task=repaired_display_task(assignment.task, repair_task),
                candidate=candidate,
            )
            if candidate_conflicts_for_repair(candidate_assignment, other_by_date_period, class_conflict_groups):
                continue
            return candidate_assignment
        return None

    for _ in range(max_moves):
        if max_seconds > 0 and time.perf_counter() - started_at > max_seconds:
            stopped_by_time = True
            break
        duplicates = duplicate_groups()
        if not duplicates:
            break
        index_by_task_id = {
            assignment.task.task_id: index
            for index, assignment in enumerate(result)
        }
        moved = False
        for (_class_id, _subject, _date_text), items, _hours in duplicates:
            items = sorted(
                items,
                key=lambda item: (
                    scheduler.period_sort_value(item.candidate.slots[0].period),
                    item.candidate.slots[0].start_time or "",
                    item.task.task_id,
                ),
                reverse=True,
            )
            for assignment in items:
                assignment_index = index_by_task_id.get(assignment.task.task_id)
                if assignment_index is None:
                    continue
                class_id = assignment.task.class_id
                if class_id not in input_cache:
                    input_cache[class_id] = repair_schedule_input_for_class(
                        data_dir,
                        class_id,
                        class_metadata,
                    )
                repair_input = input_cache[class_id]
                repair_task = flexible_repair_task_for_assignment(assignment, repair_input, class_metadata)
                other_assignments = [
                    item
                    for item in result
                    if item.task.task_id != assignment.task.task_id
                ]
                same_class_subject_dates = {
                    item.candidate.slots[0].date
                    for item in other_assignments
                    if item.task.class_id == class_id
                    and item.task.subject == assignment.task.subject
                }
                original_first = assignment.candidate.slots[0]
                try:
                    candidates = scheduler.candidate_assignments(repair_task, repair_input)
                except (KeyError, ValueError):
                    continue
                candidates.sort(
                    key=lambda candidate: (
                        week_monday(candidate.slots[0].date) != week_monday(original_first.date),
                        abs((Date.fromisoformat(candidate.slots[0].date) - Date.fromisoformat(original_first.date)).days),
                        candidate.slots[0].period != original_first.period,
                        scheduler.slot_sort_key(candidate.slots[0]),
                    )
                )
                for candidate in candidates:
                    first = candidate.slots[0]
                    if first.date == original_first.date:
                        continue
                    if first.date in same_class_subject_dates:
                        continue
                    target_replacement = scheduler.Assignment(
                        task=repaired_display_task(assignment.task, repair_task),
                        candidate=candidate,
                    )
                    if not repair_candidate_allowed_by_outer_rules(
                        assignment,
                        candidate,
                        class_metadata,
                        window_constraints,
                        blackout_dates,
                    ):
                        continue
                    if not replacement_preserves_first_lesson_module(
                        result,
                        assignment_index,
                        repair_task,
                        candidate,
                    ):
                        continue
                    if not repair_preserves_observed_stage_order(
                        repair_input,
                        result,
                        assignment_index,
                        repair_task,
                        candidate,
                    ):
                        continue

                    conflicts = assignments_conflicting_with_candidate(
                        target_replacement,
                        other_assignments,
                        class_conflict_groups,
                    )
                    target_suite_code = suite_code_for_class(class_id, class_metadata)
                    internal_conflicts = [
                        conflict
                        for conflict in conflicts
                        if suite_code_for_class(conflict.task.class_id, class_metadata) == target_suite_code
                        and conflict.task.subject != assignment.task.subject
                    ]
                    if len(conflicts) != 1 or len(internal_conflicts) != 1:
                        continue
                    blocker = internal_conflicts[0]
                    blocker_index = index_by_task_id.get(blocker.task.task_id)
                    if blocker_index is None:
                        continue
                    blocker_same_subject_on_original_date = any(
                        item.task.task_id not in {assignment.task.task_id, blocker.task.task_id}
                        and item.task.class_id == blocker.task.class_id
                        and item.task.subject == blocker.task.subject
                        and item.candidate.slots[0].date == original_first.date
                        for item in result
                    )
                    if blocker_same_subject_on_original_date:
                        continue
                    blocker_replacement = repaired_at(
                        blocker,
                        blocker_index,
                        original_first.date,
                        original_first.period,
                        {assignment.task.task_id, blocker.task.task_id},
                        [target_replacement],
                    )
                    if blocker_replacement is None:
                        continue
                    trial = scheduler.sorted_assignments(
                        [
                            item
                            for item in result
                            if item.task.task_id not in {assignment.task.task_id, blocker.task.task_id}
                        ]
                        + [target_replacement, blocker_replacement]
                    )
                    if not sequence_valid_for_class(trial, class_id):
                        continue
                    if not sequence_valid_for_class(trial, blocker.task.class_id):
                        continue
                    result = trial
                    moved_count += 1
                    moved = True
                    lines.append(
                        f"公共课同套班内部换位摊开: {assignment.task.class_id} "
                        f"{assignment.task.subject}/{assignment.task.stage or ''}/{assignment.task.course_module or ''} "
                        f"{original_first.date} {original_first.period} -> {first.date} {first.period}；"
                        f"{blocker.task.subject}/{blocker.task.stage or ''}/{blocker.task.course_module or ''} "
                        f"{blocker.candidate.slots[0].date} {blocker.candidate.slots[0].period} -> "
                        f"{blocker_replacement.candidate.slots[0].date} {blocker_replacement.candidate.slots[0].period}"
                    )
                    break
                if moved:
                    break
            if moved:
                break
        if not moved:
            break

    remaining = duplicate_groups()
    summary = (
        f"公共课同班内部换位规避 8 小时: 移动 {moved_count} 组，"
        f"剩余 {len(remaining)} 组需人工核对"
        + ("，因时间上限停止" if stopped_by_time else "")
    )
    if remaining:
        samples = []
        for (class_id, subject, date_text), _items, hours in remaining[:12]:
            suite_code = suite_code_for_class(class_id, class_metadata)
            samples.append(f"{suite_code} {class_id} {date_text} {subject} {hours:g} 小时")
        lines.append("内部换位后仍有同班同科同日高负荷: " + "；".join(samples))
    return scheduler.sorted_assignments(result), [summary, *lines[:200]]


def repair_summer_subject_week_overloads(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
    target_suite_codes: Set[str],
    max_moves: int = 160,
    max_seconds: int = 60,
) -> Tuple[List[scheduler.Assignment], List[str]]:
    if not target_suite_codes:
        return scheduler.sorted_assignments(assignments), []

    result = scheduler.sorted_assignments(list(assignments))
    window_constraints = load_summer_class_window_constraints(data_dir)
    input_cache: Dict[str, scheduler.ScheduleInput] = {}
    lines: List[str] = []
    moved_count = 0
    started_at = time.perf_counter()
    stopped_by_time = False

    def summer_public_assignment(item: scheduler.Assignment) -> bool:
        suite_code = suite_code_for_class(item.task.class_id, class_metadata)
        return (
            suite_code in target_suite_codes
            and item.task.subject in SUMMER_PUBLIC_SUBJECTS
            and assignment_matches_phase(item, SUMMER_PREPLAN_STAGES)
            and assignment_sub_product(item, class_metadata) in SUMMER_FAST_PRODUCTS
        )

    def weekly_state() -> Tuple[
        Dict[Tuple[str, str], Dict[Tuple[int, int], int]],
        Dict[Tuple[str, str], int],
        Dict[str, Set[Tuple[int, int]]],
        Dict[Tuple[str, str], int],
    ]:
        subject_weeks: Dict[Tuple[str, str], Dict[Tuple[int, int], int]] = defaultdict(lambda: defaultdict(int))
        subject_totals: Dict[Tuple[str, str], int] = defaultdict(int)
        active_weeks: Dict[str, Set[Tuple[int, int]]] = defaultdict(set)
        for item in result:
            if not summer_public_assignment(item):
                continue
            suite_code = suite_code_for_class(item.task.class_id, class_metadata)
            week = week_key(item.candidate.slots)
            active_weeks[suite_code].add(week)
            subject_weeks[(suite_code, item.task.subject)][week] += 1
            subject_totals[(suite_code, item.task.subject)] += 1
        for suite_code in target_suite_codes:
            for constraint in suite_window_constraint_items(suite_code, class_metadata, window_constraints):
                active_weeks[suite_code].update(
                    calendar_weeks_between(constraint.earliest_date, constraint.latest_date)
                )
        subject_limits: Dict[Tuple[str, str], int] = {}
        for key, total in subject_totals.items():
            suite_code, _subject = key
            week_count = max(1, len(active_weeks.get(suite_code) or set()))
            subject_limits[key] = max(1, (total + week_count - 1) // week_count)
        return subject_weeks, subject_totals, active_weeks, subject_limits

    def overload_groups() -> List[Tuple[str, str, Tuple[int, int], int, int]]:
        subject_weeks, _subject_totals, _active_weeks, subject_limits = weekly_state()
        groups: List[Tuple[str, str, Tuple[int, int], int, int]] = []
        for (suite_code, subject), loads in subject_weeks.items():
            limit = subject_limits.get((suite_code, subject))
            if not limit:
                continue
            for week, count in loads.items():
                if count > limit:
                    groups.append((suite_code, subject, week, count, limit))
        groups.sort(key=lambda item: (item[3] - item[4], item[3], item[0], item[1], item[2]), reverse=True)
        return groups

    def replacement_for_under_week(
        assignment: scheduler.Assignment,
        assignment_index: int,
        under_weeks: Sequence[Tuple[int, int]],
        subject_weeks: Dict[Tuple[str, str], Dict[Tuple[int, int], int]],
    ) -> Optional[scheduler.Assignment]:
        class_id = assignment.task.class_id
        if class_id not in input_cache:
            input_cache[class_id] = repair_schedule_input_for_class(
                data_dir,
                class_id,
                class_metadata,
            )
        repair_input = input_cache[class_id]
        repair_task = flexible_repair_task_for_assignment(assignment, repair_input, class_metadata)
        other_assignments = [
            item
            for index, item in enumerate(result)
            if index != assignment_index
        ]
        other_by_date_period: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
        for item in other_assignments:
            for key in assignment_date_period_keys(item):
                other_by_date_period[key].append(item)
        day_loads = class_teacher_day_loads_for_assignments(other_assignments, assignment.task.task_id)
        same_class_subject_dates = {
            item.candidate.slots[0].date
            for item in other_assignments
            if item.task.class_id == assignment.task.class_id
            and item.task.subject == assignment.task.subject
        }
        original_slot_key = scheduler.slot_sort_key(assignment.candidate.slots[0])
        original_date = Date.fromisoformat(assignment.candidate.slots[0].date)
        preferred_periods = {"数学": "AM", "英语": "PM", "政治": "PM"}
        preferred_period = preferred_periods.get(assignment.task.subject)
        under_week_set = set(under_weeks)
        try:
            candidates = scheduler.candidate_assignments(repair_task, repair_input)
        except (KeyError, ValueError):
            return None
        candidates.sort(
            key=lambda candidate: (
                subject_weeks.get((suite_code_for_class(assignment.task.class_id, class_metadata), assignment.task.subject), {}).get(week_key(candidate.slots), 0),
                0 if not preferred_period or candidate.slots[0].period == preferred_period else 1,
                abs((Date.fromisoformat(candidate.slots[0].date) - original_date).days),
                scheduler.slot_sort_key(candidate.slots[0]),
            )
        )
        for candidate in candidates:
            if week_key(candidate.slots) not in under_week_set:
                continue
            if scheduler.slot_sort_key(candidate.slots[0]) == original_slot_key and candidate.room_id == assignment.candidate.room_id:
                continue
            if candidate.slots[0].date in same_class_subject_dates:
                continue
            if not repair_candidate_allowed_by_outer_rules(
                assignment,
                candidate,
                class_metadata,
                window_constraints,
                blackout_dates,
            ):
                continue
            if not replacement_preserves_first_lesson_module(
                result,
                assignment_index,
                repair_task,
                candidate,
            ):
                continue
            if not repair_preserves_observed_stage_order(
                repair_input,
                result,
                assignment_index,
                repair_task,
                candidate,
            ):
                continue
            if not candidate_avoids_same_class_teacher_day_limit(day_loads, repair_task, candidate):
                continue
            candidate_assignment = scheduler.Assignment(
                task=repaired_display_task(assignment.task, repair_task),
                candidate=candidate,
            )
            if candidate_conflicts_for_repair(candidate_assignment, other_by_date_period, class_conflict_groups):
                continue
            return candidate_assignment
        return None

    for _ in range(max_moves):
        if max_seconds > 0 and time.perf_counter() - started_at > max_seconds:
            stopped_by_time = True
            break
        groups = overload_groups()
        if not groups:
            break
        subject_weeks, _subject_totals, active_weeks, subject_limits = weekly_state()
        index_by_task_id = {assignment.task.task_id: index for index, assignment in enumerate(result)}
        moved = False
        for suite_code, subject, over_week, count, limit in groups:
            under_weeks = [
                week
                for week in sorted(active_weeks.get(suite_code, set()))
                if subject_weeks.get((suite_code, subject), {}).get(week, 0) < limit
            ]
            if not under_weeks:
                continue
            candidates_to_move = [
                assignment
                for assignment in result
                if summer_public_assignment(assignment)
                and suite_code_for_class(assignment.task.class_id, class_metadata) == suite_code
                and assignment.task.subject == subject
                and week_key(assignment.candidate.slots) == over_week
            ]
            candidates_to_move.sort(
                key=lambda item: (
                    scheduler.slot_sort_key(item.candidate.slots[0]),
                    item.task.task_id,
                ),
                reverse=True,
            )
            for assignment in candidates_to_move:
                assignment_index = index_by_task_id.get(assignment.task.task_id)
                if assignment_index is None:
                    continue
                replacement = replacement_for_under_week(
                    assignment,
                    assignment_index,
                    under_weeks,
                    subject_weeks,
                )
                if replacement is None:
                    continue
                old_first = assignment.candidate.slots[0]
                new_first = replacement.candidate.slots[0]
                result[assignment_index] = replacement
                result = scheduler.sorted_assignments(result)
                moved_count += 1
                moved = True
                lines.append(
                    f"暑假科目周均衡: {suite_code} {assignment.task.class_id} {subject}/{assignment.task.course_module or ''} "
                    f"{old_first.date} {old_first.period} -> {new_first.date} {new_first.period} "
                    f"(原周 {count}>{limit})"
                )
                break
            if moved:
                break
        if not moved:
            break

    remaining = overload_groups()
    summary = (
        f"暑假科目周均衡修复: 移动 {moved_count} 节，剩余 {len(remaining)} 组"
        + ("，因时间上限停止" if stopped_by_time else "")
    )
    if remaining:
        samples = [
            f"{suite_code} {subject} {week[0]}年第{week[1]}周 {count}>{limit}"
            for suite_code, subject, week, count, limit in remaining[:16]
        ]
        lines.append("仍有暑假科目周超量: " + "；".join(samples))
    return scheduler.sorted_assignments(result), [summary, *lines[:200]]


def repair_student_experience_week_balance(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
    target_suite_codes: Optional[Set[str]] = None,
    max_moves: int = 120,
    max_seconds: int = 90,
) -> Tuple[List[scheduler.Assignment], List[str]]:
    result = scheduler.sorted_assignments(list(assignments))
    window_constraints = load_summer_class_window_constraints(data_dir)
    repair_input_cache: Dict[str, scheduler.ScheduleInput] = {}
    lines: List[str] = []
    moved_count = 0
    started_at = time.perf_counter()
    stopped_by_time = False

    def managed_assignment(assignment: scheduler.Assignment) -> bool:
        if assignment.task.subject not in SUMMER_PUBLIC_SUBJECTS:
            return False
        meta = class_metadata.get(assignment.task.class_id, {})
        if meta.get("subject_category") and meta.get("subject_category") != "公共课":
            return False
        if meta.get("is_schedule_locked") == "是":
            return False
        sub_product = assignment_sub_product(assignment, class_metadata)
        if sub_product not in PUBLIC_PRODUCT_ORDER:
            return False
        first = assignment.candidate.slots[0]
        if not student_experience_phase(first.date, sub_product):
            return False
        suite_code = suite_code_for_class(assignment.task.class_id, class_metadata)
        return bool(suite_code) and (target_suite_codes is None or suite_code in target_suite_codes)

    def group_key_for(assignment: scheduler.Assignment) -> Optional[Tuple[str, str, str]]:
        if not managed_assignment(assignment):
            return None
        sub_product = assignment_sub_product(assignment, class_metadata)
        phase = student_experience_phase(assignment.candidate.slots[0].date, sub_product)
        if not phase:
            return None
        return (sub_product, suite_code_for_class(assignment.task.class_id, class_metadata), phase)

    def week_to_monday(key: Tuple[int, int]) -> Date:
        return Date.fromisocalendar(key[0], key[1], 1)

    def group_items(key: Tuple[str, str, str]) -> List[Tuple[int, scheduler.Assignment]]:
        return [
            (index, assignment)
            for index, assignment in enumerate(result)
            if group_key_for(assignment) == key
        ]

    def group_state(key: Tuple[str, str, str]) -> Dict[str, Any]:
        sub_product, _suite_code, phase = key
        items = group_items(key)
        weeks: Dict[Tuple[int, int], Dict[str, Any]] = defaultdict(
            lambda: {"count": 0, "subjects": set(), "subject_counts": defaultdict(int), "dates": set()}
        )
        subjects: Set[str] = set()
        for _index, assignment in items:
            first = assignment.candidate.slots[0]
            week = week_key(assignment.candidate.slots)
            info = weeks[week]
            info["count"] += 1
            info["subjects"].add(assignment.task.subject)
            info["subject_counts"][assignment.task.subject] += 1
            info["dates"].add(first.date)
            subjects.add(assignment.task.subject)
        if not weeks:
            return {"items": items, "weeks": weeks, "subjects": subjects, "expected_weeks": []}
        first_week = min(weeks)
        last_week = max(weeks)
        start_monday = week_to_monday(first_week)
        end_monday = week_to_monday(last_week)
        if sub_product in WYQC_PRODUCTS and phase == "秋季":
            start_monday = week_monday(WYQC_AUTUMN_START)
            end_monday = week_monday(WYQC_AUTUMN_END)
        expected_weeks = [
            (week.isocalendar().year, week.isocalendar().week)
            for week in iter_week_mondays(start_monday, end_monday)
            if not blackout_heavy_week(week, blackout_dates)
        ]
        return {
            "items": items,
            "weeks": weeks,
            "subjects": subjects,
            "expected_weeks": expected_weeks,
        }

    def group_score(key: Tuple[str, str, str]) -> int:
        sub_product, _suite_code, phase = key
        state = group_state(key)
        weeks: Dict[Tuple[int, int], Dict[str, Any]] = state["weeks"]
        subjects: Set[str] = set(state["subjects"])
        expected_weeks: List[Tuple[int, int]] = list(state["expected_weeks"])
        if not expected_weeks:
            return 0
        comfortable_max = student_experience_weekly_max(sub_product, subjects, phase)
        balance_delta_limit = 1 if sub_product in {"全年营", "半年营", SPRINT_CAMP_PRODUCT} else 2
        score = 0
        total_values = [int(weeks.get(week, {}).get("count", 0)) for week in expected_weeks]
        if total_values and max(total_values) - min(total_values) > balance_delta_limit:
            score += (max(total_values) - min(total_values) - balance_delta_limit) * 20
        for week in expected_weeks:
            count = int(weeks.get(week, {}).get("count", 0))
            if count > comfortable_max:
                score += (count - comfortable_max) * 25
            if count == 0 and week not in {expected_weeks[0], expected_weeks[-1]}:
                score += 18
        for subject in subjects:
            values = [
                int(weeks.get(week, {}).get("subject_counts", {}).get(subject, 0))
                for week in expected_weeks
            ]
            if values and max(values) > 0 and max(values) - min(values) > balance_delta_limit:
                score += (max(values) - min(values) - balance_delta_limit) * 15
        check_subject_mix = phase == "暑假" or sub_product in {"全年营", "半年营"}
        expected_subjects = subjects & SUMMER_PUBLIC_SUBJECTS
        if check_subject_mix and len(expected_subjects) > 1:
            for week in expected_weeks[1:-1]:
                info = weeks.get(week)
                if not info:
                    continue
                week_subjects = set(info.get("subjects", set()))
                if week_subjects and week_subjects != expected_subjects:
                    score += 8 if len(week_subjects) == 1 else 5
        if sub_product in WYQC_PRODUCTS and phase == "秋季":
            subject_totals = {
                subject: sum(
                    int(info.get("subject_counts", {}).get(subject, 0))
                    for info in weeks.values()
                )
                for subject in subjects
            }
            for week in expected_weeks:
                subject_counts = weeks.get(week, {}).get("subject_counts", {})
                for subject in subjects:
                    target_weeks = set(evenly_spaced_items(expected_weeks, subject_totals.get(subject, 0)))
                    if week not in target_weeks:
                        continue
                    if int(subject_counts.get(subject, 0)) < 1:
                        score += 35
            for info in weeks.values():
                for date_text in info.get("dates", set()):
                    if date_text < WYQC_AUTUMN_START or date_text > WYQC_AUTUMN_END:
                        score += 40
        return score

    def candidate_for_week(
        assignment: scheduler.Assignment,
        assignment_index: int,
        target_week: Tuple[int, int],
        key: Tuple[str, str, str],
        current_score: int,
    ) -> Optional[scheduler.Assignment]:
        sub_product, _suite_code, phase = key
        class_id = assignment.task.class_id
        repair_input = repair_input_cache.get(class_id)
        if repair_input is None:
            repair_input = repair_schedule_input_for_class(data_dir, class_id, class_metadata)
            repair_input_cache[class_id] = repair_input
        repair_task = flexible_repair_task_for_assignment(assignment, repair_input, class_metadata)
        try:
            candidates = scheduler.candidate_assignments(repair_task, repair_input)
        except (KeyError, ValueError):
            return None
        other_assignments = [
            item for index, item in enumerate(result) if index != assignment_index
        ]
        other_by_date_period: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
        for item in other_assignments:
            for item_key in assignment_date_period_keys(item):
                other_by_date_period[item_key].append(item)
        same_class_subject_dates = {
            item.candidate.slots[0].date
            for item in other_assignments
            if item.task.class_id == assignment.task.class_id
            and item.task.subject == assignment.task.subject
        }
        day_loads = class_teacher_day_loads_for_assignments(other_assignments, assignment.task.task_id)
        original_first = assignment.candidate.slots[0]
        candidates.sort(
            key=lambda candidate: (
                candidate.slots[0].period != LONG_CAMP_PREFERRED_PERIODS.get(assignment.task.subject),
                abs((Date.fromisoformat(candidate.slots[0].date) - Date.fromisoformat(original_first.date)).days),
                scheduler.slot_sort_key(candidate.slots[0]),
            )
        )
        current_stage_violations = class_stage_order_violation_count(
            repair_input,
            result,
            assignment.task.class_id,
        )
        for candidate in candidates:
            first = candidate.slots[0]
            if week_key(candidate.slots) != target_week:
                continue
            if student_experience_phase(first.date, sub_product) != phase:
                continue
            if (
                sub_product in WYQC_PRODUCTS
                and phase == "秋季"
                and not (WYQC_AUTUMN_START <= first.date <= WYQC_AUTUMN_END)
            ):
                continue
            if scheduler.slot_sort_key(first) == scheduler.slot_sort_key(original_first) and candidate.room_id == assignment.candidate.room_id:
                continue
            if first.date in same_class_subject_dates:
                continue
            meta = class_metadata.get(assignment.task.class_id, {})
            if meta.get("start_date") and first.date < meta["start_date"]:
                continue
            if meta.get("end_date") and first.date > meta["end_date"]:
                continue
            if not repair_candidate_allowed_by_outer_rules(
                assignment,
                candidate,
                class_metadata,
                window_constraints,
                blackout_dates,
            ):
                continue
            if not replacement_preserves_first_lesson_module(
                result,
                assignment_index,
                repair_task,
                candidate,
            ):
                continue
            if not repair_preserves_observed_stage_order(
                repair_input,
                result,
                assignment_index,
                repair_task,
                candidate,
            ):
                continue
            proposed = scheduler.Assignment(
                task=repaired_display_task(assignment.task, repair_task),
                candidate=candidate,
            )
            trial = list(result)
            trial[assignment_index] = proposed
            if class_stage_order_violation_count(repair_input, trial, assignment.task.class_id) > current_stage_violations:
                continue
            if not candidate_avoids_same_class_teacher_day_limit(day_loads, repair_task, candidate):
                continue
            if candidate_conflicts_for_repair(proposed, other_by_date_period, class_conflict_groups):
                continue
            old_assignment = result[assignment_index]
            result[assignment_index] = proposed
            try:
                if group_score(key) < current_score:
                    return proposed
            finally:
                result[assignment_index] = old_assignment
        return None

    def problem_groups() -> List[Tuple[int, Tuple[str, str, str]]]:
        keys = {key for assignment in result for key in [group_key_for(assignment)] if key}
        scored = [(group_score(key), key) for key in keys]
        scored = [(score, key) for score, key in scored if score > 0]
        scored.sort(reverse=True)
        return scored

    for _ in range(max_moves):
        if max_seconds > 0 and time.perf_counter() - started_at > max_seconds:
            stopped_by_time = True
            break
        moved = False
        for score, key in problem_groups()[:16]:
            state = group_state(key)
            weeks: Dict[Tuple[int, int], Dict[str, Any]] = state["weeks"]
            expected_weeks: List[Tuple[int, int]] = list(state["expected_weeks"])
            if not expected_weeks:
                continue
            subjects: Set[str] = set(state["subjects"])
            comfortable_max = student_experience_weekly_max(key[0], subjects, key[2])
            source_items = list(state["items"])
            source_items.sort(
                key=lambda item: (
                    int(weeks.get(week_key(item[1].candidate.slots), {}).get("count", 0)),
                    int(weeks.get(week_key(item[1].candidate.slots), {}).get("subject_counts", {}).get(item[1].task.subject, 0)),
                    item[1].candidate.slots[0].date,
                    scheduler.period_sort_value(item[1].candidate.slots[0].period),
                ),
                reverse=True,
            )
            for assignment_index, assignment in source_items:
                source_week = week_key(assignment.candidate.slots)
                source_info = weeks.get(source_week, {})
                if int(source_info.get("count", 0)) <= 1:
                    continue
                source_subject_count = int(source_info.get("subject_counts", {}).get(assignment.task.subject, 0))
                if source_subject_count <= 1 and int(source_info.get("count", 0)) <= comfortable_max:
                    continue
                target_weeks = [
                    week
                    for week in expected_weeks
                    if week != source_week
                    and int(weeks.get(week, {}).get("count", 0)) < comfortable_max
                ]
                if key[0] in WYQC_PRODUCTS and key[2] == "秋季":
                    subject_total = sum(
                        int(info.get("subject_counts", {}).get(assignment.task.subject, 0))
                        for info in weeks.values()
                    )
                    preferred_subject_weeks = set(evenly_spaced_items(expected_weeks, subject_total))
                    target_weeks = [
                        week
                        for week in target_weeks
                        if week in preferred_subject_weeks
                        and int(weeks.get(week, {}).get("subject_counts", {}).get(assignment.task.subject, 0)) == 0
                    ] or [
                        week
                        for week in target_weeks
                        if int(weeks.get(week, {}).get("subject_counts", {}).get(assignment.task.subject, 0)) == 0
                    ] or target_weeks
                target_weeks.sort(
                    key=lambda week: (
                        int(weeks.get(week, {}).get("subject_counts", {}).get(assignment.task.subject, 0)),
                        int(weeks.get(week, {}).get("count", 0)),
                        abs((week_to_monday(week) - Date.fromisoformat(assignment.candidate.slots[0].date)).days),
                        week,
                    )
                )
                for target_week in target_weeks[:10]:
                    replacement = candidate_for_week(assignment, assignment_index, target_week, key, score)
                    if replacement is None:
                        continue
                    old_first = assignment.candidate.slots[0]
                    new_first = replacement.candidate.slots[0]
                    result[assignment_index] = replacement
                    result = scheduler.sorted_assignments(result)
                    moved_count += 1
                    moved = True
                    lines.append(
                        f"学生体验周均衡: {key[1]} {assignment.task.class_id} "
                        f"{assignment.task.subject}/{assignment.task.stage or ''}/{assignment.task.course_module or ''} "
                        f"{old_first.date} {old_first.period} -> {new_first.date} {new_first.period} "
                        f"({key[0]} {key[2]})"
                    )
                    break
                if moved:
                    break
            if moved:
                break
        if not moved:
            break

    final_problem_count = len(problem_groups())
    if moved_count or final_problem_count:
        lines.insert(
            0,
            f"学生体验周均衡修复: 移动 {moved_count} 节，剩余问题组 {final_problem_count}"
            + ("，因时间上限停止" if stopped_by_time else ""),
        )
    return scheduler.sorted_assignments(result), lines[:220]


def apply_2754_politics_overlap_balance(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
) -> Tuple[List[scheduler.Assignment], List[str]]:
    suite_code = "2754"
    result = scheduler.sorted_assignments(list(assignments))
    window_constraints = load_summer_class_window_constraints(data_dir)
    lines: List[str] = []

    def in_target_summer(assignment: scheduler.Assignment) -> bool:
        first = assignment.candidate.slots[0]
        return (
            suite_code_for_class(assignment.task.class_id, class_metadata) == suite_code
            and "2026-07-01" <= first.date <= "2026-08-31"
            and assignment.task.subject in {"英语", "政治"}
        )

    def week_number(assignment: scheduler.Assignment) -> int:
        return Date.fromisoformat(assignment.candidate.slots[0].date).isocalendar().week

    def standard_candidate(
        assignment: scheduler.Assignment,
        date_text: str,
        period: str,
    ) -> Optional[scheduler.Candidate]:
        slots = standard_period_slots(date_text, period)
        if sum(slot.duration_hours for slot in slots) != assignment.task.block_hours:
            return None
        return scheduler.Candidate(
            slots=slots,
            teacher_id=assignment.candidate.teacher_id,
            teacher_name=assignment.candidate.teacher_name,
            room_id=assignment.candidate.room_id,
        )

    def candidate_is_allowed(
        assignment: scheduler.Assignment,
        assignment_index: int,
        candidate: scheduler.Candidate,
        relax_politics_stage_order: bool,
    ) -> bool:
        first = candidate.slots[0]
        candidate_dates = {slot.date for slot in candidate.slots}
        if candidate_dates & blackout_dates:
            return False
        if Date.fromisoformat(first.date).weekday() == 6:
            return False
        sub_product = assignment_sub_product(assignment, class_metadata)
        if sub_product in WUYOU_PRODUCTS and candidate_dates & WUYOU_PRODUCT_BLACKOUT_DATES:
            return False
        constraint = assignment_window_constraint(assignment, class_metadata, window_constraints)
        if (
            sub_product in SUMMER_FAST_PRODUCTS
            and assignment_matches_phase(assignment, SUMMER_PREPLAN_STAGES)
            and constraint
            and not candidate_within_window_boundary(candidate, constraint)
        ):
            return False
        class_meta = class_metadata.get(assignment.task.class_id, {})
        if class_meta.get("start_date") and first.date < class_meta["start_date"]:
            return False
        if class_meta.get("end_date") and first.date > class_meta["end_date"]:
            return False

        other_assignments = [
            item for index, item in enumerate(result) if index != assignment_index
        ]
        for item in other_assignments:
            if item.task.class_id != assignment.task.class_id:
                continue
            if item.task.subject != assignment.task.subject:
                continue
            if candidate_dates & {slot.date for slot in item.candidate.slots}:
                return False

        if not relax_politics_stage_order:
            repair_input = repair_schedule_input_for_class(data_dir, assignment.task.class_id, class_metadata)
            if not repair_preserves_observed_stage_order(
                repair_input,
                result,
                assignment_index,
                assignment.task,
                candidate,
            ):
                return False

        day_loads = class_teacher_day_loads_for_assignments(
            other_assignments,
            assignment.task.task_id,
        )
        if not candidate_avoids_same_class_teacher_day_limit(
            day_loads,
            assignment.task,
            candidate,
        ):
            return False

        other_by_date_period: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
        for item in other_assignments:
            for key in assignment_date_period_keys(item):
                other_by_date_period[key].append(item)
        candidate_assignment = scheduler.Assignment(task=assignment.task, candidate=candidate)
        return not candidate_conflicts_for_repair(
            candidate_assignment,
            other_by_date_period,
            class_conflict_groups,
        )

    def move_first(
        source_filter: Any,
        target_slots: Sequence[Tuple[str, str]],
        relax_politics_stage_order: bool,
        reason: str,
    ) -> bool:
        index_by_task_id = {assignment.task.task_id: index for index, assignment in enumerate(result)}
        source_assignments = [
            assignment
            for assignment in result
            if source_filter(assignment)
        ]
        same_subject_day_counts: Dict[Tuple[str, str, str], int] = defaultdict(int)
        for item in result:
            first = item.candidate.slots[0]
            same_subject_day_counts[(item.task.class_id, item.task.subject, first.date)] += 1
        source_assignments.sort(
            key=lambda assignment: (
                same_subject_day_counts[
                    (
                        assignment.task.class_id,
                        assignment.task.subject,
                        assignment.candidate.slots[0].date,
                    )
                ],
                assignment.candidate.slots[0].period == "PM",
                scheduler.slot_sort_key(assignment.candidate.slots[0]),
                assignment.task.class_id,
                assignment.task.course_module or "",
            ),
            reverse=True,
        )
        for assignment in source_assignments:
            assignment_index = index_by_task_id.get(assignment.task.task_id)
            if assignment_index is None:
                continue
            old_first = assignment.candidate.slots[0]
            for date_text, period in target_slots:
                candidate = standard_candidate(assignment, date_text, period)
                if candidate is None:
                    continue
                if scheduler.slot_sort_key(candidate.slots[0]) == scheduler.slot_sort_key(old_first):
                    continue
                if not candidate_is_allowed(
                    assignment,
                    assignment_index,
                    candidate,
                    relax_politics_stage_order,
                ):
                    continue
                result[assignment_index] = scheduler.Assignment(
                    task=assignment.task,
                    candidate=candidate,
                )
                result[:] = scheduler.sorted_assignments(result)
                lines.append(
                    f"2754 政治阶段轻微交错均衡: {assignment.task.class_id} "
                    f"{assignment.task.subject}/{assignment.task.stage or ''}/{assignment.task.course_module or ''} "
                    f"{assignment.candidate.teacher_name or assignment.candidate.teacher_id} "
                    f"{old_first.date} {old_first.period} -> {date_text} {period}；{reason}"
                )
                return True
        return False

    def week_subject_count(subject: str, week: int) -> int:
        return sum(
            1
            for assignment in result
            if in_target_summer(assignment)
            and assignment.task.subject == subject
            and week_number(assignment) == week
        )

    def week_total(week: int) -> int:
        return sum(
            1
            for assignment in result
            if in_target_summer(assignment) and week_number(assignment) == week
        )

    politics_targets = [
        ("2026-07-28", "PM"),
        ("2026-07-30", "PM"),
        ("2026-07-31", "PM"),
    ]
    if week_subject_count("英语", 29) < 3 and week_subject_count("英语", 31) > 4:
        move_first(
            lambda assignment: (
                in_target_summer(assignment)
                and assignment.task.class_id == "KYJXY2754"
                and assignment.task.subject == "英语"
                and week_number(assignment) == 31
            ),
            [
                ("2026-07-15", "AM"),
                ("2026-07-17", "AM"),
                ("2026-07-13", "AM"),
            ],
            False,
            "补足第29周英语，拉平英政比例",
        )

    for _ in range(2):
        if week_subject_count("政治", 29) <= 5 or week_total(31) >= 8:
            break
        moved = move_first(
            lambda assignment: (
                in_target_summer(assignment)
                and assignment.task.class_id == "KYJXZ2754"
                and assignment.task.subject == "政治"
                and week_number(assignment) == 29
                and assignment.task.stage == "基础"
            ),
            politics_targets,
            True,
            "允许政治基础/强化轻微交错，把第29周政治拆到第31周",
        )
        if not moved:
            break

    if lines:
        lines.insert(0, f"2754 政治阶段轻微交错均衡: 移动 {len(lines)} 节")
    return scheduler.sorted_assignments(result), lines


def long_camp_balance_assignment(
    assignment: scheduler.Assignment,
    class_metadata: Dict[str, Dict[str, str]],
) -> bool:
    first = assignment.candidate.slots[0]
    if first.date < LONG_CAMP_BALANCE_START or first.date > AUTUMN_END:
        return False
    if assignment.task.subject not in SUMMER_PUBLIC_SUBJECTS:
        return False
    if assignment.task.task_id.startswith(("HISTORY:", "LOCKED:", "PROFESSIONAL:")):
        return False
    meta = class_metadata.get(assignment.task.class_id, {})
    sub_product = meta.get("sub_product") or assignment_sub_product(assignment, class_metadata)
    suite_code = suite_code_for_class(assignment.task.class_id, class_metadata)
    return sub_product in {"全年营", "半年营"} or suite_code in HALF_YEAR_BATCH_SUITES


def long_camp_week_allowed(
    monday: Date,
    class_ids: Sequence[str],
    class_metadata: Dict[str, Dict[str, str]],
    blackout_dates: Set[str],
) -> bool:
    for offset in range(7):
        day = monday + timedelta(days=offset)
        date_text = day.isoformat()
        if date_text < LONG_CAMP_BALANCE_START or date_text > AUTUMN_END:
            continue
        if day.weekday() == 6 or date_text in blackout_dates:
            continue
        for class_id in class_ids:
            meta = class_metadata.get(class_id, {})
            start = meta.get("start_date") or LONG_CAMP_BALANCE_START
            end = meta.get("end_date") or AUTUMN_END
            if start <= date_text <= end:
                return True
    return False


def long_camp_subject_active_weeks(
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    blackout_dates: Set[str],
) -> List[Date]:
    if not assignments:
        return []
    class_ids = sorted({assignment.task.class_id for assignment in assignments})
    item_dates_by_class: Dict[str, List[str]] = defaultdict(list)
    for assignment in assignments:
        item_dates_by_class[assignment.task.class_id].append(assignment.candidate.slots[0].date)
    start_values: List[str] = []
    end_values: List[str] = []
    for class_id in class_ids:
        class_dates = item_dates_by_class.get(class_id, [])
        meta = class_metadata.get(class_id, {})
        start = clean(meta.get("start_date")) or (min(class_dates) if class_dates else LONG_CAMP_BALANCE_START)
        end = clean(meta.get("end_date")) or (max(class_dates) if class_dates else AUTUMN_END)
        start = max(start, LONG_CAMP_BALANCE_START)
        end = min(end, AUTUMN_END)
        if start <= end:
            start_values.append(start)
            end_values.append(end)
    if not start_values or not end_values:
        start_values = [min(assignment.candidate.slots[0].date for assignment in assignments)]
        end_values = [max(assignment.candidate.slots[0].date for assignment in assignments)]

    current = week_monday(min(start_values))
    last_week = week_monday(max(end_values))
    weeks: List[Date] = []
    while current <= last_week:
        if not blackout_heavy_week(current, blackout_dates) and long_camp_week_allowed(
            current,
            class_ids,
            class_metadata,
            blackout_dates,
        ):
            weeks.append(current)
        current += timedelta(days=7)
    return weeks


def long_camp_subject_week_targets(
    weeks: Sequence[Date],
    total_count: int,
    subject: str,
) -> Dict[Tuple[int, int], int]:
    if not weeks or total_count <= 0:
        return {}
    keys = [(week.isocalendar().year, week.isocalendar().week) for week in weeks]
    weekly_min = LONG_CAMP_SUBJECT_WEEKLY_MIN.get(subject, 0)
    hard_max = LONG_CAMP_SUBJECT_WEEKLY_MAX.get(subject)
    target_max = (total_count + len(keys) - 1) // len(keys)
    if hard_max:
        target_max = max(target_max, hard_max if total_count > hard_max * len(keys) else min(target_max, hard_max))
    if weekly_min and total_count >= weekly_min * len(keys):
        targets = {key: weekly_min for key in keys}
        remaining = total_count - weekly_min * len(keys)
        target_max = max(target_max, weekly_min)
    else:
        targets = {key: 0 for key in keys}
        remaining = total_count
    while remaining > 0:
        progressed = False
        min_value = min(targets.values())
        for key in keys:
            if remaining <= 0:
                break
            if targets[key] != min_value or targets[key] >= target_max:
                continue
            targets[key] += 1
            remaining -= 1
            progressed = True
        if not progressed:
            for key in keys:
                if remaining <= 0:
                    break
                targets[key] += 1
                remaining -= 1
    return targets


def repair_long_camp_subject_week_balance(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
    target_suite_codes: Optional[Set[str]] = None,
    max_moves: int = 260,
    max_seconds: int = 90,
) -> Tuple[List[scheduler.Assignment], List[str]]:
    result = scheduler.sorted_assignments(list(assignments))
    lines: List[str] = []
    window_constraints = load_all_class_window_constraints(data_dir)
    repair_input_cache: Dict[str, scheduler.ScheduleInput] = {}
    started_at = time.perf_counter()
    stopped_by_time = False

    def target_items() -> List[scheduler.Assignment]:
        return [
            assignment
            for assignment in result
            if long_camp_balance_assignment(assignment, class_metadata)
            and (
                target_suite_codes is None
                or suite_code_for_class(assignment.task.class_id, class_metadata) in target_suite_codes
            )
        ]

    def week_key_from_date(date_text: str) -> Tuple[int, int]:
        iso = Date.fromisoformat(date_text).isocalendar()
        return iso.year, iso.week

    def group_state() -> List[Tuple[str, str, Dict[Tuple[int, int], int], Dict[Tuple[int, int], int]]]:
        grouped: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
        for assignment in target_items():
            grouped[(suite_code_for_class(assignment.task.class_id, class_metadata), assignment.task.subject)].append(assignment)
        states = []
        for (suite_code, subject), items in grouped.items():
            if len(items) < 2:
                continue
            weeks = long_camp_subject_active_weeks(items, class_metadata, blackout_dates)
            if not weeks:
                continue
            targets = long_camp_subject_week_targets(weeks, len(items), subject)
            loads: Dict[Tuple[int, int], int] = {key: 0 for key in targets}
            for assignment in items:
                key = week_key_from_date(assignment.candidate.slots[0].date)
                loads[key] = loads.get(key, 0) + 1
            states.append((suite_code, subject, loads, targets))
        return states

    def overloads() -> List[Tuple[str, str, Tuple[int, int], int, int]]:
        groups: List[Tuple[str, str, Tuple[int, int], int, int]] = []
        for suite_code, subject, loads, targets in group_state():
            for week, count in loads.items():
                target = targets.get(week, 0)
                if count > target:
                    groups.append((suite_code, subject, week, count, target))
        groups.sort(key=lambda item: (item[3] - item[4], item[0], item[1], item[2]), reverse=True)
        return groups

    def under_weeks_for(suite_code: str, subject: str) -> List[Tuple[int, int]]:
        for current_suite, current_subject, loads, targets in group_state():
            if current_suite != suite_code or current_subject != subject:
                continue
            weeks = [
                week
                for week, target in targets.items()
                if loads.get(week, 0) < target
            ]
            weeks.sort(key=lambda week: (loads.get(week, 0) - targets.get(week, 0), week))
            return weeks
        return []

    def minimum_gaps() -> List[Tuple[str, str, Tuple[int, int]]]:
        groups: List[Tuple[str, str, Tuple[int, int]]] = []
        for suite_code, subject, loads, targets in group_state():
            for week, target in targets.items():
                if target >= 1 and loads.get(week, 0) == 0:
                    groups.append((suite_code, subject, week))
        groups.sort(key=lambda item: (item[0], PUBLIC_SUBJECT_PRIORITY.get(item[1], 99), item[2]))
        return groups

    def source_weeks_for_minimum_fill(
        suite_code: str,
        subject: str,
        target_week: Tuple[int, int],
    ) -> List[Tuple[Tuple[int, int], int, int]]:
        for current_suite, current_subject, loads, targets in group_state():
            if current_suite != suite_code or current_subject != subject:
                continue
            sources: List[Tuple[Tuple[int, int], int, int]] = []
            for week, count in loads.items():
                if week == target_week:
                    continue
                floor = 1 if targets.get(week, 0) >= 1 else 0
                if count > floor:
                    sources.append((week, count, floor))
            sources.sort(
                key=lambda item: (
                    item[1] - item[2],
                    -abs((Date.fromisocalendar(item[0][0], item[0][1], 1) - Date.fromisocalendar(target_week[0], target_week[1], 1)).days),
                    item[0],
                ),
                reverse=True,
            )
            return sources
        return []

    def dates_for_week(week: Tuple[int, int]) -> List[str]:
        monday = Date.fromisocalendar(week[0], week[1], 1)
        return [(monday + timedelta(days=offset)).isoformat() for offset in range(7)]

    def standard_candidate(
        assignment: scheduler.Assignment,
        date_text: str,
        period: str,
    ) -> Optional[scheduler.Candidate]:
        slots = standard_period_slots(date_text, period)
        if sum(slot.duration_hours for slot in slots) != assignment.task.block_hours:
            return None
        return scheduler.Candidate(
            slots=slots,
            teacher_id=assignment.candidate.teacher_id,
            teacher_name=assignment.candidate.teacher_name,
            room_id=assignment.candidate.room_id,
        )

    def candidate_allowed(
        assignment: scheduler.Assignment,
        assignment_index: int,
        candidate: scheduler.Candidate,
    ) -> bool:
        first = candidate.slots[0]
        candidate_dates = {slot.date for slot in candidate.slots}
        if candidate_dates & blackout_dates:
            return False
        if Date.fromisoformat(first.date).weekday() == 6:
            return False
        meta = class_metadata.get(assignment.task.class_id, {})
        if meta.get("start_date") and first.date < meta["start_date"]:
            return False
        if meta.get("end_date") and first.date > meta["end_date"]:
            return False
        constraint = assignment_window_constraint(assignment, class_metadata, window_constraints)
        if constraint and not candidate_within_window_boundary(candidate, constraint):
            return False
        if assignment_sub_product(assignment, class_metadata) in WUYOU_PRODUCTS and candidate_dates & WUYOU_PRODUCT_BLACKOUT_DATES:
            return False
        other_assignments = [
            item for index, item in enumerate(result) if index != assignment_index
        ]
        for item in other_assignments:
            if item.task.class_id == assignment.task.class_id and item.task.subject == assignment.task.subject:
                if candidate_dates & {slot.date for slot in item.candidate.slots}:
                    return False
        repair_input = repair_input_cache.get(assignment.task.class_id)
        if repair_input is None:
            repair_input = repair_schedule_input_for_class(data_dir, assignment.task.class_id, class_metadata)
            repair_input_cache[assignment.task.class_id] = repair_input
        if not replacement_preserves_first_lesson_module(
            result,
            assignment_index,
            assignment.task,
            candidate,
        ):
            return False
        if not repair_preserves_observed_stage_order(
            repair_input,
            result,
            assignment_index,
            assignment.task,
            candidate,
        ):
            return False
        current_stage_violations = class_stage_order_violation_count(
            repair_input,
            result,
            assignment.task.class_id,
        )
        proposed_assignments = list(result)
        proposed_assignments[assignment_index] = scheduler.Assignment(task=assignment.task, candidate=candidate)
        if (
            class_stage_order_violation_count(
                repair_input,
                proposed_assignments,
                assignment.task.class_id,
            )
            > current_stage_violations
        ):
            return False
        day_loads = class_teacher_day_loads_for_assignments(
            other_assignments,
            assignment.task.task_id,
        )
        if not candidate_avoids_same_class_teacher_day_limit(day_loads, assignment.task, candidate):
            return False
        other_by_date_period: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
        for item in other_assignments:
            for key in assignment_date_period_keys(item):
                other_by_date_period[key].append(item)
        candidate_assignment = scheduler.Assignment(task=assignment.task, candidate=candidate)
        return not candidate_conflicts_for_repair(candidate_assignment, other_by_date_period, class_conflict_groups)

    moved_count = 0
    for _ in range(max_moves):
        if max_seconds > 0 and time.perf_counter() - started_at > max_seconds:
            stopped_by_time = True
            break
        moved = False
        gaps = minimum_gaps()
        for suite_code, subject, under_week in gaps:
            if max_seconds > 0 and time.perf_counter() - started_at > max_seconds:
                stopped_by_time = True
                break
            source_weeks = source_weeks_for_minimum_fill(suite_code, subject, under_week)
            if not source_weeks:
                continue
            source_week_order = {week: index for index, (week, _count, _floor) in enumerate(source_weeks)}
            candidates_to_move = [
                (index, assignment)
                for index, assignment in enumerate(result)
                if long_camp_balance_assignment(assignment, class_metadata)
                and suite_code_for_class(assignment.task.class_id, class_metadata) == suite_code
                and assignment.task.subject == subject
                and week_key_from_date(assignment.candidate.slots[0].date) in source_week_order
            ]
            candidates_to_move.sort(
                key=lambda item: (
                    source_week_order.get(week_key_from_date(item[1].candidate.slots[0].date), 999),
                    scheduler.slot_sort_key(item[1].candidate.slots[0]),
                    item[1].task.task_id,
                )
            )
            target_dates = dates_for_week(under_week)
            preferred_period = LONG_CAMP_PREFERRED_PERIODS.get(subject)
            target_slots = [
                (date_text, period)
                for date_text in target_dates
                for period in ("AM", "PM")
            ]
            target_slots.sort(
                key=lambda item: (
                    item[1] != preferred_period,
                    item[0],
                    scheduler.period_sort_value(item[1]),
                )
            )
            for assignment_index, assignment in candidates_to_move:
                if max_seconds > 0 and time.perf_counter() - started_at > max_seconds:
                    stopped_by_time = True
                    break
                old_first = assignment.candidate.slots[0]
                old_week = week_key_from_date(old_first.date)
                for date_text, period in target_slots:
                    candidate = standard_candidate(assignment, date_text, period)
                    if candidate is None or candidate.slots == assignment.candidate.slots:
                        continue
                    if not candidate_allowed(assignment, assignment_index, candidate):
                        continue
                    result[assignment_index] = scheduler.Assignment(task=assignment.task, candidate=candidate)
                    result = scheduler.sorted_assignments(result)
                    moved_count += 1
                    moved = True
                    new_first = candidate.slots[0]
                    lines.append(
                        f"长线营科目周保底: {suite_code} {assignment.task.class_id} "
                        f"{subject}/{assignment.task.stage or ''}/{assignment.task.course_module or ''} "
                        f"{old_first.date} {old_first.period} -> {new_first.date} {new_first.period} "
                        f"(填补 W{under_week[1]:02d}，来源 W{old_week[1]:02d})"
                    )
                    break
                if moved:
                    break
                if stopped_by_time:
                    break
            if moved:
                break
            if stopped_by_time:
                break
        if stopped_by_time:
            break
        if moved:
            continue

        groups = overloads()
        if not groups:
            break
        for suite_code, subject, over_week, count, target in groups:
            if max_seconds > 0 and time.perf_counter() - started_at > max_seconds:
                stopped_by_time = True
                break
            under_weeks = under_weeks_for(suite_code, subject)
            if not under_weeks:
                continue
            candidates_to_move = [
                (index, assignment)
                for index, assignment in enumerate(result)
                if long_camp_balance_assignment(assignment, class_metadata)
                and suite_code_for_class(assignment.task.class_id, class_metadata) == suite_code
                and assignment.task.subject == subject
                and week_key_from_date(assignment.candidate.slots[0].date) == over_week
            ]
            candidates_to_move.sort(
                key=lambda item: (
                    scheduler.slot_sort_key(item[1].candidate.slots[0]),
                    item[1].task.task_id,
                ),
                reverse=True,
            )
            for assignment_index, assignment in candidates_to_move:
                if max_seconds > 0 and time.perf_counter() - started_at > max_seconds:
                    stopped_by_time = True
                    break
                old_first = assignment.candidate.slots[0]
                for under_week in under_weeks:
                    target_dates = dates_for_week(under_week)
                    target_slots = [
                        (date_text, period)
                        for date_text in target_dates
                        for period in ("AM", "PM")
                    ]
                    preferred_period = LONG_CAMP_PREFERRED_PERIODS.get(subject)
                    target_slots.sort(
                        key=lambda item: (
                            item[1] != preferred_period,
                            abs((Date.fromisoformat(item[0]) - Date.fromisoformat(old_first.date)).days),
                            item[0],
                            scheduler.period_sort_value(item[1]),
                        )
                    )
                    for date_text, period in target_slots:
                        if max_seconds > 0 and time.perf_counter() - started_at > max_seconds:
                            stopped_by_time = True
                            break
                        candidate = standard_candidate(assignment, date_text, period)
                        if candidate is None:
                            continue
                        if candidate.slots == assignment.candidate.slots:
                            continue
                        if not candidate_allowed(assignment, assignment_index, candidate):
                            continue
                        result[assignment_index] = scheduler.Assignment(task=assignment.task, candidate=candidate)
                        result = scheduler.sorted_assignments(result)
                        moved_count += 1
                        moved = True
                        new_first = candidate.slots[0]
                        lines.append(
                            f"长线营科目周均衡: {suite_code} {assignment.task.class_id} "
                            f"{subject}/{assignment.task.stage or ''}/{assignment.task.course_module or ''} "
                            f"{old_first.date} {old_first.period} -> {new_first.date} {new_first.period} "
                            f"(原周 {count}>{target})"
                        )
                        break
                    if moved:
                        break
                    if stopped_by_time:
                        break
                if moved:
                    break
                if stopped_by_time:
                    break
            if moved:
                break
            if stopped_by_time:
                break
        if stopped_by_time:
            break
        if not moved:
            break

    remaining = overloads()
    remaining_gaps = minimum_gaps()
    if moved_count or remaining or remaining_gaps:
        lines.insert(
            0,
            f"长线营科目周均衡: 移动 {moved_count} 节，"
            f"剩余空周 {len(remaining_gaps)} 组，剩余超目标周 {len(remaining)} 组"
            + ("，因时间上限停止" if stopped_by_time else ""),
        )
    if remaining_gaps:
        samples = [
            f"{suite_code} {subject} W{week[1]:02d}"
            for suite_code, subject, week in remaining_gaps[:20]
        ]
        lines.append("长线营科目周保底残留: " + "；".join(samples))
    if remaining:
        samples = [
            f"{suite_code} {subject} W{week[1]:02d} {count}>{target}"
            for suite_code, subject, week, count, target in remaining[:20]
        ]
        lines.append("长线营科目周均衡残留: " + "；".join(samples))
    return scheduler.sorted_assignments(result), lines[:240]


def repair_long_camp_tail_week_chain_moves(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
) -> Tuple[List[scheduler.Assignment], List[str]]:
    result = scheduler.sorted_assignments(list(assignments))
    lines: List[str] = []
    before_conflicts = set(schedule_conflict_lines(result, class_conflict_groups))
    before_teacher_conflicts = set(teacher_time_conflict_lines(result))

    def residual_tail_gap(suite_code: str, subject: str) -> bool:
        items = [
            assignment
            for assignment in result
            if long_camp_balance_assignment(assignment, class_metadata)
            and suite_code_for_class(assignment.task.class_id, class_metadata) == suite_code
            and assignment.task.subject == subject
        ]
        if not items:
            return False
        weeks = long_camp_subject_active_weeks(items, class_metadata, blackout_dates)
        targets = long_camp_subject_week_targets(weeks, len(items), subject)
        key = (2026, 49)
        if targets.get(key, 0) < 1:
            return False
        count = sum(1 for item in items if week_key(item.candidate.slots) == key)
        return count == 0

    def find_assignment(
        class_id: str,
        subject: str,
        stage: str,
        module: str,
        date_text: str,
        period: str,
    ) -> Optional[scheduler.Assignment]:
        for assignment in result:
            first = assignment.candidate.slots[0]
            if (
                assignment.task.class_id == class_id
                and assignment.task.subject == subject
                and clean(assignment.task.stage) == stage
                and clean(assignment.task.course_module) == module
                and first.date == date_text
                and first.period == period
            ):
                return assignment
        return None

    def replacement(assignment: scheduler.Assignment, date_text: str, period: str) -> Optional[scheduler.Assignment]:
        slots = standard_period_slots(date_text, period)
        if sum(slot.duration_hours for slot in slots) != assignment.task.block_hours:
            return None
        candidate = scheduler.Candidate(
            slots=slots,
            teacher_id=assignment.candidate.teacher_id,
            teacher_name=assignment.candidate.teacher_name,
            room_id=assignment.candidate.room_id,
        )
        first = candidate.slots[0]
        candidate_dates = {slot.date for slot in candidate.slots}
        if candidate_dates & blackout_dates:
            return None
        if Date.fromisoformat(first.date).weekday() == 6:
            return None
        meta = class_metadata.get(assignment.task.class_id, {})
        if meta.get("start_date") and first.date < meta["start_date"]:
            return None
        if meta.get("end_date") and first.date > meta["end_date"]:
            return None
        return scheduler.Assignment(task=assignment.task, candidate=candidate)

    def validate_trial(trial: Sequence[scheduler.Assignment], moved_class_ids: Set[str]) -> bool:
        trial_sorted = scheduler.sorted_assignments(trial)
        if set(teacher_time_conflict_lines(trial_sorted)) - before_teacher_conflicts:
            return False
        if set(schedule_conflict_lines(trial_sorted, class_conflict_groups)) - before_conflicts:
            return False
        input_cache: Dict[str, scheduler.ScheduleInput] = {}
        for class_id in moved_class_ids:
            repair_input = input_cache.get(class_id)
            if repair_input is None:
                repair_input = repair_schedule_input_for_class(data_dir, class_id, class_metadata)
                input_cache[class_id] = repair_input
            if class_stage_order_violation_count(repair_input, trial_sorted, class_id) > class_stage_order_violation_count(
                repair_input,
                result,
                class_id,
            ):
                return False
        for moved in trial_sorted:
            if moved.task.class_id not in moved_class_ids:
                continue
            other_by_date_period: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
            for other in trial_sorted:
                if other.task.task_id == moved.task.task_id:
                    continue
                for key in assignment_date_period_keys(other):
                    other_by_date_period[key].append(other)
            if candidate_conflicts_for_repair(moved, other_by_date_period, class_conflict_groups):
                return False
        return True

    if residual_tail_gap("2765", "数学"):
        late_math = find_assignment("KYJXS2765", "数学", "四轮", "线代", "2026-10-28", "AM")
        bridge_math = find_assignment("KYJXS2765", "数学", "三轮", "真题试卷精讲", "2026-09-05", "AM")
        if late_math and bridge_math:
            late_replacement = replacement(late_math, "2026-11-30", "AM")
            bridge_replacement = replacement(bridge_math, "2026-10-28", "AM")
            if late_replacement and bridge_replacement:
                moved_task_ids = {late_math.task.task_id, bridge_math.task.task_id}
                trial = [
                    assignment
                    for assignment in result
                    if assignment.task.task_id not in moved_task_ids
                ] + [late_replacement, bridge_replacement]
                if validate_trial(trial, {"KYJXS2765"}):
                    result = scheduler.sorted_assignments(trial)
                    lines.append(
                        "尾周连锁挪课: 2765 数学 10/28 四轮线代 -> 11/30 AM；"
                        "9/5 三轮真题试卷精讲 -> 10/28 AM，补足第49周且不掏空第44周"
                    )
                else:
                    lines.append("尾周连锁挪课未执行: 2765 数学链式替换会新增冲突或阶段顺序风险")
        else:
            lines.append("尾周连锁挪课未执行: 2765 数学未找到预期的链式源课节")

    if residual_tail_gap("2766", "英语"):
        lines.append(
            "尾周连锁挪课阻塞: 2766 英语第49周可用白天在 11/30-12/1，"
            "均与固定专业课 KYXY2770 西医互斥；不移动固定西医课表或不延长 KYJXY2766 end_date 时不可排"
        )
    if residual_tail_gap("2770", "政治"):
        lines.append(
            "尾周连锁挪课阻塞: 2770 政治第49周可用白天在 11/30-12/1，"
            "均与固定专业课 KYXY2770 西医互斥；不移动固定西医课表或不延长 KYJXZ2770 end_date 时不可排"
        )
    return scheduler.sorted_assignments(result), lines


def long_camp_minimum_gap_labels(
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    blackout_dates: Set[str],
    target_suite_codes: Optional[Set[str]] = None,
) -> List[str]:
    grouped: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
    for assignment in assignments:
        if not long_camp_balance_assignment(assignment, class_metadata):
            continue
        suite_code = suite_code_for_class(assignment.task.class_id, class_metadata)
        if target_suite_codes is not None and suite_code not in target_suite_codes:
            continue
        grouped[(suite_code, assignment.task.subject)].append(assignment)

    labels: List[Tuple[str, str, Tuple[int, int]]] = []
    for (suite_code, subject), items in grouped.items():
        if len(items) < 2:
            continue
        weeks = long_camp_subject_active_weeks(items, class_metadata, blackout_dates)
        if not weeks:
            continue
        targets = long_camp_subject_week_targets(weeks, len(items), subject)
        loads: Dict[Tuple[int, int], int] = {key: 0 for key in targets}
        for assignment in items:
            loads[week_key(assignment.candidate.slots)] = loads.get(week_key(assignment.candidate.slots), 0) + 1
        for week, target in targets.items():
            if target >= 1 and loads.get(week, 0) == 0:
                labels.append((suite_code, subject, week))
    labels.sort(key=lambda item: (item[0], PUBLIC_SUBJECT_PRIORITY.get(item[1], 99), item[2]))
    return [f"{suite_code} {subject} W{week[1]:02d}" for suite_code, subject, week in labels]


def refresh_long_camp_minimum_gap_report_lines(
    lines: Sequence[str],
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    blackout_dates: Set[str],
    target_suite_codes: Optional[Set[str]] = None,
) -> List[str]:
    final_gaps = long_camp_minimum_gap_labels(
        assignments,
        class_metadata,
        blackout_dates,
        target_suite_codes=target_suite_codes,
    )
    refreshed: List[str] = []
    for line in lines:
        if line.startswith("长线营科目周保底残留:") or line.startswith("长线营科目周保底最终残留:"):
            continue
        if line.startswith("长线营科目周均衡: 移动 "):
            line = re.sub(r"剩余空周 \d+ 组", f"剩余空周 {len(final_gaps)} 组", line)
        refreshed.append(line)
    refreshed.append(
        "长线营科目周保底最终残留: " + ("；".join(final_gaps[:20]) if final_gaps else "无")
    )
    return refreshed


def date_period_in_window(
    date_text: str,
    period: str,
    start_date: str,
    start_period: str,
    end_date: str,
    end_period: str,
) -> bool:
    key = (date_text, scheduler.period_sort_value(period))
    return (
        (start_date, scheduler.period_sort_value(start_period))
        <= key
        <= (end_date, scheduler.period_sort_value(end_period))
    )


def renjie_teacher_assignment(assignment: scheduler.Assignment) -> bool:
    teacher_id = clean(assignment.candidate.teacher_id or assignment.task.teacher_id)
    teacher_name = clean(assignment.candidate.teacher_name or assignment.task.teacher_name)
    return teacher_id in RENJIE_TEACHER_IDS or teacher_name in RENJIE_TEACHER_NAMES


def renjie_must_finish_in_offline_window(assignment: scheduler.Assignment) -> bool:
    return (
        renjie_teacher_assignment(assignment)
        and assignment.task.class_id == "KYZZ2731"
        and clean(assignment.task.stage) == "基础"
    )


def renjie_offline_slot_allowed(date_text: str, period: str) -> bool:
    return any(
        date_period_in_window(
            date_text,
            period,
            start_date,
            start_period,
            end_date,
            end_period,
        )
        for start_date, start_period, end_date, end_period in RENJIE_OFFLINE_WINDOWS
    )


def renjie_offline_assignment_violation(assignment: scheduler.Assignment) -> bool:
    if not renjie_teacher_assignment(assignment):
        return False
    if renjie_must_finish_in_offline_window(assignment):
        return assignment.candidate.room_id == ONLINE_ROOM_ID or any(
            not renjie_offline_slot_allowed(slot.date, slot.period)
            for slot in assignment.candidate.slots
        )
    if assignment.candidate.room_id == ONLINE_ROOM_ID:
        return False
    return any(
        not renjie_offline_slot_allowed(slot.date, slot.period)
        for slot in assignment.candidate.slots
    )


def renjie_offline_target_candidates(
    assignment: scheduler.Assignment,
    result: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
) -> List[scheduler.Candidate]:
    target_slots: List[Tuple[str, str]] = []
    for start_date, start_period, end_date, end_period in RENJIE_OFFLINE_WINDOWS:
        current = Date.fromisoformat(start_date)
        end = Date.fromisoformat(end_date)
        while current <= end:
            date_text = current.isoformat()
            if current.weekday() != 6:
                for period in ("AM", "PM"):
                    if date_period_in_window(date_text, period, start_date, start_period, end_date, end_period):
                        target_slots.append((date_text, period))
            current += timedelta(days=1)

    same_suite = [
        item
        for item in result
        if suite_code_for_class(item.task.class_id, class_metadata)
        == suite_code_for_class(assignment.task.class_id, class_metadata)
    ]
    offline_room_id = assignment.candidate.room_id
    if offline_room_id == ONLINE_ROOM_ID:
        for item in result:
            if item.task.class_id != assignment.task.class_id:
                continue
            if item.candidate.room_id and item.candidate.room_id != ONLINE_ROOM_ID:
                offline_room_id = item.candidate.room_id
                break
    if offline_room_id == ONLINE_ROOM_ID:
        preferred_room_ids = split_pipe_values(class_metadata.get(assignment.task.class_id, {}).get("preferred_room_ids"))
        if preferred_room_ids:
            offline_room_id = sorted(preferred_room_ids)[0]

    def week_counts(candidate: scheduler.Candidate) -> Tuple[int, int]:
        week = week_monday(candidate.slots[0].date)
        suite_count = 0
        subject_count = 0
        for item in same_suite:
            if item.task.task_id == assignment.task.task_id:
                continue
            if week_monday(item.candidate.slots[0].date) != week:
                continue
            suite_count += 1
            if item.task.subject == assignment.task.subject:
                subject_count += 1
        return suite_count, subject_count

    candidates: List[scheduler.Candidate] = []
    for date_text, period in target_slots:
        slots = standard_period_slots(date_text, period)
        if sum(slot.duration_hours for slot in slots) != assignment.task.block_hours:
            continue
        candidates.append(
            scheduler.Candidate(
                slots=slots,
                teacher_id=assignment.candidate.teacher_id,
                teacher_name=assignment.candidate.teacher_name,
                room_id=offline_room_id,
            )
        )
    original_date = Date.fromisoformat(assignment.candidate.slots[0].date)
    candidates.sort(
        key=lambda candidate: (
            week_counts(candidate)[1],
            week_counts(candidate)[0],
            candidate.slots[0].period != assignment.candidate.slots[0].period,
            abs((Date.fromisoformat(candidate.slots[0].date) - original_date).days),
            scheduler.slot_sort_key(candidate.slots[0]),
        )
    )
    return candidates


def repair_renjie_offline_availability(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
    blackout_dates: Set[str],
) -> Tuple[List[scheduler.Assignment], List[str]]:
    result = scheduler.sorted_assignments(list(assignments))
    lines: List[str] = []

    for _ in range(30):
        index_by_task_id = {assignment.task.task_id: index for index, assignment in enumerate(result)}
        violations = [
            assignment
            for assignment in result
            if renjie_offline_assignment_violation(assignment)
        ]
        if not violations:
            break
        moved = False
        for assignment in violations:
            assignment_index = index_by_task_id.get(assignment.task.task_id)
            if assignment_index is None:
                continue
            other_assignments = [
                item for index, item in enumerate(result) if index != assignment_index
            ]
            other_by_date_period: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
            for item in other_assignments:
                for key in assignment_date_period_keys(item):
                    other_by_date_period[key].append(item)
            day_loads = class_teacher_day_loads_for_assignments(
                other_assignments,
                assignment.task.task_id,
            )
            repair_input = repair_schedule_input_for_class(
                data_dir,
                assignment.task.class_id,
                class_metadata,
            )
            old_first = assignment.candidate.slots[0]
            candidate_options: List[Tuple[scheduler.Candidate, str]] = []
            if (
                AUTUMN_START <= old_first.date <= AUTUMN_END
                and not renjie_must_finish_in_offline_window(assignment)
            ):
                candidate_options.append(
                    (
                        scheduler.Candidate(
                            slots=assignment.candidate.slots,
                            teacher_id=assignment.candidate.teacher_id,
                            teacher_name=assignment.candidate.teacher_name,
                            room_id=ONLINE_ROOM_ID,
                        ),
                        "秋季任洁课改为线上",
                    )
                )
            candidate_options.extend(
                (candidate, "移回任洁暑假线下可排窗口")
                for candidate in renjie_offline_target_candidates(assignment, result, class_metadata)
            )
            for candidate, reason in candidate_options:
                first = candidate.slots[0]
                if {slot.date for slot in candidate.slots} & blackout_dates:
                    continue
                class_meta = class_metadata.get(assignment.task.class_id, {})
                if class_meta.get("start_date") and first.date < class_meta["start_date"]:
                    continue
                if class_meta.get("end_date") and first.date > class_meta["end_date"]:
                    continue
                if candidate.room_id != ONLINE_ROOM_ID and not renjie_offline_slot_allowed(first.date, first.period):
                    continue
                same_class_subject_day = any(
                    item.task.class_id == assignment.task.class_id
                    and item.task.subject == assignment.task.subject
                    and {slot.date for slot in item.candidate.slots} & {slot.date for slot in candidate.slots}
                    for item in other_assignments
                )
                if same_class_subject_day:
                    continue
                if not repair_preserves_observed_stage_order(
                    repair_input,
                    result,
                    assignment_index,
                    assignment.task,
                    candidate,
                ):
                    continue
                if not candidate_avoids_same_class_teacher_day_limit(
                    day_loads,
                    assignment.task,
                    candidate,
                ):
                    continue
                candidate_assignment = scheduler.Assignment(
                    task=assignment.task,
                    candidate=candidate,
                )
                if candidate_conflicts_for_repair(
                    candidate_assignment,
                    other_by_date_period,
                    class_conflict_groups,
                ):
                    continue
                result[assignment_index] = candidate_assignment
                result = scheduler.sorted_assignments(result)
                moved = True
                lines.append(
                    f"任洁线下可排窗口修复: {assignment.task.class_id} "
                    f"{assignment.task.subject}/{assignment.task.stage or ''}/{assignment.task.course_module or ''} "
                    f"{old_first.date} {old_first.period} -> {first.date} {first.period}；{reason}"
                )
                break
            if moved:
                break
            lines.append(
                f"任洁线下可排窗口修复失败: {assignment.task.class_id} "
                f"{assignment.task.subject}/{assignment.task.course_module or ''} "
                f"{assignment.candidate.slots[0].date} {assignment.candidate.slots[0].period}"
            )
        if not moved:
            break

    week37 = week_monday("2026-09-13")
    has_2731_week37_politics = any(
        assignment.task.class_id == "KYZZ2731"
        and assignment.task.subject == "政治"
        and week_monday(assignment.candidate.slots[0].date) == week37
        for assignment in result
    )
    if not has_2731_week37_politics:
        preferred_sources = [
            ("2026-09-27", "PM"),
            ("2026-09-26", "PM"),
            ("2026-10-21", "PM"),
            ("2026-10-24", "PM"),
            ("2026-10-11", "AM"),
        ]
        index_by_task_id = {assignment.task.task_id: index for index, assignment in enumerate(result)}
        for source_date, source_period in preferred_sources:
            source_matches = [
                assignment
                for assignment in result
                if assignment.task.class_id == "KYZZ2731"
                and assignment.task.subject == "政治"
                and not renjie_teacher_assignment(assignment)
                and assignment.candidate.slots[0].date == source_date
                and assignment.candidate.slots[0].period == source_period
            ]
            if len(source_matches) != 1:
                continue
            assignment = source_matches[0]
            assignment_index = index_by_task_id.get(assignment.task.task_id)
            if assignment_index is None:
                continue
            candidate = scheduler.Candidate(
                slots=standard_period_slots("2026-09-13", "PM"),
                teacher_id=assignment.candidate.teacher_id,
                teacher_name=assignment.candidate.teacher_name,
                room_id=assignment.candidate.room_id,
            )
            other_assignments = [
                item for index, item in enumerate(result) if index != assignment_index
            ]
            if any(
                item.task.class_id == assignment.task.class_id
                and item.task.subject == assignment.task.subject
                and item.candidate.slots[0].date == "2026-09-13"
                for item in other_assignments
            ):
                continue
            repair_input = repair_schedule_input_for_class(
                data_dir,
                assignment.task.class_id,
                class_metadata,
            )
            if not repair_preserves_observed_stage_order(
                repair_input,
                result,
                assignment_index,
                assignment.task,
                candidate,
            ):
                continue
            day_loads = class_teacher_day_loads_for_assignments(
                other_assignments,
                assignment.task.task_id,
            )
            if not candidate_avoids_same_class_teacher_day_limit(
                day_loads,
                assignment.task,
                candidate,
            ):
                continue
            other_by_date_period: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
            for item in other_assignments:
                for key in assignment_date_period_keys(item):
                    other_by_date_period[key].append(item)
            candidate_assignment = scheduler.Assignment(task=assignment.task, candidate=candidate)
            if candidate_conflicts_for_repair(
                candidate_assignment,
                other_by_date_period,
                class_conflict_groups,
            ):
                continue
            old_first = assignment.candidate.slots[0]
            result[assignment_index] = candidate_assignment
            result = scheduler.sorted_assignments(result)
            lines.append(
                f"2731 秋季周节奏补位: {assignment.task.class_id} "
                f"{assignment.task.subject}/{assignment.task.stage or ''}/{assignment.task.course_module or ''} "
                f"{old_first.date} {old_first.period} -> 2026-09-13 PM"
            )
            break

    if lines:
        remaining = [
            assignment for assignment in result if renjie_offline_assignment_violation(assignment)
        ]
        lines.insert(
            0,
            f"任洁线下可排窗口修复: 剩余违规 {len(remaining)} 节",
        )
    return scheduler.sorted_assignments(result), lines


def adjust_wuyou_makeup_days_in_public_pool(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    blocked_assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    class_conflict_groups: Dict[str, Set[str]],
) -> Tuple[List[scheduler.Assignment], List[str]]:
    time_slots = load_base_time_slots(data_dir)
    result = scheduler.sorted_assignments(list(assignments))
    lines: List[str] = []
    target_task_ids = [
        assignment.task.task_id
        for assignment in result
        if assignment.candidate.slots[0].date == "2026-10-10"
        and class_metadata.get(assignment.task.class_id, {}).get("sub_product") in {"无忧秋", "无忧春", "无忧暑"}
    ]
    for task_id in target_task_ids:
        current = next((assignment for assignment in result if assignment.task.task_id == task_id), None)
        if current is None or current.candidate.slots[0].date != "2026-10-10":
            continue
        placed = False
        for target_date in ("2026-10-07", "2026-10-06"):
            for candidate in relaxed_candidates_on_date(current, time_slots, target_date):
                moved_current = scheduler.Assignment(task=current.task, candidate=candidate)
                pool_without_current = [
                    assignment for assignment in result if assignment.task.task_id != current.task.task_id
                ]
                conflicts = assignments_conflicting_with_candidate(
                    moved_current,
                    [*blocked_assignments, *pool_without_current],
                    class_conflict_groups,
                )
                if not conflicts:
                    result = scheduler.sorted_assignments([*pool_without_current, moved_current])
                    lines.append(
                        f"{current.task.class_id} {current.task.subject} 从 2026-10-10 调整到 {target_date}"
                    )
                    placed = True
                    break
                adjustable_conflicts = [
                    conflict
                    for conflict in conflicts
                    if any(item.task.task_id == conflict.task.task_id for item in result)
                ]
                if len(conflicts) != 1 or len(adjustable_conflicts) != 1:
                    continue
                conflict = adjustable_conflicts[0]
                pool_without_both = [
                    assignment
                    for assignment in result
                    if assignment.task.task_id not in {current.task.task_id, conflict.task.task_id}
                ]
                for conflict_candidate in relaxed_candidates_on_date(
                    conflict,
                    time_slots,
                    "2026-10-10",
                    preferred_period=current.candidate.slots[0].period,
                ):
                    moved_conflict = scheduler.Assignment(task=conflict.task, candidate=conflict_candidate)
                    validation_pool = [*blocked_assignments, *pool_without_both, moved_current]
                    if assignments_conflicting_with_candidate(moved_conflict, validation_pool, class_conflict_groups):
                        continue
                    result = scheduler.sorted_assignments([*pool_without_both, moved_current, moved_conflict])
                    lines.append(
                        f"{current.task.class_id} {current.task.subject} 从 2026-10-10 调整到 {target_date}；"
                        f"{conflict.task.class_id} {conflict.task.subject} 顺延到 2026-10-10"
                    )
                    placed = True
                    break
                if placed:
                    break
            if placed:
                break
        if not placed:
            lines.append(
                f"{current.task.class_id} {current.task.subject} 10/10 调班无法挪到 10/7 或 10/6，"
                "且无法将冲突公共课换到 10/10"
            )
    return scheduler.sorted_assignments(result), lines


def load_existing_output_assignments_for_classes(
    class_ids: Sequence[str],
    start_date: str,
    end_date: str,
) -> List[scheduler.Assignment]:
    if not OUTPUT_CSV.exists():
        return []
    selected_class_ids = set(class_ids)
    rows: List[dict] = []
    seen: Set[Tuple[str, str, str, str, str, str, str]] = set()
    for row in read_csv_rows(OUTPUT_CSV):
        class_id = clean(row.get("class_id"))
        date_value = clean(row.get("date"))
        if class_id not in selected_class_ids:
            continue
        if not date_value or date_value < start_date or date_value > end_date:
            continue
        key = (
            date_value,
            clean(row.get("period")),
            clean(row.get("start_time")),
            clean(row.get("end_time")),
            clean(row.get("duration_hours")),
            class_id,
            clean(row.get("subject")),
            clean(row.get("quarter")),
            clean(row.get("stage")),
            clean(row.get("course_module")),
            clean(row.get("teacher_id") or row.get("teacher_name")),
            clean(row.get("room_id")),
        )
        if key in seen:
            continue
        seen.add(key)
        rows.append(row)
    prefix_seed = "_".join(sorted(selected_class_ids)) or "UNKNOWN"
    prefix_seed = re.sub(r"[^A-Za-z0-9_]+", "_", prefix_seed)[:80]
    return scheduler.sorted_assignments(assignments_from_rows(rows, f"REUSE:{prefix_seed}"))


def build_additional_public_assignments(
    data_dir: Path,
    covered_class_ids: Set[str],
    existing_public_assignments: Sequence[scheduler.Assignment],
    history_rows: Sequence[dict],
    target_suite_codes: Optional[Set[str]] = None,
    target_sub_products: Optional[Set[str]] = None,
    allow_existing_public_adjustment: bool = True,
    fast_scope_locked_filter: bool = False,
    reuse_existing_on_block: bool = True,
) -> Tuple[List[scheduler.Assignment], List[str]]:
    class_metadata = load_class_metadata(data_dir)
    blackout_dates = load_active_blackout_dates(data_dir)
    result: List[scheduler.Assignment] = []
    strategy_lines: List[str] = []

    for sub_product in ADDITIONAL_PUBLIC_PRODUCTS:
        if target_sub_products is not None and sub_product not in target_sub_products:
            continue
        groups = suite_public_class_groups(class_metadata, sub_product, covered_class_ids)
        if target_suite_codes is not None:
            groups = [
                (suite_code, suite_class_ids)
                for suite_code, suite_class_ids in groups
                if suite_code in target_suite_codes
            ]
        if sub_product in WYQC_PRODUCTS:
            groups = sorted(
                groups,
                key=lambda item: (
                    WYQC_SUITE_PRIORITY.get(item[0], 100),
                    suite_planning_key(item[0], item[1], class_metadata),
                ),
            )
        if not groups:
            continue
        for suite_code, suite_class_ids in groups:
            class_ids = ordered_suite_class_ids(suite_class_ids, class_metadata)
            print(f"预排 {sub_product} {suite_code}: {', '.join(class_ids)}", flush=True)
            errors: List[str] = []
            scheduled: Optional[List[scheduler.Assignment]] = None
            scheduled_input: Optional[scheduler.ScheduleInput] = None
            used_label = ""

            protect_modes = (True, False) if allow_existing_public_adjustment else (True,)
            for protect_existing_public in protect_modes:
                locked_source = [*existing_public_assignments, *result]
                if fast_scope_locked_filter:
                    locked_source = relevant_locked_assignments_for_replan(
                        data_dir,
                        class_ids,
                        locked_source,
                        "2026-07-01",
                        AUTUMN_END,
                    )
                locked_public = (
                    locked_source
                    if protect_existing_public
                    else relevant_locked_assignments_for_replan(
                        data_dir,
                        class_ids,
                        locked_source,
                        "2026-07-01",
                        AUTUMN_END,
                        include_teacher=sub_product != SPRINT_CAMP_PRODUCT,
                        include_room=True,
                    )
                )
                room_attempts = (True,) if protect_existing_public else (True, False)
                for use_preferred_rooms in room_attempts:
                    source = load_schedule_input_for_classes(data_dir, class_ids)
                    if sub_product in HISTORY_DEDUCT_PRODUCTS:
                        source = apply_history_deductions(source, history_rows, "2026-07-01")
                    if sub_product in WYQC_PRODUCTS:
                        source = with_wuyou_qc_stage_windows(source)
                    if use_preferred_rooms:
                        source = with_preferred_class_rooms(source, class_metadata)
                    source = replace(
                        source,
                        locked_assignments=[
                            *source.locked_assignments,
                            *locked_public,
                        ],
                    )
                    source = with_conflict_groups_for_locked(data_dir, source, source.locked_assignments)
                    base_input = without_blackout_dates(source, blackout_dates)
                    if sub_product in WUYOU_PRODUCTS:
                        base_input = without_dates(base_input, WUYOU_PRODUCT_BLACKOUT_DATES)
                    monday_options = (False,) if sub_product in WYQC_PRODUCTS else (False, True)
                    for use_monday_exception in monday_options:
                        current_input = with_monday_exception(base_input) if use_monday_exception else base_input
                        label_parts = [
                            "锁定既有公共课" if protect_existing_public else "允许调整既有公共课",
                            "优先教室" if use_preferred_rooms else "同校区教室",
                        ]
                        if use_monday_exception:
                            label_parts.append("周一例外")
                        label = "+".join(label_parts)
                        try:
                            scheduled = run_fast_callback_with_timeout(
                                lambda current=current_input: schedule_public_suite_attempt(
                                    current,
                                    class_ids,
                                    sub_product,
                                    allow_basic_fallback=(
                                        not protect_existing_public
                                        or sub_product in WUYOU_PRODUCTS
                                    ),
                                ),
                                20,
                            )
                            scheduled_input = current_input
                            if not used_label:
                                used_label = label
                            break
                        except ValueError as exc:
                            errors.append(f"{label}: {exc}")
                    if scheduled is not None:
                        break
                if scheduled is not None:
                    break

            if scheduled is None and sub_product not in WUYOU_PRODUCTS and sub_product not in {"全年营", "半年营", SPRINT_CAMP_PRODUCT}:
                locked_source = [*existing_public_assignments, *result]
                locked_public = relevant_locked_assignments_for_replan(
                    data_dir,
                    class_ids,
                    locked_source,
                    "2026-07-01",
                    AUTUMN_END,
                )
                for use_preferred_rooms in (True, False):
                    source = load_schedule_input_for_classes(data_dir, class_ids)
                    if sub_product in HISTORY_DEDUCT_PRODUCTS:
                        source = apply_history_deductions(source, history_rows, "2026-07-01")
                    if sub_product in WYQC_PRODUCTS:
                        source = with_wuyou_qc_stage_windows(source)
                    if use_preferred_rooms:
                        source = with_preferred_class_rooms(source, class_metadata)
                    source = replace(
                        source,
                        locked_assignments=[*source.locked_assignments, *locked_public],
                    )
                    source = with_conflict_groups_for_locked(data_dir, source, source.locked_assignments)
                    base_input = without_blackout_dates(source, blackout_dates)
                    if sub_product in WUYOU_PRODUCTS:
                        base_input = without_dates(base_input, WUYOU_PRODUCT_BLACKOUT_DATES)
                    for use_monday_exception in (False, True):
                        current_input = with_monday_exception(base_input) if use_monday_exception else base_input
                        label_parts = [
                            "基础调度兜底",
                            "优先教室" if use_preferred_rooms else "同校区教室",
                        ]
                        if use_monday_exception:
                            label_parts.append("周一例外")
                        label = "+".join(label_parts)
                        try:
                            scheduled = run_fast_callback_with_timeout(
                                lambda current=current_input: schedule_public_suite_core_fallback(
                                    current,
                                    class_ids,
                                ),
                                40,
                            )
                            scheduled_input = current_input
                            used_label = label
                            break
                        except ValueError as exc:
                            errors.append(f"{label}: {exc}")
                    if scheduled is not None:
                        break
            if scheduled is None:
                detail = "；".join(errors[:8])
                if reuse_existing_on_block:
                    reused = load_existing_output_assignments_for_classes(class_ids, "2026-07-01", AUTUMN_END)
                    if reused:
                        result.extend(reused)
                        start_date = reused[0].candidate.slots[0].date
                        end_date = reused[-1].candidate.slots[0].date
                        strategy_lines.append(
                            f"{sub_product} {suite_code}: 新规则暂时阻塞，沿用上一版 {len(reused)} 节，"
                            f"{start_date} 至 {end_date}；需继续优化；{detail}"
                        )
                        continue
                strategy_lines.append(f"{sub_product} {suite_code}: 阻塞，未生成课表；{detail}")
                continue
            if scheduled_input is None:
                raise ValueError(f"{sub_product} {suite_code} 公共课预排失败: 未记录排课输入")

            result.extend(scheduled)
            start_date = scheduled[0].candidate.slots[0].date if scheduled else ""
            end_date = scheduled[-1].candidate.slots[0].date if scheduled else ""
            history_lessons = 0
            history_hours = 0
            if sub_product in HISTORY_DEDUCT_PRODUCTS:
                class_id_set = set(class_ids)
                for row in history_rows:
                    if row.get("class_id") not in class_id_set or (row.get("date") or "") >= "2026-07-01":
                        continue
                    history_lessons += 1
                    history_hours += int(float(row.get("duration_hours") or 0))
            strategy_lines.append(
                f"{sub_product} {suite_code}: {used_label}，{len(scheduled)} 节，{start_date} 至 {end_date}"
                + ("，按剩余课量周平均配额" if sub_product in {"全年营", "半年营", SPRINT_CAMP_PRODUCT} else "")
                + (f"，已读取历史 {history_lessons} 节/{history_hours} 小时并抵扣" if history_lessons else "")
            )

    return scheduler.sorted_assignments(result), strategy_lines


def assignment_capacity_warning_lines(
    assignments: Sequence[scheduler.Assignment],
    rooms: Dict[str, scheduler.Room],
    room_names: Dict[str, str],
) -> List[str]:
    seen: Set[Tuple[str, str, str, Optional[int], Optional[int]]] = set()
    lines: List[str] = []
    for assignment in scheduler.sorted_assignments(list(assignments)):
        room = rooms.get(assignment.candidate.room_id)
        shortfall = scheduler.room_capacity_shortfall(room, assignment.task.class_size)
        if shortfall <= 0:
            continue
        key = (
            assignment.task.stage or "",
            assignment.task.class_id,
            assignment.candidate.room_id,
            assignment.task.class_size,
            room.capacity if room else None,
        )
        if key in seen:
            continue
        seen.add(key)
        room_name = room_names.get(assignment.candidate.room_id, assignment.candidate.room_id)
        lines.append(
            f"{assignment.task.stage or '未分阶段'} {assignment.task.class_id} {assignment.task.class_name}: "
            f"班级人数 {assignment.task.class_size} 人，教室 {room_name} 容量 {room.capacity if room else '未知'} 人，"
            f"差异 {shortfall} 人；已继续排课，仅作容量提醒。"
        )
    return lines


def teacher_same_day_campus_warning_lines(
    assignments: Sequence[scheduler.Assignment],
    rooms: Dict[str, scheduler.Room],
    area_travel_minutes: Dict[Tuple[str, str], int],
) -> List[str]:
    by_teacher_date: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
    for assignment in assignments:
        teacher_key = candidate_teacher_key(assignment.candidate)
        if not teacher_key:
            continue
        for date_text in {slot.date for slot in assignment.candidate.slots}:
            by_teacher_date[(teacher_key, date_text)].append(assignment)

    def area_label(room: Optional[scheduler.Room], fallback_room_id: str) -> str:
        if room and room.teaching_area_name:
            return room.teaching_area_name
        return fallback_room_id

    def compact_class_ids(items: Sequence[scheduler.Assignment]) -> str:
        class_ids: List[str] = []
        seen_ids: Set[str] = set()
        for item in sorted(items, key=lambda value: value.task.class_id):
            class_id = item.task.class_id
            if class_id in seen_ids:
                continue
            seen_ids.add(class_id)
            class_ids.append(class_id)
        if len(class_ids) <= 6:
            return "、".join(class_ids)
        return "、".join(class_ids[:6]) + f" 等 {len(class_ids)} 个班"

    lines: List[str] = []
    seen: Set[Tuple[str, str, str, str, str, str]] = set()
    for (teacher_key, date_text), items in sorted(by_teacher_date.items()):
        if len(items) < 2:
            continue
        area_period_groups: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
        for item in items:
            date_slots = [slot for slot in item.candidate.slots if slot.date == date_text]
            if not date_slots:
                continue
            period = sorted({slot.period for slot in date_slots}, key=scheduler.period_sort_value)[0]
            room = rooms.get(item.candidate.room_id)
            area_id = scheduler.room_area_id(room) or item.candidate.room_id
            area_period_groups[(period, area_id)].append(item)

        groups = sorted(
            area_period_groups.items(),
            key=lambda pair: (scheduler.period_sort_value(pair[0][0]), pair[0][1]),
        )
        for left_index, ((left_period, left_area), left_items) in enumerate(groups):
            left = left_items[0]
            left_room = rooms.get(left.candidate.room_id)
            for right_period, right_area in [key for key, _items in groups[left_index + 1 :]]:
                if left_period == right_period:
                    continue
                right_items = area_period_groups[(right_period, right_area)]
                right = right_items[0]
                right_room = rooms.get(right.candidate.room_id)
                if left_area and right_area and left_area == right_area:
                    continue
                if scheduler.same_region(left_room, right_room):
                    continue
                high_risk = (
                    scheduler.is_new_station_area(left_room) and scheduler.is_new_station_avoid_target(right_room)
                ) or (
                    scheduler.is_new_station_area(right_room) and scheduler.is_new_station_avoid_target(left_room)
                )
                minutes = area_travel_minutes.get(scheduler.area_pair_key(left_area, right_area), 0)
                if not high_risk and minutes <= 30:
                    continue
                key = (teacher_key, date_text, left_period, right_period, left_area, right_area)
                if key in seen:
                    continue
                seen.add(key)
                teacher = left.candidate.teacher_name or right.candidate.teacher_name or teacher_key
                label = "新站往返滨湖/经开/翡翠湖" if high_risk else "跨区域"
                lines.append(
                    f"{date_text} {teacher}: {left_period} "
                    f"{area_label(left_room, left.candidate.room_id)}（{compact_class_ids(left_items)}）"
                    f" -> {right_period} "
                    f"{area_label(right_room, right.candidate.room_id)}（{compact_class_ids(right_items)}）"
                    f"，{label}，约 {minutes or '未知'} 分钟。"
                )
    return lines[:200]


def assignment_label(assignment: scheduler.Assignment) -> str:
    first = assignment.candidate.slots[0]
    last = assignment.candidate.slots[-1]
    return (
        f"{assignment.task.class_id} {assignment.task.subject}/"
        f"{assignment.task.stage or '未分阶段'}/{assignment.task.course_module or ''} "
        f"{assignment.candidate.teacher_name or assignment.candidate.teacher_id} "
        f"{first.start_time or ''}-{last.end_time or ''}"
    ).strip()


def load_conflict_group_lookup(data_dir: Path) -> Tuple[Dict[str, Set[str]], Dict[str, Set[str]]]:
    input_path = data_dir / "scheduler_input_draft.json"
    if not input_path.exists():
        return {}, {}
    data = load_scheduler_input_data(data_dir)
    conflict_groups: Dict[str, Set[str]] = {}
    class_conflict_groups: Dict[str, Set[str]] = defaultdict(set)
    for index, group in enumerate(data.get("conflict_groups", []), start=1):
        group_id = clean(group.get("id")) or f"CONFLICT_{index}"
        class_ids = {clean(class_id) for class_id in group.get("class_ids", []) if clean(class_id)}
        if len(class_ids) < 2:
            continue
        conflict_groups[group_id] = class_ids
        for class_id in class_ids:
            class_conflict_groups[class_id].add(group_id)
    return conflict_groups, dict(class_conflict_groups)


def schedule_conflict_lines(
    assignments: Sequence[scheduler.Assignment],
    class_conflict_groups: Dict[str, Set[str]],
) -> List[str]:
    lines: List[str] = []
    seen: Set[str] = set()

    def append_conflict(scope: str, key: Tuple[str, ...], group: List[scheduler.Assignment]) -> None:
        for left_index, left in enumerate(group):
            for right in group[left_index + 1 :]:
                if not assignments_overlap(left, right):
                    continue
                first = left.candidate.slots[0]
                detail = (
                    f"{scope} {' '.join(key)} {first.date} {first.period}: "
                    f"{assignment_label(left)}；{assignment_label(right)}"
                )
                if detail not in seen:
                    seen.add(detail)
                    lines.append(detail)

    class_groups: Dict[Tuple[str, str, str], List[scheduler.Assignment]] = defaultdict(list)
    slot_groups: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
    for assignment in assignments:
        first = assignment.candidate.slots[0]
        class_groups[(assignment.task.class_id, first.date, first.period)].append(assignment)
        slot_groups[(first.date, first.period)].append(assignment)

    for (class_id, _date, _period), group in sorted(class_groups.items()):
        if len(group) > 1:
            append_conflict("班级", (class_id,), group)
    for (_date, _period), group in sorted(slot_groups.items()):
        for left_index, left in enumerate(group):
            left_groups = class_conflict_groups.get(left.task.class_id, set())
            if not left_groups:
                continue
            for right in group[left_index + 1 :]:
                if left.task.class_id == right.task.class_id or not assignments_overlap(left, right):
                    continue
                shared_groups = left_groups & class_conflict_groups.get(right.task.class_id, set())
                if shared_groups:
                    append_conflict("互斥组", (sorted(shared_groups)[0],), [left, right])
    return lines[:300]


def class_is_locked_for_coverage(class_meta: Dict[str, str]) -> bool:
    return parse_bool(class_meta.get("is_schedule_locked"))


def public_coverage_gap_rows_from_totals(
    expected_totals: Counter[str],
    scheduled_totals: Counter[str],
    class_metadata: Dict[str, Dict[str, str]],
) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    for class_id in sorted(expected_totals):
        class_meta = class_metadata.get(class_id, {})
        if class_is_locked_for_coverage(class_meta):
            continue
        if clean(class_meta.get("subject_category")) != "公共课":
            continue
        expected = float(expected_totals[class_id])
        scheduled = float(scheduled_totals[class_id])
        gap = expected - scheduled
        if gap <= 0.01:
            continue
        rows.append(
            {
                "class_id": class_id,
                "class_name": class_meta.get("name") or class_id,
                "sub_product": class_meta.get("sub_product") or "",
                "subject": class_meta.get("subject") or "",
                "suite_code": class_meta.get("suite_code") or "",
                "expected_hours": expected,
                "scheduled_hours": scheduled,
                "gap_hours": gap,
            }
        )
    rows.sort(key=lambda row: (-float(row["gap_hours"]), str(row["sub_product"]), str(row["suite_code"]), str(row["class_id"])))
    return rows


def public_coverage_gap_rows_for_assignments(
    data_dir: Path,
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
) -> List[Dict[str, object]]:
    schedule_input = scheduler.load_input(data_dir / "scheduler_input_draft.json")
    expected_totals: Counter[str] = Counter()
    for class_id, school_class in schedule_input.classes.items():
        expected_totals[class_id] += float(sum(requirement.total_hours for requirement in school_class.requirements))
    scheduled_totals: Counter[str] = Counter()
    for assignment in assignments:
        scheduled_totals[assignment.task.class_id] += float(sum(slot.duration_hours for slot in assignment.candidate.slots))
    return public_coverage_gap_rows_from_totals(expected_totals, scheduled_totals, class_metadata)


def coverage_gap_blocking_lines(gap_rows: Sequence[Dict[str, object]], max_items: int = 30) -> List[str]:
    if not gap_rows:
        return []
    total_gap = sum(float(row["gap_hours"]) for row in gap_rows)
    lines = [f"班级总课时缺口 {len(gap_rows)} 个班，合计 {total_gap:.1f} 小时"]
    for row in gap_rows[:max_items]:
        lines.append(
            "{class_id} {sub_product}/{subject} 套班{suite_code}: 应排 {expected_hours:.1f}h，"
            "已排 {scheduled_hours:.1f}h，缺 {gap_hours:.1f}h".format(**row)
        )
    if len(gap_rows) > max_items:
        lines.append(f"... 另有 {len(gap_rows) - max_items} 个班级缺口未列出")
    return lines


def summer_distribution_warning_lines(assignments: Sequence[scheduler.Assignment]) -> List[str]:
    teacher_dates: Dict[Tuple[str, str, str], List[str]] = defaultdict(list)
    same_subject_day_counts: Dict[Tuple[str, str, str], int] = defaultdict(int)
    class_teacher_day_hours: Dict[Tuple[str, str, str, str, str], float] = defaultdict(float)
    suite_week_loads: Dict[str, Dict[Tuple[int, int], int]] = defaultdict(lambda: defaultdict(int))
    for assignment in assignments:
        if not assignment_matches_phase(assignment, SUMMER_PREPLAN_STAGES) or assignment.task.subject not in SUMMER_PUBLIC_SUBJECTS:
            continue
        suite_code = suite_code_from_class_id(assignment.task.class_id)
        if not suite_code:
            continue
        first = assignment.candidate.slots[0]
        teacher = assignment.candidate.teacher_name or assignment.candidate.teacher_id or "未分老师"
        teacher_dates[(suite_code, assignment.task.subject, teacher)].append(first.date)
        same_subject_day_counts[(suite_code, assignment.task.subject, first.date)] += 1
        suite_week_loads[suite_code][week_key(assignment.candidate.slots)] += 1
        teacher_key = candidate_teacher_key(assignment.candidate)
        if teacher_key:
            for date_text, hours in candidate_hours_by_date(assignment.candidate).items():
                class_teacher_day_hours[(suite_code, assignment.task.class_id, teacher, teacher_key, date_text)] += hours

    lines: List[str] = []
    for suite_code, subject, week, count, weekly_min, weekly_max in summer_subject_week_rule_violations(assignments):
        lines.append(
            f"{suite_code} {subject} {week[0]}年第{week[1]}周: {count} 个半天，不符合每周 {weekly_min}-{weekly_max} 个半天"
        )
    for (suite_code, class_id, teacher, _teacher_key, date_text), hours in sorted(class_teacher_day_hours.items()):
        if hours >= 8:
            lines.append(f"{suite_code} {class_id} {date_text} {teacher}: 同班同师同天 {hours:g} 小时，需重排")
    for suite_code, loads in sorted(suite_week_loads.items()):
        for week, count in sorted(loads.items()):
            if count >= 12:
                lines.append(f"{suite_code} {week[0]}年第{week[1]}周: 已排满 {count} 个半天，建议继续分散")
    for (suite_code, subject, teacher), dates in sorted(teacher_dates.items()):
        run_length = max_consecutive_dates(dates)
        if subject in {"英语", "政治"} and run_length > 3:
            lines.append(f"{suite_code} {subject} {teacher}: 连续 {run_length} 天，超过 3 天需重排")
        elif subject in {"英语", "政治"} and run_length == 3:
            lines.append(f"{suite_code} {subject} {teacher}: 连续 3 天，建议人工核对")
        elif subject == "数学" and run_length >= 4:
            lines.append(f"{suite_code} {subject} {teacher}: 连续 {run_length} 天，通常由高密度课量/资源占用导致，需人工确认")
    for (suite_code, subject, date_text), count in sorted(same_subject_day_counts.items()):
        if count > 1:
            lines.append(f"{suite_code} {date_text} {subject}: 同一天 {count} 个半天")
    return lines[:200]


def week_monday(date_text_value: str) -> Date:
    value = Date.fromisoformat(date_text_value)
    return value - timedelta(days=value.weekday())


def iter_week_mondays(start: Date, end: Date) -> Iterable[Date]:
    current = start
    while current <= end:
        yield current
        current += timedelta(days=7)


def week_display_label(monday: Date) -> str:
    iso = monday.isocalendar()
    return f"{iso.year}年第{iso.week}周({monday.isoformat()}起)"


def evenly_spaced_items(items: Sequence[Any], count: int) -> List[Any]:
    if count <= 0:
        return []
    if count >= len(items):
        return list(items)
    if count == 1:
        return [items[len(items) // 2]]
    indexes = {
        round(index * (len(items) - 1) / (count - 1))
        for index in range(count)
    }
    result = [items[index] for index in sorted(indexes)]
    cursor = 0
    while len(result) < count and cursor < len(items):
        candidate = items[cursor]
        if candidate not in result:
            result.append(candidate)
        cursor += 1
    return list(result)


def student_experience_phase(date_value: str, sub_product: str) -> Optional[str]:
    if date_value < "2026-06-25" or date_value > AUTUMN_END:
        return None
    if "2026-06-25" <= date_value <= "2026-08-31":
        return "暑假"
    if "2026-09-01" <= date_value <= AUTUMN_END:
        return "秋季"
    return None


def student_experience_weekly_max(sub_product: str, subjects: Set[str], phase: str) -> int:
    has_math = "数学" in subjects
    if sub_product in WUYOU_PRODUCTS:
        return wuyou_summer_weekly_total_max(subjects)
    if sub_product == "寒暑营":
        return 10 if has_math else 9
    if sub_product == "暑假营":
        return 10 if has_math else 8
    if sub_product in {"全年营", "半年营"}:
        return 11 if has_math else 7
    if sub_product == "冲刺营":
        return 8 if has_math else 5
    return 10 if has_math else 8


def blackout_heavy_week(monday: Date, blackout_dates: Set[str]) -> bool:
    if not blackout_dates:
        return False
    dates = {(monday + timedelta(days=offset)).isoformat() for offset in range(7)}
    return len(dates & blackout_dates) >= 4


def student_experience_warning_lines(
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Dict[str, Dict[str, str]],
    blackout_dates: Set[str],
) -> List[str]:
    grouped: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    same_class_subject_day: Dict[Tuple[str, str, str], Dict[str, Any]] = defaultdict(
        lambda: {"count": 0, "hours": 0.0, "suite_code": "", "sub_product": ""}
    )
    for assignment in assignments:
        subject = assignment.task.subject
        if subject not in SUMMER_PUBLIC_SUBJECTS:
            continue
        class_id = assignment.task.class_id
        meta = class_metadata.get(class_id, {})
        if meta.get("subject_category") and meta.get("subject_category") != "公共课":
            continue
        sub_product = assignment_sub_product(assignment, class_metadata)
        if sub_product not in PUBLIC_PRODUCT_ORDER:
            continue
        first = assignment.candidate.slots[0]
        phase = student_experience_phase(first.date, sub_product)
        if not phase:
            continue
        suite_code = suite_code_for_class(class_id, class_metadata)
        if not suite_code:
            continue
        for date_text, hours in candidate_hours_by_date(assignment.candidate).items():
            if "2026-06-25" <= date_text <= AUTUMN_END:
                day_bucket = same_class_subject_day[(class_id, subject, date_text)]
                day_bucket["count"] += 1
                day_bucket["hours"] += hours
                day_bucket["suite_code"] = suite_code
                day_bucket["sub_product"] = sub_product
        key = (sub_product, suite_code, phase)
        bucket = grouped.setdefault(
            key,
            {
                "subjects": set(),
                "weeks": defaultdict(lambda: {"count": 0, "subjects": set(), "subject_counts": defaultdict(int), "dates": set()}),
            },
        )
        bucket["subjects"].add(subject)
        week = week_monday(first.date)
        week_info = bucket["weeks"][week]
        week_info["count"] += 1
        week_info["subjects"].add(subject)
        week_info["subject_counts"][subject] += 1
        week_info["dates"].add(first.date)

    lines: List[str] = []
    for (sub_product, suite_code, phase), bucket in sorted(grouped.items()):
        subjects: Set[str] = set(bucket["subjects"])
        weeks: Dict[Date, Dict[str, Any]] = bucket["weeks"]
        if not weeks:
            continue
        comfortable_max = student_experience_weekly_max(sub_product, subjects, phase)
        active_weeks = sorted(weeks)
        first_week = active_weeks[0]
        last_week = active_weeks[-1]
        expected_subjects = subjects & SUMMER_PUBLIC_SUBJECTS
        check_subject_mix = phase == "暑假" or sub_product in {"全年营", "半年营"}
        balance_delta_limit = 1 if sub_product in {"全年营", "半年营", SPRINT_CAMP_PRODUCT} else 2
        if sub_product in WYQC_PRODUCTS and phase == "秋季":
            expected_weeks = [
                week
                for week in iter_week_mondays(
                    week_monday(WYQC_AUTUMN_START),
                    week_monday(WYQC_AUTUMN_END),
                )
                if not blackout_heavy_week(week, blackout_dates)
            ]
        else:
            expected_weeks = [
                week
                for week in iter_week_mondays(first_week, last_week)
                if not blackout_heavy_week(week, blackout_dates)
            ]
        if len(expected_weeks) >= 3:
            total_values = [int(weeks.get(week, {}).get("count", 0)) for week in expected_weeks]
            if total_values and max(total_values) - min(total_values) > balance_delta_limit:
                lines.append(
                    f"{sub_product} {suite_code} {phase}: 公共课整体周课量不均衡，"
                    f"最低 {min(total_values)}、最高 {max(total_values)} 个半天，建议继续均摊"
                )
            for subject in sorted(expected_subjects, key=lambda item: PUBLIC_SUBJECT_PRIORITY.get(item, 99)):
                subject_values = [
                    int(weeks.get(week, {}).get("subject_counts", {}).get(subject, 0))
                    for week in expected_weeks
                ]
                if subject_values and max(subject_values) > 0 and max(subject_values) - min(subject_values) > balance_delta_limit:
                    lines.append(
                        f"{sub_product} {suite_code} {phase} {subject}: 分科目周课量不均衡，"
                        f"最低 {min(subject_values)}、最高 {max(subject_values)} 个半天"
                    )
        for week in active_weeks:
            info = weeks[week]
            count = int(info["count"])
            week_subjects: Set[str] = set(info["subjects"])
            if sub_product == WYS_PRODUCT and suite_code == WYS_STAGELESS_WEEKLY_SUITE:
                if phase == "秋季":
                    if count < WYS_STAGELESS_AUTUMN_WEEKLY_MIN or count > WYS_STAGELESS_AUTUMN_WEEKLY_MAX:
                        lines.append(
                            f"{sub_product} {suite_code} {phase} {week_display_label(week)}: "
                            f"{count} 个半天，需保持每周 {WYS_STAGELESS_AUTUMN_WEEKLY_MIN}-{WYS_STAGELESS_AUTUMN_WEEKLY_MAX} 个半天"
                        )
                    continue
                if phase == "暑假" and count > WYS_STAGELESS_SUMMER_WEEKLY_MAX:
                    lines.append(
                        f"{sub_product} {suite_code} {phase} {week_display_label(week)}: "
                        f"{count} 个半天，高于上限 {WYS_STAGELESS_SUMMER_WEEKLY_MAX}"
                    )
                    continue
            if count > comfortable_max:
                lines.append(
                    f"{sub_product} {suite_code} {phase} {week_display_label(week)}: "
                    f"{count} 个半天，高于舒适上限 {comfortable_max}，建议继续分散"
                )
            if (
                check_subject_mix
                and week not in {first_week, last_week}
                and len(expected_subjects) > 1
                and week_subjects
                and week_subjects != expected_subjects
            ):
                missing = "、".join(sorted(expected_subjects - week_subjects, key=lambda item: PUBLIC_SUBJECT_PRIORITY.get(item, 99)))
                if len(week_subjects) == 1:
                    only_subject = next(iter(week_subjects))
                    lines.append(
                        f"{sub_product} {suite_code} {phase} {week_display_label(week)}: "
                        f"只有{only_subject} {count} 个半天，建议补入其他科目或前后调匀"
                    )
                elif missing:
                    lines.append(
                        f"{sub_product} {suite_code} {phase} {week_display_label(week)}: "
                        f"缺少{missing}，建议避免线下周科目过单一"
                    )
        for week in iter_week_mondays(first_week, last_week):
            if week in weeks or blackout_heavy_week(week, blackout_dates):
                continue
            lines.append(
                f"{sub_product} {suite_code} {phase} {week_display_label(week)}: "
                "中间整周无课，建议前后课节拉开填补"
            )
        if sub_product in WYQC_PRODUCTS and phase == "秋季":
            for subject in sorted(expected_subjects, key=lambda item: PUBLIC_SUBJECT_PRIORITY.get(item, 99)):
                subject_total = sum(
                    int(info.get("subject_counts", {}).get(subject, 0))
                    for info in weeks.values()
                )
                if 0 < subject_total < len(expected_weeks):
                    lines.append(
                        f"{sub_product} {suite_code} 秋季 {subject}: "
                        f"剩余 {subject_total} 个半天少于 {len(expected_weeks)} 个周段，无法做到每周都有，已按可覆盖周数均匀分布"
                    )
                target_weeks = set(evenly_spaced_items(expected_weeks, subject_total))
                for week in expected_weeks:
                    if week not in target_weeks:
                        continue
                    subject_counts = weeks.get(week, {}).get("subject_counts", {})
                    if int(subject_counts.get(subject, 0)) < 1:
                        lines.append(
                            f"{sub_product} {suite_code} 秋季 {week_display_label(week)}: "
                            f"{subject} 0 个半天，低于每周至少 1 个半天"
                        )
            late_dates = sorted(
                date_value
                for week_info in weeks.values()
                for date_value in week_info["dates"]
                if date_value > WYQC_AUTUMN_END
            )
            if late_dates:
                lines.append(
                    f"{sub_product} {suite_code} 秋季: 存在 {late_dates[0]} 之后课程，需在 {WYQC_AUTUMN_END} 前排完"
                )
            early_dates = sorted(
                date_value
                for week_info in weeks.values()
                for date_value in week_info["dates"]
                if date_value < WYQC_AUTUMN_START
            )
            if early_dates:
                lines.append(
                    f"{sub_product} {suite_code} 秋季: 存在 {early_dates[0]} 之前课程，需从 {WYQC_AUTUMN_START} 开始排"
                )
    for (class_id, subject, date_text), day_bucket in sorted(same_class_subject_day.items()):
        if day_bucket["count"] > 1 or day_bucket["hours"] >= PUBLIC_SAME_CLASS_SUBJECT_DAY_HOURS:
            lines.append(
                f"{day_bucket['sub_product']} {day_bucket['suite_code']} {class_id} {date_text} {subject}: "
                f"同班同科同日 {day_bucket['hours']:g} 小时，建议拆到其他日期"
            )
    return lines[:300]


def long_camp_subject_week_balance_warning_lines(
    assignments: Sequence[scheduler.Assignment],
    class_metadata: Optional[Dict[str, Dict[str, str]]] = None,
    blackout_dates: Optional[Set[str]] = None,
) -> List[str]:
    class_metadata = class_metadata or {}
    blackout_dates = blackout_dates or set()
    subject_weeks: Dict[Tuple[str, str], Dict[Tuple[int, int], int]] = defaultdict(lambda: defaultdict(int))
    subject_blocks: Dict[Tuple[str, str], List[Tuple[scheduler.TimeSlot, ...]]] = defaultdict(list)
    subject_assignments: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
    for assignment in assignments:
        suite_code = suite_code_from_class_id(assignment.task.class_id)
        if suite_code not in HALF_YEAR_BATCH_SUITES and "全年营" not in assignment.task.class_name:
            continue
        if assignment.task.subject not in SUMMER_PUBLIC_SUBJECTS:
            continue
        subject_weeks[(suite_code, assignment.task.subject)][week_key(assignment.candidate.slots)] += 1
        subject_blocks[(suite_code, assignment.task.subject)].append(assignment.candidate.slots)
        subject_assignments[(suite_code, assignment.task.subject)].append(assignment)

    lines: List[str] = []
    for (suite_code, subject), loads in sorted(subject_weeks.items()):
        if len(loads) < 2:
            continue
        active_weeks = long_camp_subject_active_weeks(
            subject_assignments.get((suite_code, subject), []),
            class_metadata,
            blackout_dates,
        )
        targets = long_camp_subject_week_targets(active_weeks, sum(loads.values()), subject)
        empty_required_weeks = [
            f"W{week[1]:02d}"
            for week, target in sorted(targets.items())
            if target >= 1 and loads.get(week, 0) == 0
        ]
        if empty_required_weeks:
            lines.append(
                f"{suite_code} {subject}: 可排周内存在空周，未达到每周至少 1 个半天；"
                + "，".join(empty_required_weeks)
            )
            continue
        bounds = average_subject_week_bounds_from_counts(
            {subject: subject_blocks.get((suite_code, subject), [])},
            {subject: sum(loads.values())},
            {"数学": 4, "政治": 3, "英语": 4},
        )
        weekly_min, weekly_max = bounds.get(subject, (None, None))
        values = list(loads.values())
        load_text = "，".join(
            f"W{week:02d}:{loads.get((year, week), 0)}"
            for year, week in sorted(targets or loads)
        )
        out_of_bounds_weeks = [
            f"W{week:02d}:{loads.get((year, week), 0)}"
            for year, week in sorted(targets or loads)
            if (weekly_min is not None and loads.get((year, week), 0) < weekly_min)
            or (weekly_max is not None and loads.get((year, week), 0) > weekly_max)
        ]
        if out_of_bounds_weeks:
            bound_text = f"{weekly_min or 0}-{weekly_max or '不限'}"
            lines.append(
                f"{suite_code} {subject}: 存在不符合每周 {bound_text} 个半天的周；"
                + "，".join(out_of_bounds_weeks)
            )
            continue
        span = max(values) - min(values)
        if span <= 2:
            continue
        lines.append(
            f"{suite_code} {subject}: 周课量 {min(values)}-{max(values)} 个半天，"
            f"仍有局部波动；{load_text}"
        )
    return lines[:200]


def long_camp_sequence_warning_lines(
    assignments: Sequence[scheduler.Assignment],
) -> List[str]:
    by_subject_dates: Dict[Tuple[str, str], Set[str]] = defaultdict(set)
    by_subject_assignments: Dict[Tuple[str, str], List[scheduler.Assignment]] = defaultdict(list)
    for assignment in assignments:
        suite_code = suite_code_from_class_id(assignment.task.class_id)
        if suite_code not in HALF_YEAR_BATCH_SUITES and "全年营" not in assignment.task.class_name:
            continue
        if assignment.task.subject not in {"英语", "政治", "数学"}:
            continue
        key = (suite_code, assignment.task.subject)
        by_subject_dates[key].add(assignment.candidate.slots[0].date)
        by_subject_assignments[key].append(assignment)

    lines: List[str] = []
    for (suite_code, subject), items in sorted(by_subject_assignments.items()):
        by_date: Dict[str, int] = defaultdict(int)
        for item in items:
            by_date[item.candidate.slots[0].date] += 1
        for date_text, count in sorted(by_date.items()):
            if count > 1:
                lines.append(f"{suite_code} {subject}: {date_text} 同一天排了 {count} 个半天，建议核对")

    for (suite_code, subject), dates in sorted(by_subject_dates.items()):
        if subject == "数学":
            for run in run_dates_over_limit(dates, LONG_CAMP_MATH_MAX_CONSECUTIVE_DAYS):
                lines.append(
                    f"{suite_code} 数学: {run[0]} 至 {run[-1]} 连续 {len(run)} 天上课，建议核对"
                )
            continue
        ordered_dates = sorted(Date.fromisoformat(value) for value in dates)
        for left, right in zip(ordered_dates, ordered_dates[1:]):
            if (right - left).days == 1:
                lines.append(f"{suite_code} {subject}: {left.isoformat()} 和 {right.isoformat()} 连续两天排课，建议核对")

    for (suite_code, subject), items in sorted(by_subject_assignments.items()):
        if subject not in {"英语", "政治"}:
            continue
        items = sorted(items, key=lambda item: scheduler.slot_sort_key(item.candidate.slots[0]))
        teachers = {candidate_teacher_key(item.candidate) for item in items if candidate_teacher_key(item.candidate)}
        if len(teachers) <= 1:
            continue
        for left, right in zip(items, items[1:]):
            left_teacher = candidate_teacher_key(left.candidate)
            right_teacher = candidate_teacher_key(right.candidate)
            if not left_teacher or left_teacher != right_teacher:
                continue
            lines.append(
                f"{suite_code} {subject}: {left.candidate.slots[0].date} 后接 "
                f"{right.candidate.slots[0].date} 仍为 {left.candidate.teacher_name or left_teacher}，未交替"
            )
    return lines[:200]


def write_report(
    path: Path,
    history_rows: Sequence[dict],
    locked_assignments: Sequence[scheduler.Assignment],
    summer_assignments: Sequence[scheduler.Assignment],
    halfyear_assignments: Sequence[scheduler.Assignment],
    halfyear_strategy_lines: Sequence[str],
    additional_public_assignments: Sequence[scheduler.Assignment],
    additional_public_strategy_lines: Sequence[str],
    conflict_rebuilt_summer_suites: Sequence[str],
    distribution_rebuilt_summer_suites: Sequence[str],
    missing_summer_suites: Sequence[str],
    autumn_actual_assignments: Sequence[scheduler.Assignment],
    autumn_display_assignments: Sequence[scheduler.Assignment],
    final_assignments: Sequence[scheduler.Assignment],
    class_conflict_groups: Dict[str, Set[str]],
    warnings: Sequence[str],
    ignored: Sequence[str],
    online_merge_lines: Sequence[str],
    capacity_warnings: Sequence[str],
    teacher_campus_warnings: Sequence[str],
    global_repair_lines: Sequence[str] = (),
    class_metadata: Optional[Dict[str, Dict[str, str]]] = None,
    blackout_dates: Optional[Set[str]] = None,
) -> None:
    class_metadata = class_metadata or load_class_metadata(Path("data"))
    blackout_dates = blackout_dates or set()
    counts: Dict[Tuple[str, str], int] = defaultdict(int)
    for row in history_rows:
        counts[(row["stage"], row["class_id"])] += 1
    overlap_lines = autumn_teacher_overlap_lines(autumn_actual_assignments)
    conflict_lines = schedule_conflict_lines(final_assignments, class_conflict_groups)
    teacher_conflict_lines = teacher_time_conflict_lines(final_assignments)
    summer_distribution_lines = summer_distribution_warning_lines(summer_assignments)
    long_camp_assignments = [*halfyear_assignments, *additional_public_assignments]
    long_camp_balance_lines = long_camp_subject_week_balance_warning_lines(
        long_camp_assignments,
        class_metadata,
        blackout_dates,
    )
    long_camp_sequence_lines = long_camp_sequence_warning_lines(long_camp_assignments)
    student_experience_lines = student_experience_warning_lines(final_assignments, class_metadata, blackout_dates)
    lines = [
        "# 课表维护总表构建报告",
        "",
        f"- 历史/已排课节: {len(history_rows)}",
        f"- 专业课固定课表课节: {assignment_standard_lesson_count(locked_assignments)}",
        f"- 暑假预排课节: {assignment_standard_lesson_count(summer_assignments)}",
        f"- 半年营 2770-2779 预排课节: {assignment_standard_lesson_count(halfyear_assignments)}",
        f"- 其他公共课产品预排课节: {assignment_standard_lesson_count(additional_public_assignments)}",
        f"- 暑假冲突自动重排套班: {len(conflict_rebuilt_summer_suites)}",
        f"- 暑假均衡/班级窗口自动重排套班: {len(distribution_rebuilt_summer_suites)}",
        f"- 暑假缺班级排课窗口套班: {len(missing_summer_suites)}",
        f"- 秋季实际排课课节: {assignment_standard_lesson_count(autumn_actual_assignments)}",
        f"- 秋季合并显示课节: {assignment_standard_lesson_count(autumn_display_assignments)}",
        f"- 班级/套班同时间冲突: {len(conflict_lines)}",
        f"- 老师同时间硬冲突: {len(teacher_conflict_lines)}",
        f"- 全局消冲突记录: {len(global_repair_lines)}",
        f"- 暑假均衡残留提示: {len(summer_distribution_lines)}",
        f"- 长线营科目周均衡提醒: {len(long_camp_balance_lines)}",
        f"- 长线营科目/老师连续提醒: {len(long_camp_sequence_lines)}",
        f"- 学生体验提醒: {len(student_experience_lines)}",
        f"- 已忽略历史错位缺课记录: {len(ignored)}",
        f"- 秋季同师同时间记录: {len(overlap_lines)}",
        f"- 春秋线上合班显示记录: {len(online_merge_lines)}",
        f"- 教室容量提醒: {len(capacity_warnings)}",
        f"- 老师同日跨校区提醒: {len(teacher_campus_warnings)}",
        f"- 待核对提示: {len(warnings)}",
        "",
        "## 历史课节计数",
        "",
    ]
    for (stage, class_id), count in sorted(counts.items()):
        lines.append(f"- {stage} {class_id}: {count}")
    if missing_summer_suites:
        lines.extend(["", "## 暑假缺班级排课窗口", ""])
        lines.append("以下套班存在维护范围内的公共课班级，但 class_window_boundaries.csv 中没有可用暑假窗口，本次未预排暑假。")
        lines.extend(f"- {suite_code}" for suite_code in missing_summer_suites)
    if conflict_rebuilt_summer_suites:
        lines.extend(["", "## 暑假冲突自动重排套班", ""])
        lines.append("以下套班在旧预排结果中存在同一班级同一时间叠课，已按当前产品课时和固定课表重新生成暑假排课。")
        lines.extend(f"- {suite_code}" for suite_code in conflict_rebuilt_summer_suites)
    if distribution_rebuilt_summer_suites:
        lines.extend(["", "## 暑假均衡/班级窗口自动重排套班", ""])
        lines.append("以下套班在旧预排结果中存在老师连续集中、科目周分布偏斜，或与班级排课窗口表中的场地不一致，已优先使用均衡策略重新生成暑假排课。")
        lines.extend(f"- {suite_code}" for suite_code in distribution_rebuilt_summer_suites)
    if halfyear_strategy_lines:
        lines.extend(["", "## 半年营 2770-2779 预排", ""])
        lines.append("本批次在不改动已确认暑假/秋季课表和专业课固定课表的前提下逐套预排；均衡策略不可行时使用轮排兜底；严格周二到周六排不通时，使用产品规则中的周一例外。")
        lines.extend(f"- {item}" for item in halfyear_strategy_lines)
    if additional_public_strategy_lines:
        lines.extend(["", "## 其他公共课产品预排", ""])
        lines.append("本批次按产品顺序继续处理全年营、无忧暑、无忧秋、无忧春、冲刺营；先锁定既有公共课尝试，排不通时按“允许调整既有公共课”口径放宽，并在本节标记。")
        lines.extend(f"- {item}" for item in additional_public_strategy_lines)
    if conflict_lines:
        lines.extend(["", "## 班级/套班同时间冲突", ""])
        lines.append("以下冲突仍需人工核对；同一半天内前后两节不重叠的 2 小时课不会计入。")
        lines.extend(f"- {item}" for item in conflict_lines)
    if teacher_conflict_lines:
        lines.extend(["", "## 老师同时间硬冲突", ""])
        lines.append("以下为同一老师在同一时间段被安排到多个不同教室的硬冲突，必须重排。")
        lines.extend(f"- {item}" for item in teacher_conflict_lines)
    if global_repair_lines:
        lines.extend(["", "## 全局消冲突记录", ""])
        lines.append("本节记录全局后处理移动的公共课课节；历史课表、专业课固定课表和线上合班展示课表保持锁定。")
        lines.extend(f"- {item}" for item in global_repair_lines[:200])
    if summer_distribution_lines:
        lines.extend(["", "## 暑假均衡残留提示", ""])
        lines.append("以下不是硬冲突，但说明在当前班级排课窗口、老师/教室资源和固定课表下仍有局部集中，需要人工确认是否接受。")
        lines.extend(f"- {item}" for item in summer_distribution_lines)
    if long_camp_balance_lines:
        lines.extend(["", "## 长线营科目周均衡提醒", ""])
        lines.append("全年营/半年营公共课已按科目周均衡优先排课；以下是在当前固定课表、互斥、停课日期和周一例外口径下仍存在的局部波动。")
        lines.extend(f"- {item}" for item in long_camp_balance_lines)
    if long_camp_sequence_lines:
        lines.extend(["", "## 长线营科目/老师连续提醒", ""])
        lines.append("全年营/半年营英语、政治已优先按老师交替和同科目隔天排课，数学已优先避免连续超过 3 天；以下是在资源约束下仍需人工核对的残留项。")
        lines.extend(f"- {item}" for item in long_camp_sequence_lines)
    if student_experience_lines:
        lines.extend(["", "## 学生体验提醒", ""])
        lines.append("以下不是硬冲突，而是从学生视角看可能不舒服的周分布：过满、单科过集中或中间空周。")
        lines.extend(f"- {item}" for item in student_experience_lines)
    if warnings:
        lines.extend(["", "## 待核对提示", ""])
        lines.extend(f"- {warning}" for warning in warnings[:200])
    if capacity_warnings:
        lines.extend(["", "## 教室容量提醒", ""])
        lines.append("教室容量小于班级人数时不阻塞排课，仅提示人工核对。")
        lines.extend(f"- {warning}" for warning in capacity_warnings[:200])
    if teacher_campus_warnings:
        lines.extend(["", "## 老师同日跨校区提醒", ""])
        lines.append("排课已优先同校区、其次同区域；以下为在当前规则和资源下仍出现的跨区域/高风险往返，需要人工核对是否接受。")
        lines.extend(f"- {warning}" for warning in teacher_campus_warnings[:200])
    if ignored:
        lines.extend(["", "## 已忽略记录", ""])
        lines.extend(f"- {item}" for item in ignored[:200])
    if online_merge_lines:
        lines.extend(["", "## 春秋线上合班显示关系", ""])
        lines.extend(f"- {item}" for item in online_merge_lines[:300])
    if overlap_lines:
        lines.extend(["", "## 秋季同师同时间记录", ""])
        lines.append("以下为实际排课班级之间的同师同时间记录。")
        lines.extend(f"- {item}" for item in overlap_lines[:200])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def append_run_summary(path: Path, mode: str, perf: PerfLog, extra_lines: Sequence[str]) -> None:
    lines = ["", "## 运行摘要", "", f"- 运行模式: {mode}", *extra_lines, "", "## 性能日志", "", *perf.markdown_lines()]
    with path.open("a", encoding="utf-8") as handle:
        handle.write("\n".join(lines) + "\n")


def split_arg_values(values: Sequence[str]) -> Set[str]:
    result: Set[str] = set()
    for value in values:
        for part in split_pipe_values(value, include_whitespace=True):
            cleaned = clean(part)
            if cleaned:
                result.add(cleaned)
    return result


def run_fast_callback_with_timeout(callback, seconds: int = FAST_ATTEMPT_TIMEOUT_SECONDS):
    if seconds == FAST_ATTEMPT_TIMEOUT_SECONDS:
        try:
            seconds = int(os.environ.get("FAST_ATTEMPT_TIMEOUT_SECONDS", str(seconds)) or seconds)
        except ValueError:
            seconds = FAST_ATTEMPT_TIMEOUT_SECONDS
    if seconds <= 0 or not hasattr(signal, "SIGALRM"):
        return callback()
    previous_handler = signal.getsignal(signal.SIGALRM)

    def timeout_handler(_signum, _frame) -> None:
        raise FastAttemptTimeout(f"快速尝试超过 {seconds} 秒，已停止本次局部策略")

    signal.signal(signal.SIGALRM, timeout_handler)
    signal.alarm(seconds)
    try:
        return callback()
    finally:
        signal.alarm(0)
        signal.signal(signal.SIGALRM, previous_handler)


def normalize_suite_code(value: str) -> str:
    cleaned = clean(value)
    match = re.search(r"(\d{4})", cleaned)
    return match.group(1) if match else cleaned


def truthy_text(value: object) -> bool:
    return clean(value) in {"是", "1", "true", "True", "yes", "Y", "y"}


def is_public_schedulable_meta(meta: Dict[str, str]) -> bool:
    return meta.get("subject_category") == "公共课" and not truthy_text(meta.get("is_schedule_locked"))


def suite_code_for_class(class_id: str, class_metadata: Dict[str, Dict[str, str]]) -> str:
    return clean(class_metadata.get(class_id, {}).get("suite_code")) or suite_code_from_class_id(class_id)


def load_output_rows(path: Path) -> List[dict]:
    return read_csv_rows(path)


def resolve_fast_scope(
    data_dir: Path,
    suite_values: Sequence[str],
    class_values: Sequence[str],
    sub_product_values: Sequence[str],
) -> Tuple[Dict[str, Dict[str, str]], Set[str], Set[str], Set[str], Set[str], Set[str]]:
    class_metadata = load_class_metadata(data_dir)
    requested_class_ids = split_arg_values(class_values)
    requested_suite_codes = {normalize_suite_code(value) for value in split_arg_values(suite_values)}
    requested_sub_products = split_arg_values(sub_product_values)

    affected_suites = set(requested_suite_codes)
    for class_id in requested_class_ids:
        suite_code = suite_code_for_class(class_id, class_metadata)
        if suite_code:
            affected_suites.add(suite_code)

    affected_class_ids: Set[str] = set()
    for class_id, meta in class_metadata.items():
        if not is_public_schedulable_meta(meta):
            continue
        suite_code = suite_code_for_class(class_id, class_metadata)
        if (
            class_id in requested_class_ids
            or (suite_code and suite_code in affected_suites)
            or (requested_sub_products and meta.get("sub_product") in requested_sub_products)
        ):
            affected_class_ids.add(class_id)
            if suite_code:
                affected_suites.add(suite_code)

    affected_sub_products = {
        class_metadata.get(class_id, {}).get("sub_product", "")
        for class_id in affected_class_ids
        if class_metadata.get(class_id, {}).get("sub_product")
    }
    summer_suites = {
        suite_code_for_class(class_id, class_metadata)
        for class_id in affected_class_ids
        if class_metadata.get(class_id, {}).get("sub_product") in SUMMER_FAST_PRODUCTS
        and suite_code_for_class(class_id, class_metadata) in MAINTENANCE_SUITES
    }
    halfyear_suites = {
        suite_code_for_class(class_id, class_metadata)
        for class_id in affected_class_ids
        if suite_code_for_class(class_id, class_metadata) in HALF_YEAR_BATCH_SUITES
    }
    additional_suites = {
        suite_code_for_class(class_id, class_metadata)
        for class_id in affected_class_ids
        if class_metadata.get(class_id, {}).get("sub_product") in ADDITIONAL_PUBLIC_PRODUCTS
    }
    return (
        class_metadata,
        affected_class_ids,
        affected_suites,
        affected_sub_products,
        summer_suites,
        halfyear_suites,
        additional_suites,
    )


def should_remove_fast_row(
    row: dict,
    class_metadata: Dict[str, Dict[str, str]],
    affected_class_ids: Set[str],
    summer_suites: Set[str],
    halfyear_suites: Set[str],
    additional_suites: Set[str],
    additional_phase: str = "",
) -> bool:
    class_id = clean(row.get("class_id"))
    if class_id not in affected_class_ids:
        return False
    meta = class_metadata.get(class_id, {})
    suite_code = suite_code_for_class(class_id, class_metadata)
    sub_product = meta.get("sub_product", "")
    date_text = clean(row.get("date"))
    if suite_code in summer_suites and sub_product in SUMMER_FAST_PRODUCTS:
        return row_matches_phase(row, SUMMER_PREPLAN_STAGES)
    if suite_code in halfyear_suites:
        return date_text >= "2026-07-01"
    if suite_code in additional_suites:
        if additional_phase in {"autumn", "fall", "秋季"} and sub_product in WYQC_PRODUCTS:
            return date_text >= WYQC_AUTUMN_START
        return date_text >= "2026-07-01"
    return False


def build_fast_summer_assignments(
    data_dir: Path,
    target_suite_codes: Set[str],
    protected_assignments: Sequence[scheduler.Assignment],
    protect_prior_target_suites: bool = True,
) -> Tuple[List[scheduler.Assignment], List[str]]:
    if not target_suite_codes:
        return [], []
    class_metadata = load_class_metadata(data_dir)
    suite_window_constraints = load_summer_suite_window_constraints(data_dir, class_metadata)
    priority_suite_code_list = split_pipe_values(
        os.environ.get("FAST_PRIORITY_SUITE_CODES")
        or os.environ.get("FAST_PROTECTED_SUITE_CODES", "")
    )
    priority_suite_codes = set(priority_suite_code_list)
    priority_order = {suite_code: index for index, suite_code in enumerate(priority_suite_code_list)}
    result: List[scheduler.Assignment] = []
    lines: List[str] = []
    remaining_suite_codes = set(target_suite_codes)
    paired_core_suites = {"2757", "2793"}
    if not protect_prior_target_suites and paired_core_suites.issubset(remaining_suite_codes):
        print("快速重排 暑假 2757+2793 联合调度", flush=True)
        class_ids = class_ids_for_suite_codes(data_dir, sorted(paired_core_suites), None)
        selected_constraints = [
            suite_window_constraints[suite_code]
            for suite_code in paired_core_suites
            if suite_code in suite_window_constraints
        ]
        if len(selected_constraints) != len(paired_core_suites):
            lines.append("暑假 2757+2793 联合核心调度未启用: 缺班级暑假排课窗口")
        else:
            scoped_protected_assignments = relevant_locked_assignments_for_replan(
                data_dir,
                class_ids,
                protected_assignments,
                min(constraint.earliest_date for constraint in selected_constraints),
                max(constraint.latest_date for constraint in selected_constraints),
            )
            try:
                scheduled = run_fast_callback_with_timeout(
                    lambda: rebuild_summer_suites_with_core_scheduler(
                        data_dir,
                        sorted(paired_core_suites),
                        class_metadata,
                        suite_window_constraints,
                        scoped_protected_assignments,
                    ),
                    25,
                )
                result.extend(scheduled)
                lines.append(f"暑假 2757+2793: 联合核心调度，{len(scheduled)} 节")
                remaining_suite_codes -= paired_core_suites
            except ValueError as exc:
                lines.append(f"暑假 2757+2793 联合核心调度未启用: {exc}")
    for suite_code in sorted(
        target_suite_codes,
        key=lambda item: (
            item not in priority_suite_codes,
            priority_order.get(item, 999),
            item not in SUMMER_CAMP_SUITES,
            item,
        ),
    ):
        if suite_code not in remaining_suite_codes:
            continue
        if suite_code not in suite_window_constraints:
            raise ValueError(f"{suite_code} 缺少班级暑假排课窗口，快速模式无法重排")
        print(f"快速重排 暑假 {suite_code}", flush=True)
        class_ids = class_ids_for_suite_codes(data_dir, [suite_code], None)
        constraint = suite_window_constraints[suite_code]
        planning_protected_assignments = (
            [*protected_assignments, *result]
            if protect_prior_target_suites
            else list(protected_assignments)
        )
        scoped_protected_assignments = relevant_locked_assignments_for_replan(
            data_dir,
            class_ids,
            planning_protected_assignments,
            constraint.earliest_date,
            constraint.latest_date,
        )
        try:
            scheduled, used_label, _errors = run_summer_rebuild_attempts(
                (
                    (
                        "快速轮排",
                        lambda suite=suite_code: rebuild_summer_suite_round_robin(
                            data_dir,
                            suite,
                            class_metadata,
                            suite_window_constraints,
                            scoped_protected_assignments,
                            0,
                        ),
                        FAST_ATTEMPT_TIMEOUT_SECONDS,
                    ),
                    (
                        "均衡",
                        lambda suite=suite_code: rebuild_summer_suite_balanced(
                            data_dir,
                            suite,
                            class_metadata,
                            suite_window_constraints,
                            scoped_protected_assignments,
                        ),
                        FAST_ATTEMPT_TIMEOUT_SECONDS,
                    ),
                    (
                        "基础调度器",
                        lambda suite=suite_code: rebuild_summer_suite_with_core_scheduler(
                            data_dir,
                            suite,
                            class_metadata,
                            suite_window_constraints,
                            scoped_protected_assignments,
                        ),
                        FAST_ATTEMPT_TIMEOUT_SECONDS,
                    ),
                )
            )
        except ValueError as exc:
            raise ValueError(f"{suite_code} 暑假快速重排失败: {exc}") from exc
        result.extend(scheduled)
        lines.append(f"暑假 {suite_code}: {used_label}，{len(scheduled)} 节")
    return scheduler.sorted_assignments(result), lines


def write_fast_report(
    path: Path,
    *,
    target_suites: Set[str],
    target_sub_products: Set[str],
    affected_class_ids: Set[str],
    reused_count: int,
    removed_count: int,
    new_count: int,
    final_assignments: Sequence[scheduler.Assignment],
    class_conflict_groups: Dict[str, Set[str]],
    strategy_lines: Sequence[str],
    warning_lines: Sequence[str],
    perf: PerfLog,
    class_metadata: Dict[str, Dict[str, str]],
    blackout_dates: Set[str],
    affected_start_date: Optional[str] = None,
) -> None:
    conflict_lines = schedule_conflict_lines(final_assignments, class_conflict_groups)
    student_experience_lines = student_experience_warning_lines(final_assignments, class_metadata, blackout_dates)
    def conflict_line_in_scope(line: str) -> bool:
        if not affected_start_date:
            return True
        match = re.search(r"\d{4}-\d{2}-\d{2}", line)
        return bool(match and match.group(0) >= affected_start_date)

    affected_conflict_lines = [
        line for line in conflict_lines
        if any(class_id and class_id in line for class_id in affected_class_ids)
        and conflict_line_in_scope(line)
    ]
    existing_conflict_lines = [
        line for line in conflict_lines
        if line not in set(affected_conflict_lines)
    ]
    needs_full = bool(affected_conflict_lines)
    lines = [
        "# 课表维护快速更新报告",
        "",
        "## 运行摘要",
        "",
        "- 运行模式: fast",
        f"- 重排套班: {', '.join(sorted(target_suites)) if target_suites else '未指定'}",
        f"- 涉及子产品: {', '.join(sorted(target_sub_products)) if target_sub_products else '未识别'}",
        f"- 涉及班级: {len(affected_class_ids)}",
        f"- 复用课节数: {reused_count}",
        f"- 删除旧课节数: {removed_count}",
        f"- 新生成课节数: {new_count}",
        f"- 是否需要全量重算: {'是' if needs_full else '否'}",
        f"- 学生体验提醒: {len(student_experience_lines)}",
        "",
        "## 重排明细",
        "",
        *(f"- {line}" for line in strategy_lines),
    ]
    if warning_lines:
        lines.extend(["", "## 快速模式提示", ""])
        lines.extend(f"- {line}" for line in warning_lines[:200])
    if student_experience_lines:
        lines.extend(["", "## 学生体验提醒", ""])
        lines.append("以下不是硬冲突，而是从学生视角看可能不舒服的周分布：过满、单科过集中或中间空周。")
        lines.extend(f"- {line}" for line in student_experience_lines[:200])
    if affected_conflict_lines:
        lines.extend(["", "## 受影响范围冲突提示", ""])
        lines.append("快速模式已写出结果，但受影响班级仍存在以下冲突；最终发布前请运行全量重算确认。")
        lines.extend(f"- {line}" for line in affected_conflict_lines[:200])
    if existing_conflict_lines:
        lines.extend(["", "## 既有冲突提示", ""])
        lines.append("以下冲突来自未重排的上一版课表，本次快速更新未扩大范围处理。")
        lines.extend(f"- {line}" for line in existing_conflict_lines[:200])
    lines.extend(["", "## 性能日志", "", *perf.markdown_lines()])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_fast_failure_report(
    path: Path,
    *,
    suite_values: Sequence[str],
    class_values: Sequence[str],
    sub_product_values: Sequence[str],
    error: Exception,
    perf: PerfLog,
) -> None:
    lines = [
        "# 课表维护快速更新报告",
        "",
        "## 运行摘要",
        "",
        "- 运行模式: fast",
        f"- 重排套班参数: {', '.join(split_arg_values(suite_values)) or '未指定'}",
        f"- 重排班级参数: {', '.join(split_arg_values(class_values)) or '未指定'}",
        f"- 子产品参数: {', '.join(split_arg_values(sub_product_values)) or '未指定'}",
        "- 是否需要全量重算: 是",
        "",
        "## 失败原因",
        "",
        str(error),
        "",
        "## 处理建议",
        "",
        "- 这是快速局部重排的保护性退出，说明只移动指定范围时排不通，或需要移动上游已排班级。",
        "- 可先扩大套班/子产品范围再试一次；最终发布前请使用“全量重算”。",
        "",
        "## 性能日志",
        "",
        *perf.markdown_lines(),
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_full(data_dir: Path, perf: PerfLog) -> None:
    print("读取历史课表...", flush=True)
    raw_history_rows, warnings, ignored = normalize_history_rows(HISTORY_PATH, data_dir)
    history_rows, spring_merge_lines = expand_online_merge_rows(raw_history_rows, data_dir)
    history_assignments = assignments_from_rows(history_rows, "HISTORY")
    perf.mark("读取历史课表")
    print("读取专业课固定课表...", flush=True)
    locked_assignments, locked_rows = load_locked_schedule_assignments(data_dir)
    perf.mark("读取专业课固定课表")
    print("生成寒暑营/无忧寒/暑假营暑假课表...", flush=True)
    (
        summer_assignments,
        missing_summer_suites,
        conflict_rebuilt_summer_suites,
        distribution_rebuilt_summer_suites,
    ) = build_summer_assignments(data_dir, locked_assignments)
    perf.mark("生成暑假批次")
    print("生成春秋线上合班课表...", flush=True)
    autumn_actual_assignments, autumn_room_names = build_autumn_assignments(
        data_dir,
        [*locked_assignments, *summer_assignments],
    )
    autumn_assignments, autumn_merge_lines = expand_online_merge_assignments(autumn_actual_assignments, data_dir)
    perf.mark("生成春秋线上合班")
    print("生成半年营 2770-2779 课表...", flush=True)
    halfyear_assignments, halfyear_strategy_lines = build_halfyear_batch_assignments(
        data_dir,
        [*locked_assignments, *summer_assignments, *autumn_actual_assignments],
        history_rows,
    )
    perf.mark("生成半年营")
    print("生成其他公共课产品课表...", flush=True)
    covered_public_class_ids = {
        assignment.task.class_id
        for assignment in [*summer_assignments, *autumn_actual_assignments, *halfyear_assignments]
    }
    additional_public_assignments, additional_public_strategy_lines = build_additional_public_assignments(
        data_dir,
        covered_public_class_ids,
        [
            *locked_assignments,
            *summer_assignments,
            *autumn_actual_assignments,
            *halfyear_assignments,
        ],
        history_rows,
    )
    perf.mark("生成其他公共课产品")
    _conflict_groups, class_conflict_groups = load_conflict_group_lookup(data_dir)
    class_metadata = load_class_metadata(data_dir)
    blackout_dates = load_active_blackout_dates(data_dir)
    adjusted_public_assignments, makeup_lines = adjust_wuyou_makeup_days_in_public_pool(
        data_dir,
        [*halfyear_assignments, *additional_public_assignments],
        [*history_assignments, *locked_assignments, *summer_assignments, *autumn_assignments],
        class_metadata,
        class_conflict_groups,
    )
    halfyear_class_ids = halfyear_batch_class_ids(data_dir)
    halfyear_assignments = [
        assignment for assignment in adjusted_public_assignments if assignment.task.class_id in halfyear_class_ids
    ]
    additional_public_assignments = [
        assignment for assignment in adjusted_public_assignments if assignment.task.class_id not in halfyear_class_ids
    ]
    repaired_public_assignments, global_repair_lines = repair_global_teacher_time_conflicts(
        data_dir,
        [*halfyear_assignments, *additional_public_assignments],
        [*history_assignments, *locked_assignments, *autumn_assignments, *summer_assignments],
        class_conflict_groups,
    )
    repaired_by_task_id = {
        assignment.task.task_id: assignment
        for assignment in repaired_public_assignments
    }
    halfyear_assignments = [repaired_by_task_id.get(assignment.task.task_id, assignment) for assignment in halfyear_assignments]
    additional_public_assignments = [
        repaired_by_task_id.get(assignment.task.task_id, assignment)
        for assignment in additional_public_assignments
    ]
    summer_assignments = deduplicate_assignments(summer_assignments)
    halfyear_assignments = deduplicate_assignments(halfyear_assignments)
    additional_public_assignments = deduplicate_assignments(additional_public_assignments)
    autumn_assignments = deduplicate_assignments(autumn_assignments)
    post_global_summer_rebuild_suites = summer_distribution_rebuild_suite_codes(summer_assignments)
    if post_global_summer_rebuild_suites:
        suite_window_constraints = load_summer_suite_window_constraints(data_dir, class_metadata)
        summer_rebuild_lines = [
            "全局修复后暑假均衡回补: " + "、".join(post_global_summer_rebuild_suites)
        ]
        successful_post_global_summer_rebuild_suites: List[str] = []
        for suite_code in post_global_summer_rebuild_suites:
            if suite_code not in suite_window_constraints:
                summer_rebuild_lines.append(f"{suite_code}: 缺班级暑假排课窗口，未回补")
                continue
            suite_class_ids = class_ids_for_suite_codes(data_dir, [suite_code], None)
            constraint = suite_window_constraints[suite_code]
            non_target_summer = [
                assignment
                for assignment in summer_assignments
                if suite_code_from_class_id(assignment.task.class_id) != suite_code
            ]
            protected_for_suite = relevant_locked_assignments_for_replan(
                data_dir,
                suite_class_ids,
                [
                    *history_assignments,
                    *locked_assignments,
                    *autumn_assignments,
                    *halfyear_assignments,
                    *additional_public_assignments,
                    *non_target_summer,
                ],
                constraint.earliest_date,
                constraint.latest_date,
            )
            try:
                rebuilt = rebuild_summer_suite_round_robin(
                    data_dir,
                    suite_code,
                    class_metadata,
                    suite_window_constraints,
                    protected_for_suite,
                    0,
                )
            except ValueError as exc:
                summer_rebuild_lines.append(f"{suite_code}: 回补失败，保留全局修复结果；{exc}")
                continue
            summer_assignments = scheduler.sorted_assignments([*non_target_summer, *rebuilt])
            successful_post_global_summer_rebuild_suites.append(suite_code)
            summer_rebuild_lines.append(f"{suite_code}: 已回补 {len(rebuilt)} 节")
        distribution_rebuilt_summer_suites = sorted(
            set(distribution_rebuilt_summer_suites) | set(successful_post_global_summer_rebuild_suites)
        )
        repaired_public_assignments, post_summer_repair_lines = repair_global_teacher_time_conflicts(
            data_dir,
            [*summer_assignments, *halfyear_assignments, *additional_public_assignments],
            [*history_assignments, *locked_assignments, *autumn_assignments],
            class_conflict_groups,
        )
        repaired_by_task_id = {
            assignment.task.task_id: assignment
            for assignment in repaired_public_assignments
        }
        summer_assignments = [repaired_by_task_id.get(assignment.task.task_id, assignment) for assignment in summer_assignments]
        halfyear_assignments = [repaired_by_task_id.get(assignment.task.task_id, assignment) for assignment in halfyear_assignments]
        additional_public_assignments = [
            repaired_by_task_id.get(assignment.task.task_id, assignment)
            for assignment in additional_public_assignments
        ]
        global_repair_lines = [
            *global_repair_lines,
            *summer_rebuild_lines,
            "暑假均衡回补后全局老师硬冲突复检:",
            *post_summer_repair_lines,
        ]
        summer_assignments = deduplicate_assignments(summer_assignments)
        halfyear_assignments = deduplicate_assignments(halfyear_assignments)
        additional_public_assignments = deduplicate_assignments(additional_public_assignments)
    perf.mark("全局老师硬冲突修复")

    assignments = deduplicate_assignments([
        *history_assignments,
        *locked_assignments,
        *summer_assignments,
        *autumn_assignments,
        *halfyear_assignments,
        *additional_public_assignments,
    ])
    perf.mark("合并与调班处理")
    assignments, wys_stageless_lines = rebuild_2731_stage_priority_schedule(
        data_dir,
        assignments,
        class_metadata,
        class_conflict_groups,
        blackout_dates,
    )
    if wys_stageless_lines:
        global_repair_lines = [*global_repair_lines, *wys_stageless_lines]
    perf.mark("2731阶段优先重建")
    try:
        assignments, wys_no_math_lines = rebuild_2727_no_math_sequence_schedule(
            data_dir,
            assignments,
            class_metadata,
            class_conflict_groups,
            blackout_dates,
        )
    except ValueError as exc:
        wys_no_math_lines = [f"2727 无忧暑英政顺序重排暂未执行: {exc}"]
    if wys_no_math_lines:
        global_repair_lines = [*global_repair_lines, *wys_no_math_lines]
    perf.mark("2727无忧暑英政顺序重排")
    assignments, public_experience_repair_lines = repair_public_same_subject_day_overloads(
        data_dir,
        assignments,
        class_metadata,
        class_conflict_groups,
        blackout_dates,
    )
    if public_experience_repair_lines:
        global_repair_lines = [*global_repair_lines, *public_experience_repair_lines]
    perf.mark("公共课同班同科同日规避")
    assignments, overlap_2754_lines = apply_2754_politics_overlap_balance(
        data_dir,
        assignments,
        class_metadata,
        class_conflict_groups,
        blackout_dates,
    )
    if overlap_2754_lines:
        global_repair_lines = [*global_repair_lines, *overlap_2754_lines]
    perf.mark("2754政治阶段交错均衡")
    post_special_movable_assignments = [
        assignment
        for assignment in assignments
        if movable_public_experience_assignment(assignment, class_metadata)
        and not (
            assignment_matches_phase(assignment, SUMMER_PREPLAN_STAGES)
            and suite_code_from_class_id(assignment.task.class_id) in MAINTENANCE_SUITES
        )
    ]
    post_special_movable_task_ids = {
        assignment.task.task_id for assignment in post_special_movable_assignments
    }
    post_special_protected_assignments = [
        assignment
        for assignment in assignments
        if assignment.task.task_id not in post_special_movable_task_ids
    ]
    post_special_repaired_assignments, post_special_repair_lines = repair_global_teacher_time_conflicts(
        data_dir,
        post_special_movable_assignments,
        post_special_protected_assignments,
        class_conflict_groups,
    )
    assignments = deduplicate_assignments([
        *post_special_protected_assignments,
        *post_special_repaired_assignments,
    ])
    if post_special_repair_lines:
        global_repair_lines = [*global_repair_lines, "专项重排后全局老师硬冲突复检:", *post_special_repair_lines]
    perf.mark("专项后全局老师硬冲突复修")
    assignments, renjie_lines = repair_renjie_offline_availability(
        data_dir,
        assignments,
        class_metadata,
        class_conflict_groups,
        blackout_dates,
    )
    if renjie_lines:
        global_repair_lines = [*global_repair_lines, *renjie_lines]
    perf.mark("任洁线下可排窗口修复")
    assignments, long_camp_week_balance_lines = repair_long_camp_subject_week_balance(
        data_dir,
        assignments,
        class_metadata,
        class_conflict_groups,
        blackout_dates,
    )
    if long_camp_week_balance_lines:
        global_repair_lines = [*global_repair_lines, *long_camp_week_balance_lines]
    perf.mark("长线营科目周均衡修复")
    assignments, tail_chain_lines = repair_long_camp_tail_week_chain_moves(
        data_dir,
        assignments,
        class_metadata,
        class_conflict_groups,
        blackout_dates,
    )
    if tail_chain_lines:
        global_repair_lines = [*global_repair_lines, *tail_chain_lines]
    perf.mark("长线营尾周连锁挪课")
    assignments, student_week_balance_lines = repair_student_experience_week_balance(
        data_dir,
        assignments,
        class_metadata,
        class_conflict_groups,
        blackout_dates,
        max_moves=180,
        max_seconds=150,
    )
    if student_week_balance_lines:
        global_repair_lines = [*global_repair_lines, *student_week_balance_lines]
    if long_camp_week_balance_lines or tail_chain_lines or student_week_balance_lines:
        global_repair_lines = refresh_long_camp_minimum_gap_report_lines(
            global_repair_lines,
            assignments,
            class_metadata,
            blackout_dates,
        )
    perf.mark("学生体验周均衡修复")
    final_movable_assignments = [
        assignment
        for assignment in assignments
        if movable_public_experience_assignment(assignment, class_metadata)
    ]
    final_movable_task_ids = {assignment.task.task_id for assignment in final_movable_assignments}
    final_protected_assignments = [
        assignment
        for assignment in assignments
        if assignment.task.task_id not in final_movable_task_ids
    ]
    final_repaired_assignments, final_repair_lines = repair_global_teacher_time_conflicts(
        data_dir,
        final_movable_assignments,
        final_protected_assignments,
        class_conflict_groups,
        max_seconds=max(GLOBAL_REPAIR_MAX_SECONDS, 90),
    )
    assignments = deduplicate_assignments([
        *final_protected_assignments,
        *final_repaired_assignments,
    ])
    if final_repair_lines:
        global_repair_lines = [*global_repair_lines, "发布前全局老师硬冲突兜底复修:", *final_repair_lines]
    perf.mark("发布前全局老师硬冲突兜底复修")
    assignments, politics_segment_lines = repair_politics_segments_for_teacher_conflicts(
        data_dir,
        assignments,
        class_metadata,
        class_conflict_groups,
        blackout_dates,
    )
    if politics_segment_lines:
        global_repair_lines = [*global_repair_lines, "政治老师冲突整段重排:", *politics_segment_lines]
    perf.mark("政治整段重排消冲突")
    if politics_segment_lines and teacher_time_conflict_groups(assignments):
        final_movable_assignments = [
            assignment
            for assignment in assignments
            if movable_public_experience_assignment(assignment, class_metadata)
        ]
        final_movable_task_ids = {assignment.task.task_id for assignment in final_movable_assignments}
        final_protected_assignments = [
            assignment
            for assignment in assignments
            if assignment.task.task_id not in final_movable_task_ids
        ]
        final_repaired_assignments, final_repair_lines = repair_global_teacher_time_conflicts(
            data_dir,
            final_movable_assignments,
            final_protected_assignments,
            class_conflict_groups,
            max_seconds=max(GLOBAL_REPAIR_MAX_SECONDS, 90),
        )
        assignments = deduplicate_assignments([
            *final_protected_assignments,
            *final_repaired_assignments,
        ])
        if final_repair_lines:
            global_repair_lines = [*global_repair_lines, "政治整段后硬冲突复修:", *final_repair_lines]
    perf.mark("政治整段后硬冲突复修")
    room_names = load_room_names(data_dir)
    room_names.update({row["room_id"]: row["room_name"] for row in history_rows if row.get("room_id")})
    room_names.update({row["room_id"]: row["room_name"] for row in locked_rows if row.get("room_id")})
    room_names.update(autumn_room_names)
    rooms = load_rooms_for_capacity(data_dir / "scheduler_input_draft.json")
    capacity_warnings = assignment_capacity_warning_lines(assignments, rooms, room_names)
    teacher_campus_warnings = teacher_same_day_campus_warning_lines(
        assignments,
        rooms,
        load_area_travel_minutes(data_dir),
    )
    perf.mark("生成容量与冲突提示")
    publish_blocking_teacher_conflicts = teacher_time_conflict_lines(assignments)
    if publish_blocking_teacher_conflicts and not ALLOW_PUBLISH_WITH_TEACHER_CONFLICTS:
        rejected_dir = Path("outputs") / f"rejected_full_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        rejected_dir.mkdir(parents=True, exist_ok=True)
        rejected_csv = rejected_dir / OUTPUT_CSV.name
        rejected_html = rejected_dir / OUTPUT_HTML.name
        rejected_report = rejected_dir / OUTPUT_REPORT.name
        rejected_teacher_conflicts = rejected_dir / TEACHER_CONFLICT_CSV.name
        write_batch_csv(assignments, rejected_csv, room_names)
        write_day_table_html(
            assignments,
            rejected_html,
            "课表维护总表（未发布候选）",
            ["AM", "PM", "EVENING"],
            room_names,
            assignments[0].candidate.slots[0].date if assignments else None,
            assignments[-1].candidate.slots[0].date if assignments else None,
            class_metadata,
            load_all_class_window_constraint_items(data_dir),
        )
        write_teacher_time_conflicts_csv(assignments, rejected_teacher_conflicts, room_names)
        write_report(
            rejected_report,
            history_rows,
            locked_assignments,
            summer_assignments,
            halfyear_assignments,
            halfyear_strategy_lines,
            additional_public_assignments,
            additional_public_strategy_lines,
            conflict_rebuilt_summer_suites,
            distribution_rebuilt_summer_suites,
            missing_summer_suites,
            autumn_actual_assignments,
            autumn_assignments,
            assignments,
            class_conflict_groups,
            [*makeup_lines, *warnings],
            ignored,
            [*spring_merge_lines, *autumn_merge_lines],
            capacity_warnings,
            teacher_campus_warnings,
            global_repair_lines,
            class_metadata,
            blackout_dates,
        )
        append_run_summary(
            rejected_report,
            "full-rejected",
            perf,
            [
                f"- 未发布原因: 老师同时间硬冲突 {len(publish_blocking_teacher_conflicts)} 条",
                f"- 候选输出 CSV: {rejected_csv}",
                f"- 候选输出 HTML: {rejected_html}",
                f"- 继续保留正式课表维护总表: {OUTPUT_HTML}",
            ],
        )
        print(
            f"全量候选未发布: 仍有 {len(publish_blocking_teacher_conflicts)} 条老师同时间硬冲突，"
            f"候选已保存到 {rejected_dir}"
        )
        return
    publish_blocking_coverage_gaps = public_coverage_gap_rows_for_assignments(
        data_dir,
        assignments,
        class_metadata,
    )
    if publish_blocking_coverage_gaps and not ALLOW_PUBLISH_WITH_COVERAGE_GAPS:
        coverage_lines = coverage_gap_blocking_lines(publish_blocking_coverage_gaps)
        rejected_dir = Path("outputs") / f"rejected_full_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        rejected_dir.mkdir(parents=True, exist_ok=True)
        rejected_csv = rejected_dir / OUTPUT_CSV.name
        rejected_html = rejected_dir / OUTPUT_HTML.name
        rejected_report = rejected_dir / OUTPUT_REPORT.name
        rejected_teacher_conflicts = rejected_dir / TEACHER_CONFLICT_CSV.name
        write_batch_csv(assignments, rejected_csv, room_names)
        write_day_table_html(
            assignments,
            rejected_html,
            "课表维护总表（未发布候选）",
            ["AM", "PM", "EVENING"],
            room_names,
            assignments[0].candidate.slots[0].date if assignments else None,
            assignments[-1].candidate.slots[0].date if assignments else None,
            class_metadata,
            load_all_class_window_constraint_items(data_dir),
        )
        write_teacher_time_conflicts_csv(assignments, rejected_teacher_conflicts, room_names)
        write_report(
            rejected_report,
            history_rows,
            locked_assignments,
            summer_assignments,
            halfyear_assignments,
            halfyear_strategy_lines,
            additional_public_assignments,
            additional_public_strategy_lines,
            conflict_rebuilt_summer_suites,
            distribution_rebuilt_summer_suites,
            missing_summer_suites,
            autumn_actual_assignments,
            autumn_assignments,
            assignments,
            class_conflict_groups,
            [*makeup_lines, *warnings, "课时覆盖缺口发布阻断:", *coverage_lines],
            ignored,
            [*spring_merge_lines, *autumn_merge_lines],
            capacity_warnings,
            teacher_campus_warnings,
            global_repair_lines,
            class_metadata,
            blackout_dates,
        )
        append_run_summary(
            rejected_report,
            "full-rejected",
            perf,
            [
                f"- 未发布原因: {coverage_lines[0]}",
                f"- 候选输出 CSV: {rejected_csv}",
                f"- 候选输出 HTML: {rejected_html}",
                f"- 继续保留正式课表维护总表: {OUTPUT_HTML}",
            ],
        )
        print(
            f"全量候选未发布: {coverage_lines[0]}，"
            f"候选已保存到 {rejected_dir}"
        )
        return
    write_batch_csv(assignments, OUTPUT_CSV, room_names)
    write_day_table_html(
        assignments,
        OUTPUT_HTML,
        "课表维护总表",
        ["AM", "PM", "EVENING"],
        room_names,
        assignments[0].candidate.slots[0].date if assignments else None,
        assignments[-1].candidate.slots[0].date if assignments else None,
        class_metadata,
        load_all_class_window_constraint_items(data_dir),
    )
    write_teacher_time_conflicts_csv(assignments, TEACHER_CONFLICT_CSV, room_names)
    perf.mark("写出 CSV/HTML")
    write_report(
        OUTPUT_REPORT,
        history_rows,
        locked_assignments,
        summer_assignments,
        halfyear_assignments,
        halfyear_strategy_lines,
        additional_public_assignments,
        additional_public_strategy_lines,
        conflict_rebuilt_summer_suites,
        distribution_rebuilt_summer_suites,
        missing_summer_suites,
        autumn_actual_assignments,
        autumn_assignments,
        assignments,
        class_conflict_groups,
        [*makeup_lines, *warnings],
        ignored,
        [*spring_merge_lines, *autumn_merge_lines],
        capacity_warnings,
        teacher_campus_warnings,
        global_repair_lines,
        class_metadata,
        blackout_dates,
    )
    append_run_summary(
        OUTPUT_REPORT,
        "full",
        perf,
        [
            f"- 输出课节数: {assignment_standard_lesson_count(assignments)}",
            f"- 间隔/交替微调轮数: {FULL_SPACING_IMPROVEMENT_PASSES}",
            f"- 输出 CSV: {OUTPUT_CSV}",
            f"- 输出 HTML: {OUTPUT_HTML}",
        ],
    )
    shutil.copyfile(OUTPUT_CSV, LEGACY_OUTPUT_CSV)
    shutil.copyfile(OUTPUT_HTML, LEGACY_OUTPUT_HTML)
    shutil.copyfile(OUTPUT_REPORT, LEGACY_OUTPUT_REPORT)
    print(f"已更新课表维护总表: {assignment_standard_lesson_count(assignments)} 条课节")
    print(OUTPUT_CSV)
    print(OUTPUT_HTML)


def run_fast(
    data_dir: Path,
    suite_values: Sequence[str],
    class_values: Sequence[str],
    sub_product_values: Sequence[str],
    perf: PerfLog,
) -> None:
    if not OUTPUT_CSV.exists():
        print("快速模式找不到上一版稳定总表，自动回退全量重算。", flush=True)
        run_full(data_dir, perf)
        return
    baseline_rows = load_output_rows(OUTPUT_CSV)
    if not baseline_rows:
        print("快速模式发现上一版总表为空，自动回退全量重算。", flush=True)
        run_full(data_dir, perf)
        return
    perf.mark("读取上一版总表")

    (
        class_metadata,
        affected_class_ids,
        affected_suites,
        affected_sub_products,
        summer_suites,
        halfyear_suites,
        additional_suites,
    ) = resolve_fast_scope(data_dir, suite_values, class_values, sub_product_values)
    supported_suites = set(summer_suites) | set(halfyear_suites) | set(additional_suites)
    if not affected_class_ids or not supported_suites:
        raise ValueError("快速模式没有匹配到可局部重排的公共课班级，请指定 suite_code、class_id 或 sub_product，或改用 --mode full")
    unsupported_suites = affected_suites - supported_suites
    warning_lines = [
        f"以下套班不在当前快速重排批次内，已保留上一版课表；如需调整请全量重算: {', '.join(sorted(unsupported_suites))}"
    ] if unsupported_suites else []
    allow_previous_public_adjustment = parse_bool(
        os.environ.get("FAST_ALLOW_PREVIOUS_PUBLIC_ADJUSTMENT", "")
    )
    protect_fast_target_classes = parse_enabled(
        os.environ.get("FAST_PROTECT_TARGET_CLASSES", "1"),
        default=True,
    )
    protected_suite_codes_override = set(
        split_pipe_values(os.environ.get("FAST_PROTECTED_SUITE_CODES", ""))
    )
    if allow_previous_public_adjustment:
        warning_lines.append("本次快速重排允许移动非目标公共课，目标套班排好后作为保护课表。")
    if allow_previous_public_adjustment and not protect_fast_target_classes:
        warning_lines.append("本次快速重排允许目标套班参与老师硬冲突修复。")
    if protected_suite_codes_override:
        warning_lines.append(
            "本次快速重排仅保护指定优先套班: "
            + ", ".join(sorted(protected_suite_codes_override))
        )
    protect_prior_target_suites_value = os.environ.get("FAST_PROTECT_PRIOR_TARGET_SUITES")
    if protect_prior_target_suites_value is None:
        protect_prior_target_suites_value = "1" if protect_fast_target_classes else "0"
    protect_prior_target_suites_for_generation = parse_enabled(
        protect_prior_target_suites_value,
        default=True,
    )
    if protect_prior_target_suites_for_generation and not protect_fast_target_classes:
        warning_lines.append("本次快速重排生成阶段保护前序目标套班，冲突修复阶段仍允许目标套班微调。")
    additional_phase = fast_additional_phase()
    autumn_only_additional = False
    if additional_phase in {"autumn", "fall", "秋季"}:
        warning_lines.append("本次快速重排仅处理无忧秋/无忧春秋季段，保留 7-8 月既有课表。")
    priority_context_suite_codes = set(
        split_pipe_values(os.environ.get("FAST_PRIORITY_CONTEXT_SUITE_CODES", ""))
    )
    locked_reused_suite_codes = set(split_pipe_values(os.environ.get("FAST_LOCKED_SUITE_CODES", "")))
    repair_locked_suite_codes = set(split_pipe_values(os.environ.get("FAST_REPAIR_LOCKED_SUITE_CODES", "")))
    last_resort_repair_suite_codes = set(split_pipe_values(os.environ.get("FAST_LAST_RESORT_REPAIR_SUITE_CODES", "")))
    if priority_context_suite_codes:
        warning_lines.append(
            "本次快速重排在生成优先套班时参考上下文套班: "
            + ", ".join(sorted(priority_context_suite_codes))
        )
    if locked_reused_suite_codes:
        warning_lines.append(
            "本次快速重排锁定已确认套班: "
            + ", ".join(sorted(locked_reused_suite_codes))
        )
    if repair_locked_suite_codes:
        warning_lines.append(
            "本次快速重排仅在冲突修复阶段锁定套班: "
            + ", ".join(sorted(repair_locked_suite_codes))
        )
    if last_resort_repair_suite_codes:
        warning_lines.append(
            "若最终仍有硬冲突，将仅对这些套班做最小兜底让位: "
            + ", ".join(sorted(last_resort_repair_suite_codes))
        )
    perf.mark("识别快速重排范围")

    reused_rows: List[dict] = []
    removed_rows: List[dict] = []
    for row in baseline_rows:
        if should_remove_fast_row(
            row,
            class_metadata,
            affected_class_ids,
            summer_suites,
            halfyear_suites,
            additional_suites,
            additional_phase,
        ):
            removed_rows.append(row)
        else:
            reused_rows.append(row)
    reused_assignments = assignments_from_rows(reused_rows, "FAST_BASE")
    perf.mark("转换复用课节")
    protected_for_target_replan = (
        [
            assignment
            for assignment in reused_assignments
            if (
                not movable_public_experience_assignment(assignment, class_metadata)
                or suite_code_for_class(assignment.task.class_id, class_metadata) in priority_context_suite_codes
                or suite_code_for_class(assignment.task.class_id, class_metadata) in locked_reused_suite_codes
            )
        ]
        if allow_previous_public_adjustment
        else reused_assignments
    )

    history_rows: List[dict] = []
    if halfyear_suites or (affected_sub_products & HISTORY_DEDUCT_PRODUCTS):
        raw_history_rows, history_warnings, ignored = normalize_history_rows(HISTORY_PATH, data_dir)
        history_rows, spring_merge_lines = expand_online_merge_rows(raw_history_rows, data_dir)
        warning_lines.extend(history_warnings)
        warning_lines.extend(ignored[:20])
        warning_lines.extend(spring_merge_lines[:20])
        perf.mark("读取历史课表")

    strategy_lines: List[str] = []
    new_assignments: List[scheduler.Assignment] = []

    if summer_suites:
        summer_assignments, summer_lines = build_fast_summer_assignments(
            data_dir,
            summer_suites,
            protected_for_target_replan,
            protect_prior_target_suites=protect_prior_target_suites_for_generation,
        )
        new_assignments.extend(summer_assignments)
        strategy_lines.extend(summer_lines)
        perf.mark("快速重排暑假批次")

    public_replan_assignments: List[scheduler.Assignment] = []
    if halfyear_suites:
        halfyear_planning_protected = (
            [*protected_for_target_replan, *new_assignments]
            if protect_fast_target_classes
            else list(protected_for_target_replan)
        )
        halfyear_assignments, halfyear_lines = build_halfyear_batch_assignments(
            data_dir,
            halfyear_planning_protected,
            history_rows,
            target_suite_codes=halfyear_suites,
            fast_scope_locked_filter=True,
        )
        public_replan_assignments.extend(halfyear_assignments)
        strategy_lines.extend(f"半年营 {line}" for line in halfyear_lines)
        perf.mark("快速重排半年营")

    additional_sub_products = {
        class_metadata.get(class_id, {}).get("sub_product", "")
        for class_id in affected_class_ids
        if suite_code_for_class(class_id, class_metadata) in additional_suites
    } & set(ADDITIONAL_PUBLIC_PRODUCTS)
    autumn_only_additional = (
        additional_phase in {"autumn", "fall", "秋季"}
        and bool(additional_suites)
        and additional_sub_products <= WYQC_PRODUCTS
        and not summer_suites
        and not halfyear_suites
    )
    if additional_suites:
        additional_planning_protected = (
            [*protected_for_target_replan, *new_assignments, *public_replan_assignments]
            if protect_fast_target_classes
            else list(protected_for_target_replan)
        )
        all_public_class_ids = {
            class_id
            for class_id, meta in class_metadata.items()
            if is_public_schedulable_meta(meta)
        }
        target_additional_class_ids = {
            class_id
            for class_id in affected_class_ids
            if suite_code_for_class(class_id, class_metadata) in additional_suites
        }
        if additional_phase in {"autumn", "fall", "秋季"} and additional_sub_products <= WYQC_PRODUCTS:
            additional_assignments, additional_lines = build_wyqc_autumn_assignments(
                data_dir,
                additional_planning_protected,
                history_rows,
                target_suite_codes=additional_suites,
                target_sub_products=additional_sub_products,
                allow_existing_public_adjustment=allow_previous_public_adjustment,
                fast_scope_locked_filter=True,
            )
        else:
            additional_assignments, additional_lines = build_additional_public_assignments(
                data_dir,
                all_public_class_ids - target_additional_class_ids,
                additional_planning_protected,
                history_rows,
                target_suite_codes=additional_suites,
                target_sub_products=additional_sub_products,
                allow_existing_public_adjustment=allow_previous_public_adjustment,
                fast_scope_locked_filter=True,
                reuse_existing_on_block=not allow_previous_public_adjustment,
            )
        public_replan_assignments.extend(additional_assignments)
        strategy_lines.extend(additional_lines)
        perf.mark("快速重排其他公共课")

    _conflict_groups, class_conflict_groups = load_conflict_group_lookup(data_dir)
    if public_replan_assignments:
        adjusted_public_assignments, makeup_lines = adjust_wuyou_makeup_days_in_public_pool(
            data_dir,
            public_replan_assignments,
            [*reused_assignments, *new_assignments],
            class_metadata,
            class_conflict_groups,
        )
        public_replan_assignments = adjusted_public_assignments
        strategy_lines.extend(makeup_lines)
        perf.mark("快速调班与冲突保护")

    new_assignments.extend(public_replan_assignments)
    if not new_assignments:
        raise ValueError("快速模式没有生成新的课节，请检查指定范围是否属于已支持的公共课批次")

    assignments = deduplicate_assignments([*reused_assignments, *new_assignments])
    if not autumn_only_additional:
        assignments, wys_stageless_lines = rebuild_2731_stage_priority_schedule(
            data_dir,
            assignments,
            class_metadata,
            class_conflict_groups,
            load_active_blackout_dates(data_dir),
        )
        if wys_stageless_lines:
            strategy_lines.extend(wys_stageless_lines)
    else:
        strategy_lines.append("秋季-only 快速模式: 跳过 2731 无忧暑阶段优先重建。")
    perf.mark("快速2731阶段优先重建")
    if not autumn_only_additional:
        try:
            assignments, wys_no_math_lines = rebuild_2727_no_math_sequence_schedule(
                data_dir,
                assignments,
                class_metadata,
                class_conflict_groups,
                load_active_blackout_dates(data_dir),
            )
        except ValueError as exc:
            wys_no_math_lines = [f"2727 无忧暑英政顺序重排暂未执行: {exc}"]
        if wys_no_math_lines:
            strategy_lines.extend(wys_no_math_lines)
    else:
        strategy_lines.append("秋季-only 快速模式: 跳过 2727 无忧暑英政顺序重排。")
    perf.mark("快速2727无忧暑英政顺序重排")
    try:
        fast_same_day_repair_max_seconds = int(os.environ.get("FAST_SAME_DAY_REPAIR_MAX_SECONDS", "25") or "25")
    except ValueError:
        fast_same_day_repair_max_seconds = 25
    if not autumn_only_additional:
        assignments, public_experience_repair_lines = repair_public_same_subject_day_overloads(
            data_dir,
            assignments,
            class_metadata,
            class_conflict_groups,
            load_active_blackout_dates(data_dir),
            target_class_ids=affected_class_ids,
            max_seconds=fast_same_day_repair_max_seconds,
        )
        if public_experience_repair_lines:
            strategy_lines.extend(public_experience_repair_lines)
    else:
        strategy_lines.append("秋季-only 快速模式: 跳过同科同日体验修复，优先保护周分布。")
    perf.mark("快速公共课同班同科同日规避")
    try:
        fast_repair_max_seconds = int(os.environ.get("FAST_GLOBAL_REPAIR_MAX_SECONDS", "25") or "25")
    except ValueError:
        fast_repair_max_seconds = 25
    protected_class_ids = (
        {
            class_id
            for class_id in affected_class_ids
            if suite_code_for_class(class_id, class_metadata) in protected_suite_codes_override
        }
        if protected_suite_codes_override
        else affected_class_ids
    )
    target_protected_task_ids = {
        assignment.task.task_id
        for assignment in new_assignments
        if assignment.task.class_id in protected_class_ids
    } if protect_fast_target_classes else set()
    if allow_previous_public_adjustment:
        fast_movable_assignments = [
            assignment
            for assignment in assignments
            if assignment.task.task_id not in target_protected_task_ids
            and movable_public_experience_assignment(assignment, class_metadata)
            and suite_code_for_class(assignment.task.class_id, class_metadata) not in locked_reused_suite_codes
            and suite_code_for_class(assignment.task.class_id, class_metadata) not in repair_locked_suite_codes
        ]
    else:
        fast_movable_assignments = [
            assignment
            for assignment in assignments
            if assignment.task.class_id in affected_class_ids
            and movable_public_experience_assignment(assignment, class_metadata)
        ]
    fast_movable_task_ids = {assignment.task.task_id for assignment in fast_movable_assignments}
    fast_protected_assignments = [
        assignment
        for assignment in assignments
        if assignment.task.task_id not in fast_movable_task_ids
    ]
    fast_repaired_assignments, fast_repair_lines = repair_global_teacher_time_conflicts(
        data_dir,
        fast_movable_assignments,
        fast_protected_assignments,
        class_conflict_groups,
        max_seconds=fast_repair_max_seconds,
    )
    assignments = deduplicate_assignments([*fast_protected_assignments, *fast_repaired_assignments])
    if fast_repair_lines:
        strategy_lines.extend(["快速专项后老师硬冲突复检:", *fast_repair_lines])
    perf.mark("快速专项后老师硬冲突复修")
    if not autumn_only_additional:
        assignments, post_conflict_experience_lines = repair_public_same_subject_day_overloads(
            data_dir,
            assignments,
            class_metadata,
            class_conflict_groups,
            load_active_blackout_dates(data_dir),
            target_class_ids=affected_class_ids,
            max_seconds=fast_same_day_repair_max_seconds,
        )
        if post_conflict_experience_lines:
            strategy_lines.extend(["老师冲突修复后同科同日复检:", *post_conflict_experience_lines])
    else:
        strategy_lines.append("秋季-only 快速模式: 老师冲突后跳过同科同日复检，优先保护周分布。")
    perf.mark("快速老师修复后同科同日复检")
    if not autumn_only_additional:
        assignments, internal_swap_lines = repair_public_same_subject_day_overloads_with_internal_swap(
            data_dir,
            assignments,
            class_metadata,
            class_conflict_groups,
            load_active_blackout_dates(data_dir),
            target_class_ids=affected_class_ids,
            max_seconds=fast_same_day_repair_max_seconds,
        )
        if internal_swap_lines:
            strategy_lines.extend(["同班内部换位后同科同日复检:", *internal_swap_lines])
    else:
        strategy_lines.append("秋季-only 快速模式: 跳过同班内部换位复检，优先保护周分布。")
    perf.mark("快速同班内部换位复检")
    assignments, summer_week_balance_lines = repair_summer_subject_week_overloads(
        data_dir,
        assignments,
        class_metadata,
        class_conflict_groups,
        load_active_blackout_dates(data_dir),
        target_suite_codes=summer_suites,
        max_seconds=90,
    )
    if summer_week_balance_lines:
        strategy_lines.extend(summer_week_balance_lines)
    perf.mark("快速暑假科目周均衡修复")
    if "2754" in summer_suites:
        assignments, overlap_2754_lines = apply_2754_politics_overlap_balance(
            data_dir,
            assignments,
            class_metadata,
            class_conflict_groups,
            load_active_blackout_dates(data_dir),
        )
        if overlap_2754_lines:
            strategy_lines.extend(overlap_2754_lines)
    perf.mark("快速2754政治阶段交错均衡")
    if last_resort_repair_suite_codes:
        remaining_conflict_groups = teacher_time_conflict_groups(assignments)
        last_resort_task_ids = {
            assignment.task.task_id
            for group in remaining_conflict_groups
            for assignment in group
            if suite_code_for_class(assignment.task.class_id, class_metadata) in last_resort_repair_suite_codes
            and movable_public_experience_assignment(assignment, class_metadata)
        }
        if last_resort_task_ids:
            last_resort_movable = [
                assignment
                for assignment in assignments
                if assignment.task.task_id in last_resort_task_ids
            ]
            last_resort_protected = [
                assignment
                for assignment in assignments
                if assignment.task.task_id not in last_resort_task_ids
            ]
            last_resort_repaired, last_resort_lines = repair_global_teacher_time_conflicts(
                data_dir,
                last_resort_movable,
                last_resort_protected,
                class_conflict_groups,
                max_seconds=fast_repair_max_seconds,
            )
            assignments = deduplicate_assignments([*last_resort_protected, *last_resort_repaired])
            if last_resort_lines:
                strategy_lines.extend(["锁定套班最小兜底让位:", *last_resort_lines])
            last_resort_class_ids = {assignment.task.class_id for assignment in last_resort_repaired}
            assignments, last_resort_experience_lines = repair_public_same_subject_day_overloads(
                data_dir,
                assignments,
                class_metadata,
                class_conflict_groups,
                load_active_blackout_dates(data_dir),
                target_class_ids=last_resort_class_ids,
                max_seconds=fast_same_day_repair_max_seconds,
            )
            if last_resort_experience_lines:
                strategy_lines.extend(["锁定套班让位后同科同日复检:", *last_resort_experience_lines])
        perf.mark("快速锁定套班兜底让位")
    assignments, renjie_lines = repair_renjie_offline_availability(
        data_dir,
        assignments,
        class_metadata,
        class_conflict_groups,
        load_active_blackout_dates(data_dir),
    )
    if renjie_lines:
        strategy_lines.extend(renjie_lines)
    perf.mark("快速任洁线下可排窗口修复")
    long_camp_target_suites = (halfyear_suites | additional_suites) & {
        suite_code_for_class(class_id, class_metadata)
        for class_id, meta in class_metadata.items()
        if meta.get("sub_product") in {"全年营", "半年营"}
        or suite_code_for_class(class_id, class_metadata) in HALF_YEAR_BATCH_SUITES
    }
    long_camp_blackout_dates = load_active_blackout_dates(data_dir)
    long_camp_week_balance_lines: List[str] = []
    if long_camp_target_suites:
        assignments, long_camp_week_balance_lines = repair_long_camp_subject_week_balance(
            data_dir,
            assignments,
            class_metadata,
            class_conflict_groups,
            long_camp_blackout_dates,
            target_suite_codes=long_camp_target_suites,
            max_moves=320,
        )
        if long_camp_week_balance_lines:
            strategy_lines.extend(long_camp_week_balance_lines)
    perf.mark("快速长线营科目周均衡修复")
    assignments, tail_chain_lines = repair_long_camp_tail_week_chain_moves(
        data_dir,
        assignments,
        class_metadata,
        class_conflict_groups,
        long_camp_blackout_dates,
    )
    if tail_chain_lines:
        strategy_lines.extend(tail_chain_lines)
    perf.mark("快速长线营尾周连锁挪课")
    student_week_balance_lines: List[str] = []
    if not autumn_only_additional:
        assignments, student_week_balance_lines = repair_student_experience_week_balance(
            data_dir,
            assignments,
            class_metadata,
            class_conflict_groups,
            long_camp_blackout_dates,
            target_suite_codes=supported_suites,
            max_moves=160,
            max_seconds=120,
        )
        if student_week_balance_lines:
            strategy_lines.extend(student_week_balance_lines)
    else:
        strategy_lines.append("秋季-only 快速模式: 跳过全局学生体验周均衡修复。")
    if long_camp_week_balance_lines or tail_chain_lines or student_week_balance_lines:
        strategy_lines = refresh_long_camp_minimum_gap_report_lines(
            strategy_lines,
            assignments,
            class_metadata,
            long_camp_blackout_dates,
            target_suite_codes=long_camp_target_suites or None,
        )
    perf.mark("快速学生体验周均衡修复")
    if not autumn_only_additional:
        assignments, politics_segment_lines = repair_politics_segments_for_teacher_conflicts(
            data_dir,
            assignments,
            class_metadata,
            class_conflict_groups,
            long_camp_blackout_dates,
        )
        if politics_segment_lines:
            strategy_lines.extend(["政治老师冲突整段重排:", *politics_segment_lines])
    else:
        strategy_lines.append("秋季-only 快速模式: 跳过全局政治整段重排。")
    perf.mark("快速政治整段重排消冲突")
    room_names = load_room_names(data_dir)
    room_names.update({row["room_id"]: row.get("room_name") or row["room_id"] for row in baseline_rows if row.get("room_id")})
    publish_blocking_teacher_conflicts = teacher_time_conflict_lines(assignments)
    if publish_blocking_teacher_conflicts and not ALLOW_PUBLISH_WITH_TEACHER_CONFLICTS:
        rejected_dir = Path("outputs") / f"rejected_fast_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        rejected_dir.mkdir(parents=True, exist_ok=True)
        rejected_csv = rejected_dir / OUTPUT_CSV.name
        rejected_html = rejected_dir / OUTPUT_HTML.name
        rejected_conflicts = rejected_dir / TEACHER_CONFLICT_CSV.name
        write_batch_csv(assignments, rejected_csv, room_names)
        write_day_table_html(
            assignments,
            rejected_html,
            "课表维护总表（快速候选未发布）",
            ["AM", "PM", "EVENING"],
            room_names,
            assignments[0].candidate.slots[0].date if assignments else None,
            assignments[-1].candidate.slots[0].date if assignments else None,
            class_metadata,
            load_all_class_window_constraint_items(data_dir),
        )
        write_teacher_time_conflicts_csv(assignments, rejected_conflicts, room_names)
        raise ValueError(
            "快速模式候选存在老师硬冲突，已拒绝发布并另存到 "
            f"{rejected_dir}: {'；'.join(publish_blocking_teacher_conflicts[:5])}"
        )
    publish_blocking_coverage_gaps = public_coverage_gap_rows_for_assignments(
        data_dir,
        assignments,
        class_metadata,
    )
    if publish_blocking_coverage_gaps and not ALLOW_PUBLISH_WITH_COVERAGE_GAPS:
        coverage_lines = coverage_gap_blocking_lines(publish_blocking_coverage_gaps)
        rejected_dir = Path("outputs") / f"rejected_fast_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        rejected_dir.mkdir(parents=True, exist_ok=True)
        rejected_csv = rejected_dir / OUTPUT_CSV.name
        rejected_html = rejected_dir / OUTPUT_HTML.name
        rejected_conflicts = rejected_dir / TEACHER_CONFLICT_CSV.name
        rejected_report = rejected_dir / OUTPUT_REPORT.name
        write_batch_csv(assignments, rejected_csv, room_names)
        write_day_table_html(
            assignments,
            rejected_html,
            "课表维护总表（快速候选未发布）",
            ["AM", "PM", "EVENING"],
            room_names,
            assignments[0].candidate.slots[0].date if assignments else None,
            assignments[-1].candidate.slots[0].date if assignments else None,
            class_metadata,
            load_all_class_window_constraint_items(data_dir),
        )
        write_teacher_time_conflicts_csv(assignments, rejected_conflicts, room_names)
        rejected_report.write_text(
            "# 课表维护快速候选未发布\n\n"
            "## 阻断原因\n\n"
            + "\n".join(f"- {line}" for line in coverage_lines)
            + "\n\n"
            f"- 候选输出 CSV: {rejected_csv}\n"
            f"- 候选输出 HTML: {rejected_html}\n"
            f"- 继续保留正式课表维护总表: {OUTPUT_HTML}\n",
            encoding="utf-8",
        )
        raise ValueError(
            "快速模式候选存在班级总课时缺口，已拒绝发布并另存到 "
            f"{rejected_dir}: {coverage_lines[0]}"
        )
    rooms = load_rooms_for_capacity(data_dir / "scheduler_input_draft.json")
    capacity_warnings = assignment_capacity_warning_lines(new_assignments, rooms, room_names)
    warning_lines.extend(capacity_warnings)
    warning_lines.extend(
        teacher_same_day_campus_warning_lines(
            new_assignments,
            rooms,
            load_area_travel_minutes(data_dir),
        )
    )
    write_batch_csv(assignments, OUTPUT_CSV, room_names)
    write_day_table_html(
        assignments,
        OUTPUT_HTML,
        "课表维护总表",
        ["AM", "PM", "EVENING"],
        room_names,
        assignments[0].candidate.slots[0].date if assignments else None,
        assignments[-1].candidate.slots[0].date if assignments else None,
        class_metadata,
        load_all_class_window_constraint_items(data_dir),
    )
    write_teacher_time_conflicts_csv(assignments, TEACHER_CONFLICT_CSV, room_names)
    perf.mark("写出快速 CSV/HTML")
    write_fast_report(
        OUTPUT_REPORT,
        target_suites=supported_suites,
        target_sub_products=affected_sub_products,
        affected_class_ids=affected_class_ids,
        reused_count=assignment_standard_lesson_count(reused_assignments),
        removed_count=len(removed_rows),
        new_count=assignment_standard_lesson_count(new_assignments),
        final_assignments=assignments,
        class_conflict_groups=class_conflict_groups,
        strategy_lines=strategy_lines,
        warning_lines=warning_lines,
        perf=perf,
        class_metadata=class_metadata,
        blackout_dates=load_active_blackout_dates(data_dir),
        affected_start_date=WYQC_AUTUMN_START if autumn_only_additional else None,
    )
    shutil.copyfile(OUTPUT_CSV, LEGACY_OUTPUT_CSV)
    shutil.copyfile(OUTPUT_HTML, LEGACY_OUTPUT_HTML)
    shutil.copyfile(OUTPUT_REPORT, LEGACY_OUTPUT_REPORT)
    print(
        "已快速更新课表维护总表: "
        f"复用 {assignment_standard_lesson_count(reused_assignments)} 条，"
        f"重排 {assignment_standard_lesson_count(new_assignments)} 条"
    )
    print(OUTPUT_CSV)
    print(OUTPUT_HTML)


def main() -> None:
    parser = argparse.ArgumentParser(description="构建课表维护总表")
    parser.add_argument("--data-dir", default="data")
    parser.add_argument("--mode", choices=("full", "fast"), default="full")
    parser.add_argument("--suite-code", action="append", default=[], help="快速模式重排套班编码，可重复或逗号分隔")
    parser.add_argument("--class-id", action="append", default=[], help="快速模式重排班级编码，可重复或逗号分隔")
    parser.add_argument("--sub-product", action="append", default=[], help="快速模式重排子产品，可重复或逗号分隔")
    args = parser.parse_args()
    data_dir = Path(args.data_dir)
    perf = PerfLog()
    try:
        if args.mode == "fast":
            run_fast(data_dir, args.suite_code, args.class_id, args.sub_product, perf)
        else:
            run_full(data_dir, perf)
    except Exception as exc:
        if args.mode == "fast":
            write_fast_failure_report(
                OUTPUT_REPORT,
                suite_values=args.suite_code,
                class_values=args.class_id,
                sub_product_values=args.sub_product,
                error=exc,
                perf=perf,
            )
        raise


if __name__ == "__main__":
    main()
