#!/usr/bin/env python3
"""Build an ERP lesson-id map by class-level date/time sequence."""

from __future__ import annotations

import argparse
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.csv_utils import read_csv_rows, write_csv_rows
from scripts.field_utils import normalize_excel_text as clean


DEFAULT_SCHEDULE = Path("outputs/batch_schedule_maintenance.csv")
DEFAULT_OUTPUT_DIR = Path("outputs")
PUBLIC_SUBJECTS = {"英语", "政治", "数学", "语文"}


def split_values(value: str) -> List[str]:
    return [item.strip() for item in value.replace(",", "|").replace("，", "|").split("|") if item.strip()]


def normalize_date(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = clean(value).replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return text[:10]


def normalize_time(value: object) -> Tuple[str, str]:
    text = clean(value)
    if "~" in text:
        start, end = text.split("~", 1)
        return normalize_one_time(start), normalize_one_time(end)
    return normalize_one_time(text), ""


def normalize_one_time(value: str) -> str:
    text = clean(value)
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%H:%M")
        except ValueError:
            pass
    return text


def schedule_rows(path: Path, start_date: str, end_date: str, include_subjects: set[str]) -> List[Dict[str, str]]:
    rows = []
    for row in read_csv_rows(path):
        date = normalize_date(row.get("date"))
        subject = clean(row.get("subject"))
        if start_date <= date <= end_date and (not include_subjects or subject in include_subjects):
            rows.append(row)
    rows.sort(key=schedule_sort_key)
    return rows


def schedule_sort_key(row: Dict[str, str]) -> Tuple[str, str, str, str, str, str]:
    return (
        clean(row.get("class_id")),
        normalize_date(row.get("date")),
        normalize_one_time(clean(row.get("start_time"))),
        normalize_one_time(clean(row.get("end_time"))),
        clean(row.get("lesson_slot")),
        clean(row.get("course_code")) or clean(row.get("course_name")),
    )


def erp_sort_key(row: Dict[str, str]) -> Tuple[str, str, str, str, str]:
    start_time, end_time = normalize_time(row.get("时间"))
    return (
        clean(row.get("班级编码")),
        normalize_date(row.get("日期")),
        start_time,
        end_time,
        clean(row.get("课次ID")),
    )


def load_erp_rows(path: Path) -> List[Dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    worksheet.reset_dimensions()
    raw_rows = worksheet.iter_rows(values_only=True)
    # Row 1 is the title.
    next(raw_rows, None)
    headers = [clean(value) for value in next(raw_rows)]
    rows: List[Dict[str, str]] = []
    for raw in raw_rows:
        row = {headers[index]: clean(raw[index]) if index < len(raw) else "" for index in range(len(headers))}
        if any(row.values()):
            rows.append(row)
    return rows


def group_by_class(rows: Iterable[Dict[str, str]], class_field: str) -> Dict[str, List[Dict[str, str]]]:
    grouped: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for row in rows:
        class_id = clean(row.get(class_field))
        if class_id:
            grouped[class_id].append(row)
    return grouped


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Match ERP lesson IDs to generated schedule rows by class-level date/time sequence.")
    parser.add_argument("--erp-export", type=Path, required=True)
    parser.add_argument("--schedule-csv", type=Path, default=DEFAULT_SCHEDULE)
    parser.add_argument("--start-date", default="2026-07-01")
    parser.add_argument("--end-date", default="2026-12-31")
    parser.add_argument("--include-subjects", default="|".join(sorted(PUBLIC_SUBJECTS)))
    parser.add_argument("--output-map", type=Path)
    parser.add_argument("--report", type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_map = args.output_map or DEFAULT_OUTPUT_DIR / f"erp_lesson_id_map_{timestamp}.csv"
    report = args.report or DEFAULT_OUTPUT_DIR / f"erp_lesson_id_map_report_{timestamp}.csv"
    include_subjects = set(split_values(args.include_subjects))

    schedule = schedule_rows(args.schedule_csv, args.start_date, args.end_date, include_subjects)
    schedule_class_ids = {clean(row.get("class_id")) for row in schedule}
    erp_rows = [
        row
        for row in load_erp_rows(args.erp_export)
        if clean(row.get("班级编码")) in schedule_class_ids
        and clean(row.get("授课内容类型")) == "正课"
    ]

    schedule_by_class = group_by_class(schedule, "class_id")
    erp_by_class = group_by_class(erp_rows, "班级编码")
    for items in schedule_by_class.values():
        items.sort(key=schedule_sort_key)
    for items in erp_by_class.values():
        items.sort(key=erp_sort_key)

    mapped_rows: List[Dict[str, str]] = []
    report_rows: List[Dict[str, str]] = []
    for class_id in sorted(schedule_class_ids):
        schedule_items = schedule_by_class.get(class_id, [])
        erp_items = erp_by_class.get(class_id, [])
        pair_count = min(len(schedule_items), len(erp_items))
        for index in range(pair_count):
            schedule_row = schedule_items[index]
            erp_row = erp_items[index]
            mapped_rows.append(
                {
                    "class_id": class_id,
                    "date": normalize_date(schedule_row.get("date")),
                    "start_time": normalize_one_time(clean(schedule_row.get("start_time"))),
                    "end_time": normalize_one_time(clean(schedule_row.get("end_time"))),
                    "erp_lesson_id": clean(erp_row.get("课次ID")),
                    "sequence": str(index + 1),
                    "source_erp_date": normalize_date(erp_row.get("日期")),
                    "source_erp_time": clean(erp_row.get("时间")),
                    "subject": clean(schedule_row.get("subject")),
                    "stage": clean(schedule_row.get("stage")),
                    "course_module": clean(schedule_row.get("course_module")),
                }
            )
        if len(schedule_items) != len(erp_items):
            report_rows.append(
                {
                    "type": "count_mismatch",
                    "class_id": class_id,
                    "schedule_count": str(len(schedule_items)),
                    "erp_count": str(len(erp_items)),
                    "diff_schedule_minus_erp": str(len(schedule_items) - len(erp_items)),
                    "message": "ERP课次数量多于预排课表，剩余课次ID未使用。" if len(erp_items) > len(schedule_items) else "ERP课次数量少于预排课表，部分课表无法回填课次ID。",
                }
            )
        for index, erp_row in enumerate(erp_items[pair_count:], start=pair_count + 1):
            report_rows.append(
                {
                    "type": "unused_erp_lesson",
                    "class_id": class_id,
                    "schedule_count": str(len(schedule_items)),
                    "erp_count": str(len(erp_items)),
                    "diff_schedule_minus_erp": str(len(schedule_items) - len(erp_items)),
                    "message": f"第 {index} 个ERP课次未使用: {clean(erp_row.get('课次ID'))} {normalize_date(erp_row.get('日期'))} {clean(erp_row.get('时间'))}",
                }
            )
        if len(schedule_items) > len(erp_items):
            for index, schedule_row in enumerate(schedule_items[pair_count:], start=pair_count + 1):
                report_rows.append(
                    {
                        "type": "missing_erp_lesson",
                        "class_id": class_id,
                        "schedule_count": str(len(schedule_items)),
                        "erp_count": str(len(erp_items)),
                        "diff_schedule_minus_erp": str(len(schedule_items) - len(erp_items)),
                        "message": f"第 {index} 个预排课节缺ERP课次ID: {normalize_date(schedule_row.get('date'))} {clean(schedule_row.get('start_time'))}~{clean(schedule_row.get('end_time'))}",
                    }
                )

    write_csv_rows(
        output_map,
        ["class_id", "date", "start_time", "end_time", "erp_lesson_id", "sequence", "source_erp_date", "source_erp_time", "subject", "stage", "course_module"],
        mapped_rows,
    )
    write_csv_rows(
        report,
        ["type", "class_id", "schedule_count", "erp_count", "diff_schedule_minus_erp", "message"],
        report_rows,
    )

    counts = Counter(row["type"] for row in report_rows)
    print(f"schedule_rows={len(schedule)}")
    print(f"erp_rows={len(erp_rows)}")
    print(f"mapped_rows={len(mapped_rows)}")
    print(f"output_map={output_map}")
    print(f"report={report}")
    for key in sorted(counts):
        print(f"{key}={counts[key]}")


if __name__ == "__main__":
    main()
