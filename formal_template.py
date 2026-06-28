#!/usr/bin/env python3
from __future__ import annotations

import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

import data_admin_server
from business_class_import import (
    BUSINESS_EXAM_MONTH,
    BUSINESS_EXAM_SEASON,
    BUSINESS_PROJECT,
    WINDOW_END,
    WINDOW_START,
    compact_text,
    merge_candidate_rows,
    normalize_scheduled_lessons,
    product_course_hour_rows,
    row_value,
    parse_business_date,
    learn_teacher_assignments,
    teacher_employee_ids_from_business_rows,
)
from scripts.csv_utils import csv_rows_text, serialize_csv_value
from scripts.table_schema import BUSINESS_PRODUCT_MAPPING_FIELDNAMES, TEACHER_ASSIGNMENT_FIELDNAMES


@dataclass
class TemplateGenerationResult:
    xlsx_path: Path
    zip_path: Path
    report_path: Path
    upload_dir: Path
    row_counts: Dict[str, int]
    warnings: List[str]

def csv_text(rows: Sequence[Mapping[str, Any]], fieldnames: Sequence[str]) -> str:
    return csv_rows_text(fieldnames, rows, bom=True, value_formatter=serialize_csv_value)


def selected_business_rows(rows: Iterable[Mapping[str, Any]]) -> Tuple[List[Mapping[str, Any]], List[str]]:
    selected: List[Mapping[str, Any]] = []
    warnings: List[str] = []
    skipped_project = 0
    skipped_exam = 0
    skipped_window = 0
    for row in rows:
        class_id = row_value(row, "班级编码")
        if row_value(row, "管理项目") != BUSINESS_PROJECT:
            skipped_project += 1
            continue
        if compact_text(row.get("考试月份")) != BUSINESS_EXAM_MONTH:
            skipped_exam += 1
            continue
        try:
            start_date = parse_business_date(row.get("实际开课日期"), f"班级 {class_id}/实际开课日期")
            end_date = parse_business_date(row.get("实际结课日期"), f"班级 {class_id}/实际结课日期")
        except ValueError as exc:
            warnings.append(str(exc))
            continue
        if start_date > WINDOW_END or end_date < WINDOW_START:
            skipped_window += 1
            continue
        selected.append(row)
    warnings.append(
        f"模板筛选业务班级: 纳入 {len(selected)} 行，跳过非考研 {skipped_project} 行，"
        f"跳过非 {BUSINESS_EXAM_MONTH} 考试月份 {skipped_exam} 行，跳过窗口外 {skipped_window} 行。"
    )
    return selected, warnings


def product_catalog_rows(base_payload: Mapping[str, Any]) -> List[Dict[str, Any]]:
    products = data_admin_server.product_catalog(
        list(base_payload.get("products", [])),
        list(base_payload.get("product_courses", [])),
    )
    course_counts: Dict[str, int] = {}
    for course in base_payload.get("product_courses", []):
        product_id = data_admin_server.normalize_text(course.get("product_id"))
        if product_id:
            course_counts[product_id] = course_counts.get(product_id, 0) + 1
    rows: List[Dict[str, Any]] = []
    for product_id, product in sorted(products.items()):
        rows.append(
            {
                "local_product_id": product_id,
                "product_name": product.get("name", ""),
                "project": product.get("project", ""),
                "product_line": product.get("product_line", ""),
                "sub_product": product.get("sub_product", ""),
                "product_system": product.get("product_system", ""),
                "subject": product.get("subject", ""),
                "subject_category": product.get("subject_category", ""),
                "course_nature": product.get("course_nature", ""),
                "course_count": course_counts.get(product_id, 0),
            }
        )
    return rows


def candidate_products(product_name: str, product_rows: Sequence[Mapping[str, Any]]) -> str:
    text = data_admin_server.normalize_text(product_name)
    if not text:
        return ""
    scored: List[Tuple[int, str]] = []
    for row in product_rows:
        product_id = data_admin_server.normalize_text(row.get("local_product_id") or row.get("canonical_product_id"))
        name = data_admin_server.normalize_text(row.get("product_name"))
        if not product_id:
            continue
        score = 0
        if text == name:
            score += 100
        if text and name and (text in name or name in text):
            score += 40
        for keyword in ("英语", "政治", "数学", "管综", "计算机", "西医", "全年", "半年", "寒暑", "暑假", "冲刺", "无忧", "导学"):
            if keyword in text and keyword in name:
                score += 5
        if score:
            scored.append((score, f"{product_id}:{name}"))
    scored.sort(reverse=True)
    return "|".join(item for _, item in scored[:5])


def existing_product_map(rows: Iterable[Mapping[str, Any]]) -> Dict[str, str]:
    result: Dict[str, str] = {}
    for row in rows:
        business_product_id = row_value(row, "business_product_id", "erp_course_code", "课程产品编号", "product_id")
        local_product_id = row_value(row, "local_product_id", "canonical_product_id", "系统产品ID", "标准产品ID", "canonical_id")
        if business_product_id and local_product_id:
            result[business_product_id] = local_product_id
    return result


def build_product_map_rows(
    business_rows: Sequence[Mapping[str, Any]],
    product_rows: Sequence[Mapping[str, Any]],
    existing_map: Mapping[str, str],
) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    for row in business_rows:
        business_product_id = row_value(row, "课程产品编号")
        if not business_product_id:
            continue
        item = grouped.setdefault(
            business_product_id,
            {
                "local_product_id": existing_map.get(business_product_id, ""),
                "local_product_name": "",
                "local_product_line": "",
                "local_sub_product": "",
                "local_product_system": "",
                "local_course_nature": "",
                "local_subject": "",
                "erp_product_key": "",
                "erp_course_code": business_product_id,
                "erp_course_name": row_value(row, "课程产品名称"),
                "erp_version_code": "",
                "erp_version_name": "",
                "erp_product_system": row_value(row, "产品体系"),
                "erp_product_category": "",
                "erp_project_name": row_value(row, "管理项目"),
                "erp_subject": "",
                "erp_class_type": "",
                "erp_duration_minutes": "",
                "erp_lesson_count": "",
                "erp_single_lesson_minutes": "",
                "erp_class_form": "",
                "erp_teaching_method": "",
                "match_status": "已匹配" if existing_map.get(business_product_id) else "待确认",
                "match_confidence": "",
                "match_source": "预填模板",
                "business_product_id": business_product_id,
                "business_product_name": row_value(row, "课程产品名称"),
                "product_system": row_value(row, "产品体系"),
                "class_count": 0,
                "class_name_keywords": "",
                "candidate_local_products": "",
                "notes": "",
            },
        )
        item["class_count"] += 1
    for item in grouped.values():
        if item["product_system"] == "计费体系":
            item["candidate_local_products"] = ""
            item["notes"] = "计费体系自动排除，无需填写"
            item["match_status"] = "自动排除"
            continue
        item["candidate_local_products"] = candidate_products(item["business_product_name"], product_rows)
        if not item["local_product_id"]:
            candidate_note = (
                f"；候选本地产品：{item['candidate_local_products']}"
                if item["candidate_local_products"]
                else ""
            )
            item["notes"] = f"请确认 local_product_id{candidate_note}"
    return [grouped[key] for key in sorted(grouped)]


def build_business_selection_rows(business_rows: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for row in business_rows:
        rows.append(
            {
                "class_id": row_value(row, "班级编码"),
                "class_name": row_value(row, "班级名称（外）"),
                "business_product_id": row_value(row, "课程产品编号"),
                "business_product_name": row_value(row, "课程产品名称"),
                "product_system": row_value(row, "产品体系"),
                "exam_season": BUSINESS_EXAM_SEASON,
                "exam_month": row_value(row, "考试月份"),
                "actual_start_date": row_value(row, "实际开课日期"),
                "actual_end_date": row_value(row, "实际结课日期"),
                "room_id": row_value(row, "教室编码"),
                "room_name": row_value(row, "上课教室"),
                "teacher_raw": row_value(row, "授课教师"),
                "merge_detail": row_value(row, "合班详情"),
                "schedule_status": row_value(row, "排课完成状态"),
            }
        )
    return rows


def build_gap_rows(
    product_map_rows: Sequence[Mapping[str, Any]],
    lesson_warnings: Sequence[str],
    learning_warnings: Sequence[str],
) -> List[Dict[str, Any]]:
    gaps: List[Dict[str, Any]] = []
    for row in product_map_rows:
        if row.get("product_system") == "计费体系":
            continue
        if not data_admin_server.normalize_text(row.get("local_product_id")):
            gaps.append(
                {
                    "gap_type": "产品映射待确认",
                    "target_id": row.get("business_product_id", ""),
                    "description": f"{row.get('business_product_name', '')} 需要填写 local_product_id",
                    "severity": "阻塞",
                }
            )
    for warning in [*lesson_warnings, *learning_warnings]:
        gaps.append({"gap_type": "历史课表提示", "target_id": "", "description": warning, "severity": "核对"})
    return gaps


def autosize_sheet(sheet: Any) -> None:
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells[:200]:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)


def write_workbook(path: Path, sheets: Sequence[Tuple[str, Sequence[Mapping[str, Any]], Sequence[str]]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ModuleNotFoundError as exc:
        raise RuntimeError("生成 Excel 模板需要 openpyxl，请先安装 requirements.txt") from exc

    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    for title, rows, fieldnames in sheets:
        sheet = workbook.create_sheet(title[:31])
        sheet.append(list(fieldnames))
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for row in rows:
            sheet.append([serialize_csv_value(row.get(field, "")) for field in fieldnames])
        autosize_sheet(sheet)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_zip(path: Path, sheets: Sequence[Tuple[str, Sequence[Mapping[str, Any]], Sequence[str]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for title, rows, fieldnames in sheets:
            archive.writestr(f"{title}.csv", csv_text(rows, fieldnames).encode("utf-8"))


def write_template_report(
    path: Path,
    *,
    source: Path,
    row_counts: Mapping[str, int],
    warnings: Sequence[str],
    xlsx_path: Path,
    zip_path: Path,
    gap_count: int,
) -> None:
    lines = [
        "# 排课预填模板生成报告",
        "",
        f"- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 源数据: {source}",
        f"- Excel 模板: {xlsx_path}",
        f"- CSV 模板包: {zip_path}",
        f"- 缺口数量: {gap_count}",
        "",
        "## 识别到的数据表",
    ]
    for table, count in sorted(row_counts.items()):
        lines.append(f"- {table}: {count} 行")
    if warnings:
        lines.extend(["", "## 提示"])
        lines.extend(f"- {warning}" for warning in warnings)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def generate_formal_launch_template(source: Path, output_dir: Path, timestamp: str) -> TemplateGenerationResult:
    import run_scheduling_pipeline

    tables = run_scheduling_pipeline.load_source_tables(source)
    row_counts = {name: len(table.rows) for name, table in tables.items()}
    business_table = tables.get("business_classes")
    if not business_table:
        raise ValueError("生成预填模板需要上传业务班级导出表，例如 20260429班级查询导出.csv")

    base_payload = run_scheduling_pipeline.overlay_standard_tables_on_state(tables)
    business_rows, filter_warnings = selected_business_rows(business_table.rows)
    employee_ids_by_name = teacher_employee_ids_from_business_rows(business_table.rows)
    scheduled_rows = list(tables.get("scheduled_lessons").rows) if "scheduled_lessons" in tables else []
    lessons, lesson_warnings = normalize_scheduled_lessons(scheduled_rows)
    _, learned_assignment_rows, learning_warnings = learn_teacher_assignments(lessons, {}, employee_ids_by_name)
    explicit_merge_rows, inferred_merge_rows = merge_candidate_rows(lessons)

    product_rows = product_catalog_rows(base_payload)
    existing_mapping_rows = list(tables["business_product_mappings"].rows) if "business_product_mappings" in tables else []
    product_map_rows = build_product_map_rows(
        business_rows,
        product_rows,
        existing_product_map(existing_mapping_rows),
    )
    business_selection_rows = build_business_selection_rows(business_rows)
    history_rows = [
        {
            "class_id": row.get("class_id", ""),
            "product_id": row.get("product_id", ""),
            "product_name": row.get("product_name", ""),
            "subject": row.get("subject", ""),
            "stage": row.get("stage", ""),
            "course_group": row.get("course_group", ""),
            "teacher_id": row.get("teacher_id", ""),
            "teacher_name": row.get("teacher_name", ""),
            "source": row.get("notes", ""),
        }
        for row in learned_assignment_rows
    ]
    for row in product_course_hour_rows(lessons):
        history_rows.append({"class_id": "", "product_id": row["business_product_id"], "product_name": row["business_product_name"], "subject": row["subject"], "stage": row["stage"], "course_group": row["course_group"], "teacher_id": "", "teacher_name": "", "source": f"{row['course_module']} 历史课时 {row['scheduled_hours']} 小时 / {row['lesson_count']} 节"})
    for row in explicit_merge_rows:
        history_rows.append({"class_id": row["source_class_id"], "product_id": "", "product_name": "", "subject": row["subject"], "stage": row["stage"], "course_group": row["course_group"], "teacher_id": "", "teacher_name": "", "source": f"历史明确合班: {row['source_class_id']} -> {row['scheduled_class_id']}，如需部分合班例外请在班级老师安排表维护共享关系"})
    for row in inferred_merge_rows:
        history_rows.append({"class_id": row["class_ids"], "product_id": "", "product_name": "", "subject": row["subject"], "stage": row["stage"], "course_group": row["course_group"], "teacher_id": row["teacher_id"], "teacher_name": row["teacher_name"], "source": f"{row['course_module']} 疑似合班候选"})

    gap_rows = build_gap_rows(product_map_rows, lesson_warnings, learning_warnings)
    instruction_rows = [
        {"item": "流程", "description": "先补齐 business_product_mappings 和 class_teacher_assignments，再回到网页做上传前校验。"},
        {"item": "首批范围", "description": f"系统只纳入 考研/考博、考试月份 {BUSINESS_EXAM_MONTH}、考季 {BUSINESS_EXAM_SEASON}，且与 2026-07-01 至 2026-12-31 有交集的班级。"},
        {"item": "计费体系", "description": "计费体系自动排除，不需要填写准入规则。"},
        {"item": "ERP产品对应", "description": "同一 ERP 课程产品可能对应多个本地排课产品时，在 local_product_id 填本地产品 ID，多个用 | 分隔；需要按班级名称区分时，拆成多行并填写 class_name_keywords。"},
        {"item": "老师安排", "description": "按班级、产品、阶段、课程类别填写；合班共享课表时填写 class_schedule_mode=共享实际排课班级，并填写 actual_scheduled_class_id。"},
        {"item": "共享课表关系", "description": "业务导出自带合班详情时，在班级老师安排表确认实际排课班级；共享班只维护 actual_scheduled_class_id，不单独生成课次。"},
        {"item": "历史课表", "description": "已排课明细用于学习老师和抵扣 2026-06-30 前已排课时，不会直接生成未来排课结果。"},
    ]
    scheduled_field_rows = [
        {"field": "class_id", "required": "是", "description": "班级编码"},
        {"field": "date", "required": "是", "description": "上课日期，支持 YYYY/MM/DD 或 YYYY-MM-DD"},
        {"field": "start_time", "required": "建议", "description": "开始时间，例如 08:00"},
        {"field": "end_time", "required": "建议", "description": "结束时间，例如 10:00"},
        {"field": "duration_hours", "required": "是", "description": "单次课时，可由开始/结束时间推断，但建议直接提供"},
        {"field": "teacher_id", "required": "建议", "description": "教师员工ID"},
        {"field": "teacher_name", "required": "建议", "description": "教师姓名"},
        {"field": "room_id", "required": "建议", "description": "教室编码"},
        {"field": "business_product_id", "required": "建议", "description": "业务课程产品编号"},
        {"field": "subject", "required": "是", "description": "科目"},
        {"field": "quarter", "required": "可选", "description": "季度标签，例如 寒假、春季、暑假、秋季"},
        {"field": "stage", "required": "是", "description": "阶段"},
        {"field": "course_module", "required": "是", "description": "课程模块"},
        {"field": "course_group", "required": "是", "description": "课程组"},
        {"field": "merge_group", "required": "可选", "description": "明确合班班级编码，多个用 | 分隔"},
    ]

    sheets: List[Tuple[str, Sequence[Mapping[str, Any]], Sequence[str]]] = [
        ("填写说明", instruction_rows, ["item", "description"]),
        ("business_product_mappings", product_map_rows, BUSINESS_PRODUCT_MAPPING_FIELDNAMES),
        ("class_teacher_assignments", learned_assignment_rows, TEACHER_ASSIGNMENT_FIELDNAMES),
        ("scheduled_lessons字段说明", scheduled_field_rows, ["field", "required", "description"]),
        ("产品目录参考", product_rows, ["local_product_id", "product_name", "project", "product_line", "sub_product", "product_system", "subject", "subject_category", "course_nature", "course_count"]),
        ("业务班级筛选结果", business_selection_rows, ["class_id", "class_name", "business_product_id", "business_product_name", "product_system", "exam_season", "exam_month", "actual_start_date", "actual_end_date", "room_id", "room_name", "teacher_raw", "merge_detail", "schedule_status"]),
        ("历史学习结果", history_rows, ["class_id", "product_id", "product_name", "subject", "stage", "course_group", "teacher_id", "teacher_name", "source"]),
        ("缺口清单", gap_rows, ["gap_type", "target_id", "description", "severity"]),
    ]

    template_dir = output_dir / "templates" / timestamp
    xlsx_path = template_dir / f"formal_launch_template_{timestamp}.xlsx"
    zip_path = template_dir / f"formal_launch_csv_templates_{timestamp}.zip"
    report_path = template_dir / f"template_report_{timestamp}.md"
    write_workbook(xlsx_path, sheets)
    write_zip(zip_path, sheets[1:])
    warnings = [*filter_warnings, *lesson_warnings, *learning_warnings]
    write_template_report(
        report_path,
        source=source,
        row_counts=row_counts,
        warnings=warnings,
        xlsx_path=xlsx_path,
        zip_path=zip_path,
        gap_count=len(gap_rows),
    )
    return TemplateGenerationResult(
        xlsx_path=xlsx_path,
        zip_path=zip_path,
        report_path=report_path,
        upload_dir=source,
        row_counts=row_counts,
        warnings=warnings,
    )
