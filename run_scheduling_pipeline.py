#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import shutil
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set

import data_admin_server
import scheduler
from business_class_import import BusinessDataError, convert_business_tables
from generate_time_slots import generate_time_slots, parse_weekdays
from scripts.csv_utils import clean_csv_rows, read_csv_with_fieldnames, write_csv_rows
from scripts.field_utils import normalize_excel_cell_text, normalize_text
from scripts.product_catalog import product_catalog as shared_product_catalog
from scripts.template_tables import (
    BUSINESS_SOURCE_TABLES,
    build_table_aliases,
    table_name_for as shared_table_name_for,
)
from scripts.table_schema import STANDARD_TABLE_FIELDNAMES, TEACHER_ASSIGNMENT_FIELDNAMES


TABLES = list(STANDARD_TABLE_FIELDNAMES)

BUSINESS_TABLES = list(BUSINESS_SOURCE_TABLES)

SOURCE_TABLES = [*TABLES, *BUSINESS_TABLES]
REPORT_TABLES = [*TABLES, "business_classes", "scheduled_lessons"]
COMPATIBILITY_REPORT_TABLES: List[str] = []

TABLE_ALIASES = build_table_aliases(SOURCE_TABLES)

TABLE_FIELDNAMES = STANDARD_TABLE_FIELDNAMES


@dataclass
class LoadedTable:
    name: str
    source: str
    rows: List[Dict[str, Any]]


@dataclass
class PipelineResult:
    scheduler_input_path: Path
    schedule_csv_path: Path
    schedule_html_path: Path
    report_path: Path
    backup_path: Optional[Path]
    row_counts: Dict[str, int]
    warnings: List[str]
    generated_files: List[Path]


@dataclass
class PreflightResult:
    passed: bool
    report_path: Path
    row_counts: Dict[str, int]
    warnings: List[str]
    generated_files: List[Path]
    missing_teacher_requirements: List[MissingTeacherRequirement]
    missing_teacher_rows: List[Dict[str, str]]
    error: str = ""


@dataclass(frozen=True)
class MissingTeacherRequirement:
    class_id: str
    product_id: str
    subject: str
    stage: str
    course_group: str


@dataclass
class MissingTeacherDiagnostics:
    requirements: List[MissingTeacherRequirement] = field(default_factory=list)
    rows: List[Dict[str, str]] = field(default_factory=list)


class PipelineError(RuntimeError):
    pass


def empty_source_row_counts() -> Dict[str, int]:
    return {table: 0 for table in SOURCE_TABLES}


@dataclass
class PipelineRunContext:
    timestamp: str
    source: Path
    data_dir: Path
    output_dir: Path
    report_path: Path
    tables: Dict[str, LoadedTable] = field(default_factory=dict)
    row_counts: Dict[str, int] = field(default_factory=empty_source_row_counts)
    warnings: List[str] = field(default_factory=list)
    generated_files: List[Path] = field(default_factory=list)
    state: Optional[Dict[str, Any]] = None
    time_slots: Optional[List[Dict[str, Any]]] = None


def row_counts_for_tables(tables: Dict[str, LoadedTable]) -> Dict[str, int]:
    row_counts = {table: len(loaded.rows) for table, loaded in tables.items()}
    for table in SOURCE_TABLES:
        row_counts.setdefault(table, 0)
    return row_counts


def table_name_for(value: str) -> Optional[str]:
    return shared_table_name_for(value, TABLE_ALIASES)


def source_files(source: Path) -> List[Path]:
    if source.is_file():
        return [source]
    if not source.exists():
        raise PipelineError(f"源数据路径不存在: {source}")
    files = [
        path for path in sorted(source.iterdir())
        if path.suffix.lower() in {".csv", ".xlsx", ".xlsm"} and not path.name.startswith("~$")
    ]
    if not files:
        raise PipelineError(f"源数据目录中没有找到 CSV/XLSX 文件: {source}")
    return files


def load_csv_table_rows(path: Path) -> List[Dict[str, str]]:
    try:
        fieldnames, rows = read_csv_with_fieldnames(path)
    except UnicodeDecodeError as exc:
        raise PipelineError(f"无法识别 CSV 编码: {path}") from exc
    if not fieldnames:
        return []
    validate_headers([header or "" for header in fieldnames], str(path))
    return clean_csv_rows(rows)


def cell_text(cell: Any) -> str:
    return normalize_excel_cell_text(cell.value, getattr(cell, "number_format", ""))


def read_xlsx_tables(path: Path) -> List[LoadedTable]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise PipelineError("读取 Excel 需要先安装 openpyxl: pip install -r requirements.txt") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    tables: List[LoadedTable] = []
    for sheet in workbook.worksheets:
        if sheet.max_row == 1 and sheet.max_column == 1:
            try:
                sheet.reset_dimensions()
            except AttributeError:
                pass
        rows = list(sheet.iter_rows())
        if not rows:
            continue
        header_row_index = first_non_empty_row_index(rows)
        if header_row_index is None:
            continue
        table_name = table_name_for(sheet.title)
        if not table_name:
            data_rows = rows[header_row_index + 1 :]
            if any(any(cell_text(cell) for cell in row) for row in data_rows):
                print(f"跳过未识别的 Excel sheet: {path.name}/{sheet.title}", file=sys.stderr)
            continue
        header_row_index = header_row_index_for_table(rows, table_name, header_row_index)
        headers = [cell_text(cell) for cell in rows[header_row_index]]
        validate_headers(headers, f"{path.name}/{sheet.title}")
        records = []
        for row in rows[header_row_index + 1 :]:
            record = {
                header: cell_text(cell)
                for header, cell in zip(headers, row)
                if header
            }
            records.append(record)
        tables.append(LoadedTable(table_name, f"{path.name}/{sheet.title}", clean_csv_rows(records)))
    return tables


def first_non_empty_row_index(rows: Sequence[Sequence[Any]]) -> Optional[int]:
    for index, row in enumerate(rows):
        if any(cell_text(cell) for cell in row):
            return index
    return None


def header_row_index_for_table(
    rows: Sequence[Sequence[Any]],
    table_name: str,
    fallback_index: int,
) -> int:
    expected = set(TABLE_FIELDNAMES.get(table_name, []))
    if not expected:
        return fallback_index
    for index, row in enumerate(rows[:20]):
        values = [cell_text(cell) for cell in row if cell_text(cell)]
        if not values:
            continue
        matching = [value for value in values if value in expected]
        if len(matching) >= min(2, len(expected)):
            return index
    return fallback_index


def validate_headers(headers: Sequence[str], label: str) -> None:
    seen: Set[str] = set()
    for header in headers:
        if not header:
            continue
        if header in seen:
            raise PipelineError(f"{label} 存在重复表头: {header}")
        seen.add(header)


def load_source_tables(source: Path) -> Dict[str, LoadedTable]:
    loaded: Dict[str, LoadedTable] = {}
    for path in source_files(source):
        suffix = path.suffix.lower()
        if suffix == ".csv":
            table_name = table_name_for(path.name)
            if not table_name:
                print(f"跳过未识别的 CSV 文件: {path.name}", file=sys.stderr)
                continue
            tables = [LoadedTable(table_name, path.name, load_csv_table_rows(path))]
        else:
            tables = read_xlsx_tables(path)

        for table in tables:
            if table.name in loaded:
                first = loaded[table.name].source
                raise PipelineError(f"重复的数据表 {table.name}: {first} 和 {table.source}")
            loaded[table.name] = table

    if not loaded:
        raise PipelineError("没有识别到任何标准数据表")
    return loaded


def tables_with_business_aliases(tables: Dict[str, LoadedTable]) -> Dict[str, LoadedTable]:
    aliased = dict(tables)
    if "scheduled_lessons" not in aliased and "historical_scheduled_lessons" in tables:
        source = tables["historical_scheduled_lessons"]
        aliased["scheduled_lessons"] = LoadedTable(
            "scheduled_lessons",
            source.source,
            list(source.rows),
        )
    return aliased


def payload_from_tables(tables: Dict[str, LoadedTable]) -> Dict[str, Any]:
    payload = {table: list(tables.get(table, LoadedTable(table, "", [])).rows) for table in TABLES}
    assignments_by_class: Dict[str, List[Dict[str, Any]]] = {}
    known_class_ids = {
        normalize_text(row.get("id") or row.get("class_id"))
        for row in payload["classes"]
    }

    for assignment in payload["class_teacher_assignments"]:
        class_id = normalize_text(assignment.get("class_id"))
        if not class_id:
            raise PipelineError("班级老师安排表存在缺少 class_id 的行")
        if known_class_ids and class_id not in known_class_ids:
            raise PipelineError(f"班级老师安排引用了不存在的班级: {class_id}")
        assignments_by_class.setdefault(class_id, []).append(assignment)

    for row in payload["classes"]:
        class_id = normalize_text(row.get("id") or row.get("class_id"))
        row["teacher_assignments"] = assignments_by_class.get(class_id, [])

    payload.pop("class_teacher_assignments", None)
    return payload


def standard_payload_from_state(state: Dict[str, Any]) -> Dict[str, Any]:
    return {table: list(state.get(table, [])) for table in TABLES}


def overlay_standard_tables_on_state(tables: Dict[str, LoadedTable]) -> Dict[str, Any]:
    payload = standard_payload_from_state(data_admin_server.load_state())
    payload["classes"] = []
    payload["class_teacher_assignments"] = []
    for table in TABLES:
        if table in tables:
            payload[table] = list(tables[table].rows)
    payload["classes"] = []
    return payload


def build_payload_from_tables(
    tables: Dict[str, LoadedTable],
    output_dir: Optional[Path] = None,
    timestamp: Optional[str] = None,
) -> tuple[Dict[str, Any], List[str], List[Path]]:
    if "business_classes" in tables:
        conversion_tables = tables_with_business_aliases(tables)
        base_payload = overlay_standard_tables_on_state(conversion_tables)
        result = convert_business_tables(conversion_tables, base_payload, output_dir=output_dir, timestamp=timestamp)
        return result.payload, result.warnings, result.generated_files
    return payload_from_tables(tables), [], []


def backup_data_dir(data_dir: Path, output_dir: Path, timestamp: str) -> Optional[Path]:
    if not data_dir.exists():
        return None
    backup_path = output_dir / "backups" / f"data_{timestamp}"
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copytree(data_dir, backup_path)
    return backup_path


def validate_required_data(state: Dict[str, Any]) -> None:
    if not state["classes"]:
        raise PipelineError("缺少班级基础信息表，无法排课")
    if not state["rooms"]:
        raise PipelineError("缺少教室表，无法排课")
    if any(cls.get("product_id") for cls in state["classes"]) and not state["product_courses"]:
        raise PipelineError("班级引用了产品，但缺少产品课程课时表")


def validate_scheduler_input(state: Dict[str, Any], time_slots: List[Dict[str, Any]]) -> None:
    try:
        scheduler.load_input_data(data_admin_server.build_scheduler_input(state, time_slots=time_slots))
    except ValueError as exc:
        raise PipelineError(str(exc)) from exc


def prepare_state_for_scheduling(
    *,
    args: argparse.Namespace,
    tables: Dict[str, LoadedTable],
    output_dir: Path,
    timestamp: str,
    warnings: List[str],
    generated_files: List[Path],
    no_time_slots_error: str,
) -> tuple[Dict[str, Any], List[Dict[str, Any]]]:
    payload, conversion_warnings, conversion_files = build_payload_from_tables(tables, output_dir, timestamp)
    warnings.extend(conversion_warnings)
    generated_files.extend(conversion_files)
    state = data_admin_server.normalize_payload(payload)
    validate_required_data(state)

    expanded_rules = expanded_rules_for_state(state)
    warnings.extend(prepare_class_windows(state, expanded_rules))
    time_slots, time_slot_warning = time_slots_for_state(
        state,
        parse_weekdays(args.exclude_weekdays),
        args.slot_set,
        getattr(args, "sunday_policy", "summer-only"),
    )
    if time_slot_warning:
        warnings.append(time_slot_warning)
    if not available_time_slots(state, time_slots):
        raise PipelineError(no_time_slots_error)
    return state, time_slots


def parse_iso_date(value: str, label: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise PipelineError(f"{label} 需要使用 YYYY-MM-DD 格式: {value}") from exc


def expanded_rules_for_state(state: Dict[str, Any]) -> List[Dict[str, Any]]:
    referenced_product_ids = {
        cls["product_id"] for cls in state["classes"] if cls.get("product_id")
    }
    catalog = shared_product_catalog(state["products"], state["product_courses"])
    return data_admin_server.scheduler_rules(
        state["product_schedule_rules"],
        referenced_product_ids,
        catalog,
    )


def latest_rule_end_date(product_id: str, expanded_rules: List[Dict[str, Any]]) -> str:
    candidates = [
        rule["end_date"]
        for rule in expanded_rules
        if rule.get("product_id") == product_id and rule.get("end_date")
    ]
    return max(candidates) if candidates else ""


def prepare_class_windows(state: Dict[str, Any], expanded_rules: List[Dict[str, Any]]) -> List[str]:
    warnings: List[str] = []
    for cls in state["classes"]:
        class_label = cls.get("name") or cls["id"]
        if not cls.get("start_date"):
            raise PipelineError(f"班级 {class_label} 缺少 start_date")
        parse_iso_date(cls["start_date"], f"班级 {class_label}/start_date")
        if not cls.get("start_period"):
            cls["start_period"] = "AM"
            warnings.append(f"班级 {class_label} 缺少 start_period，已按 AM 处理。")

        if not cls.get("end_date"):
            inferred_end_date = latest_rule_end_date(cls.get("product_id", ""), expanded_rules)
            if not inferred_end_date:
                raise PipelineError(f"班级 {class_label} 缺少 end_date，且没有可用的产品排课规则 end_date")
            cls["end_date"] = inferred_end_date
            warnings.append(f"班级 {class_label} 缺少 end_date，已按产品排课规则最晚日期 {inferred_end_date} 处理。")
        parse_iso_date(cls["end_date"], f"班级 {class_label}/end_date")
        if cls["end_date"] < cls["start_date"]:
            raise PipelineError(f"班级 {class_label} 的 end_date 早于 start_date")
        if not cls.get("end_period"):
            cls["end_period"] = "EVENING"
            warnings.append(f"班级 {class_label} 缺少 end_period，已按 EVENING 处理。")
    return warnings


def build_time_slots(
    state: Dict[str, Any],
    excluded_weekdays: Set[int],
    slot_set: str,
    sunday_policy: str = "summer-only",
) -> List[Dict[str, Any]]:
    start = min(parse_iso_date(cls["start_date"], f"班级 {cls['id']}/start_date") for cls in state["classes"])
    end = max(parse_iso_date(cls["end_date"], f"班级 {cls['id']}/end_date") for cls in state["classes"])
    return generate_time_slots(start, end, excluded_weekdays, slot_set, sunday_policy)


def time_slots_for_state(
    state: Dict[str, Any],
    excluded_weekdays: Set[int],
    slot_set: str,
    sunday_policy: str = "summer-only",
) -> tuple[List[Dict[str, Any]], Optional[str]]:
    if state.get("time_slots"):
        return list(state["time_slots"]), f"已使用上传/后台课节表 {len(state['time_slots'])} 行。"
    return build_time_slots(state, excluded_weekdays, slot_set, sunday_policy), None


def available_time_slots(
    state: Dict[str, Any],
    time_slots: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    return [
        slot for slot in time_slots
        if data_admin_server.slot_is_usable(slot)
        and not data_admin_server.slot_is_blackout(slot, state["global_blackout_dates"])
    ]


def report_table_names(tables: Dict[str, LoadedTable]) -> List[str]:
    names = list(REPORT_TABLES)
    names.extend(table for table in COMPATIBILITY_REPORT_TABLES if table in tables)
    return names


def sanitize_markdown_table_cell(value: object) -> str:
    return normalize_text(value).replace("|", "\\|").replace("\n", " ")


def report_header_lines(source: Path, backup_path: Optional[Path], error: Optional[str]) -> List[str]:
    lines = [
        "# 排课导入报告",
        "",
        f"- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 源数据: {source}",
    ]
    if backup_path:
        lines.append(f"- 数据备份: {backup_path}")
    if error:
        lines.append("- 状态: 失败")
        lines.append(f"- 错误: {error}")
    else:
        lines.append("- 状态: 成功")
    return lines


def report_data_table_lines(tables: Dict[str, LoadedTable], row_counts: Dict[str, int]) -> List[str]:
    lines = ["", "## 数据表"]
    for table in report_table_names(tables):
        loaded = tables.get(table)
        source_text = loaded.source if loaded else "未提供"
        lines.append(f"- {table}: {row_counts.get(table, 0)} 行 ({source_text})")
    return lines


def report_warning_lines(warnings: List[str]) -> List[str]:
    if not warnings:
        return []
    return ["", "## 提示", *(f"- {warning}" for warning in warnings)]


def report_missing_teacher_lines(missing_teacher_rows: Optional[List[Dict[str, str]]]) -> List[str]:
    if not missing_teacher_rows:
        return []
    lines = [
        "",
        "## 缺老师补录摘要",
        f"- 缺口数量: {len(missing_teacher_rows)}",
        "- 请下载 `missing_class_teacher_assignments_*.csv` 补齐老师后重新校验。",
        "",
        "| 班级 | 产品 | 科目 | 阶段 | 课程组 |",
        "| --- | --- | --- | --- | --- |",
    ]
    for row in missing_teacher_rows[:30]:
        lines.append(
            "| "
            + " | ".join(
                sanitize_markdown_table_cell(row.get(field, ""))
                for field in ("class_name", "product_name", "subject", "stage", "course_group")
            )
            + " |"
        )
    if len(missing_teacher_rows) > 30:
        lines.append(f"- 仅展示前 30 条，另有 {len(missing_teacher_rows) - 30} 条请查看补录 CSV。")
    return lines


def report_output_lines(
    scheduler_input_path: Optional[Path],
    schedule_csv_path: Optional[Path],
    schedule_html_path: Optional[Path],
) -> List[str]:
    if not (scheduler_input_path or schedule_csv_path or schedule_html_path):
        return []
    lines = ["", "## 输出"]
    if scheduler_input_path:
        lines.append(f"- 排课输入: {scheduler_input_path}")
    if schedule_csv_path:
        lines.append(f"- CSV 明细: {schedule_csv_path}")
    if schedule_html_path:
        lines.append(f"- HTML 甘特图: {schedule_html_path}")
    return lines


def report_generated_file_lines(generated_files: Optional[List[Path]]) -> List[str]:
    if not generated_files:
        return []
    return ["", "## 生成参考文件", *(f"- {file_path}" for file_path in generated_files)]


def write_report(
    path: Path,
    *,
    source: Path,
    tables: Dict[str, LoadedTable],
    row_counts: Dict[str, int],
    warnings: List[str],
    backup_path: Optional[Path],
    scheduler_input_path: Optional[Path],
    schedule_csv_path: Optional[Path],
    schedule_html_path: Optional[Path],
    generated_files: Optional[List[Path]] = None,
    missing_teacher_rows: Optional[List[Dict[str, str]]] = None,
    error: Optional[str] = None,
) -> None:
    lines: List[str] = []
    lines.extend(report_header_lines(source, backup_path, error))
    lines.extend(report_data_table_lines(tables, row_counts))
    lines.extend(report_warning_lines(warnings))
    lines.extend(report_missing_teacher_lines(missing_teacher_rows))
    lines.extend(report_output_lines(scheduler_input_path, schedule_csv_path, schedule_html_path))
    lines.extend(report_generated_file_lines(generated_files))
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def class_teacher_template_context(state: Optional[Dict[str, Any]]) -> Dict[str, Dict[str, str]]:
    if not state:
        return {}
    product_meta = shared_product_catalog(state.get("products", []), state.get("product_courses", []))
    context: Dict[str, Dict[str, str]] = {}
    for cls in state.get("classes", []):
        class_id = normalize_text(cls.get("id"))
        if not class_id:
            continue
        product_id = normalize_text(cls.get("product_id"))
        meta = product_meta.get(product_id, {})
        context[class_id] = {
            "class_name": normalize_text(cls.get("name")),
            "product_id": product_id,
            "product_name": normalize_text(meta.get("name")) or normalize_text(cls.get("product_name")),
        }
    return context


def parse_missing_teacher_requirements(error: str) -> List[MissingTeacherRequirement]:
    requirements: List[MissingTeacherRequirement] = []
    seen: Set[tuple[str, str, str, str, str]] = set()
    for line in error.splitlines():
        line = line.strip()
        class_id = ""
        product_id = ""
        labels_text = ""
        if line.startswith("班级 ") and "缺少课程老师安排:" in line:
            class_text, _, labels_text = line.partition("缺少课程老师安排:")
            class_text = class_text.removeprefix("班级 ").strip()
        elif line.startswith("班级 ") and " 缺少 " in line and " 的老师安排" in line:
            class_text, _, labels_text = line.partition(" 缺少 ")
            class_text = class_text.removeprefix("班级 ").strip()
            labels_text = labels_text.removesuffix(" 的老师安排")
        else:
            continue
        if " 的产品 " in class_text:
            class_id, _, product_id = class_text.partition(" 的产品 ")
            class_id = class_id.strip()
            product_id = product_id.strip()
        else:
            class_id = class_text.strip()
        for label in re.split(r"[、，,；;]+", labels_text):
            parts = [part.strip() for part in label.strip().split("/") if part.strip()]
            if len(parts) < 3:
                continue
            subject, stage = parts[0], parts[1]
            course_group = "/".join(parts[2:])
            key = (class_id, product_id, subject, stage, course_group)
            if key in seen:
                continue
            seen.add(key)
            requirements.append(
                MissingTeacherRequirement(
                    class_id=class_id,
                    product_id=product_id,
                    subject=subject,
                    stage=stage,
                    course_group=course_group,
                )
            )
    return requirements


def missing_teacher_template_rows(
    error: str,
    state: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, str]]:
    return missing_teacher_rows_for_requirements(
        parse_missing_teacher_requirements(error),
        state,
    )


def missing_teacher_rows_for_requirements(
    requirements: Sequence[MissingTeacherRequirement],
    state: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, str]]:
    context = class_teacher_template_context(state)
    product_names = {
        product_id: normalize_text(meta.get("name"))
        for product_id, meta in shared_product_catalog(
            state.get("products", []),
            state.get("product_courses", []),
        ).items()
    } if state else {}
    rows: List[Dict[str, str]] = []
    for requirement in requirements:
        class_context = context.get(requirement.class_id, {})
        class_product_id = class_context.get("product_id", "")
        product_id = requirement.product_id or class_product_id
        if requirement.product_id:
            product_name = product_names.get(product_id, "")
            if not product_name and product_id == class_product_id:
                product_name = class_context.get("product_name", "")
        else:
            product_name = class_context.get("product_name", "") or product_names.get(product_id, "")
        rows.append(
            {
                "class_id": requirement.class_id,
                "class_name": class_context.get("class_name", ""),
                "product_id": product_id,
                "product_name": product_name,
                "subject": requirement.subject,
                "stage": requirement.stage,
                "course_group": requirement.course_group,
                "class_schedule_mode": "本班实际排课",
                "actual_scheduled_class_id": requirement.class_id,
                "teacher_id": "",
                "teacher_name": "",
                "assignment_extra_time_requirement": "",
                "notes": "上传前校验自动生成，请补齐老师后重新上传",
            }
        )
    return rows


def write_missing_teacher_template(
    output_dir: Path,
    timestamp: str,
    error: str,
    state: Optional[Dict[str, Any]] = None,
) -> Optional[Path]:
    return write_missing_teacher_rows_template(
        output_dir,
        timestamp,
        missing_teacher_template_rows(error, state),
    )


def write_missing_teacher_rows_template(
    output_dir: Path,
    timestamp: str,
    rows: List[Dict[str, str]],
) -> Optional[Path]:
    if not rows:
        return None

    path = output_dir / f"missing_class_teacher_assignments_{timestamp}.csv"
    write_csv_rows(path, TEACHER_ASSIGNMENT_FIELDNAMES, rows, encoding="utf-8")
    return path


def build_pipeline_context(args: argparse.Namespace, report_prefix: str) -> PipelineRunContext:
    timestamp = args.timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    source = Path(args.source).resolve()
    data_dir = Path(args.data_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    data_admin_server.DATA_DIR = data_dir
    return PipelineRunContext(
        timestamp=timestamp,
        source=source,
        data_dir=data_dir,
        output_dir=output_dir,
        report_path=output_dir / f"{report_prefix}_{timestamp}.md",
    )


def prepare_pipeline_context(
    context: PipelineRunContext,
    args: argparse.Namespace,
    *,
    no_time_slots_error: str,
) -> None:
    context.tables = load_source_tables(context.source)
    context.row_counts = row_counts_for_tables(context.tables)
    state, time_slots = prepare_state_for_scheduling(
        args=args,
        tables=context.tables,
        output_dir=context.output_dir,
        timestamp=context.timestamp,
        warnings=context.warnings,
        generated_files=context.generated_files,
        no_time_slots_error=no_time_slots_error,
    )
    context.state = state
    context.time_slots = time_slots
    validate_scheduler_input(state, time_slots)


def merge_business_warnings(exc: Exception, warnings: List[str]) -> None:
    if isinstance(exc, BusinessDataError):
        warnings.extend(warning for warning in exc.warnings if warning not in warnings)


def missing_teacher_diagnostics_for_error(
    error: str,
    context: PipelineRunContext,
) -> MissingTeacherDiagnostics:
    requirements = parse_missing_teacher_requirements(error)
    rows = missing_teacher_rows_for_requirements(requirements, context.state)
    if rows:
        missing_teacher_path = write_missing_teacher_rows_template(
            context.output_dir,
            context.timestamp,
            rows,
        )
        if missing_teacher_path:
            context.generated_files.append(missing_teacher_path)
    return MissingTeacherDiagnostics(requirements=requirements, rows=rows)


def write_pipeline_report(
    context: PipelineRunContext,
    *,
    backup_path: Optional[Path] = None,
    scheduler_input_path: Optional[Path] = None,
    schedule_csv_path: Optional[Path] = None,
    schedule_html_path: Optional[Path] = None,
    missing_teacher_rows: Optional[List[Dict[str, str]]] = None,
    error: Optional[str] = None,
) -> None:
    write_report(
        context.report_path,
        source=context.source,
        tables=context.tables,
        row_counts=context.row_counts,
        warnings=context.warnings,
        backup_path=backup_path,
        scheduler_input_path=scheduler_input_path,
        schedule_csv_path=schedule_csv_path,
        schedule_html_path=schedule_html_path,
        generated_files=context.generated_files,
        missing_teacher_rows=missing_teacher_rows,
        error=error,
    )


def run_pipeline(args: argparse.Namespace) -> PipelineResult:
    context = build_pipeline_context(args, "import_report")
    backup_path: Optional[Path] = None
    scheduler_input_path: Optional[Path] = None
    schedule_csv_path: Optional[Path] = None
    schedule_html_path: Optional[Path] = None

    try:
        prepare_pipeline_context(
            context,
            args=args,
            no_time_slots_error="没有可用课节，请检查 02_课节表 is_usable 或 16_全局停课日期表。",
        )
        if context.state is None or context.time_slots is None:
            raise PipelineError("排课上下文未初始化")
        backup_path = backup_data_dir(context.data_dir, context.output_dir, context.timestamp)

        data_admin_server.save_state(context.state)
        export_result = data_admin_server.export_scheduler_input(context.state, time_slots=context.time_slots)
        scheduler_input_path = Path(export_result["path"])
        pending_schedule_csv_path = context.output_dir / f"schedule_{context.timestamp}.csv"
        pending_schedule_html_path = context.output_dir / f"schedule_{context.timestamp}.html"

        schedule_input = scheduler.load_input(scheduler_input_path)
        assignments = scheduler.schedule(schedule_input)
        scheduler.write_csv(assignments, pending_schedule_csv_path, schedule_input)
        scheduler.write_html(assignments, schedule_input, pending_schedule_html_path)
        schedule_csv_path = pending_schedule_csv_path
        schedule_html_path = pending_schedule_html_path

        write_pipeline_report(
            context,
            backup_path=backup_path,
            scheduler_input_path=scheduler_input_path,
            schedule_csv_path=schedule_csv_path,
            schedule_html_path=schedule_html_path,
        )
    except Exception as exc:
        merge_business_warnings(exc, context.warnings)
        diagnostics = missing_teacher_diagnostics_for_error(str(exc), context)
        write_pipeline_report(
            context,
            backup_path=backup_path,
            scheduler_input_path=scheduler_input_path,
            schedule_csv_path=schedule_csv_path,
            schedule_html_path=schedule_html_path,
            missing_teacher_rows=diagnostics.rows,
            error=str(exc),
        )
        raise

    return PipelineResult(
        scheduler_input_path=scheduler_input_path,
        schedule_csv_path=schedule_csv_path,
        schedule_html_path=schedule_html_path,
        report_path=context.report_path,
        backup_path=backup_path,
        row_counts=context.row_counts,
        warnings=context.warnings,
        generated_files=context.generated_files,
    )


def run_preflight(args: argparse.Namespace) -> PreflightResult:
    context = build_pipeline_context(args, "preflight_report")
    error = ""
    passed = False
    diagnostics = MissingTeacherDiagnostics()

    try:
        prepare_pipeline_context(
            context,
            args=args,
            no_time_slots_error="上传前校验未找到任何可用课节，请检查 02_课节表 is_usable 或 16_全局停课日期表。",
        )
        passed = True
    except Exception as exc:
        merge_business_warnings(exc, context.warnings)
        error = str(exc)

    if error:
        diagnostics = missing_teacher_diagnostics_for_error(error, context)

    write_pipeline_report(
        context,
        missing_teacher_rows=diagnostics.rows,
        error=None if passed else error,
    )
    return PreflightResult(
        passed=passed,
        report_path=context.report_path,
        row_counts=context.row_counts,
        warnings=context.warnings,
        generated_files=context.generated_files,
        missing_teacher_requirements=diagnostics.requirements,
        missing_teacher_rows=diagnostics.rows,
        error=error,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="从混合 Excel/CSV 源数据导入并生成排课结果")
    parser.add_argument("--source", required=True, help="源数据目录或单个 Excel/CSV 文件")
    parser.add_argument("--data-dir", default="data", help="写入的数据目录，默认 data")
    parser.add_argument("--output-dir", default="outputs", help="输出目录，默认 outputs")
    parser.add_argument("--timestamp", help="可选：固定输出时间戳，便于测试")
    parser.add_argument(
        "--preflight",
        action="store_true",
        help="只执行上传前校验并生成预检报告，不写入 data/，不正式排课",
    )
    parser.add_argument(
        "--exclude-weekdays",
        default="Sun",
        help="生成课节时排除的星期，用逗号分隔，默认 Sun",
    )
    parser.add_argument(
        "--slot-set",
        choices=["all", "day", "evening"],
        default="all",
        help="生成课节范围，默认 all",
    )
    parser.add_argument(
        "--sunday-policy",
        choices=["always", "summer-only"],
        default="summer-only",
        help="周日课节策略：always=全程排除周日，summer-only=仅 7-8 月排除周日，默认 summer-only",
    )
    return parser


def ensure_failure_report(report_path: Path, args: argparse.Namespace, error: str) -> None:
    if report_path.exists():
        return
    write_report(
        report_path,
        source=Path(args.source).resolve(),
        tables={},
        row_counts=empty_source_row_counts(),
        warnings=[],
        backup_path=None,
        scheduler_input_path=None,
        schedule_csv_path=None,
        schedule_html_path=None,
        generated_files=[],
        error=error,
    )


def print_generated_files(paths: Sequence[Path]) -> None:
    if not paths:
        return
    print("参考文件:")
    for path in paths:
        print(f"- {path}")


def print_preflight_result(result: PreflightResult) -> None:
    if result.passed:
        print("上传前校验通过")
    else:
        print("上传前校验未通过", file=sys.stderr)
    print(f"预检报告: {result.report_path}")
    print_generated_files(result.generated_files)
    if result.missing_teacher_rows:
        print(f"缺老师补录: {len(result.missing_teacher_rows)} 条")
    if result.error:
        print(f"错误摘要: {result.error.splitlines()[0]}", file=sys.stderr)


def run_preflight_cli(args: argparse.Namespace, report_path: Path) -> int:
    try:
        result = run_preflight(args)
    except Exception as exc:
        ensure_failure_report(report_path, args, str(exc))
        print(f"上传前校验失败: {exc}", file=sys.stderr)
        print(f"预检报告: {report_path}", file=sys.stderr)
        return 1

    print_preflight_result(result)
    return 0 if result.passed else 1


def print_pipeline_result(result: PipelineResult) -> None:
    print(f"排课输入: {result.scheduler_input_path}")
    print(f"CSV 明细: {result.schedule_csv_path}")
    print(f"HTML 甘特图: {result.schedule_html_path}")
    print(f"导入报告: {result.report_path}")
    if result.backup_path:
        print(f"数据备份: {result.backup_path}")


def run_pipeline_cli(args: argparse.Namespace, report_path: Path) -> int:
    try:
        result = run_pipeline(args)
    except Exception as exc:
        ensure_failure_report(report_path, args, str(exc))
        print(f"排课闭环失败: {exc}", file=sys.stderr)
        print(f"失败报告: {report_path}", file=sys.stderr)
        return 1

    print_pipeline_result(result)
    return 0


def main(argv: Optional[Sequence[str]] = None) -> None:
    parser = build_parser()
    args = parser.parse_args(argv)
    timestamp = args.timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    args.timestamp = timestamp
    output_dir = Path(args.output_dir).resolve()
    report_path = output_dir / f"preflight_report_{timestamp}.md" if args.preflight else output_dir / f"import_report_{timestamp}.md"
    if args.preflight:
        raise SystemExit(run_preflight_cli(args, report_path)) from None

    raise SystemExit(run_pipeline_cli(args, report_path)) from None


if __name__ == "__main__":
    main()
